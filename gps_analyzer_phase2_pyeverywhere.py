import os
import sys
import re
import glob
import math
import time
import uuid
import pickle
import chardet
import datetime
import traceback
import threading
import requests
import osmium
import folium
import numpy as np
import pandas as pd
import osmnx as ox
from shapely.geometry import Point, LineString
from geopy.distance import geodesic
import networkx as nx

from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from shapely.geometry import Point, LineString
from shapely.ops import transform
from pyproj import Transformer

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from folium import plugins
from datetime import datetime
from openpyxl.chart.series import Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import BarChart3D
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.colors import ColorChoice

from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.radar_chart import RadarChart
from openpyxl.cell.cell import MergedCell
from geopy.distance import geodesic

import traceback


""" New Tracker__ Version 5

Integrating Tracker Parameters

addressing problem with zigzag issue..
new building graph function to add junction


New reconstruct_route_geom

Changed Graph Building (reverted to old logic to build roundabouts correctly)
Key Fixes Applied:

Grid reduced to 1m (0.00001°) - industry standard
Self-loop elimination - deduplicate canonical sequence before edge creation
Bidirectional edges - add reverse edges (key=1) for non-oneway roads
Name preservation in bridging - reject paths that change road name
Topology-first transition probability - check connectivity before applying same-name bonus, verify path length for disconnected segments
Use None instead of '' for missing names to avoid false matches
changed multi week analyzer

"""
# Version 20: Fixed _calculate_bearing() method signature conflict - removed duplicate definition and updated all calls to use tuple format


class UnifiedOfflineMapDataManager:
    
    def __init__(self, cache_dir="map_cache", max_workers=5, enable_parallel=False, 
                 enable_road_context=True, session_id=None, pbf_file_path="alsace-latest.osm.pbf"):
        self.pbf_file_path = pbf_file_path
        self.session_id = session_id

        # Validate PBF file exists
        #if not os.path.exists(self.pbf_file_path):
            #raise ValueError(f"PBF file not found at {self.pbf_file_path}. Download from Geofabrik or similar.")

        self.cache_dir = r"C:\Users\DELL\OneDrive\Upwork\Working_Directory\Claude\map_cache"
        #if not os.path.exists(self.cache_dir):
            #os.makedirs(self.cache_dir)
                    
        # NEW: Master cache file path
        self.master_cache_file = os.path.join(cache_dir, "alsace_master_cache.pkl")
        self.master_data = None  # Will hold complete Alsace data
            
        self.map_data = {
            'speed_limits': {},
            'roundabouts': [],
            'stop_signs': [],
            'schools': [],
            'roads': [],
            'speed_zones_30': [],
            'traffic_lights': []            
        }
        
        # Parallel functionality (kept for compatibility, but PBF parsing is sequential)
        self.max_workers = max_workers
        self.enable_parallel = enable_parallel
        if enable_parallel:
            self.download_lock = threading.Lock()
        else:
            self.download_lock = None
        
        # Road context functionality 
        self.enable_road_context = enable_road_context
        if enable_road_context:
            self.driven_road_ids = set()
            self.route_geometry = None
            self.driven_road_ids = set()

    # ========================================================================
    # NEW: MASTER CACHE METHODS
    # ========================================================================
    
    def build_master_cache(self, force_rebuild=False):
        """Build complete Alsace map cache - run once only"""
        
        if os.path.exists(self.master_cache_file) and not force_rebuild:
            print(f"Master cache already exists at {self.master_cache_file}")
            print("Use force_rebuild=True to recreate, or call load_master_cache() to use existing")
            return True
            
        print("=" * 60)
        print("BUILDING MASTER CACHE FOR ENTIRE ALSACE REGION")
        print("This will take 5-7 minutes but only needs to be done ONCE")
        print("=" * 60)
        
        start_time = time.time()
        
        try:
            # Create master cache handler (no bbox filtering)
            master_handler = self._MasterCacheHandler(self)
            
            print("Processing entire PBF file - this may take several minutes...")
            master_handler.apply_file(self.pbf_file_path, locations=True, idx='flex_mem')
            
            print(f"PBF processing complete!")
            print(f"- Nodes processed: {master_handler.nodes_processed:,}")
            print(f"- Ways processed: {master_handler.ways_processed:,}")
            print(f"- Relations processed: {master_handler.relations_processed:,}")
            print(f"- Features extracted: {master_handler.features_extracted:,}")
            
            # Build master data structure
            master_data = {
                'speed_limits': master_handler.local_speed_limits,
                'roads': master_handler.local_roads,
                'roundabouts': master_handler.local_roundabouts,
                'stop_signs': master_handler.local_stop_signs,
                'traffic_lights': master_handler.local_traffic_lights,
                'schools': master_handler.local_schools,
                'speed_zones_30': master_handler.local_speed_zones,
                'metadata': {
                    'created': time.time(),
                    'pbf_file': self.pbf_file_path,
                    'total_features': master_handler.features_extracted,
                    'processing_time': time.time() - start_time
                }
            }
            
            # Save master cache
            print(f"Saving master cache to {self.master_cache_file}...")
            with open(self.master_cache_file, 'wb') as f:
                pickle.dump(master_data, f, protocol=pickle.HIGHEST_PROTOCOL)
            
            self.master_data = master_data
            
            elapsed = time.time() - start_time
            print("=" * 60)
            print("MASTER CACHE CREATED SUCCESSFULLY!")
            print(f"Processing time: {elapsed:.1f} seconds")
            print(f"Cache size: {os.path.getsize(self.master_cache_file) / 1024 / 1024:.1f} MB")
            print("\nExtracted features:")
            print(f"- Roads with speed limits: {len(master_data['roads'])}")
            print(f"- Speed limit points: {len(master_data['speed_limits'])}")
            print(f"- Roundabouts: {len(master_data['roundabouts'])}")
            print(f"- Schools: {len(master_data['schools'])}")
            print(f"- Stop signs: {len(master_data['stop_signs'])}")
            print(f"- Traffic lights: {len(master_data['traffic_lights'])}")
            print(f"- Speed zones: {len(master_data['speed_zones_30'])}")
            print("=" * 60)
            
            return True
            
        except Exception as e:
            print(f"Master cache creation failed: {e}")
            traceback.print_exc()
            return False
    
    def load_master_cache(self):
        """Load the master cache into memory"""
        if not os.path.exists(self.master_cache_file):
            print(f"Master cache not found at {self.master_cache_file}")
            print("Run build_master_cache() first")
            return False
            
        try:
            print(f"Loading master cache from {self.master_cache_file}...")
            start_time = time.time()
            
            with open(self.master_cache_file, 'rb') as f:
                self.master_data = pickle.load(f)
            
            load_time = time.time() - start_time
            cache_size = os.path.getsize(self.master_cache_file) / 1024 / 1024
            
            print(f"Master cache loaded in {load_time:.2f} seconds ({cache_size:.1f} MB)")
            print(f"Available features:")
            print(f"- Roads: {len(self.master_data['roads'])}")
            print(f"- Speed limits: {len(self.master_data['speed_limits'])}")
            print(f"- Roundabouts: {len(self.master_data['roundabouts'])}")
            print(f"- Schools: {len(self.master_data['schools'])}")
            print(f"- Stop signs: {len(self.master_data['stop_signs'])}")
            print(f"- Traffic lights: {len(self.master_data['traffic_lights'])}")
            print(f"- Speed zones: {len(self.master_data['speed_zones_30'])}")

          
            
            return True
            
        except Exception as e:
            print(f"Failed to load master cache: {e}")
            return False

    # ========================================================================
    # MODIFIED: MAIN PUBLIC METHODS - NOW USE MASTER CACHE
    # ========================================================================
    
    def download_area_data(self, bbox, area_name="route_area"):
        """Extract map data for bbox - now uses master cache approach"""
        
        # Step 1: Ensure master cache exists and is loaded
        if self.master_data is None:
            print("Master cache not loaded. Loading now...")
            if not self.load_master_cache():
                print("Master cache not found. Building now...")
                if not self.build_master_cache():
                    print("Failed to build master cache. Using fallback data.")
                    #self._load_fallback_data(bbox)
                    return False
        
        # Step 2: Check for existing filtered cache
        min_lat, min_lon, max_lat, max_lon = bbox
        geo_hash = f"{min_lat:.4f}_{min_lon:.4f}_{max_lat:.4f}_{max_lon:.4f}"
        cache_suffix = f"_{self.session_id}" if self.session_id else ""
        cache_file = os.path.join(self.cache_dir, f"{area_name}_{geo_hash}{cache_suffix}_data.pkl")
        
        if os.path.exists(cache_file):
            print(f"Loading cached area data from {cache_file}")
            with open(cache_file, 'rb') as f:
                self.map_data = pickle.load(f)
            return True
        
        # Step 3: Filter master cache by bbox (fast in-memory operation)
        print(f"Filtering master cache for area: {area_name}")
        start_time = time.time()
        
        #print(f" ---> [DEBUG] Filtering bbox: {bbox}")
        try:
            self._filter_master_cache_by_bbox(bbox)
            
            # Cache the filtered result
            with open(cache_file, 'wb') as f:
                pickle.dump(self.map_data, f)
            
            elapsed = time.time() - start_time
            print(f"Area data filtered and cached in {elapsed:.2f} seconds")
            print(f"Extracted for bbox: {len(self.map_data['roads'])} roads, {len(self.map_data['speed_limits'])} speed points")
            print(f" --> Filtering bbox: lat [{min_lat}, {max_lat}], lon [{min_lon}, {max_lon}]")
            return True
            
        except Exception as e:
            print(f"Error filtering master cache: {e}")
            #self._load_fallback_data(bbox)
            return False
    
    def _filter_master_cache_by_bbox(self, bbox):
        """Filter master cache data by bounding box - fast in-memory operation"""
        min_lat, min_lon, max_lat, max_lon = bbox
        
        # Reset map_data
        self.map_data = {
            'speed_limits': {},
            'roundabouts': [],
            'stop_signs': [],
            'schools': [],
            'roads': [],
            'speed_zones_30': [],
            'traffic_lights': []
        }
        
        # Filter speed limits
        for coord_key, speed_limit in self.master_data['speed_limits'].items():
            try:
                lat, lon = map(float, coord_key.split(','))
                if min_lat <= lat <= max_lat and min_lon <= lon <= max_lon:
                    self.map_data['speed_limits'][coord_key] = speed_limit
            except:
                continue
        
        # Filter roads
        for road in self.master_data['roads']:
            # Check if any geometry point is in bbox

            
            if self.master_data['roads']:
                sample_road = self.master_data['roads'][0]
                
                
            geometry = road.get('geometry', [])
            
            if any(min_lat <= p['lat'] <= max_lat and min_lon <= p['lon'] <= max_lon 
                   for p in geometry):
                self.map_data['roads'].append(road)
        print(f"Filtered roads count: {len(self.map_data['roads'])}")
        
        # Filter point features (schools, stop signs, traffic lights)
        for feature_type in ['schools', 'stop_signs', 'traffic_lights', 'roundabouts', 'speed_zones_30']:
            for feature in self.master_data[feature_type]:
                lat, lon = feature['lat'], feature['lon']
                if min_lat <= lat <= max_lat and min_lon <= lon <= max_lon:
                    self.map_data[feature_type].append(feature)
    
    def set_road_context(self, driven_road_ids, route_geometry=None):
        """Set the context of roads actually driven (unchanged)"""
        if not self.enable_road_context:
            print("Road context not enabled. Enable with enable_road_context=True")
            return
            
        self.driven_road_ids = driven_road_ids
        self.route_geometry = route_geometry
        print(f"Set road context: {len(driven_road_ids)} driven road segments")


    def get_speed_limit(self, matched_edge):
        """Extract speed limit from NetworkX edge tuple"""
        try:
            if not matched_edge or not hasattr(self, 'road_graph') or not self.road_graph:
                return 50
                
            # matched_edge is a tuple like (node_u, node_v, key)
            if isinstance(matched_edge, tuple) and len(matched_edge) >= 2:
                try:
                    # Get edge data from NetworkX graph
                    if len(matched_edge) == 2:
                        u, v = matched_edge
                        edge_data = self.road_graph.edges[u, v, 0]  # Default key=0
                    else:
                        edge_data = self.road_graph.edges[matched_edge]
                    
                    # Extract speed limit from edge data
                    if 'maxspeed' in edge_data:
                        return int(edge_data['maxspeed'])
                        
                    # Fallback to highway type
                    highway_defaults = {
                        'motorway': 130, 'trunk': 110, 'primary': 90,
                        'secondary': 90, 'tertiary': 50, 'residential': 50,
                        'living_street': 30, 'unclassified': 50
                    }
                    return highway_defaults.get(edge_data.get('highway', 'unclassified'), 50)
                    
                except KeyError:
                    print(f"DEBUG: Edge {matched_edge} not found in road_graph")
                    return 50
            else:
                print(f"DEBUG: Invalid matched_edge format: {matched_edge}")
                return 50
                    
        except Exception as e:
            print(f"DEBUG: get_speed_limit failed for {matched_edge}: {e}")
            return 50
        

    
    def get_nearby_features(self, lat, lon, feature_type, radius=200):
        """Get nearby features of specified type (unchanged)"""
        features = self.map_data.get(feature_type, [])
        nearby = []
        
        for feature in features:
            distance = geodesic((lat, lon), (feature['lat'], feature['lon'])).meters
            if distance <= radius:
                feature_copy = feature.copy()
                feature_copy['distance'] = distance
                nearby.append(feature_copy)
        
        return sorted(nearby, key=lambda x: x['distance'])

        # ========================================================================
    def build_osmid_road_feature_lookup(self):
        """
        FULLY OPTIMIZED: Build fast lookup for road-linked features
        """
        if hasattr(self, '_road_lookup_cache'):
            return self._road_lookup_cache
        
        print("Building road feature lookup (one-time operation)...")
        
        road_way_ids = set()
        road_osmids = set() 
        road_node_refs = set()
        
        roads_data = self.map_data.get('roads', [])
        
        # OPTIMIZED: Batch operations and avoid repeated type checks
        for road in roads_data:
            # Fast ID extraction
            road_id = road.get('id')
            if road_id:
                road_way_ids.add(road_id)
            
            road_osmid = road.get('osmid')
            if road_osmid:
                road_osmids.add(road_osmid)
            
            # OPTIMIZED node extraction - avoid string operations
            # OPTIMIZED node extraction - avoid string operations
            # SIMPLIFIED node extraction - now we have reliable list data
            node_refs = road.get('node_refs')
            if node_refs and isinstance(node_refs, (list, tuple)):
                road_node_refs.update(node_refs)
                        
            # Check tags only if node_refs wasn't found
            elif 'tags' in road:
                tags = road['tags']
                nodes_tag = tags.get('nodes')
                if nodes_tag and isinstance(nodes_tag, str):
                    try:
                        nodes = [int(x) for x in nodes_tag.split(',') if x.strip().isdigit()]
                        road_node_refs.update(nodes)
                    except:
                        pass
        
        self._road_lookup_cache = {
            'road_way_ids': road_way_ids,
            'road_osmids': road_osmids, 
            'road_node_refs': road_node_refs,
            'stats': {
                'total_roads': len(roads_data),
                'way_ids_count': len(road_way_ids),
                'osmids_count': len(road_osmids),
                'node_refs_count': len(road_node_refs)
            }
        }
        
        print(f"Road lookup built: {len(road_way_ids)} way IDs, {len(road_osmids)} OSM IDs, {len(road_node_refs)} node refs")
        return self._road_lookup_cache
    
    def get_osmid_road_linked_features(self, feature_list, feature_type="feature"):
        """
        CENTRAL FILTERING FUNCTION: Filter any feature list to only road-linked items
        
        This works for stop signs, roundabouts, schools, or any feature with id/osmid
        
        Args:
            feature_list: List of features (stop signs, roundabouts, etc.)
            feature_type: String for logging ("stops", "roundabouts", etc.)
        
        Returns:
            dict with filtered features and relationship details
        """
        # Ensure lookup is built
        if not hasattr(self, '_road_lookup_cache'):
            self.build_osmid_road_feature_lookup()
        
        lookup = self._road_lookup_cache
        road_way_ids = lookup['road_way_ids']
        road_osmids = lookup['road_osmids'] 
        road_node_refs = lookup['road_node_refs']
        
        # Categorize features by relationship type
        relationships = {
            'same_way_id': [],      # Feature ID matches road way ID
            'same_osmid': [],       # Feature OSM ID matches road OSM ID  
            'node_in_road': [],     # Feature node ID is part of road geometry
            'no_relationship': []   # No direct ID relationship found
        }
        
        # Fast O(1) lookup for each feature
        for feature in feature_list:
            feature_id = feature.get('id')
            feature_osmid = feature.get('osmid')
            
            relationship_found = False
            
            # Check 1: Feature ID matches road way ID
            if feature_id is not None and feature_id in road_way_ids:
                relationships['same_way_id'].append(feature)
                relationship_found = True
                
            # Check 2: Feature OSM ID matches road OSM ID
            elif feature_osmid is not None and feature_osmid in road_osmids:
                relationships['same_osmid'].append(feature)  
                relationship_found = True
                
            # Check 3: Feature node ID is part of road way
            elif feature_osmid is not None and feature_osmid in road_node_refs:
                relationships['node_in_road'].append(feature)
                relationship_found = True
                
            # No relationship found
            if not relationship_found:
                relationships['no_relationship'].append(feature)
        
        # Combine all road-linked features
        road_linked_features = (
            relationships['same_way_id'] + 
            relationships['same_osmid'] + 
            relationships['node_in_road']
        )
        
        # Statistics
        total_features = len(feature_list)
        linked_count = len(road_linked_features)
        filter_percentage = (linked_count / total_features * 100) if total_features > 0 else 0
        
        print(f"Road-linked {feature_type} filtering:")
        print(f"  Total {feature_type}: {total_features}")
        print(f"  Road-linked: {linked_count} ({filter_percentage:.1f}%)")
        print(f"    - Same way ID: {len(relationships['same_way_id'])}")
        print(f"    - Same OSM ID: {len(relationships['same_osmid'])}")  
        print(f"    - Node in road: {len(relationships['node_in_road'])}")
        print(f"    - No relationship: {len(relationships['no_relationship'])}")
        
        return {
            'road_linked_features': road_linked_features,
            'relationships': relationships,
            'statistics': {
                'total_count': total_features,
                'linked_count': linked_count,
                'filter_percentage': filter_percentage
            }
        }
    
    def get_driven_road_linked_features(self, feature_list, feature_type):
        if not self.driven_road_ids:
            result = self.get_osmid_road_linked_features(feature_list, feature_type)
            return result['road_linked_features']
            
        result = self.get_osmid_road_linked_features(feature_list, feature_type)
        road_linked_features = result['road_linked_features']
        
        # Get all node IDs from driven road segments
        driven_node_ids = set()
        if hasattr(self, 'road_graph') and self.road_graph:
            for edge_tuple in self.driven_road_ids:
                if len(edge_tuple) >= 2:  # (u, v, key)
                    u, v = edge_tuple[:2]
                    driven_node_ids.add(u)
                    driven_node_ids.add(v)
        
        print(f"DEBUG: Extracted {len(driven_node_ids)} node IDs from {len(self.driven_road_ids)} driven edges")
        print(f"DEBUG: Sample node IDs: {list(driven_node_ids)[:5]}")

        # NEW: Debug sample stop sign OSM IDs
        sample_features = road_linked_features[:3]
        print(f"DEBUG: Sample {feature_type} OSM IDs: {[f.get('osmid') for f in sample_features]}")
        
        # Check for overlap
        feature_osmids = {f.get('osmid') for f in road_linked_features if f.get('osmid')}
        overlap = driven_node_ids.intersection(feature_osmids)
        print(f"DEBUG: {len(overlap)} {feature_type} OSM IDs overlap with driven node IDs")
        print(f"DEBUG: Sample overlapping IDs: {sorted(list(overlap))[:3]}")
        
        # Filter stop signs by node IDs (since they're linked via nodes)
        driven_features = []
        for feature in road_linked_features:
            feature_osmid = feature.get('osmid')  # Stop signs are nodes
            
            if feature_osmid in driven_node_ids:
                driven_features.append(feature)
        
        print(f"Filtered {len(road_linked_features)} road-linked {feature_type} to {len(driven_features)} on driven roads")
        return driven_features
        
            
    def set_driven_road_ids(self, road_ids):
        """Called after map matching completes"""
        self.driven_road_ids = road_ids            

    # ========================================================================
    # NEW: MASTER CACHE PBF HANDLER - NO BBOX FILTERING
    # ========================================================================
    
    class _MasterCacheHandler(osmium.SimpleHandler):
        """PBF handler for master cache - processes ENTIRE file without bbox filtering"""
        
        def __init__(self, parent_manager):
            super().__init__()
            self.parent = parent_manager
            
            # Storage for ALL collected data
            self.local_roads = []
            self.local_speed_limits = {}
            self.local_roundabouts = []
            self.local_stop_signs = []
            self.local_traffic_lights = []
            self.local_schools = []
            self.local_speed_zones = []
            # Store ALL nodes and ways (no bbox filtering)
            self.nodes = {}
            self.ways = {}
            
            # Progress tracking
            self.nodes_processed = 0
            self.ways_processed = 0
            self.relations_processed = 0
            self.features_extracted = 0
        
        
        def node(self, n):
            """Enhanced node processing with missing Overpass features"""
            self.nodes_processed += 1
            
            if self.nodes_processed % 200000 == 0:
                print(f"Processed {self.nodes_processed:,} nodes...")
            
            self.nodes[n.id] = (n.location.lat, n.location.lon)
            
            tags = {tag.k: tag.v for tag in n.tags}
            lat, lon = n.location.lat, n.location.lon
            
            # SCHOOLS - Enhanced coverage
            is_traffic_school = False
            school_type = None
            
            if (tags.get('amenity') == 'school' and  
                tags.get('school:type', 'primary') in ['primary', 'secondary'] and  
                not tags.get('disused') and  
                not tags.get('access') == 'private'):
                
                is_traffic_school = True
                school_type = 'school'
                
            elif tags.get('highway') == 'school_crossing' or tags.get('crossing') == 'school':
                is_traffic_school = True
                school_type = 'school_crossing'
            
            # Only create school_data if it's traffic-relevant
            if is_traffic_school:
                school_data = {
                    'lat': lat,
                    'lon': lon,
                    'id': n.id,
                    'name': tags.get('name', school_type.title()),
                    'school_type': school_type
                }
                
                if self.parent.enable_road_context:
                    school_data['osmid'] = n.id
                    school_data['tags'] = tags
                
                self.local_schools.append(school_data)
                self.features_extracted += 1
        
            # STOP SIGNS - Enhanced coverage
            
            # STOP SIGNS - Only traffic-relevant ones
            is_relevant_stop = False
            
            if tags.get('highway') == 'stop':
                # Direct highway stop - usually legitimate
                is_relevant_stop = True
            elif tags.get('traffic_sign') == 'FR:B1':
                # Official French stop sign - check context
                landuse = tags.get('landuse', '')
                amenity = tags.get('amenity', '')
                
                # Exclude private/commercial contexts
                if not any(x in landuse.lower() + amenity.lower() for x in 
                          ['commercial', 'industrial', 'retail', 'private']):
                    is_relevant_stop = True
            
            # Exclude traffic calming stops (usually residential)
            # tags.get('traffic_calming') == 'stop_sign' -> EXCLUDE
            
            if is_relevant_stop:
                self.local_stop_signs.append({
                    'lat': lat,
                    'lon': lon,
                    'id': n.id,
                    'osmid': n.id,  # Add this line                
                    'type': 'stop',
                    'tags': tags  # Keep for debugging
                })
                self.features_extracted += 1
                
                # TRAFFIC LIGHTS - Enhanced coverage
                if (tags.get('highway') == 'traffic_signals' or
                    tags.get('traffic_signals') in ['yes', 'signal', 'traffic_lights'] or
                    tags.get('crossing') == 'traffic_signals' or  # NEW: Pedestrian crossings
                    tags.get('amenity') == 'traffic_light' or
                    tags.get('traffic_signals:direction') in ['forward', 'backward', 'both'] or  # NEW: Directional
                    tags.get('traffic_signals:sound') == 'yes' or  # NEW: Audio signals
                    'traffic_signals' in tags):
                    
                    self.local_traffic_lights.append({
                        'lat': lat,
                        'lon': lon,
                        'id': n.id,
                        'type': 'traffic_signals',
                        'osmid': n.id,
                        'tags': tags                        
                    })
                    self.features_extracted += 1
                
            # ROUNDABOUTS & TRAFFIC CALMING - Enhanced coverage
            if (tags.get('highway') == 'mini_roundabout' or
                tags.get('traffic_calming') == 'island'):  # NEW: Traffic calming islands
                
                roundabout_type = 'mini_roundabout' if tags.get('highway') == 'mini_roundabout' else 'traffic_island'
                
                roundabout_data = {
                    'lat': lat,
                    'lon': lon,
                    'id': f"node_{n.id}",
                    'osmid': n.id,
                    'type': 'node',
                    'subtype': roundabout_type,  # NEW: Distinguish types
                    'tags': tags
                }
                
                if self.parent.enable_road_context:
                    roundabout_data['geometry'] = [{'lat': lat, 'lon': lon}]
                
                self.local_roundabouts.append(roundabout_data)
                self.features_extracted += 1
            
        
        def way(self, w):
            """Enhanced way processing with missing Overpass features"""
            self.ways_processed += 1
            
            if self.ways_processed % 50000 == 0:
                print(f"Processed {self.ways_processed:,} ways, extracted {self.features_extracted} features...")
            
            tags = {tag.k: tag.v for tag in w.tags}
            
            # Build geometry
            geometry = []
            available_nodes = 0
            node_refs = []
            
            for node_ref in w.nodes:
                node_refs.append(node_ref.ref) 
                if node_ref.ref in self.nodes:
                    lat, lon = self.nodes[node_ref.ref]
                    geometry.append({'lat': lat, 'lon': lon})
                    available_nodes += 1
                
            if len(geometry) < 2:
                return
            
            self.ways[w.id] = geometry
            
            # ROADS - Unchanged (already comprehensive)
            if tags.get('highway') in ['motorway', 'trunk', 'primary', 'secondary', 
                                     'tertiary', 'residential', 'living_street', 
                                     'unclassified', 'service']:
                speed_limit = self.parent.extract_speed_limit(tags)
                
                self.local_roads.append({
                    'id': w.id,
                    'osmid': w.id, 
                    'tags': tags,
                    'speed_limit': speed_limit,
                    'geometry': geometry,
                    'node_refs': [node_ref.ref for node_ref in w.nodes]  # Add this lin
                })
                
                for point in geometry:
                    coord_key = f"{point['lat']:.4f},{point['lon']:.4f}"
                    self.local_speed_limits[coord_key] = speed_limit
                
                self.features_extracted += 1
            
            # ROUNDABOUTS - Enhanced coverage
            if (tags.get('junction') in ['roundabout', 'circular'] or
                tags.get('highway') == 'mini_roundabout' or
                tags.get('traffic_calming') == 'island'):  # NEW: Added traffic islands
                
                center_lat = sum(p['lat'] for p in geometry) / len(geometry)
                center_lon = sum(p['lon'] for p in geometry) / len(geometry)
                
                # Determine type
                roundabout_type = 'way'
                if tags.get('highway') == 'mini_roundabout':
                    roundabout_type = 'mini_roundabout'
                elif tags.get('traffic_calming') == 'island':
                    roundabout_type = 'traffic_island'
                
                roundabout_data = {
                    'lat': center_lat,
                    'lon': center_lon,
                    'id': w.id,
                    'osmid': w.id,
                    'type': roundabout_type,  # Enhanced: More specific typing
                    'tags': tags
                }
                
                if self.parent.enable_road_context:
                    roundabout_data['geometry'] = geometry
                
                self.local_roundabouts.append(roundabout_data)
                self.features_extracted += 1
            
            # TRAFFIC SIGNALS - Enhanced coverage for way-based signals
            if (tags.get('highway') == 'traffic_signals' or
                tags.get('traffic_signals') == 'yes'):
                
                center_lat = sum(p['lat'] for p in geometry) / len(geometry)
                center_lon = sum(p['lon'] for p in geometry) / len(geometry)
                
                self.local_traffic_lights.append({
                    'lat': center_lat,
                    'lon': center_lon,
                    'id': w.id,
                    'type': 'traffic_signals'
                })
                self.features_extracted += 1
            
            # SPEED ZONES - Unchanged (already comprehensive)
            if (tags.get('maxspeed') in ['30', '20'] or 
                tags.get('zone:maxspeed') in ['30', '20'] or
                tags.get('highway') == 'living_street'):
                
                speed_limit = tags.get('maxspeed') or tags.get('zone:maxspeed') or 'living_street'
                center_lat = sum(p['lat'] for p in geometry) / len(geometry)
                center_lon = sum(p['lon'] for p in geometry) / len(geometry)
                
                self.local_speed_zones.append({
                    'id': w.id,
                    'type': 'way',
                    'speed_limit': speed_limit,
                    'tags': tags,
                    'geometry': geometry,
                    'lat': center_lat,
                    'lon': center_lon
                })
                self.features_extracted += 1
            
            # SCHOOLS - Enhanced coverage
            if (tags.get('amenity') in ['school', 'kindergarten', 'college', 'university'] or
                tags.get('landuse') == 'education' or
                tags.get('building') == 'school' or
                tags.get('leisure') in ['schoolyard'] or
                # NEW: Enhanced playground detection
                (tags.get('leisure') == 'playground' and 
                 tags.get('operator', '').lower().find('school') >= 0)):
                
                center_lat = sum(p['lat'] for p in geometry) / len(geometry)
                center_lon = sum(p['lon'] for p in geometry) / len(geometry)
                
                school_type = 'school'
                if tags.get('amenity'):
                    school_type = tags.get('amenity')
                elif tags.get('landuse') == 'education':
                    school_type = 'education_area'
                elif tags.get('leisure') == 'schoolyard':
                    school_type = 'schoolyard'
                elif tags.get('leisure') == 'playground':
                    school_type = 'school_playground'  # NEW: Distinguish school playgrounds
                
                school_data = {
                    'lat': center_lat,
                    'lon': center_lon,
                    'id': w.id,
                    'name': tags.get('name', school_type.title()),
                    'school_type': school_type
                }
                
                if self.parent.enable_road_context:
                    school_data['osmid'] = w.id
                    school_data['tags'] = tags
                
                self.local_schools.append(school_data)
                self.features_extracted += 1        
            
        def relation(self, r):
            """Process ALL relations - no bbox filtering"""
            self.relations_processed += 1
            
            if self.relations_processed % 5000 == 0:
                print(f"Processed {self.relations_processed:,} relations...")
            
            tags = {tag.k: tag.v for tag in r.tags}
            
            # Complex roundabouts - Extract ALL
            if tags.get('junction') in ['roundabout', 'circular']:
                # Get member ways with available geometry
                member_ways = []
                all_geometry = []
                
                for member in r.members:
                    if member.type == 'w' and member.role in ['', 'outer']:
                        if member.ref in self.ways:
                            way_geometry = self.ways[member.ref]
                            member_ways.append(member.ref)
                            all_geometry.extend(way_geometry)
                
                if all_geometry:
                    # Calculate center
                    center_lat = sum(p['lat'] for p in all_geometry) / len(all_geometry)
                    center_lon = sum(p['lon'] for p in all_geometry) / len(all_geometry)
                    
                    roundabout_data = {
                        'lat': center_lat,
                        'lon': center_lon,
                        'id': f"rel_{r.id}",
                        'osmid': r.id,
                        'type': 'relation',
                        'member_ways': member_ways,
                        'tags': tags
                    }
                
                if self.parent.enable_road_context:
                    roundabout_data['geometry'] = all_geometry
                
                self.local_roundabouts.append(roundabout_data)
                self.features_extracted += 1
        
            # Speed zones (relation-based) - Extract ALL
            if (tags.get('type') == 'multipolygon' and 
                tags.get('zone:maxspeed') in ['30', '20']):
                
                # Collect geometry from member ways
                all_geometry = []
                for member in r.members:
                    if member.type == 'w' and member.role in ['', 'outer']:
                        if member.ref in self.ways:
                            all_geometry.extend(self.ways[member.ref])
                
                if all_geometry:
                    # Calculate center
                    center_lat = sum(p['lat'] for p in all_geometry) / len(all_geometry)
                    center_lon = sum(p['lon'] for p in all_geometry) / len(all_geometry)
                    
                    self.local_speed_zones.append({
                        'id': r.id,
                        'type': 'relation',
                        'speed_limit': tags.get('zone:maxspeed'),
                        'tags': tags,
                        'geometry': all_geometry,
                        'lat': center_lat,
                        'lon': center_lon
                    })
                    self.features_extracted += 1
    
        # ========================================================================
        # UTILITY METHODS - UNCHANGED FROM ORIGINAL
        # ========================================================================
        
       
        def extract_speed_limit(self, tags):
            """
            Extracts speed limit from OSM tags specifically for France.
            It handles standard French implicit limits and explicit numeric values.
            """
            
            # First check for explicit maxspeed tags
            if 'maxspeed' in tags:
                maxspeed = str(tags['maxspeed']).strip().upper()
            
                # --- PRIORITY 1: Handle explicit numeric speed limits FIRST ---
                try:
                    match = re.search(r'(\d+)', maxspeed)
                    if match:
                        speed = int(match.group(1))
                        # Validate for a reasonable range of speeds in France.
                        if 10 <= speed <= 131:
                            return speed
                except ValueError:
                    pass  # Continue to French implicit tags if parsing fails.
            
                # --- PRIORITY 2: Handle implicit French speed limits ---
                if maxspeed == 'FR:URBAN':
                    return 50  # Default in built-up areas.
            
                if maxspeed == 'FR:RURAL':
                    return 80
            
                if maxspeed == 'FR:TRUNK':
                    return 110
            
                if maxspeed == 'FR:MOTORWAY':
                    return 130
            
            # --- PRIORITY 3: Default speed limits based on road type ---
            highway_type = tags.get('highway', '')
            speed_defaults = {
                'motorway': 130,
                'trunk': 110, 
                'primary': 90,
                'secondary': 90,
                'tertiary': 90,
                'residential': 50,
                'living_street': 30,
                'unclassified': 50,
                'service': 30
            }
            
            return speed_defaults.get(highway_type, 82)
    
        def _load_fallback_data(self, bbox):
            """Load fallback data when extraction fails (unchanged)"""
            print("Loading fallback map data...")
            self._generate_fallback_speed_limits(bbox)
            print(f"Loaded fallback data with {len(self.map_data['speed_limits'])} speed limit points")
    
        def _generate_fallback_speed_limits(self, bbox):
            """Generate fallback speed limit grid when extraction fails (unchanged)"""
            min_lat, min_lon, max_lat, max_lon = bbox
            
            # Create a grid of speed limits (simplified approach)
            lat_steps = np.linspace(min_lat, max_lat, 20)
            lon_steps = np.linspace(min_lon, max_lon, 20)
            
            for lat in lat_steps:
                for lon in lon_steps:
                    coord_key = f"{lat:.4f},{lon:.4f}"
                    # Assume 50 km/h for most areas, 30 km/h for some residential
                    self.map_data['speed_limits'][coord_key] = 50


# ========================================================================
# UPDATED TEST CODE - NOW TESTS MASTER CACHE FUNCTIONALITY
# ========================================================================

# ======================================================================
# **********************************************************************
# SECTION old bookmark
# ======================================================================


class ParallelSpeedLimitCalculator:
    """Calculate speed limits for GPS points using graph edges or coordinate cache"""
    
    def __init__(self, map_manager, road_graph=None, max_workers=4):
        self.map_manager = map_manager
        self.road_graph = road_graph  # Use graph for edge-based lookups
        self.max_workers = max_workers
        self.grid_size = 0.002  # ~200m grid cells
        self.spatial_grid = {}
        self.parsed_coords = {}
        
        # Build coordinate cache for fallback lookups
        self._build_coordinate_cache()
    
    def _build_coordinate_cache(self):
        """Build coordinate cache with spatial index for faster lookups"""
        print("Building coordinate cache for speed limit lookups...")
        
        speed_limits_data = self.map_manager.map_data.get('speed_limits', {})
        total_coords = len(speed_limits_data)
        
        if total_coords == 0:
            print("  ⚠️ No speed limits in coordinate cache, will use graph edges only")
            return
        
        coords_for_index = []
        processed = 0
        
        for cached_coord, speed_limit in speed_limits_data.items():
            processed += 1
            if processed % 50000 == 0:
                print(f"  Processing coordinate {processed}/{total_coords}")
            
            try:
                cached_lat, cached_lon = map(float, cached_coord.split(','))
                self.parsed_coords[cached_coord] = (cached_lat, cached_lon, speed_limit)
                coords_for_index.append((cached_lat, cached_lon, speed_limit))
            except:
                continue
        
        print(f"  Parsed {len(self.parsed_coords)} speed limit coordinates")
        
        # Build spatial grid index
        for cached_lat, cached_lon, speed_limit in coords_for_index:
            grid_lat = round(cached_lat / self.grid_size) * self.grid_size
            grid_lon = round(cached_lon / self.grid_size) * self.grid_size
            grid_key = f"{grid_lat:.3f},{grid_lon:.3f}"
            
            if grid_key not in self.spatial_grid:
                self.spatial_grid[grid_key] = []
            self.spatial_grid[grid_key].append((cached_lat, cached_lon, speed_limit))
        
        print(f"  Built spatial grid with {len(self.spatial_grid)} cells")
    
    def calculate_speed_limits_for_df(self, df):
        """
        Calculate speed limits for all GPS points in DataFrame.
        
        Priority:
        1. Use matched_edge's maxspeed from graph (most accurate)
        2. Fall back to coordinate cache lookup
        3. Fall back to road type defaults
        """
        print(f"Calculating speed limits for {len(df)} GPS points...")
        
        speed_limits = []
        method_counts = {'graph': 0, 'cache': 0, 'default': 0}
        
        for idx, row in df.iterrows():
            speed_limit = None
            
            # METHOD 1: Use matched edge from graph (BEST)
            if self.road_graph is not None and 'matched_edge' in df.columns:
                edge = row.get('matched_edge')
                if edge is not None and isinstance(edge, tuple) and len(edge) == 3:
                    try:
                        edge_data = self.road_graph.edges.get(edge, {})
                        maxspeed = edge_data.get('maxspeed')
                        if maxspeed is not None and maxspeed > 0:
                            speed_limit = maxspeed
                            method_counts['graph'] += 1
                    except:
                        pass
            
            # METHOD 2: Use coordinate cache lookup
            if speed_limit is None:
                if 'road_matched_lat' in df.columns and pd.notna(row.get('road_matched_lat')):
                    lat, lon = row['road_matched_lat'], row['road_matched_lon']
                else:
                    lat, lon = row['lat'], row['lon']
                
                speed_limit = self._get_speed_limit_from_cache(lat, lon)
                if speed_limit is not None:
                    method_counts['cache'] += 1
            
            # METHOD 3: Default based on road type or fallback
            if speed_limit is None:
                speed_limit = self._get_default_speed_limit(row)
                method_counts['default'] += 1
            
            speed_limits.append(speed_limit)
        
        print(f"  Speed limit sources: graph={method_counts['graph']}, cache={method_counts['cache']}, default={method_counts['default']}")
        
        return speed_limits
    
        
    def _get_speed_limit_from_cache(self, lat, lon):
        """Look up speed limit from coordinate cache using spatial index"""
        
        if not self.spatial_grid:
            return None
        
        # Quick exact match check
        coord_key = f"{lat:.4f},{lon:.4f}"
        speed_limits_data = self.map_manager.map_data.get('speed_limits', {})
        if coord_key in speed_limits_data:
            return speed_limits_data[coord_key]
        
        # Search spatial grid
        grid_lat = round(lat / self.grid_size) * self.grid_size
        grid_lon = round(lon / self.grid_size) * self.grid_size
        
        # Check current and surrounding cells
        candidates = []
        for lat_offset in [-self.grid_size, 0, self.grid_size]:
            for lon_offset in [-self.grid_size, 0, self.grid_size]:
                check_grid_key = f"{grid_lat + lat_offset:.3f},{grid_lon + lon_offset:.3f}"
                if check_grid_key in self.spatial_grid:
                    candidates.extend(self.spatial_grid[check_grid_key])
        
        if not candidates:
            return None
        
        # Find closest candidate
        min_distance = float('inf')
        closest_limit = None
        
        for cached_lat, cached_lon, speed_limit in candidates:
            # Fast approximate distance
            lat_diff = lat - cached_lat
            lon_diff = lon - cached_lon
            approx_dist_sq = lat_diff * lat_diff + lon_diff * lon_diff
            
            # Very close match - return immediately
            if approx_dist_sq < 0.000001:  # ~10m
                return speed_limit
            
            # Reasonable candidate - calculate exact distance
            if approx_dist_sq < 0.0001:  # ~100m threshold
                distance = geodesic((lat, lon), (cached_lat, cached_lon)).meters
                if distance < min_distance and distance < 50:  # Max 50m
                    min_distance = distance
                    closest_limit = speed_limit
        
        return closest_limit
    
    
    def _get_default_speed_limit(self, row):
        """Get default speed limit based on road type or use fallback"""
        
        # If we have matched edge, try to get highway type
        if self.road_graph is not None and 'matched_edge' in row.index:
            edge = row.get('matched_edge')
            if edge is not None and isinstance(edge, tuple) and len(edge) == 3:
                try:
                    edge_data = self.road_graph.edges.get(edge, {})
                    highway = edge_data.get('highway', '')
                    
                    # Default speeds by road type (France)
                    road_type_defaults = {
                        'motorway': 130,
                        'motorway_link': 90,
                        'trunk': 110,
                        'trunk_link': 70,
                        'primary': 80,
                        'primary_link': 50,
                        'secondary': 90,
                        'secondary_link': 50,
                        'tertiary': 70,
                        'tertiary_link': 50,
                        'unclassified': 50,
                        'residential': 50,
                        'living_street': 20,
                        'service': 30
                    }
                    
                    if highway in road_type_defaults:
                        return road_type_defaults[highway]
                except:
                    pass
        
        # Ultimate fallback
        return 50

    

    
 
# ============================================================================
# 2. OPTIMIZED DATA PROCESSOR
# ============================================================================
# ============================================================================
# FAST REPORT GENERATOR
# ============================================================================


class UnifiedGPSDataProcessor:
    """
    Unified GPS Data Processor that combines functionality of:
    - RoadContextGPSDataProcessor (single trip processing with road matching)
    - MultiWeekGPSProcessor (multi-week data handling)
    
    Maintains ALL original functionality and method names while eliminating code redundancy.
    """
    
    def __init__(self, map_manager = None):
        # Single trip properties (from RoadContextGPSDataProcessor)
        self.map_manager = map_manager
        self.cache_dir = getattr(map_manager, 'cache_dir', 'temp_cache') if map_manager else 'temp_cache'
        self.df = None
        self.processed_df = None
        self.road_graph = None
        self.driven_edges = set()
        self.route_geometry = None

        # Optimization: Spatial indexing and caching for map matching
        self._edge_spatial_index = None  # Will store STRtree for fast spatial queries
        self._edge_index_map = {}  # Maps STRtree indices back to edge keys
        self._candidate_cache = {}  # Cache for identical GPS coordinates

        # Map Matching Configuration (industry-standard parameters)
        # These can be tuned based on GPS quality and road network characteristics
        # FIX: Increase connectivity bonus to 15x for urban areas (industry standard)
        self.connectivity_bonus = 3.0  # Connected edges get 3x probability boost
        self.disconnected_penalty = 0.05  # Disconnected edges get 95% reduction
        self.enable_gap_bridging = True  # Fill gaps with shortest path (recommended)

        self.distance_priority_thresholds = {
            'very_close': 3.0,      # meters - almost certainly correct
            'close': 7.0,           # meters - very likely correct  
            'plausible': 15.0       # meters - plausible match
        }
        self.distance_priority_overrides = {
            'very_close': 0.95,      # override transition prob if < 0.1
            'close': 0.85,           # override transition prob if < 0.1
            'plausible': 0.5        # override transition prob if < 0.05
        }

        # Multi-week properties (from MultiWeekGPSProcessor)
        self.weekly_data = {}  # Store data by week
        self.combined_data = None
        self.weekly_results = {}
        self.is_multi_week_mode = False
    
    # ========================================================================
    # SINGLE TRIP METHODS - Original RoadContextGPSDataProcessor functionality
    # ========================================================================
    def test_feature_filtering_export(self, output_filename="filtered_features_test.csv"):
        """
        Test function to extract and export filtered stop signs and roundabouts
        This helps verify which features are being filtered and why
        """

        
        print("Testing feature filtering and exporting to CSV...")
        
        # Step 1: Get driven road IDs from map_manager
        driven_road_ids = set()
        
        if hasattr(self.map_manager, 'driven_road_ids') and self.map_manager.driven_road_ids:
            driven_road_ids = self.map_manager.driven_road_ids
            print(f"Found driven_road_ids via map_manager.driven_road_ids: {len(driven_road_ids)}")
        elif hasattr(self.map_manager, 'get_driven_road_ids'):
            driven_road_ids = self.map_manager.get_driven_road_ids()
            print(f"Found driven_road_ids via map_manager.get_driven_road_ids(): {len(driven_road_ids)}")
        else:
            print("ERROR: No driven road IDs found in map_manager")
            return
        
        if not driven_road_ids:
            print("WARNING: driven_road_ids is empty - no filtering possible")
            return
        
        # FIXED: Extract individual node IDs from edge tuples
        driven_node_ids = set()
        for edge_tuple in driven_road_ids:
            if isinstance(edge_tuple, tuple) and len(edge_tuple) >= 2:
                u, v = edge_tuple[:2]  # Extract node IDs from (u, v, key) tuple
                driven_node_ids.add(u)
                driven_node_ids.add(v)
            else:
                driven_node_ids.add(edge_tuple)  # In case it's already a node ID
        
        print(f"Extracted {len(driven_node_ids)} node IDs from {len(driven_road_ids)} driven edges")
        print(f"Sample driven node IDs: {list(driven_node_ids)[:10]}")
        
        # Step 2: Get all features
        all_stop_signs = self.map_manager.map_data.get('stop_signs', [])
        all_roundabouts = self.map_manager.map_data.get('roundabouts', [])
        
        print(f"Total features: {len(all_stop_signs)} stop signs, {len(all_roundabouts)} roundabouts")
        
        # Step 3: Filter and categorize features
        results = []
        
        # Process stop signs
        for i, stop in enumerate(all_stop_signs):
            stop_id = stop.get('id')
            stop_osmid = stop.get('osmid')
            
            # FIXED: Check if stop sign node ID is in driven node IDs
            is_filtered = stop_osmid in driven_node_ids if stop_osmid else False
            
            results.append({
                'feature_type': 'stop_sign',
                'feature_index': i,
                'id': stop_id,
                'osmid': stop_osmid,
                'lat': stop.get('lat'),
                'lon': stop.get('lon'),
                'is_filtered': is_filtered,
                'filter_reason': 'osmid_in_driven_nodes' if is_filtered else 'osmid_not_in_driven_nodes',
                'name': stop.get('name', 'Unknown')
            })
        
        # Process roundabouts
        for i, roundabout in enumerate(all_roundabouts):
            rb_id = roundabout.get('id')
            rb_osmid = roundabout.get('osmid')
            
            # FIXED: Check if roundabout node ID is in driven node IDs
            is_filtered = rb_osmid in driven_node_ids if rb_osmid else False
            
            results.append({
                'feature_type': 'roundabout',
                'feature_index': i,
                'id': rb_id,
                'osmid': rb_osmid,
                'lat': roundabout.get('lat'),
                'lon': roundabout.get('lon'),
                'is_filtered': is_filtered,
                'filter_reason': 'osmid_in_driven_nodes' if is_filtered else 'osmid_not_in_driven_nodes',
                'name': roundabout.get('name', 'Unknown')
            })
        
        
        # Step 4: Create DataFrame and export
        df = pd.DataFrame(results)
        
        # Summary statistics
        total_features = len(results)
        filtered_features = len(df[df['is_filtered'] == True])
        stop_signs_filtered = len(df[(df['feature_type'] == 'stop_sign') & (df['is_filtered'] == True)])
        roundabouts_filtered = len(df[(df['feature_type'] == 'roundabout') & (df['is_filtered'] == True)])
        
        print(f"Filtering Results:")
        print(f"  Total features: {total_features}")
        print(f"  Filtered (kept): {filtered_features} ({filtered_features/total_features*100:.1f}%)")
        print(f"  Stop signs filtered: {stop_signs_filtered}/{len(all_stop_signs)} ({stop_signs_filtered/len(all_stop_signs)*100:.1f}%)")
        print(f"  Roundabouts filtered: {roundabouts_filtered}/{len(all_roundabouts)} ({roundabouts_filtered/len(all_roundabouts)*100:.1f}%)")
        
        # Export to CSV
        df.to_csv(output_filename, index=False)
        print(f"Results exported to: {output_filename}")
        
        # Show sample of filtered features
        filtered_df = df[df['is_filtered'] == True]
        if len(filtered_df) > 0:
            print(f"\nSample of filtered features:")
            print(filtered_df[['feature_type', 'osmid', 'lat', 'lon', 'name']].head(10))
        
        # Show sample of rejected features
        rejected_df = df[df['is_filtered'] == False]
        if len(rejected_df) > 0:
            print(f"\nSample of rejected features:")
            print(rejected_df[['feature_type', 'osmid', 'lat', 'lon', 'name']].head(5))
        
        return {
            'total_features': total_features,
            'filtered_count': filtered_features,
            'stop_signs_filtered': stop_signs_filtered,
            'roundabouts_filtered': roundabouts_filtered,
            'driven_road_ids_count': len(driven_road_ids),
            'csv_file': output_filename
        }


    def load_and_process_csv(self, filepath):
        """Load and process single CSV/Excel with road matching (original method)"""
        
        try:
            print(f"📂 Loading {filepath}...")

            
            # Check if file is Excel or CSV
            if filepath.lower().endswith(('.xlsx', '.xls')):
                # Excel file processing
                xl_file = pd.ExcelFile(filepath)
                sheet_names = xl_file.sheet_names
                
                # Find data sheet
                data_sheet = sheet_names[0]  # Default to first sheet
                for sheet in sheet_names:
                    if any(keyword in sheet.lower() for keyword in ['gps', 'data', 'track', 'location']):
                        data_sheet = sheet
                        break
                
                print(f"   📄 Using Excel sheet: {data_sheet}")
                df = pd.read_excel(filepath, sheet_name=data_sheet)
                
            else:
                # CSV file processing (original logic)
                # Read only first 10KB for encoding detection
                with open(filepath, 'rb') as f:
                    sample = f.read(10240)
                
                detected = chardet.detect(sample)
                encoding = detected.get('encoding', 'utf-8')
                
                # Fallback encodings if detection fails
                if not encoding or detected.get('confidence', 0) < 0.7:
                    encoding = 'utf-8'
                
                df = pd.read_csv(filepath, encoding=encoding)
            
            # Apply common CSV processing logic
            processed_df = self._apply_standard_csv_processing(df)

            # NEW: Assess GPS quality
            self._assess_gps_quality(processed_df)
            
            
            self.processed_df = processed_df
            self.df = processed_df  # Maintain compatibility
            print(f"✅ Processed {len(processed_df)} GPS points with road context")
            return True
            
        except Exception as e:
            print(f"❌ Error processing file: {e}")
            return False


    def get_bounding_box(self):
        """Get optimized bounding box based on actual route (original method)"""
        if self.processed_df is None:
            return None
        
        # Use smaller, speed-adaptive buffer
        avg_speed = self.processed_df['speed_kmh'].mean()
        buffer = max(0.001, min(0.003, avg_speed / 2000))  # Adaptive: 100-300m
        
        return (
            self.processed_df['lat'].min() - buffer,
            self.processed_df['lon'].min() - buffer,
            self.processed_df['lat'].max() + buffer,
            self.processed_df['lon'].max() + buffer
        )
        
    def get_week_specific_bbox(self, week_data):
        """Get bounding box for a specific week's data"""
        if week_data is None or len(week_data) == 0:
            return None
            
        buffer = 0.005  # Smaller buffer for individual weeks (500m)
        return (
            week_data['lat'].min() - buffer,
            week_data['lon'].min() - buffer,
            week_data['lat'].max() + buffer,
            week_data['lon'].max() + buffer
        )
        
    def should_use_combined_download(self, max_area_km2=1):
        """Check if all weeks are close enough to use single download (France-optimized)"""
        if not self.weekly_data:
            return False
        
        # Calculate combined bounding box
        all_lats = []
        all_lons = []
        for week_data in self.weekly_data.values():
            all_lats.extend(week_data['lat'])
            all_lons.extend(week_data['lon'])
        
        lat_range = max(all_lats) - min(all_lats)
        lon_range = max(all_lons) - min(all_lons)
        
        # France-specific conversion (more accurate than generic formula)
        # At France's latitude (~46°N), 1° lat ≈ 111 km, 1° lon ≈ 78 km
        area_km2 = lat_range * 111 * lon_range * 78
        
        print(f"Combined area: {area_km2:.1f} km² (threshold: {max_area_km2} km²)")
        return area_km2 <= max_area_km2

    def get_driven_road_ids(self):
        """Get OSM way IDs of roads actually driven (original method)"""
        if not self.driven_edges:
            return set()
        
        road_ids = set()
        for edge in self.driven_edges:
            if edge and len(edge) >= 3:
                edge_data = self.road_graph.edges[edge]
                osmid = edge_data.get('osmid', None)
                if osmid:
                    if isinstance(osmid, list):
                        road_ids.update(osmid)
                    else:
                        road_ids.add(osmid)
        
        return road_ids
    
    # ========================================================================
    # MULTI-WEEK METHODS - Original MultiWeekGPSProcessor functionality  
    # ========================================================================

    
    def load_multiple_csvs(self, csv_files_or_pattern, driver_name=None):
        """
        Load multiple CSV/Excel files for multi-week analysis with encoding fallback.
    
        Args:
            csv_files_or_pattern: List of files OR pattern like "driver1_week*.*"
            driver_name: Driver name for reports
        """
        self.is_multi_week_mode = True
    
        # Resolve file list
        if isinstance(csv_files_or_pattern, str):
            csv_files = sorted(glob.glob(csv_files_or_pattern)) if '*' in csv_files_or_pattern else [csv_files_or_pattern]
        else:
            csv_files = csv_files_or_pattern
    
        if not csv_files:
            print("❌ No files found!")
            return False
    
        print(f"📁 Loading {len(csv_files)} files for multi-week analysis...")
    
        all_dataframes = []
        for i, csv_file in enumerate(csv_files, 1):
            print(f"  📂 Processing Week {i}: {os.path.basename(csv_file)}")
    
            try:
                # Check if file is Excel or CSV
                if csv_file.lower().endswith(('.xlsx', '.xls')):
                    # Excel file processing
                    xl_file = pd.ExcelFile(csv_file)
                    sheet_names = xl_file.sheet_names
                    
                    # Find data sheet
                    data_sheet = sheet_names[0]  # Default to first sheet
                    for sheet in sheet_names:
                        if any(keyword in sheet.lower() for keyword in ['gps', 'data', 'track', 'location']):
                            data_sheet = sheet
                            break
                    
                    df = pd.read_excel(csv_file, sheet_name=data_sheet)
                    print(f"    📄 Using Excel sheet: {data_sheet}")
                    
                else:
                    # CSV file processing (original logic)
                    # Detect encoding
                    with open(csv_file, 'rb') as f:
                        sample = f.read(10240)
                    detected = chardet.detect(sample)
                    encoding = detected.get('encoding', 'utf-8')
    
                    # Read CSV with detected encoding
                    df = pd.read_csv(csv_file, encoding=encoding)
    
                # Apply standard processing
                processed_df = self._apply_standard_csv_processing(df)
                processed_df['week_number'] = i
                processed_df['week_label'] = f"Week {i}"
                processed_df['source_file'] = os.path.basename(csv_file)
    
                self.weekly_data[f"Week {i}"] = processed_df.copy()
                all_dataframes.append(processed_df)
    
                file_type = "Excel" if csv_file.lower().endswith(('.xlsx', '.xls')) else "CSV"
                encoding_info = "" if file_type == "Excel" else f", encoding: {encoding}"
                print(f"    ✅ Week {i}: {len(processed_df)} GPS points ({file_type}{encoding_info})")
    
            except Exception as e:
                print(f"    ❌ Failed to process: {csv_file} - {e}")
    
        if not all_dataframes:
            print("❌ No valid data found in any files!")
            return False
    
        self.combined_data = pd.concat(all_dataframes, ignore_index=True)
        self.combined_data = self.combined_data.sort_values(['week_number', 'timestamp']).reset_index(drop=True)

        # NEW: Assess GPS quality on combined data
        self._assess_gps_quality(self.combined_data)
    
        print(f"✅ Combined data: {len(self.combined_data)} total GPS points across {len(csv_files)} files")
        return True
    
    
        
    def get_weekly_bounding_boxes(self):
        """Get bounding box for all weeks combined (original method)"""
        if self.combined_data is None:
            return None
            
        buffer = 0.01  # Larger buffer for multi-week data
        return (
            self.combined_data['lat'].min() - buffer,
            self.combined_data['lon'].min() - buffer,
            self.combined_data['lat'].max() + buffer,
            self.combined_data['lon'].max() + buffer
        )
    
    # ========================================================================
    # UNIFIED PRIVATE METHODS - Common processing logic
    # ========================================================================
    
    def _apply_standard_csv_processing(self, df):
        """
        Apply standard CSV processing logic used by both single and multi-week processors.
        Supports multiple tracker formats with auto-detection.
        
        Tracker 1: Uses 'Date Heure' column with format 'DD.MM.YYYY HH:MM:SS'
        Tracker 2: Uses 'dt_tracker' column with format 'YYYY-MM-DD HH:MM:SS'
        
        Enhanced: Parses Teltonika IO parameters for advanced behavior detection.
        """
        
        # Detect tracker type BEFORE renaming
        is_tracker1 = 'Date Heure' in df.columns or any('Date Heure' in str(col) for col in df.columns)
        is_tracker2 = 'dt_tracker' in df.columns or any('dt_tracker' in str(col) for col in df.columns)
    
        # Column mapping (unified for both trackers)
        column_mapping = {
            'Latitude': 'lat',
            'lat': 'lat',
            'Longitude': 'lon',
            'lng': 'lon',
            'Date Heure': 'timestamp',
            'dt_tracker': 'timestamp',
            'La rapidité': 'speed',
            'speed': 'speed',
            'angle': 'heading',
            'altitude': 'altitude'
        }
    
        df = df.rename(columns=column_mapping)
    
        # Parse timestamps correctly
        if is_tracker1:
            # Tracker 1: DD.MM.YYYY HH:MM:SS
            df['timestamp'] = pd.to_datetime(df['timestamp'], format='%d.%m.%Y %H:%M:%S')
        
        elif is_tracker2:
            # Tracker 2: YYYY-MM-DD HH:MM:SS
            try:
                df['timestamp'] = pd.to_datetime(df['timestamp'], format='%Y-%m-%d %H:%M:%S')
            except Exception:
                # Fallback for other formats
                try:
                    df['timestamp'] = pd.to_datetime(df['timestamp'], format='%m/%d/%Y %H:%M')
                except Exception:
                    df['timestamp'] = pd.to_datetime(df['timestamp'])
        
        else:
            raise ValueError("Unknown tracker format. No recognizable date column found.")
    
        # Handle speed column (with or without 'km/h' suffix)
        if df['speed'].dtype == 'object':
            # String type - remove 'km/h' if present (tracker 1)
            df['speed_kmh'] = df['speed'].str.replace(' km/h', '').astype(float)
        else:
            # Numeric type - use directly (tracker 2)
            df['speed_kmh'] = df['speed'].astype(float)
        
        # ================================================================
        # NEW: Parse Teltonika IO Parameters (with fallback)
        # ================================================================
        df = self._parse_teltonika_io_params(df)
        

        
        # Find first non-zero speed index
        start_idx = 0
        for idx, speed in enumerate(df['speed_kmh']):
            if speed > 0:
                start_idx = idx
                break
        
        # Find last non-zero speed index
        end_idx = len(df) - 1
        for i in range(len(df) - 1, -1, -1):
            if df.iloc[i]['speed_kmh'] > 0:
                end_idx = i
                break
        
        # Filter data between first and last movement
        df = df.iloc[start_idx:end_idx + 1].copy().reset_index(drop=True)
        df = df.sort_values('timestamp').reset_index(drop=True)
        
        # Calculate time differences
        df['time_diff_s'] = df['timestamp'].diff().dt.total_seconds().fillna(0)
        
        # Calculate distances between consecutive points
        distances = [0]
        for i in range(1, len(df)):
            dist = geodesic(
                (df.iloc[i - 1]['lat'], df.iloc[i - 1]['lon']),
                (df.iloc[i]['lat'], df.iloc[i]['lon'])
            ).meters
            distances.append(dist)
        df['distance_m'] = distances
        
        return df
    
    
    def _parse_teltonika_io_params(self, df):
        """
        Parse Teltonika IO parameters from 'params' column.
        Extracts key parameters for enhanced behavior detection.
        
        Returns DataFrame with new columns (or None values if params not available).
        """
        
        # Check if params column exists
        has_params = 'params' in df.columns
        
        # Initialize new columns with default values
        df['io_movement'] = None       # io240: 0=stationary, 1=moving
        df['io_ignition'] = None       # io239: 0=off, 1=on
        df['io_engine_load'] = None    # io1440: engine load %
        df['io_green_driving_type'] = None   # io253: 0=none, 1=harsh_accel, 2=harsh_brake, 3=harsh_corner
        df['io_green_driving_value'] = None  # io254: severity value
        df['io_crash_detection'] = None      # io247: crash detected
        df['io_odometer'] = None       # io16: total odometer (meters)
        df['io_trip_odometer'] = None  # io800: trip odometer (meters)
        
        if not has_params:
            print("   ℹ️  No 'params' column found - using basic GPS data only")
            df['has_io_params'] = False
            return df
        
        # Parse params for each row
        io_data = []
        
        for idx, row in df.iterrows():
            params_str = row.get('params', '')
            parsed = self._parse_single_params_string(params_str)
            io_data.append(parsed)
        
        # Convert to DataFrame and merge
        io_df = pd.DataFrame(io_data)
        
        # Assign parsed values
        df['io_movement'] = io_df['io240'].values
        df['io_ignition'] = io_df['io239'].values
        df['io_engine_load'] = io_df['io1440'].values
        df['io_green_driving_type'] = io_df['io253'].values
        df['io_green_driving_value'] = io_df['io254'].values
        df['io_crash_detection'] = io_df['io247'].values
        df['io_odometer'] = io_df['io16'].values
        df['io_trip_odometer'] = io_df['io800'].values
        
        # Flag to indicate IO params are available
        df['has_io_params'] = io_df['has_data'].values
        
        # Count successful parses
        params_count = df['has_io_params'].sum()
        print(f"   ✅ Parsed IO parameters for {params_count}/{len(df)} points")
        
        return df
    
    
    def _parse_single_params_string(self, params_str):
        """
        Parse a single params string like:
        "gpslev=24, gsmlev=4, hdop=0.5, io240=1, io253=0, ..."
        
        Returns dict with extracted values.
        """
        
        result = {
            'io239': None,    # Ignition
            'io240': None,    # Movement
            'io247': None,    # Crash
            'io253': None,    # Green driving type
            'io254': None,    # Green driving value
            'io1440': None,   # Engine load
            'io16': None,     # Odometer
            'io800': None,    # Trip odometer
            'has_data': False
        }
        
        if not params_str or pd.isna(params_str):
            return result
        
        try:
            # Remove quotes if present
            params_str = str(params_str).strip('"').strip("'")
            
            # Split by comma and parse each key=value pair
            for param in params_str.split(','):
                param = param.strip()
                if '=' in param:
                    key, value = param.split('=', 1)
                    key = key.strip().lower()
                    value = value.strip()
                    
                    # Parse numeric values
                    try:
                        if '.' in value:
                            value = float(value)
                        else:
                            value = int(value)
                    except ValueError:
                        pass  # Keep as string if not numeric
                    
                    # Map to result keys
                    if key == 'io239':
                        result['io239'] = value
                    elif key == 'io240':
                        result['io240'] = value
                    elif key == 'io247':
                        result['io247'] = value
                    elif key == 'io253':
                        result['io253'] = value
                    elif key == 'io254':
                        result['io254'] = value
                    elif key == 'io1440':
                        result['io1440'] = value
                    elif key == 'io16':
                        result['io16'] = value
                    elif key == 'io800':
                        result['io800'] = value
            
            # Mark as having data if at least some key params exist
            if result['io240'] is not None or result['io239'] is not None:
                result['has_data'] = True
                
        except Exception as e:
            # Silently fail - just return empty result
            pass
        
        return result

    def _assess_gps_quality(self, df):
        """
        Assess GPS data quality to determine detection strategy.
        Call this after _apply_standard_csv_processing().
        
        Returns dict with quality metrics.
        """
        
        quality = {
            'avg_interval_seconds': None,
            'min_interval_seconds': None,
            'max_interval_seconds': None,
            'is_sparse': False,           # True if avg interval > 4 seconds
            'has_io_params': False,       # True if Teltonika params available
            'has_movement_status': False, # True if io240 available
            'has_engine_data': False,     # True if io1440 available
            'has_harsh_events': False,    # True if io253 available
            'detection_mode': 'basic'     # 'enhanced', 'standard', 'basic'
        }
        
        # Calculate interval statistics
        if 'time_diff_s' in df.columns:
            intervals = df['time_diff_s'][df['time_diff_s'] > 0]
            if len(intervals) > 0:
                quality['avg_interval_seconds'] = intervals.mean()
                quality['min_interval_seconds'] = intervals.min()
                quality['max_interval_seconds'] = intervals.max()
                quality['is_sparse'] = quality['avg_interval_seconds'] > 4
        
        # Check IO params availability
        if 'has_io_params' in df.columns:
            quality['has_io_params'] = df['has_io_params'].any()
        
        if 'io_movement' in df.columns:
            quality['has_movement_status'] = df['io_movement'].notna().any()
        
        if 'io_engine_load' in df.columns:
            quality['has_engine_data'] = df['io_engine_load'].notna().any()
        
        if 'io_green_driving_type' in df.columns:
            quality['has_harsh_events'] = df['io_green_driving_type'].notna().any()

        
        # Determine detection mode
        if quality['has_movement_status'] and quality['has_engine_data']:
            quality['detection_mode'] = 'enhanced'
        elif quality['has_io_params']:
            quality['detection_mode'] = 'standard'
        else:
            quality['detection_mode'] = 'basic'
        
        # Store as instance attribute for use in analysis
        self.gps_quality = quality
        
        # Print summary
        print(f"\n📊 GPS Quality Assessment:")
        print(f"   Average interval: {quality['avg_interval_seconds']:.1f}s {'(SPARSE)' if quality['is_sparse'] else '(OK)'}")
        print(f"   Detection mode: {quality['detection_mode'].upper()}")
        print(f"   IO params available: {'✅' if quality['has_io_params'] else '❌'}")
        print(f"   Movement status (io240): {'✅' if quality['has_movement_status'] else '❌'}")
        print(f"   Engine data (io1440): {'✅' if quality['has_engine_data'] else '❌'}")
       
        return quality
        
    

    def debug_bbox_calculation(self):
        """Debug where the bounding box comes from"""
        
        print("🔍 DEBUGGING BOUNDING BOX CALCULATION")
        print("="*60)
        
        # 1. Check combined_data
        print("\n1️⃣ COMBINED DATA:")
        if self.combined_data is not None:
            print(f"   Rows: {len(self.combined_data)}")
            print(f"   Lat range: [{self.combined_data['lat'].min():.4f}, {self.combined_data['lat'].max():.4f}]")
            print(f"   Lon range: [{self.combined_data['lon'].min():.4f}, {self.combined_data['lon'].max():.4f}]")
        else:
            print("   ❌ combined_data is None!")
        
        # 2. Check weekly_data totals
        print("\n2️⃣ WEEKLY DATA:")
        all_lats = []
        all_lons = []
        for week, df in self.weekly_data.items():
            print(f"   {week}: {len(df)} rows, lat [{df['lat'].min():.4f}, {df['lat'].max():.4f}]")
            all_lats.extend(df['lat'].tolist())
            all_lons.extend(df['lon'].tolist())
        
        print(f"\n   Combined from weekly_data:")
        print(f"   Lat range: [{min(all_lats):.4f}, {max(all_lats):.4f}]")
        print(f"   Lon range: [{min(all_lons):.4f}, {max(all_lons):.4f}]")
        
        # 3. Calculate what bbox SHOULD be
        buffer = 0.01
        correct_bbox = (
            min(all_lats) - buffer,
            min(all_lons) - buffer,
            max(all_lats) + buffer,
            max(all_lons) + buffer
        )
        print(f"\n3️⃣ CORRECT BBOX SHOULD BE:")
        print(f"   ({correct_bbox[0]:.4f}, {correct_bbox[1]:.4f}, {correct_bbox[2]:.4f}, {correct_bbox[3]:.4f})")
        
        # 4. Check what get_weekly_bounding_boxes returns
        print("\n4️⃣ get_weekly_bounding_boxes() RETURNS:")
        bbox = self.get_weekly_bounding_boxes()
        if bbox:
            print(f"   ({bbox[0]:.4f}, {bbox[1]:.4f}, {bbox[2]:.4f}, {bbox[3]:.4f})")
            
            # Compare
            if abs(bbox[0] - correct_bbox[0]) > 0.001:
                print(f"   ⚠️ Min lat mismatch! Got {bbox[0]:.4f}, expected {correct_bbox[0]:.4f}")
        else:
            print("   ❌ Returns None!")
    


    def _build_optimized_graph(self, corridor_polygon, area_hash):
        """
        FIXED: Proper node merging + bidirectional edges + chain preservation + junction tag
        """
        print(f"DEBUG: Starting optimized graph building with area_hash: {area_hash}")
        
        import networkx as nx
        from collections import defaultdict
        
        G = nx.MultiDiGraph()
        
        print(f"  🔧 Processing {len(self.map_manager.map_data['roads'])} roads...")
    
        # Get expanded bounding box
        bbox = corridor_polygon.bounds
        buffer_deg = 0.002
        expanded_bbox = (
            bbox[0] - buffer_deg,
            bbox[1] - buffer_deg,
            bbox[2] + buffer_deg,
            bbox[3] + buffer_deg
        )
    
        # Filter relevant roads
        relevant_roads = []
        for road in self.map_manager.map_data['roads']:
            geometry = road.get('geometry', [])
            if len(geometry) < 2:
                continue
                
            road_minx = min(p['lon'] for p in geometry)
            road_maxx = max(p['lon'] for p in geometry)
            road_miny = min(p['lat'] for p in geometry)
            road_maxy = max(p['lat'] for p in geometry)
    
            if (road_maxx >= expanded_bbox[0] and road_minx <= expanded_bbox[2] and
                road_maxy >= expanded_bbox[1] and road_miny <= expanded_bbox[3]):
                relevant_roads.append(road)
    
        print(f"  ✅ {len(relevant_roads)} roads in expanded bounding box")
        
        # FIXED: Use 1m grid precision (0.00001 degrees ≈ 1.1m)
        GRID_SIZE = 0.00001
        
        def get_grid_key(lat, lon):
            grid_lat = round(lat / GRID_SIZE) * GRID_SIZE
            grid_lon = round(lon / GRID_SIZE) * GRID_SIZE
            return (round(grid_lat, 6), round(grid_lon, 6))
        
        # Collect all nodes with grid keys
        grid_to_nodes = defaultdict(list)
        
        for road in relevant_roads:
            geometry = road.get('geometry', [])
            node_refs = road.get('node_refs', [])
            
            if len(node_refs) != len(geometry):
                continue
                
            for i, point in enumerate(geometry):
                grid_key = get_grid_key(point['lat'], point['lon'])
                grid_to_nodes[grid_key].append((node_refs[i], point['lat'], point['lon']))
        
        # Pick canonical node per grid cell
        grid_to_canonical = {}
        original_to_canonical = {}
        
        merged_count = 0
        for grid_key, nodes in grid_to_nodes.items():
            if len(nodes) > 1:
                merged_count += len(nodes) - 1
            
            canonical_id = min(n[0] for n in nodes)
            avg_lat = sum(n[1] for n in nodes) / len(nodes)
            avg_lon = sum(n[2] for n in nodes) / len(nodes)
            
            grid_to_canonical[grid_key] = (canonical_id, avg_lat, avg_lon)
            
            for node_id, _, _ in nodes:
                original_to_canonical[node_id] = canonical_id
        
        print(f"  🔗 Merged {merged_count} duplicate nodes at intersections")
        print(f"  📍 Total unique nodes: {len(grid_to_canonical)}")
        
        # Add nodes to graph
        nodes_added = set()
        for canonical_id, avg_lat, avg_lon in grid_to_canonical.values():
            if canonical_id not in nodes_added:
                G.add_node(canonical_id, y=avg_lat, x=avg_lon)
                nodes_added.add(canonical_id)
        
        # FIXED: Add edges with self-loop handling + bidirectional support
        edges_added = set()
        
        for road in relevant_roads:
            geometry = road.get('geometry', [])
            node_refs = road.get('node_refs', [])
            
            if len(node_refs) != len(geometry):
                continue
            
            # Get canonical sequence
            canonical_sequence = [
                original_to_canonical.get(node_refs[i], node_refs[i])
                for i in range(len(geometry))
            ]
            
            # Remove consecutive duplicates caused by merging
            deduplicated = [canonical_sequence[0]]
            for node_id in canonical_sequence[1:]:
                if node_id != deduplicated[-1]:
                    deduplicated.append(node_id)
            
            # Check if oneway
            tags = road.get('tags', {})
            is_oneway = tags.get('oneway') == 'yes'
            
            # FIX: Extract junction tag from road tags
            junction_tag = tags.get('junction', None)
            
            # Add edges
            for i in range(len(deduplicated) - 1):
                u_id = deduplicated[i]
                v_id = deduplicated[i + 1]
                
                # Calculate edge length
                from geopy.distance import geodesic
                u_data = G.nodes[u_id]
                v_data = G.nodes[v_id]
                length_m = geodesic((u_data['y'], u_data['x']), (v_data['y'], v_data['x'])).meters
                
                # FIX: Edge attributes now include junction tag
                edge_attrs = {
                    'osmid': road.get('id'),
                    'highway': tags.get('highway', 'unclassified'),
                    'maxspeed': road.get('speed_limit', 50),
                    'name': tags.get('name', None),
                    'length': length_m,
                    'junction': junction_tag,  # NEW: Preserve junction tag
                    'oneway': is_oneway  # NEW: Also preserve oneway info
                }
                
                # Forward edge
                edge_key = (u_id, v_id, 0)
                if edge_key not in edges_added:
                    G.add_edge(u_id, v_id, 0, **edge_attrs)
                    edges_added.add(edge_key)
                
                # Reverse edge if not oneway
                if not is_oneway:
                    rev_key = (v_id, u_id, 1)
                    if rev_key not in edges_added:
                        G.add_edge(v_id, u_id, 1, **edge_attrs)
                        edges_added.add(rev_key)
        
        G.graph['crs'] = 'EPSG:4326'
        self.road_graph = G
        
        print(f"  📊 Graph built: {G.number_of_nodes()} nodes, {G.number_of_edges()} edges")
        
        # NEW: Count junction types
        junction_counts = {}
        for u, v, k, data in G.edges(keys=True, data=True):
            junc = data.get('junction', 'none')
            if junc:
                junction_counts[junc] = junction_counts.get(junc, 0) + 1
        if junction_counts:
            print(f"  🔄 Junction types: {junction_counts}")
        
        # Verify connectivity
        num_components = nx.number_weakly_connected_components(G)
        if num_components > 1:
            components = list(nx.weakly_connected_components(G))
            sizes = sorted([len(c) for c in components], reverse=True)
            print(f"  ⚠️  Graph has {num_components} components. Sizes: {sizes[:5]}")
        else:
            print(f"  ✅ Graph is fully connected!")
        
        # Cache the graph
        cache_file = os.path.join(self.cache_dir, f"road_graph_{area_hash}.pkl")
        os.makedirs(self.cache_dir, exist_ok=True)
        try:
            with open(cache_file, 'wb') as f:
                pickle.dump({'graph': G}, f)
            print(f"  💾 Cached graph to {cache_file}")
        except Exception:
            pass
        
        return G




    
    # Add this diagnostic after _build_optimized_graph completes
    def diagnose_specific_gap(self, gap_edge1, gap_edge2):
        """Diagnose why two edges are not connected"""
        u1, v1, k1 = gap_edge1
        u2, v2, k2 = gap_edge2
        
        print(f"\n🔍 GAP DIAGNOSTIC:")
        print(f"   Edge 1: {gap_edge1}")
        print(f"   Edge 2: {gap_edge2}")
        
        # Get edge data
        try:
            e1_data = self.road_graph.edges[gap_edge1]
            e2_data = self.road_graph.edges[gap_edge2]
            
            print(f"   Edge 1 osmid: {e1_data.get('osmid')}, name: {e1_data.get('name', 'unnamed')}")
            print(f"   Edge 2 osmid: {e2_data.get('osmid')}, name: {e2_data.get('name', 'unnamed')}")
            
            # Check if they share any nodes
            nodes_e1 = {u1, v1}
            nodes_e2 = {u2, v2}
            shared = nodes_e1.intersection(nodes_e2)
            
            print(f"   Edge 1 nodes: {u1} → {v1}")
            print(f"   Edge 2 nodes: {u2} → {v2}")
            print(f"   Shared nodes: {shared}")
            
            # Get node coordinates
            n1_coords = (self.road_graph.nodes[v1]['y'], self.road_graph.nodes[v1]['x'])
            n2_coords = (self.road_graph.nodes[u2]['y'], self.road_graph.nodes[u2]['x'])
            
            from geopy.distance import geodesic
            node_dist = geodesic(n1_coords, n2_coords).meters
            
            print(f"   Distance between v1 and u2: {node_dist:.1f}m")
            
            if node_dist < 10:
                print(f"   ⚠️  Nodes are very close but NOT connected! This is the bug.")
            
        except Exception as e:
            print(f"   Error: {e}")
            
    # Add this after the graph building completes, before map matching starts
    def diagnose_graph_connectivity(self):
        """Deep diagnostic of graph connectivity issues"""
        import networkx as nx
        
        print("\n🔍 GRAPH CONNECTIVITY DIAGNOSTIC:")
        print(f"   Total nodes: {self.road_graph.number_of_nodes()}")
        print(f"   Total edges: {self.road_graph.number_of_edges()}")
        
        # Check connected components
        num_components = nx.number_weakly_connected_components(self.road_graph)
        print(f"   Weakly connected components: {num_components}")
        
        if num_components > 1:
            components = list(nx.weakly_connected_components(self.road_graph))
            sizes = sorted([len(c) for c in components], reverse=True)
            print(f"   Component sizes (top 5): {sizes[:5]}")
            print(f"   ⚠️  Graph is fragmented into {num_components} pieces!")
        
        # Sample some edges to check node sharing
        print("\n   Checking node sharing at intersections...")
        
        # Get all unique node IDs
        all_nodes = set(self.road_graph.nodes())
        
        # Count how many edges each node participates in
        node_degree = {}
        for node in all_nodes:
            in_deg = self.road_graph.in_degree(node)
            out_deg = self.road_graph.out_degree(node)
            node_degree[node] = in_deg + out_deg
        
        # Intersection nodes should have degree >= 3
        degree_1_nodes = sum(1 for d in node_degree.values() if d == 1)
        degree_2_nodes = sum(1 for d in node_degree.values() if d == 2)
        degree_3plus_nodes = sum(1 for d in node_degree.values() if d >= 3)
        
        print(f"   Dead-end nodes (degree 1): {degree_1_nodes}")
        print(f"   Through nodes (degree 2): {degree_2_nodes}")
        print(f"   Intersection nodes (degree 3+): {degree_3plus_nodes}")
        
        # Check a sample of consecutive roads to see if they share nodes
        print("\n   Checking if roads share intersection nodes...")
        
        roads = self.map_manager.map_data.get('roads', [])[:20]  # Sample first 20 roads
        
        shared_count = 0
        not_shared_count = 0
        
        for i in range(len(roads) - 1):
            road1_nodes = set(roads[i].get('node_refs', []))
            road2_nodes = set(roads[i+1].get('node_refs', []))
            
            shared = road1_nodes.intersection(road2_nodes)
            if shared:
                shared_count += 1
            else:
                not_shared_count += 1
        
        print(f"   Roads sharing nodes: {shared_count}")
        print(f"   Roads NOT sharing nodes: {not_shared_count}")
        
        return num_components
    
    # Call it after graph building:
    # self.diagnose_graph_connectivity()


    def deep_check_road_coverage(self, lat, lon):
        """Deep check why no roads are found at a specific point"""
        from geopy.distance import geodesic
        from shapely.geometry import Point
        
        print(f"\n🔍 DEEP CHECK AT ({lat}, {lon})")
        print("="*60)
        
        # 1. Find nearest node in graph
        print("\n1️⃣ NEAREST NODES IN ROAD GRAPH:")
        node_distances = []
        for node, data in self.road_graph.nodes(data=True):
            if 'y' in data and 'x' in data:
                dist = geodesic((lat, lon), (data['y'], data['x'])).meters
                node_distances.append((node, data['y'], data['x'], dist))
        
        node_distances.sort(key=lambda x: x[3])
        
        for node, n_lat, n_lon, dist in node_distances[:5]:
            print(f"   Node {node}: ({n_lat:.6f}, {n_lon:.6f}) - {dist:.0f}m away")
        
        nearest_dist = node_distances[0][3] if node_distances else float('inf')
        print(f"\n   ➡️ Nearest node is {nearest_dist:.0f}m away")
        
        # 2. Find nearest edge in graph
        print("\n2️⃣ NEAREST EDGES IN ROAD GRAPH:")
        gps_point = Point(lon, lat)
        edge_distances = []
        
        for u, v, key, data in self.road_graph.edges(keys=True, data=True):
            edge = (u, v, key)
            edge_geom = self._get_edge_geometry_fast(edge)
            if edge_geom is None:
                continue
            
            dist_m = gps_point.distance(edge_geom) * 111000
            edge_distances.append({
                'edge': edge,
                'distance': dist_m,
                'name': data.get('name', ''),
                'osmid': data.get('osmid')
            })
        
        edge_distances.sort(key=lambda x: x['distance'])
        
        for e in edge_distances[:5]:
            print(f"   '{e['name']}' (OSM:{e['osmid']}): {e['distance']:.0f}m away")
        
        nearest_edge_dist = edge_distances[0]['distance'] if edge_distances else float('inf')
        print(f"\n   ➡️ Nearest edge is {nearest_edge_dist:.0f}m away")
        
        # 3. Check map_data for roads near this point
        print("\n3️⃣ ROADS IN MAP_DATA NEAR THIS POINT:")
        roads_found = []
        
        for road in self.map_manager.map_data.get('roads', []):
            geometry = road.get('geometry', [])
            for pt in geometry:
                dist = geodesic((lat, lon), (pt['lat'], pt['lon'])).meters
                if dist < 100:
                    roads_found.append({
                        'name': road.get('tags', {}).get('name', ''),
                        'osmid': road.get('id'),
                        'distance': dist,
                        'highway': road.get('tags', {}).get('highway', '')
                    })
                    break
        
        roads_found.sort(key=lambda x: x['distance'])
        
        if roads_found:
            print(f"   Found {len(roads_found)} roads within 100m in map_data:")
            for r in roads_found[:10]:
                print(f"   '{r['name']}' (OSM:{r['osmid']}, {r['highway']}): {r['distance']:.0f}m")
        else:
            print("   ❌ NO roads within 100m in map_data either!")
        
        # 4. Diagnosis
        print("\n4️⃣ DIAGNOSIS:")
        if nearest_edge_dist > 100 and roads_found:
            print(f"   ⚠️ Roads exist in map_data but NOT in road_graph!")
            print(f"   The _build_optimized_graph() function filtered them out.")
            print(f"   Check the corridor_polygon or bbox used during graph building.")
        elif nearest_edge_dist > 100 and not roads_found:
            print(f"   ❌ No roads in map_data either - this area has no OSM road data")
            print(f"   Or the bbox filter excluded this area during master cache filtering.")
        else:
            print(f"   ✅ Roads exist - issue might be in candidate search logic.")
    
    
        
    def trace_data_flow(self):
        """Trace where different data comes from"""
        
        print("🔍 TRACING DATA FLOW")
        print("="*60)
        
        # 1. GPS Data bounds
        print("\n1️⃣ GPS DATA BOUNDS:")
        for week, df in self.weekly_data.items():
            print(f"   {week}: lat [{df['lat'].min():.4f}, {df['lat'].max():.4f}], lon [{df['lon'].min():.4f}, {df['lon'].max():.4f}]")
        
        # 2. Map data bounds (from map_manager)
        print("\n2️⃣ MAP DATA (from map_manager.map_data):")
        if hasattr(self, 'map_manager') and self.map_manager:
            roads = self.map_manager.map_data.get('roads', [])
            if roads:
                all_lats = []
                all_lons = []
                for road in roads:
                    for pt in road.get('geometry', []):
                        all_lats.append(pt['lat'])
                        all_lons.append(pt['lon'])
                if all_lats:
                    print(f"   Roads: {len(roads)}")
                    print(f"   Lat range: [{min(all_lats):.4f}, {max(all_lats):.4f}]")
                    print(f"   Lon range: [{min(all_lons):.4f}, {max(all_lons):.4f}]")
            else:
                print("   No roads in map_data")
        else:
            print("   No map_manager reference")
        
        # 3. Road graph bounds
        print("\n3️⃣ ROAD GRAPH BOUNDS:")
        if self.road_graph:
            lats = [data['y'] for _, data in self.road_graph.nodes(data=True) if 'y' in data]
            lons = [data['x'] for _, data in self.road_graph.nodes(data=True) if 'x' in data]
            print(f"   Nodes: {len(lats)}")
            print(f"   Lat range: [{min(lats):.4f}, {max(lats):.4f}]")
            print(f"   Lon range: [{min(lons):.4f}, {max(lons):.4f}]")
        else:
            print("   No road_graph")
        
        # 4. Check for mismatch
        print("\n4️⃣ MISMATCH ANALYSIS:")
        if self.weekly_data and self.road_graph:
            gps_min_lat = min(df['lat'].min() for df in self.weekly_data.values())
            gps_max_lat = max(df['lat'].max() for df in self.weekly_data.values())
            
            graph_lats = [data['y'] for _, data in self.road_graph.nodes(data=True) if 'y' in data]
            graph_min_lat = min(graph_lats)
            graph_max_lat = max(graph_lats)
            
            if gps_min_lat < graph_min_lat:
                gap = (graph_min_lat - gps_min_lat) * 111  # km
                print(f"   ⚠️ GPS goes {gap:.1f} km SOUTH of road graph!")
            if gps_max_lat > graph_max_lat:
                gap = (gps_max_lat - graph_max_lat) * 111  # km
                print(f"   ⚠️ GPS goes {gap:.1f} km NORTH of road graph!")
        
        
    # Add this after _build_optimized_graph() completes
    def debug_graph_connectivity(self):
        """Check graph connectivity"""
        import networkx as nx
        
        # Check if graph is connected
        if self.road_graph.is_directed():
            # For directed graph, check weak connectivity
            num_components = nx.number_weakly_connected_components(self.road_graph)
            largest_cc = max(nx.weakly_connected_components(self.road_graph), key=len)
        else:
            num_components = nx.number_connected_components(self.road_graph)
            largest_cc = max(nx.connected_components(self.road_graph), key=len)
        
        total_nodes = self.road_graph.number_of_nodes()
        total_edges = self.road_graph.number_of_edges()
        largest_cc_size = len(largest_cc)
        
        print(f"\n🔍 GRAPH CONNECTIVITY DEBUG:")
        print(f"   Total nodes: {total_nodes}")
        print(f"   Total edges: {total_edges}")
        print(f"   Connected components: {num_components}")
        print(f"   Largest component: {largest_cc_size} nodes ({100*largest_cc_size/total_nodes:.1f}%)")
        
        if num_components > 1:
            print(f"   ⚠️  WARNING: Graph has {num_components} disconnected components!")
            print(f"   This explains why gaps are being detected.")
    
    def _perform_map_matching(self, df, build_graph_only=False, rebuild_graph=True):
        """
        GOLD STANDARD: HMM-based map matching with Viterbi algorithm
        
        Args:
            df: DataFrame with GPS points
            build_graph_only: If True, only build the graph, don't do matching
            rebuild_graph: If False, use existing road_graph instead of rebuilding
        """
        try:
            # 1) Build corridor and road network
            coords = list(zip(df['lon'].tolist(), df['lat'].tolist()))
            raw_line = LineString(coords)
            buffer_degrees = 0.002  # ~200 meters at France latitude
            corridor_ll = raw_line.buffer(buffer_degrees)
    
            # 2) Get or build cached graph for this area
            bbox = corridor_ll.bounds
            area_hash = f"{bbox[0]:.4f}_{bbox[1]:.4f}_{bbox[2]:.4f}_{bbox[3]:.4f}"
    
            # NEW: Only rebuild graph if requested
            if rebuild_graph or self.road_graph is None:
                print("  📥 Building optimized road network from PBF data...")
                if not self._load_cached_graph(area_hash):
                    self.road_graph = self._build_optimized_graph(corridor_ll, area_hash)
            else:
                print(f"  📍 Reusing existing road graph ({self.road_graph.number_of_nodes()} nodes, {self.road_graph.number_of_edges()} edges)")
    
            # NEW: If only building graph, stop here
            if build_graph_only:
                print(f"  ✅ Graph building complete (build_graph_only=True)")
                return
    
            # Build spatial index for fast candidate queries (do this every time)
            self._build_edge_spatial_index()
    
            # 3) HMM-based map matching
            print("  🎯 Performing HMM-based map matching with Viterbi algorithm...")
            X = df['lon'].to_numpy()
            Y = df['lat'].to_numpy()
    
            # Step 3a: Find candidate edges for each GPS point
            print("  🔍 Finding candidate edges for each GPS point...")
            candidates_per_point = self._find_candidates_hmm(X, Y, max_candidates=10, search_radius=100)
    
            # Step 3b: Run Viterbi algorithm to find optimal path
            print("  🧮 Running Viterbi algorithm to find optimal road sequence...")
            matched_edges = self._viterbi_map_matching(X, Y, candidates_per_point, df)

            # Post-process to fix reverse edge zig-zags
            matched_edges = self._fix_reverse_edge_zigzag(matched_edges)
            
            # Apply trajectory smoothing for junction edge switches
            matched_edges = self._trajectory_smoothing(matched_edges, X, Y)

    
            # 4) Reconstruct route geometry following actual road network
            print("  🛣️  Reconstructing route geometry from road network...")
            route_points = self._reconstruct_route_geometry(X, Y, matched_edges)
    
            # Add matched edges to driven edges set
            for edge in matched_edges:
                if edge:
                    self.driven_edges.add(edge)
    
            # 5) Write results to dataframe
            simple_projected = []
            for i, selected_edge in enumerate(matched_edges):
                if selected_edge:
                    geom = self._get_edge_geometry_fast(selected_edge)
                    if geom is not None:
                        pt = Point(X[i], Y[i])
                        try:
                            projected = geom.interpolate(geom.project(pt))
                            simple_projected.append((projected.y, projected.x))
                        except Exception:
                            simple_projected.append((Y[i], X[i]))
                    else:
                        simple_projected.append((Y[i], X[i]))
                else:
                    simple_projected.append((Y[i], X[i]))
    
            df['matched_edge'] = matched_edges
            df['road_matched_lat'] = [p[0] for p in simple_projected]
            df['road_matched_lon'] = [p[1] for p in simple_projected]
    
            # 6) Create route geometry from reconstructed points
            valid_points = []
            for lat, lon in route_points:
                if lat and lon and not (lat == 0 or lon == 0):
                    valid_points.append((lon, lat))
    
            if len(valid_points) <= 1:
                print("  ⚠️  No valid route points found, using original GPS coordinates")
                fallback_points = []
                for i in range(len(df)):
                    if pd.notna(df.iloc[i]['lat']) and pd.notna(df.iloc[i]['lon']):
                        fallback_points.append((df.iloc[i]['lon'], df.iloc[i]['lat']))
    
                if len(fallback_points) > 1:
                    self.route_geometry = LineString(fallback_points)
                    print(f"  ✅ Created fallback route geometry with {len(fallback_points)} GPS points")
    
            if len(valid_points) > 1:
                try:
                    self.route_geometry = LineString(valid_points)
                    print(f"  ✅ Created route geometry with {len(valid_points)} points")
                    self.route_geometry = self.route_geometry.simplify(0.000001, preserve_topology=True)
    
                    # Update road context
                    if hasattr(self, 'map_manager') and self.map_manager:
                        driven_road_ids = {self.road_graph.edges[edge].get('osmid')
                                         for edge in self.driven_edges if edge}
                        driven_road_ids.discard(None)
                        if driven_road_ids:
                            self.map_manager.set_road_context(driven_road_ids, self.route_geometry)
                            print(f"  ✅ Identified {len(driven_road_ids)} unique road segments")
    
                except Exception as e:
                    print(f"  ⚠️  Could not create route geometry: {type(e).__name__}: {str(e)}")
                    self.route_geometry = None
    
        except Exception as e:
            print(f"  ⚠️  Map matching failed: {e}. Using GPS coordinates.")
            import traceback
            traceback.print_exc()
            df['matched_edge'] = [None] * len(df)
            df['road_matched_lat'] = df['lat']
            df['road_matched_lon'] = df['lon']
    
        
        
        # After map matching, add this debug
    def debug_edge_continuity(self, df, sample_size=20):
        """Check if consecutive matched edges share nodes"""
        print("\n🔍 EDGE CONTINUITY DEBUG:")
        
        gaps_found = 0
        for i in range(1, min(sample_size, len(df))):
            prev_edge = df.iloc[i-1].get('matched_edge')
            curr_edge = df.iloc[i].get('matched_edge')
            
            if prev_edge and curr_edge:
                u1, v1, k1 = prev_edge
                u2, v2, k2 = curr_edge
                
                is_connected = (v1 == u2) or (v1 == v2) or (u1 == u2) or (u1 == v2)
                is_same = prev_edge == curr_edge
                
                if not is_connected and not is_same:
                    gaps_found += 1
                    # Get road names
                    prev_name = self.road_graph.edges[prev_edge].get('name', 'unnamed')
                    curr_name = self.road_graph.edges[curr_edge].get('name', 'unnamed')
                    
                    print(f"   Point {i}: GAP - '{prev_name}' → '{curr_name}'")
                    print(f"      Prev edge nodes: {u1} → {v1}")
                    print(f"      Curr edge nodes: {u2} → {v2}")
        
        print(f"\n   Total gaps in first {sample_size} points: {gaps_found}")

    
    def _find_candidates_hmm(self, X, Y, max_candidates=10, search_radius=100):
        """
        IMPROVED: Find candidates with route-awareness for sparse GPS data
        """
        candidates_per_point = []
        n_points = len(X)
        
        prev_best_candidates = None
        
        show_progress = n_points > 500
        progress_interval = max(100, n_points // 20)
    
        for i in range(n_points):
            lon, lat = X[i], Y[i]
    
            if show_progress and i > 0 and i % progress_interval == 0:
                print(f"    Progress: {i}/{n_points} points ({i/n_points*100:.1f}%)")
    
            try:
                # Get spatially nearby candidates
                if self._edge_spatial_index is not None:
                    spatial_candidates = self._find_candidates_with_spatial_index(
                        lon, lat, max_candidates * 2, search_radius
                    )
                else:
                    spatial_candidates = self._find_candidates_fallback(
                        lon, lat, max_candidates * 2, search_radius
                    )
                
                # ============ KEY IMPROVEMENT: Route-aware candidate expansion ============
                # If we have previous candidates, try to find connected paths
                
                if prev_best_candidates and i > 0:
                    from geopy.distance import geodesic
                    gps_dist = geodesic((Y[i-1], X[i-1]), (lat, lon)).meters
                    
                    # For sparse GPS (>50m gap), add candidates reachable from previous
                    if gps_dist > 50:
                        connected_candidates = []
                        
                        for prev_edge, _ in prev_best_candidates[:3]:  # Top 3 previous
                            # Find edges connected to previous edge's end node
                            u_prev, v_prev, k_prev = prev_edge
                            
                            # Get neighbors of v_prev (the "exit" node)
                            if v_prev in self.road_graph:
                                for neighbor in self.road_graph.neighbors(v_prev):
                                    for key in self.road_graph[v_prev][neighbor]:
                                        edge = (v_prev, neighbor, key)
                                        
                                        # Check if this edge is close to current GPS point
                                        edge_dist = self._get_distance_to_edge(lon, lat, edge)
                                        
                                        if edge_dist < search_radius * 1.5:  # Slightly larger radius
                                            connected_candidates.append((edge, edge_dist))
                                            
                                            # Also check edges connected to THIS edge
                                            # (to handle 2-hop gaps)
                                            for next_neighbor in self.road_graph.neighbors(neighbor):
                                                for next_key in self.road_graph[neighbor][next_neighbor]:
                                                    next_edge = (neighbor, next_neighbor, next_key)
                                                    next_dist = self._get_distance_to_edge(lon, lat, next_edge)
                                                    if next_dist < search_radius * 1.5:
                                                        connected_candidates.append((next_edge, next_dist))
                        
                        # Merge with spatial candidates, prioritizing connected ones
                        all_candidates = {}
                        
                        # Add connected candidates with distance bonus
                        for edge, dist in connected_candidates:
                            if edge not in all_candidates:
                                all_candidates[edge] = dist * 0.8  # 20% bonus for connectivity
                        
                        # Add spatial candidates
                        for edge, dist in spatial_candidates:
                            if edge not in all_candidates:
                                all_candidates[edge] = dist
                            else:
                                # If already present (connected), keep lower distance
                                all_candidates[edge] = min(all_candidates[edge], dist * 0.8)
                        
                        # Convert back to sorted list
                        candidates = sorted(all_candidates.items(), key=lambda x: x[1])
                        candidates = [(e, d) for e, d in candidates[:max_candidates]]
                    else:
                        candidates = spatial_candidates[:max_candidates]
                else:
                    candidates = spatial_candidates[:max_candidates]
                
                # Store for next iteration
                prev_best_candidates = candidates[:3] if candidates else None
                
                candidates_per_point.append(candidates)
    
            except Exception as e:
                print(f"    ⚠️  Error finding candidates at point {i}: {e}")
                candidates_per_point.append([])
                prev_best_candidates = None
    
        if show_progress:
            print(f"    ✅ Processed {n_points} GPS points")
    
        return candidates_per_point

    def debug_graph_building_bbox(self):
        """Debug what bbox is used for graph building"""
        
        print("🔍 DEBUGGING GRAPH BUILDING BBOX")
        print("="*60)
        
        # Check the problematic point
        test_lat, test_lon = 48.810827, 7.583511
        
        # 1. Get current road graph bbox
        lats = [data['y'] for _, data in self.road_graph.nodes(data=True) if 'y' in data]
        lons = [data['x'] for _, data in self.road_graph.nodes(data=True) if 'x' in data]
        
        graph_bbox = (min(lons), min(lats), max(lons), max(lats))
        print(f"\n1️⃣ Current Road Graph BBox:")
        print(f"   Min: ({graph_bbox[1]:.6f}, {graph_bbox[0]:.6f})")
        print(f"   Max: ({graph_bbox[3]:.6f}, {graph_bbox[2]:.6f})")
        
        # 2. Check if test point is in bbox
        in_bbox = (graph_bbox[0] <= test_lon <= graph_bbox[2] and 
                   graph_bbox[1] <= test_lat <= graph_bbox[3])
        print(f"\n   Test point ({test_lat}, {test_lon}) in bbox: {in_bbox}")
        
        # 3. Get GPS data bbox for each week
        print(f"\n2️⃣ GPS Data BBox per Week:")
        for week, df in self.weekly_data.items():
            week_bbox = (
                df['lon'].min(), df['lat'].min(),
                df['lon'].max(), df['lat'].max()
            )
            
            test_in_week = (week_bbox[0] <= test_lon <= week_bbox[2] and 
                            week_bbox[1] <= test_lat <= week_bbox[3])
            
            print(f"   {week}:")
            print(f"      Lat: [{week_bbox[1]:.6f}, {week_bbox[3]:.6f}]")
            print(f"      Lon: [{week_bbox[0]:.6f}, {week_bbox[2]:.6f}]")
            print(f"      Contains test point: {'✅ YES' if test_in_week else '❌ NO'}")
        
        # 4. Calculate what bbox SHOULD be used
        print(f"\n3️⃣ Combined GPS Data BBox (what should be used):")
        all_lats = []
        all_lons = []
        for df in self.weekly_data.values():
            all_lats.extend(df['lat'].tolist())
            all_lons.extend(df['lon'].tolist())
        
        combined_bbox = (min(all_lons), min(all_lats), max(all_lons), max(all_lats))
        print(f"   Lat: [{combined_bbox[1]:.6f}, {combined_bbox[3]:.6f}]")
        print(f"   Lon: [{combined_bbox[0]:.6f}, {combined_bbox[2]:.6f}]")
        
        test_in_combined = (combined_bbox[0] <= test_lon <= combined_bbox[2] and 
                            combined_bbox[1] <= test_lat <= combined_bbox[3])
        print(f"   Contains test point: {'✅ YES' if test_in_combined else '❌ NO'}")


    def verify_graph_coverage_at_points(self, coordinates):
        """Check if road graph has ANY roads near these GPS points"""
        
        print("🔍 VERIFYING ROAD GRAPH COVERAGE AT SPECIFIC POINTS")
        print("="*60)
        
        # 1. Check road graph bounds
        lats = [data['y'] for _, data in self.road_graph.nodes(data=True) if 'y' in data]
        lons = [data['x'] for _, data in self.road_graph.nodes(data=True) if 'x' in data]
        
        print(f"\n📊 Road Graph Stats:")
        print(f"   Nodes: {len(lats)}")
        print(f"   Lat range: [{min(lats):.6f}, {max(lats):.6f}]")
        print(f"   Lon range: [{min(lons):.6f}, {max(lons):.6f}]")
        
        # 2. Check if GPS points are within bounds
        print(f"\n📍 GPS Points vs Graph Bounds:")
        for i, (lat, lon) in enumerate(coordinates[:5]):
            lat_in = min(lats) <= lat <= max(lats)
            lon_in = min(lons) <= lon <= max(lons)
            status = "✅ IN BOUNDS" if (lat_in and lon_in) else "❌ OUT OF BOUNDS"
            print(f"   Point {i} ({lat:.6f}, {lon:.6f}): {status}")
            if not lat_in:
                if lat < min(lats):
                    print(f"      → Lat is {(min(lats) - lat) * 111000:.0f}m SOUTH of graph")
                else:
                    print(f"      → Lat is {(lat - max(lats)) * 111000:.0f}m NORTH of graph")
        
        # 3. Find nearest node to first GPS point
        test_lat, test_lon = coordinates[0]
        min_dist = float('inf')
        nearest_node = None
        
        for node, data in self.road_graph.nodes(data=True):
            if 'y' in data and 'x' in data:
                from geopy.distance import geodesic
                dist = geodesic((test_lat, test_lon), (data['y'], data['x'])).meters
                if dist < min_dist:
                    min_dist = dist
                    nearest_node = (node, data['y'], data['x'])
        
        print(f"\n📏 Nearest graph node to Point 0:")
        print(f"   Node: {nearest_node[0]}")
        print(f"   Location: ({nearest_node[1]:.6f}, {nearest_node[2]:.6f})")
        print(f"   Distance: {min_dist:.0f}m")
        
        # 4. Check map_manager.map_data for roads in this area
        print(f"\n📊 Checking map_manager.map_data for roads near Point 0:")
        roads_nearby = 0
        for road in self.map_manager.map_data.get('roads', []):
            for pt in road.get('geometry', []):
                if abs(pt['lat'] - test_lat) < 0.01 and abs(pt['lon'] - test_lon) < 0.01:
                    roads_nearby += 1
                    if roads_nearby <= 3:
                        print(f"   Found: '{road.get('tags', {}).get('name', 'unnamed')}' (OSM: {road.get('id')})")
                    break
        
        print(f"   Total roads within ~1km: {roads_nearby}")
        
        if roads_nearby > 0 and min_dist > 1000:
            print(f"\n   ⚠️ PROBLEM: Roads exist in map_data but NOT in road_graph!")
            print(f"   The graph building step is filtering out these roads.")
    
    
    
        
    def diagnose_specific_points(self, lat1, lon1, lat2, lon2):
        """
        Diagnose what roads are available near specific GPS coordinates
        """
        from geopy.distance import geodesic
        from shapely.geometry import Point
        
        print("="*70)
        print("🔍 DETAILED DIAGNOSTIC FOR SPECIFIC GPS POINTS")
        print("="*70)
        
        if self.road_graph is None:
            print("❌ No road_graph found!")
            return
        
        print(f"\n📊 Road Graph: {self.road_graph.number_of_nodes()} nodes, {self.road_graph.number_of_edges()} edges")
        print(f"📊 Driven Edges: {len(self.driven_edges)} edges")
        
        points = [
            ("Point 1", lat1, lon1),
            ("Point 2", lat2, lon2)
        ]
        
        for label, lat, lon in points:
            print(f"\n{'='*70}")
            print(f"📍 {label}: ({lat:.6f}, {lon:.6f})")
            print(f"{'='*70}")
            
            # Find all edges within 100m
            print(f"\n🔍 Finding all road edges within 100m...")
            
            candidates = []
            gps_point = Point(lon, lat)
            
            for u, v, key, data in self.road_graph.edges(keys=True, data=True):
                edge = (u, v, key)
                
                # Get edge geometry
                edge_geom = self._get_edge_geometry_fast(edge)
                if edge_geom is None:
                    continue
                
                # Calculate distance
                dist_deg = gps_point.distance(edge_geom)
                dist_m = dist_deg * 111000  # Approximate conversion
                
                if dist_m < 100:
                    candidates.append({
                        'edge': edge,
                        'distance': dist_m,
                        'name': data.get('name', ''),
                        'osmid': data.get('osmid', 'N/A'),
                        'highway': data.get('highway', 'N/A'),
                        'in_driven': edge in self.driven_edges
                    })
            
            # Sort by distance
            candidates.sort(key=lambda x: x['distance'])
            
            print(f"\n   Found {len(candidates)} candidate edges:\n")
            print(f"   {'#':<3} {'Dist':<8} {'Name':<30} {'OSM ID':<12} {'Type':<12} {'Driven?':<8}")
            print(f"   {'-'*3} {'-'*8} {'-'*30} {'-'*12} {'-'*12} {'-'*8}")
            
            for j, cand in enumerate(candidates[:15]):  # Show top 15
                name = cand['name'][:28] if cand['name'] else '(unnamed)'
                driven = "✅ YES" if cand['in_driven'] else "❌ NO"
                print(f"   {j+1:<3} {cand['distance']:<8.1f} {name:<30} {cand['osmid']:<12} {cand['highway']:<12} {driven}")
            
            # Highlight the issue
            driven_candidates = [c for c in candidates if c['in_driven']]
            not_driven = [c for c in candidates if not c['in_driven']]
            
            print(f"\n   📊 Summary:")
            print(f"      Edges in driven set: {len(driven_candidates)}")
            print(f"      Edges NOT in driven set: {len(not_driven)}")
            
            if driven_candidates:
                closest_driven = driven_candidates[0]
                print(f"\n   ➡️  Closest DRIVEN edge: '{closest_driven['name']}' at {closest_driven['distance']:.1f}m")
            
            if not_driven and candidates:
                closest_overall = candidates[0]
                if not closest_overall['in_driven']:
                    print(f"   ⚠️  Closest OVERALL edge: '{closest_overall['name']}' at {closest_overall['distance']:.1f}m (NOT DRIVEN)")
                    print(f"   ⚠️  THIS MAY BE THE CORRECT ROAD THAT WAS MISSED!")
        
        # Check distance between the two points
        print(f"\n{'='*70}")
        print("📏 GPS POINTS ANALYSIS")
        print(f"{'='*70}")
        
        dist_between = geodesic((lat1, lon1), (lat2, lon2)).meters
        print(f"\n   Distance between Point 1 and Point 2: {dist_between:.1f}m")
        
        print(f"\n{'='*70}")
        print("END DIAGNOSTIC")
        print(f"{'='*70}")        


    
    def _find_candidates_fallback(self, lon, lat, max_candidates, search_radius):
        """
        Fallback candidate finding without spatial index (slower).
        Uses single nearest_edges call + spatial search.
        """
        import osmnx as ox

        candidates = []
        seen_edges = set()

        # Get the nearest edge (single call, not redundant)
        nearest_edge, nearest_dist = ox.distance.nearest_edges(
            self.road_graph, lon, lat, return_dist=True
        )

        if nearest_dist <= search_radius:
            candidates.append((nearest_edge, nearest_dist))
            seen_edges.add(nearest_edge)

        # Find additional nearby candidates
        additional = self._get_edges_near_point(lon, lat, max_dist=search_radius)
        for edge, dist in additional:
            if edge not in seen_edges and len(candidates) < max_candidates:
                candidates.append((edge, dist))
                seen_edges.add(edge)

        # If no candidates within radius, use nearest edge anyway
        if not candidates:
            candidates = [(nearest_edge, nearest_dist)]

        return candidates

    
    def _viterbi_map_matching(self, X, Y, candidates_per_point, df):
        """
        IMPROVED Viterbi algorithm with road continuity bias
        
        Key improvement: When GPS is sparse, prefer staying on the same road
        (same OSM ID) over jumping to parallel/adjacent roads.
        """

    
        n_points = len(X)
        if n_points == 0:
            return []
    
        # Viterbi tables
        V = [{} for _ in range(n_points)]
    
        # Initialize first observation
        first_candidates = candidates_per_point[0]
        if not first_candidates:
            return [None] * n_points
    
        for cand_idx, (edge, dist) in enumerate(first_candidates):
            emission_prob = self._emission_probability(dist)
            V[0][cand_idx] = (math.log(emission_prob), None)
    
        # Forward pass
        for t in range(1, n_points):
            curr_candidates = candidates_per_point[t]
            prev_candidates = candidates_per_point[t-1]
    
            if not curr_candidates:
                V[t] = {}
                continue
    
            if not prev_candidates:
                for cand_idx, (edge, dist) in enumerate(curr_candidates):
                    emission_prob = self._emission_probability(dist)
                    V[t][cand_idx] = (math.log(emission_prob), None)
                continue
    
            # Calculate GPS distance for this step
            from geopy.distance import geodesic
            gps_distance = geodesic(
                (Y[t-1], X[t-1]), (Y[t], X[t])
            ).meters
    
            # For each current candidate
            for curr_cand_idx, (curr_edge, curr_dist) in enumerate(curr_candidates):
                emission_prob = self._emission_probability(curr_dist)
                emission_log_prob = math.log(emission_prob)
    
                best_prob = -float('inf')
                best_prev_idx = None
    
                # Get current edge's road ID (OSM way ID)
                curr_road_id = self._get_osmid_for_edge(curr_edge)
    
                # Find best previous candidate
                for prev_cand_idx, (prev_edge, prev_dist) in enumerate(prev_candidates):
                    if prev_cand_idx not in V[t-1]:
                        continue
    
                    prev_prob = V[t-1][prev_cand_idx][0]

                    # NEW: Set distance for _transition_probability to use
                    self._current_candidate_distance = curr_dist
        
                    # Base transition probability
                    transition_prob = self._transition_probability(
                        prev_edge, curr_edge,
                        (X[t-1], Y[t-1]), (X[t], Y[t]),
                        df.iloc[t-1]['timestamp'], df.iloc[t]['timestamp']
                    )


                    if curr_dist < self.distance_priority_thresholds['very_close']:
                        if transition_prob < 0.1:
                            transition_prob = self.distance_priority_overrides['very_close']
                            
                    elif curr_dist < self.distance_priority_thresholds['close']:
                        if transition_prob < 0.1:
                            transition_prob = self.distance_priority_overrides['close']
                    elif curr_dist < self.distance_priority_thresholds['plausible']:
                        if transition_prob < 0.05:
                            transition_prob = self.distance_priority_overrides['plausible']
    
                    # ============ KEY IMPROVEMENT: ROAD CONTINUITY BIAS ============
                    # If GPS points are far apart (sparse data), add strong bias 
                    # to stay on the same road
                    
                    prev_road_id = self._get_osmid_for_edge(prev_edge)
                    
                    if curr_road_id and prev_road_id and curr_road_id == prev_road_id:
                        # Same road - apply continuity bonus
                        # Stronger bonus for larger GPS gaps (more uncertainty)
                        if gps_distance > 100:
                            continuity_bonus = 5.0  # Very strong for sparse GPS
                        elif gps_distance > 50:
                            continuity_bonus = 3.0  # Strong
                        else:
                            continuity_bonus = 1.   # Moderate
                        
                        transition_prob = min(transition_prob * continuity_bonus, 0.95)
                    
                    # ============ END IMPROVEMENT ============
    
                    transition_log_prob = math.log(max(transition_prob, 1e-15))
                    total_prob = prev_prob + transition_log_prob + emission_log_prob
    
                    if total_prob > best_prob:
                        best_prob = total_prob
                        best_prev_idx = prev_cand_idx
    
                V[t][curr_cand_idx] = (best_prob, best_prev_idx)
    
        # Backward pass - reconstruct best path
        best_path_indices = [None] * n_points
    
        if V[n_points-1]:
            best_final_idx = max(V[n_points-1].keys(),
                                key=lambda idx: V[n_points-1][idx][0])
            best_path_indices[n_points-1] = best_final_idx
    
            for t in range(n_points-2, -1, -1):
                if best_path_indices[t+1] is not None and V[t+1]:
                    prev_idx = V[t+1][best_path_indices[t+1]][1]
                    best_path_indices[t] = prev_idx
                else:
                    if V[t]:
                        best_path_indices[t] = max(V[t].keys(),
                                                  key=lambda idx: V[t][idx][0])
    
        # Convert indices to edges
        matched_edges = []
        for t in range(n_points):
            if best_path_indices[t] is not None and candidates_per_point[t]:
                try:
                    edge = candidates_per_point[t][best_path_indices[t]][0]
                    matched_edges.append(edge)
                except IndexError:
                    matched_edges.append(None)
            else:
                matched_edges.append(None)
    
        return matched_edges

    def debug_viterbi_matching_sequence(self, coordinates):
        """
        Debug the Viterbi map matching decision for a sequence of GPS points.
        Shows exactly what candidates are found and how decisions are made.
        
        coordinates: list of (lat, lon) tuples
        """
        from geopy.distance import geodesic
        from shapely.geometry import Point
        
        print("="*80)
        print("🔍 VITERBI MAP MATCHING DEBUG - STEP BY STEP ANALYSIS")
        print("="*80)
        
        if self.road_graph is None:
            print("❌ No road graph loaded!")
            return
        
        print(f"\n📊 Road Graph: {self.road_graph.number_of_nodes()} nodes, {self.road_graph.number_of_edges()} edges")
        print(f"📍 Analyzing {len(coordinates)} GPS points\n")
        
        prev_best_edge = None
        prev_best_name = None
        
        for i, (lat, lon) in enumerate(coordinates):
            print(f"\n{'='*80}")
            print(f"📍 POINT {i}: ({lat:.6f}, {lon:.6f})")
            print(f"{'='*80}")
            
            # Find all candidate edges within 100m
            gps_point = Point(lon, lat)
            candidates = []
            
            for u, v, key, data in self.road_graph.edges(keys=True, data=True):
                edge = (u, v, key)
                edge_geom = self._get_edge_geometry_fast(edge)
                if edge_geom is None:
                    continue
                
                dist_m = gps_point.distance(edge_geom) * 111000
                
                if dist_m < 100:
                    # Check connectivity to previous edge
                    is_connected = False
                    if prev_best_edge:
                        prev_u, prev_v, prev_k = prev_best_edge
                        is_connected = (prev_v == u) or (prev_v == v) or (prev_u == u) or (prev_u == v)
                    
                    # Calculate emission probability
                    emission_prob = self._emission_probability(dist_m)
                    
                    candidates.append({
                        'edge': edge,
                        'distance': dist_m,
                        'name': data.get('name', ''),
                        'osmid': data.get('osmid'),
                        'highway': data.get('highway', ''),
                        'emission_prob': emission_prob,
                        'connected_to_prev': is_connected
                    })
            
            # Sort by distance
            candidates.sort(key=lambda x: x['distance'])
            
            if not candidates:
                print("   ❌ NO CANDIDATES FOUND WITHIN 100m!")
                prev_best_edge = None
                prev_best_name = None
                continue
            
            print(f"\n   📋 Found {len(candidates)} candidate edges:")
            print(f"   {'#':<3} {'Dist':<8} {'Emission':<10} {'Connected':<10} {'Name':<25} {'Type':<12} {'OSM ID'}")
            print(f"   {'-'*3} {'-'*8} {'-'*10} {'-'*10} {'-'*25} {'-'*12} {'-'*12}")
            
            for j, cand in enumerate(candidates[:10]):  # Show top 10
                conn_str = "✅ YES" if cand['connected_to_prev'] else "❌ NO"
                name = cand['name'][:24] if cand['name'] else '(unnamed)'
                print(f"   {j+1:<3} {cand['distance']:<8.1f} {cand['emission_prob']:<10.6f} {conn_str:<10} {name:<25} {cand['highway']:<12} {cand['osmid']}")
            
            # Analyze the decision
            closest = candidates[0]
            closest_connected = next((c for c in candidates if c['connected_to_prev']), None)
            
            print(f"\n   🎯 DECISION ANALYSIS:")
            print(f"      Closest edge: '{closest['name']}' at {closest['distance']:.1f}m")
            
            if closest_connected and closest_connected != closest:
                print(f"      Closest CONNECTED edge: '{closest_connected['name']}' at {closest_connected['distance']:.1f}m")
                
                # This is where the problem happens!
                if closest_connected['distance'] > closest['distance'] + 5:
                    print(f"\n      ⚠️  POTENTIAL PROBLEM DETECTED!")
                    print(f"         The closest edge ({closest['distance']:.1f}m) is NOT connected to previous")
                    print(f"         The closest CONNECTED edge is {closest_connected['distance']:.1f}m away")
                    print(f"         Difference: {closest_connected['distance'] - closest['distance']:.1f}m")
                    
                    if closest_connected['distance'] > closest['distance'] * 2:
                        print(f"         🚨 VITERBI WILL LIKELY CHOOSE THE WRONG ROAD!")
            
            # Simulate what Viterbi would choose
            # In reality, Viterbi considers the full sequence, but we can estimate
            if prev_best_edge:
                # Calculate transition probabilities
                print(f"\n   🔄 TRANSITION FROM PREVIOUS EDGE:")
                print(f"      Previous: '{prev_best_name}'")
                
                best_combined_score = 0
                best_choice = None
                
                for cand in candidates[:5]:
                    emission = cand['emission_prob']
                    
                    if cand['connected_to_prev']:
                        transition = 0.9  # High for connected
                    else:
                        transition = 1e-15  # Very low for disconnected
                    
                    combined = emission * transition
                    
                    if combined > best_combined_score:
                        best_combined_score = combined
                        best_choice = cand
                    
                    conn_str = "CONNECTED" if cand['connected_to_prev'] else "disconnected"
                    print(f"      → '{cand['name'][:20]}': emission={emission:.6f} × transition={transition:.2e} = {combined:.2e} ({conn_str})")
                
                if best_choice:
                    print(f"\n   ➡️  VITERBI WOULD CHOOSE: '{best_choice['name']}' at {best_choice['distance']:.1f}m")
                    
                    if best_choice != closest:
                        print(f"   ⚠️  THIS IS NOT THE CLOSEST EDGE! (closest was '{closest['name']}' at {closest['distance']:.1f}m)")
                    
                    prev_best_edge = best_choice['edge']
                    prev_best_name = best_choice['name']
                else:
                    prev_best_edge = closest['edge']
                    prev_best_name = closest['name']
            else:
                # First point - just pick closest
                print(f"\n   ➡️  FIRST POINT - Selecting closest: '{closest['name']}'")
                prev_best_edge = closest['edge']
                prev_best_name = closest['name']
            
            # Calculate GPS distance from previous point
            if i > 0:
                prev_lat, prev_lon = coordinates[i-1]
                gps_dist = geodesic((prev_lat, prev_lon), (lat, lon)).meters
                print(f"\n   📏 GPS distance from previous point: {gps_dist:.1f}m")
        
        print(f"\n{'='*80}")
        print("END DEBUG")
        print(f"{'='*80}")

    
    def _get_osmid_for_edge(self, edge):
        """Get OSM way ID for an edge"""
        if not edge:
            return None
        try:
            edge_data = self.road_graph.edges[edge]
            osmid = edge_data.get('osmid')
            if isinstance(osmid, list):
                return osmid[0] if osmid else None
            return osmid
        except:
            return None

        
        
    def _bridge_gap_with_shortest_path(self, prev_edge, curr_edge, prev_coord=None, curr_coord=None):
        """
        FIXED: Direction-aware bridge path - rejects paths going backwards
        """
        import networkx as nx
        from geopy.distance import geodesic
    
        try:
            u1, v1, k1 = prev_edge
            u2, v2, k2 = curr_edge
    
            # Calculate travel bearing from GPS coordinates
            travel_bearing = None
            if prev_coord and curr_coord:
                travel_bearing = self._calculate_bearing(prev_coord, curr_coord)
    
            # Try different node combinations
            node_pairs = [(v1, u2), (v1, v2), (u1, u2), (u1, v2)]
            
            best_path = None
            best_length = float('inf')
            best_bearing_diff = 180
    
            for source, target in node_pairs:
                if source == target:
                    continue
    
                try:
                    path_nodes = nx.shortest_path(
                        self.road_graph, source=source, target=target, weight='length'
                    )
                    
                    path_length = nx.shortest_path_length(
                        self.road_graph, source=source, target=target, weight='length'
                    )
    
                    # Calculate path bearing and check direction
                    if len(path_nodes) >= 2 and travel_bearing is not None:
                        first_node = path_nodes[0]
                        last_node = path_nodes[-1]
                        
                        first_coord = (
                            self.road_graph.nodes[first_node]['y'],
                            self.road_graph.nodes[first_node]['x']
                        )
                        last_coord = (
                            self.road_graph.nodes[last_node]['y'],
                            self.road_graph.nodes[last_node]['x']
                        )
                        
                        path_bearing = self._calculate_bearing(first_coord, last_coord)
                        
                        # Calculate bearing difference
                        bearing_diff = abs(travel_bearing - path_bearing)
                        if bearing_diff > 180:
                            bearing_diff = 360 - bearing_diff
                        
                        # CRITICAL FIX: Reject paths that go backwards (> 90 degrees off)
                        if bearing_diff > 90:
                            continue  # Skip this path - wrong direction!
                        
                        # Prefer paths that are shorter AND more aligned with travel direction
                        # Score combines length and direction (lower is better)
                        score = path_length + (bearing_diff * 2)  # Penalize direction mismatch
                        
                        if bearing_diff < best_bearing_diff or (bearing_diff == best_bearing_diff and path_length < best_length):
                            best_path = path_nodes
                            best_length = path_length
                            best_bearing_diff = bearing_diff
                    else:
                        # No bearing info - just use shortest path
                        if path_length < best_length:
                            best_path = path_nodes
                            best_length = path_length
    
                except (nx.NetworkXNoPath, nx.NodeNotFound):
                    continue
    
            # Build bridge coordinates from best path
            if best_path and len(best_path) > 1:
                bridge_coords = []
                
                for i in range(len(best_path) - 1):
                    node_u, node_v = best_path[i], best_path[i + 1]
                    
                    bridge_edge = None
                    if self.road_graph.has_edge(node_u, node_v, 0):
                        bridge_edge = (node_u, node_v, 0)
                    else:
                        for key in self.road_graph[node_u][node_v]:
                            bridge_edge = (node_u, node_v, key)
                            break
    
                    if bridge_edge:
                        edge_geom = self._get_edge_geometry_fast(bridge_edge)
                        if edge_geom:
                            for coord in edge_geom.coords:
                                lat, lon = coord[1], coord[0]
                                if not bridge_coords or bridge_coords[-1] != (lat, lon):
                                    bridge_coords.append((lat, lon))
    
                return bridge_coords if bridge_coords else None
            
            return None
    
        except Exception as e:
            return None        


    def _fix_reverse_edge_zigzag(self, matched_edges):
        """
        Post-process to fix zig-zag caused by alternating between reverse edges.
        If consecutive edges are reverse of each other (same road segment, opposite direction),
        keep consistent direction.
        """
        if len(matched_edges) < 2:
            return matched_edges
        
        fixed = matched_edges.copy()
        fixes_made = 0
        
        for i in range(1, len(fixed)):
            prev = fixed[i-1]
            curr = fixed[i]
            
            if prev and curr:
                u1, v1, k1 = prev
                u2, v2, k2 = curr
                
                # If current is reverse of previous (same segment, opposite direction)
                if u1 == v2 and v1 == u2:
                    fixed[i] = prev  # Keep same direction as previous
                    fixes_made += 1
        
        if fixes_made > 0:
            print(f"  ✅ Fixed {fixes_made} reverse-edge zig-zags")
        
        return fixed
    
        
    def _reconstruct_route_geometry(self, X, Y, matched_edges):
        """
        Reconstruct route geometry with gap bridging
        """
        from shapely.geometry import Point, LineString
        from geopy.distance import geodesic
        import networkx as nx
    
        if not matched_edges or len(matched_edges) == 0:
            return []
    
        route_coords = []
        prev_edge = None
        gap_count = 0
        bridge_count = 0
    
        print("\n🔧 DEBUG: Route Geometry Reconstruction (Point-by-Point Projection)")
        print(f"   Total GPS points: {len(matched_edges)}")
        print(f"   Unique edges: {len(set(e for e in matched_edges if e is not None))}")
    
        for i, curr_edge in enumerate(matched_edges):
            if curr_edge is None:
                route_coords.append((Y[i], X[i]))
                prev_edge = None
                continue
    
            edge_geom = self._get_edge_geometry_fast(curr_edge)
            if edge_geom is None:
                route_coords.append((Y[i], X[i]))
                prev_edge = curr_edge
                continue
    
            # Check for gap with previous edge
            if prev_edge is not None and prev_edge != curr_edge:
                u1, v1, k1 = prev_edge
                u2, v2, k2 = curr_edge
                is_connected = (v1 == u2) or (v1 == v2) or (u1 == u2) or (u1 == v2)
    
                if not is_connected and route_coords:
                    gap_count += 1
                    prev_coord = route_coords[-1]
                    curr_coord = (Y[i], X[i])
                    gap_dist = geodesic(prev_coord, curr_coord).meters
    
                    if gap_dist > 10.0:
                        #print(f"   ⚠️  GAP {gap_count}: Checking bridge for {gap_dist:.1f}m gap...")
                        
                        # Try to bridge with shortest path
                        try:
                            path_length = nx.shortest_path_length(
                                self.road_graph, v1, u2, weight='length'
                            )
                            
                            # Only bridge if path is reasonable (< 2x GPS distance)
                            if path_length < gap_dist * 2.0:
                                bridge_coords = self._bridge_gap_with_shortest_path(
                                    prev_edge, curr_edge, prev_coord, curr_coord
                                )
                                if bridge_coords:
                                    coords_added = 0
                                    for coord in bridge_coords:
                                        if not route_coords or geodesic(route_coords[-1], coord).meters > 1.0:
                                            route_coords.append(coord)
                                            coords_added += 1
                                    if coords_added > 0:
                                        bridge_count += 1
                                        #print(f"      ✅ Bridged with {coords_added} points (path: {path_length:.0f}m)")
                            else:
                                print(f"      ⚠️  Path too long ({path_length:.0f}m > {gap_dist*2:.0f}m), skipping bridge")
                            
                        except (nx.NetworkXNoPath, nx.NodeNotFound):
                            print(f"      ⚠️  No path found, skipping bridge")
    
            # Project GPS point onto matched edge
            try:
                gps_point = Point(X[i], Y[i])
                projected = edge_geom.interpolate(edge_geom.project(gps_point))
                lat, lon = projected.y, projected.x
    
                if not route_coords or geodesic(route_coords[-1], (lat, lon)).meters > 1.0:
                    route_coords.append((lat, lon))
    
            except Exception as e:
                route_coords.append((Y[i], X[i]))
    
            prev_edge = curr_edge
    
        # Simplify route geometry
        if len(route_coords) > 2:
            try:
                line = LineString([(lon, lat) for lat, lon in route_coords])
                simplified = line.simplify(0.00001, preserve_topology=True)
                route_coords = [(p[1], p[0]) for p in simplified.coords]
                print(f"   🔄 Simplified from {len(line.coords)} to {len(route_coords)} points")
            except Exception as e:
                print(f"   ⚠️  Simplification failed: {e}")
    
        print(f"\n   📊 Summary:")
        print(f"      Total GPS points processed: {len(matched_edges)}")
        print(f"      Total route coordinates: {len(route_coords)}")
        print(f"      Gaps detected: {gap_count}")
        print(f"      Gaps bridged: {bridge_count}")
        if gap_count > bridge_count:
            print(f"      ⚠️  {gap_count - bridge_count} gaps not bridged (path too long or no path)")
    
        return route_coords
    def analyze_road_distribution(self):
        """Analyze where roads actually exist in the graph"""
        
        print(f"\n🔍 ANALYZING ROAD DISTRIBUTION IN GRAPH")
        print("="*60)
        
        # Collect all node coordinates
        lats = []
        lons = []
        for node, data in self.road_graph.nodes(data=True):
            if 'y' in data and 'x' in data:
                lats.append(data['y'])
                lons.append(data['x'])
        
        print(f"\n📊 Total nodes: {len(lats)}")
        
        # Create latitude bands
        lat_min, lat_max = min(lats), max(lats)
        band_size = 0.01  # ~1.1 km bands
        
        print(f"\n📊 Road density by latitude band (~1.1km each):")
        print(f"{'Latitude Range':<25} {'Nodes':<10} {'Density'}")
        print(f"{'-'*25} {'-'*10} {'-'*20}")
        
        current_lat = lat_min
        while current_lat < lat_max:
            band_end = current_lat + band_size
            count = sum(1 for lat in lats if current_lat <= lat < band_end)
            bar = '█' * (count // 100) if count > 0 else '░'
            print(f"{current_lat:.4f} - {band_end:.4f}    {count:<10} {bar}")
            current_lat = band_end
        
        # Check specific area around problematic point
        test_lat, test_lon = 48.810827, 7.583511
        nearby_radius = 0.01  # ~1.1 km
        
        nearby_nodes = sum(1 for lat, lon in zip(lats, lons) 
                           if abs(lat - test_lat) < nearby_radius and abs(lon - test_lon) < nearby_radius)
        
        print(f"\n📍 Nodes within 1km of test point ({test_lat}, {test_lon}): {nearby_nodes}")
        
        if nearby_nodes == 0:
            print("   ⚠️  NO ROAD DATA EXISTS IN THIS AREA!")


    def diagnose_connectivity_at_point(self, lat, lon):
        """
        Check why a closer road was not selected - analyze graph connectivity
        """
        from shapely.geometry import Point
        
        print("="*70)
        print(f"🔍 CONNECTIVITY DIAGNOSTIC AT ({lat}, {lon})")
        print("="*70)
        
        gps_point = Point(lon, lat)
        
        # Find all edges within 50m
        candidates = []
        for u, v, key, data in self.road_graph.edges(keys=True, data=True):
            edge = (u, v, key)
            edge_geom = self._get_edge_geometry_fast(edge)
            if edge_geom is None:
                continue
            dist_m = gps_point.distance(edge_geom) * 111000
            if dist_m < 50:
                candidates.append({
                    'edge': edge,
                    'u': u,
                    'v': v,
                    'distance': dist_m,
                    'name': data.get('name', ''),
                    'osmid': data.get('osmid'),
                    'in_driven': edge in self.driven_edges
                })
        
        candidates.sort(key=lambda x: x['distance'])
        
        # Get the closest edge and the closest DRIVEN edge
        closest = candidates[0] if candidates else None
        closest_driven = next((c for c in candidates if c['in_driven']), None)
        
        if not closest or not closest_driven:
            print("❌ Could not find edges for comparison")
            return
        
        print(f"\n📍 CLOSEST EDGE (should be selected):")
        print(f"   Name: '{closest['name']}'")
        print(f"   Distance: {closest['distance']:.1f}m")
        print(f"   Edge: {closest['edge']}")
        print(f"   Nodes: {closest['u']} → {closest['v']}")
        print(f"   In Driven Set: {'✅ YES' if closest['in_driven'] else '❌ NO'}")
        
        print(f"\n📍 CLOSEST DRIVEN EDGE (was selected):")
        print(f"   Name: '{closest_driven['name']}'")
        print(f"   Distance: {closest_driven['distance']:.1f}m")
        print(f"   Edge: {closest_driven['edge']}")
        print(f"   Nodes: {closest_driven['u']} → {closest_driven['v']}")
        
        # Check if they share any nodes (direct connectivity)
        shared_nodes = set([closest['u'], closest['v']]) & set([closest_driven['u'], closest_driven['v']])
        print(f"\n🔗 DIRECT CONNECTIVITY:")
        if shared_nodes:
            print(f"   ✅ Share nodes: {shared_nodes}")
        else:
            print(f"   ❌ No shared nodes - NOT directly connected")
        
        # Check if there's a path between them
        print(f"\n🛤️  PATH ANALYSIS:")
        import networkx as nx
        
        # Try to find path from closest to driven
        for start_node in [closest['u'], closest['v']]:
            for end_node in [closest_driven['u'], closest_driven['v']]:
                try:
                    path_length = nx.shortest_path_length(
                        self.road_graph, start_node, end_node, weight='length'
                    )
                    path = nx.shortest_path(self.road_graph, start_node, end_node)
                    print(f"   Path {start_node} → {end_node}: {path_length:.1f}m ({len(path)} nodes)")
                except nx.NetworkXNoPath:
                    print(f"   Path {start_node} → {end_node}: ❌ NO PATH EXISTS")
                except nx.NodeNotFound as e:
                    print(f"   Path {start_node} → {end_node}: ❌ Node not found: {e}")
        
        # Check what edges connect TO the closest edge
        print(f"\n🔀 EDGES CONNECTING TO CLOSEST EDGE ('{closest['name']}'):")
        
        # Incoming edges to u
        incoming_to_u = list(self.road_graph.in_edges(closest['u'], data=True))
        print(f"   Incoming to node {closest['u']}: {len(incoming_to_u)} edges")
        for u_in, v_in, data in incoming_to_u[:5]:
            name = data.get('name', 'unnamed')
            in_driven = (u_in, v_in, 0) in self.driven_edges
            print(f"      ← '{name}' {'✅ DRIVEN' if in_driven else ''}")
        
        # Outgoing edges from v
        outgoing_from_v = list(self.road_graph.out_edges(closest['v'], data=True))
        print(f"   Outgoing from node {closest['v']}: {len(outgoing_from_v)} edges")
        for u_out, v_out, data in outgoing_from_v[:5]:
            name = data.get('name', 'unnamed')
            in_driven = (u_out, v_out, 0) in self.driven_edges
            print(f"      → '{name}' {'✅ DRIVEN' if in_driven else ''}")
        
        # Check what edges connect TO the driven edge
        print(f"\n🔀 EDGES CONNECTING TO DRIVEN EDGE ('{closest_driven['name']}'):")
        
        incoming_to_driven = list(self.road_graph.in_edges(closest_driven['u'], data=True))
        print(f"   Incoming to node {closest_driven['u']}: {len(incoming_to_driven)} edges")
        for u_in, v_in, data in incoming_to_driven[:5]:
            name = data.get('name', 'unnamed')
            in_driven = (u_in, v_in, 0) in self.driven_edges
            print(f"      ← '{name}' {'✅ DRIVEN' if in_driven else ''}")
        
        print(f"\n{'='*70}")
        print("END CONNECTIVITY DIAGNOSTIC")
        print(f"{'='*70}")
    
    def diagnose_node_connectivity(self, node_id):
        """Check what's happening at a specific node"""
        print(f"\n🔍 NODE CONNECTIVITY: {node_id}")
        print("="*50)
        
        if node_id not in self.road_graph:
            print(f"❌ Node {node_id} not in graph!")
            return
        
        # Get node position
        node_data = self.road_graph.nodes[node_id]
        lat = node_data.get('y', 'N/A')
        lon = node_data.get('x', 'N/A')
        print(f"📍 Position: ({lat}, {lon})")
        
        # Incoming edges
        incoming = list(self.road_graph.in_edges(node_id, data=True))
        print(f"\n⬅️  INCOMING EDGES ({len(incoming)}):")
        for u, v, data in incoming:
            name = data.get('name', 'unnamed')
            osmid = data.get('osmid', 'N/A')
            edge = (u, v, 0)
            driven = "✅ DRIVEN" if edge in self.driven_edges else ""
            print(f"   {u} → {v}: '{name}' (OSM:{osmid}) {driven}")
        
        # Outgoing edges
        outgoing = list(self.road_graph.out_edges(node_id, data=True))
        print(f"\n➡️  OUTGOING EDGES ({len(outgoing)}):")
        for u, v, data in outgoing:
            name = data.get('name', 'unnamed')
            osmid = data.get('osmid', 'N/A')
            edge = (u, v, 0)
            driven = "✅ DRIVEN" if edge in self.driven_edges else ""
            print(f"   {u} → {v}: '{name}' (OSM:{osmid}) {driven}")
        
        # Check degree
        in_deg = self.road_graph.in_degree(node_id)
        out_deg = self.road_graph.out_degree(node_id)
        print(f"\n📊 Degree: in={in_deg}, out={out_deg}, total={in_deg + out_deg}")
        
        if out_deg == 0:
            print("⚠️  WARNING: This is a DEAD END node (no outgoing edges)!")
        if in_deg == 0:
            print("⚠️  WARNING: This node has no incoming edges!")

    def diagnose_driven_edges_near_point(self, lat, lon, radius=100):
        """Show all DRIVEN edges near a point"""
        from shapely.geometry import Point
        
        print(f"\n🔍 DRIVEN EDGES NEAR ({lat}, {lon})")
        print("="*60)
        
        gps_point = Point(lon, lat)
        
        driven_nearby = []
        for edge in self.driven_edges:
            edge_geom = self._get_edge_geometry_fast(edge)
            if edge_geom is None:
                continue
            dist_m = gps_point.distance(edge_geom) * 111000
            if dist_m < radius:
                u, v, k = edge
                data = self.road_graph.edges[edge]
                driven_nearby.append({
                    'edge': edge,
                    'distance': dist_m,
                    'name': data.get('name', ''),
                    'osmid': data.get('osmid'),
                    'u': u,
                    'v': v
                })
        
        driven_nearby.sort(key=lambda x: x['distance'])
        
        print(f"\n📊 Found {len(driven_nearby)} DRIVEN edges within {radius}m:\n")
        print(f"{'#':<3} {'Dist':<8} {'Name':<25} {'OSM ID':<12} {'Edge Nodes'}")
        print(f"{'-'*3} {'-'*8} {'-'*25} {'-'*12} {'-'*30}")
        
        for i, e in enumerate(driven_nearby):
            print(f"{i+1:<3} {e['distance']:<8.1f} {e['name'][:24]:<25} {e['osmid']:<12} {e['u']} → {e['v']}")
        
        # Check if Rue Principale edges are in driven set
        print(f"\n🛣️  Checking 'Rue Principale' edges:")
        rue_principale_edges = [
            (1342273213, 292875390, 0),   # Incoming to intersection
            (292875390, 1727025617, 0),   # Through intersection
            (1727025617, 292875388, 0),   # Continuing south
        ]
        
        for edge in rue_principale_edges:
            in_driven = edge in self.driven_edges
            print(f"   {edge}: {'✅ DRIVEN' if in_driven else '❌ NOT DRIVEN'}")

    def check_graph_coverage_for_gps(self, week_label="Week 1"):
        """Check if road graph covers the GPS data"""
        
        print(f"\n🔍 CHECKING ROAD GRAPH COVERAGE FOR {week_label}")
        print("="*60)
        
        # Get GPS bounds
        df = self.weekly_data.get(week_label)
        if df is None:
            print("❌ No data for this week")
            return
        
        gps_min_lat, gps_max_lat = df['lat'].min(), df['lat'].max()
        gps_min_lon, gps_max_lon = df['lon'].min(), df['lon'].max()
        
        print(f"\n📍 GPS Data Bounds ({len(df)} points):")
        print(f"   Latitude:  {gps_min_lat:.6f} to {gps_max_lat:.6f}")
        print(f"   Longitude: {gps_min_lon:.6f} to {gps_max_lon:.6f}")
        
        # Get graph bounds
        all_lats = [data['y'] for _, data in self.road_graph.nodes(data=True) if 'y' in data]
        all_lons = [data['x'] for _, data in self.road_graph.nodes(data=True) if 'x' in data]
        
        graph_min_lat, graph_max_lat = min(all_lats), max(all_lats)
        graph_min_lon, graph_max_lon = min(all_lons), max(all_lons)
        
        print(f"\n🗺️  Road Graph Bounds ({self.road_graph.number_of_nodes()} nodes):")
        print(f"   Latitude:  {graph_min_lat:.6f} to {graph_max_lat:.6f}")
        print(f"   Longitude: {graph_min_lon:.6f} to {graph_max_lon:.6f}")
        
        # Check how many GPS points are within graph bounds
        in_bounds = df[
            (df['lat'] >= graph_min_lat) & (df['lat'] <= graph_max_lat) &
            (df['lon'] >= graph_min_lon) & (df['lon'] <= graph_max_lon)
        ]
        
        out_of_bounds = df[
            (df['lat'] < graph_min_lat) | (df['lat'] > graph_max_lat) |
            (df['lon'] < graph_min_lon) | (df['lon'] > graph_max_lon)
        ]
        
        print(f"\n📊 GPS Points Coverage:")
        print(f"   ✅ Within graph bounds: {len(in_bounds)} ({len(in_bounds)/len(df)*100:.1f}%)")
        print(f"   ❌ Outside graph bounds: {len(out_of_bounds)} ({len(out_of_bounds)/len(df)*100:.1f}%)")
        
        if len(out_of_bounds) > 0:
            print(f"\n⚠️  PROBLEM: {len(out_of_bounds)} GPS points are outside the road graph!")
            print(f"\n   Sample out-of-bounds points:")
            for idx, row in out_of_bounds.head(5).iterrows():
                print(f"      Index {idx}: ({row['lat']:.6f}, {row['lon']:.6f})")
            
            # Check the specific problematic point
            test_lat, test_lon = 48.810827, 7.583511
            is_in = (graph_min_lat <= test_lat <= graph_max_lat) and (graph_min_lon <= test_lon <= graph_max_lon)
            print(f"\n   Your test point ({test_lat}, {test_lon}):")
            print(f"      Within graph bounds: {'✅ YES' if is_in else '❌ NO'}")
    
    def find_nearest_edges_anywhere(self, lat, lon, max_results=10):
        """Find the nearest edges regardless of distance"""
        from shapely.geometry import Point
        
        print(f"\n🔍 FINDING NEAREST EDGES TO ({lat}, {lon})")
        print("="*60)
        
        gps_point = Point(lon, lat)
        
        all_edges = []
        for u, v, key, data in self.road_graph.edges(keys=True, data=True):
            edge = (u, v, key)
            edge_geom = self._get_edge_geometry_fast(edge)
            if edge_geom is None:
                continue
            dist_m = gps_point.distance(edge_geom) * 111000
            all_edges.append({
                'edge': edge,
                'distance': dist_m,
                'name': data.get('name', ''),
                'osmid': data.get('osmid')
            })
        
        all_edges.sort(key=lambda x: x['distance'])
        
        print(f"\n📊 Top {max_results} nearest edges (out of {len(all_edges)} total):\n")
        print(f"{'#':<3} {'Distance':<12} {'Name':<30} {'OSM ID'}")
        print(f"{'-'*3} {'-'*12} {'-'*30} {'-'*15}")
        
        for i, e in enumerate(all_edges[:max_results]):
            print(f"{i+1:<3} {e['distance']:<12.1f} {e['name'][:29]:<30} {e['osmid']}")
        
        if all_edges:
            closest = all_edges[0]
            print(f"\n📍 Closest edge is {closest['distance']:.1f}m away")
            print(f"   Name: '{closest['name']}'")
            print(f"   Edge: {closest['edge']}")
            
            # Get coordinates of closest edge
            u, v, k = closest['edge']
            u_data = self.road_graph.nodes[u]
            v_data = self.road_graph.nodes[v]
            print(f"   Start node: ({u_data['y']:.6f}, {u_data['x']:.6f})")
            print(f"   End node: ({v_data['y']:.6f}, {v_data['x']:.6f})")
    



    def find_wrong_turn_origin(self, target_osmid=29583561):
        """
        Find where a specific road (by OSM ID) first appears in driven edges
        """
        print(f"\n🔍 Finding first occurrence of OSM ID {target_osmid} in driven edges")
        print("="*60)
        
        matches = []
        for edge in self.driven_edges:
            data = self.road_graph.edges[edge]
            if data.get('osmid') == target_osmid:
                u, v, k = edge
                u_data = self.road_graph.nodes[u]
                v_data = self.road_graph.nodes[v]
                matches.append({
                    'edge': edge,
                    'name': data.get('name', ''),
                    'u_lat': u_data.get('y'),
                    'u_lon': u_data.get('x'),
                    'v_lat': v_data.get('y'),
                    'v_lon': v_data.get('x'),
                })
        
        print(f"\n📊 Found {len(matches)} driven edges with OSM ID {target_osmid}:")
        for i, m in enumerate(matches):
            print(f"\n   Edge {i+1}: {m['edge']}")
            print(f"   Name: '{m['name']}'")
            print(f"   Start: ({m['u_lat']:.6f}, {m['u_lon']:.6f})")
            print(f"   End:   ({m['v_lat']:.6f}, {m['v_lon']:.6f})")
        
    def debug_driven_edges_detail(self):
        """Debug the driven_edges set"""
        print(f"\n🔍 DEBUGGING DRIVEN EDGES SET")
        print("="*60)
        
        print(f"\n📊 Total driven edges: {len(self.driven_edges)}")
        print(f"📊 Type: {type(self.driven_edges)}")
        
        # Show first 10 driven edges
        print(f"\n📋 First 10 driven edges:")
        for i, edge in enumerate(list(self.driven_edges)[:10]):
            try:
                data = self.road_graph.edges[edge]
                name = data.get('name', 'unnamed')
                osmid = data.get('osmid', 'N/A')
                print(f"   {i+1}. {edge} - '{name}' (OSM: {osmid})")
            except Exception as e:
                print(f"   {i+1}. {edge} - Error: {e}")
        
        # Check the specific edges we found earlier
        print(f"\n🔍 Checking specific edges:")
        test_edges = [
            (4800949776, 4450438621, 0),
            (4450438621, 2100350117, 0),
        ]
        
        for edge in test_edges:
            in_set = edge in self.driven_edges
            print(f"   {edge}: {'✅ IN SET' if in_set else '❌ NOT IN SET'}")
            
            if in_set:
                try:
                    data = self.road_graph.edges[edge]
                    print(f"      Name: '{data.get('name', '')}'")
                    print(f"      OSM ID: {data.get('osmid')} (type: {type(data.get('osmid'))})")
                except:
                    pass
        
        # Count unique OSM IDs in driven edges
        print(f"\n📊 Unique road names in driven edges:")
        names = {}
        for edge in self.driven_edges:
            try:
                data = self.road_graph.edges[edge]
                name = data.get('name', 'unnamed')
                osmid = data.get('osmid', 'N/A')
                if name not in names:
                    names[name] = {'count': 0, 'osmid': osmid}
                names[name]['count'] += 1
            except:
                pass
        
        for name, info in sorted(names.items(), key=lambda x: -x[1]['count'])[:15]:
            print(f"   '{name}': {info['count']} edges (OSM: {info['osmid']})")
    
    def verify_edge_membership(self, lat, lon):
        """Verify exactly what's happening with edge detection"""
        from shapely.geometry import Point
        
        print(f"\n🔍 VERIFYING EDGE DETECTION AT ({lat}, {lon})")
        print("="*60)
        
        gps_point = Point(lon, lat)
        
        # Find ALL edges within 50m (not just driven)
        print(f"\n📊 All edges within 50m:")
        
        all_nearby = []
        for u, v, key, data in self.road_graph.edges(keys=True, data=True):
            edge = (u, v, key)
            edge_geom = self._get_edge_geometry_fast(edge)
            if edge_geom is None:
                continue
            dist_m = gps_point.distance(edge_geom) * 111000
            if dist_m < 50:
                # Check membership explicitly
                in_driven = edge in self.driven_edges
                
                all_nearby.append({
                    'edge': edge,
                    'distance': dist_m,
                    'name': data.get('name', ''),
                    'osmid': data.get('osmid'),
                    'in_driven': in_driven
                })
        
        all_nearby.sort(key=lambda x: x['distance'])
        
        print(f"\n{'#':<3} {'Dist':<8} {'Name':<25} {'In Driven?':<12} {'Edge'}")
        print(f"{'-'*3} {'-'*8} {'-'*25} {'-'*12} {'-'*40}")
        
        for i, e in enumerate(all_nearby[:20]):
            driven_str = "✅ YES" if e['in_driven'] else "❌ NO"
            print(f"{i+1:<3} {e['distance']:<8.1f} {e['name'][:24]:<25} {driven_str:<12} {e['edge']}")
        
        # Summary
        driven_count = sum(1 for e in all_nearby if e['in_driven'])
        print(f"\n📊 Summary: {driven_count} driven edges out of {len(all_nearby)} total within 50m")
        
        # Now check what the route_geometry looks like near this point
        print(f"\n🛤️  Checking route_geometry near this point:")
        if self.route_geometry is not None:
            from shapely.geometry import Point
            from shapely.ops import nearest_points
            
            nearest_pt = nearest_points(gps_point, self.route_geometry)[1]
            dist_to_route = gps_point.distance(nearest_pt) * 111000
            
            print(f"   Nearest point on route: ({nearest_pt.y:.6f}, {nearest_pt.x:.6f})")
            print(f"   Distance to route: {dist_to_route:.1f}m")
        else:
            print("   ❌ No route_geometry available")

    
    def _extract_edge_segment(self, edge_coords, start_dist, end_dist, edge_geom):
        """
        Extract a segment of an edge's geometry between two distances along the edge.

        Args:
            edge_coords: List of (x, y) coordinates for the edge
            start_dist: Distance along edge where segment starts
            end_dist: Distance along edge where segment ends
            edge_geom: Shapely LineString of the edge

        Returns:
            List of (x, y) coordinates for the segment
        """
        from shapely.geometry import Point

        if len(edge_coords) <= 2:
            # Simple edge - return all coordinates
            return edge_coords

        segment_coords = []
        cumulative_dist = 0.0

        for j in range(len(edge_coords) - 1):
            p1 = Point(edge_coords[j])
            p2 = Point(edge_coords[j + 1])
            segment_length = p1.distance(p2)
            next_cumulative = cumulative_dist + segment_length

            # Check if this segment overlaps with our desired range
            if next_cumulative >= start_dist and cumulative_dist <= end_dist:
                # Add first point if not already added
                if not segment_coords:
                    segment_coords.append(edge_coords[j])

                # Add second point
                segment_coords.append(edge_coords[j + 1])

            cumulative_dist = next_cumulative

            # Stop if we've passed the end
            if cumulative_dist > end_dist:
                break

        # Ensure we have at least some coordinates
        if not segment_coords:
            # Return interpolated points if extraction failed
            try:
                start_point = edge_geom.interpolate(start_dist)
                end_point = edge_geom.interpolate(end_dist)
                segment_coords = [(start_point.x, start_point.y), (end_point.x, end_point.y)]
            except:
                segment_coords = edge_coords

        return segment_coords

    def _emission_probability(self, distance_meters, sigma=10.0, gps_bearing=None,
                              edge=None, prev_gps=None, curr_gps=None):
        """
        Emission probability: P(GPS observation | road edge)

        Based on:
        1. GPS measurement noise (distance), modeled as Gaussian distribution
        2. Direction alignment (bearing difference) - prevents matching to wrong-direction roads

        sigma = GPS measurement standard deviation (typically 5-15 meters)
        """
        import math

        # Base probability from distance (Gaussian PDF)
        prob = math.exp(-0.5 * (distance_meters / sigma) ** 2)
        prob = prob / (sigma * math.sqrt(2 * math.pi))

        # FIX: Add bearing-aware penalty to prevent matching to wrong-direction roads
        # This is critical for roundabouts and parallel roads
        if gps_bearing is not None and edge is not None and prev_gps is not None:
            try:
                # Get edge geometry and calculate its bearing
                edge_geom = self._get_edge_geometry_fast(edge)
                if edge_geom and edge_geom.length > 0:
                    # Calculate edge bearing at the point closest to current GPS
                    from shapely.geometry import Point
                    gps_point = Point(curr_gps[0], curr_gps[1])
                    projection_distance = edge_geom.project(gps_point)

                    # Get two points on edge to calculate bearing
                    # Sample slightly ahead and behind the projection point
                    sample_dist = min(10 / 111000.0, edge_geom.length * 0.1)  # 10m or 10% of edge

                    point1 = edge_geom.interpolate(max(0, projection_distance - sample_dist))
                    point2 = edge_geom.interpolate(min(edge_geom.length, projection_distance + sample_dist))

                    # Calculate edge bearing
                    edge_bearing = self._calculate_bearing(
                        (point1.y, point1.x), (point2.y, point2.x)
                    )

                    # Calculate bearing difference (handle 0/360 wraparound)
                    bearing_diff = abs(gps_bearing - edge_bearing)
                    if bearing_diff > 180:
                        bearing_diff = 360 - bearing_diff

                    # Penalize if bearing difference is large (> 45 degrees)
                    if bearing_diff > 45:
                        # Strong penalty for opposite direction (> 135 degrees)
                        if bearing_diff > 135:
                            prob *= 0.01  # 99% penalty
                        else:
                            # Gradual penalty for moderate misalignment
                            penalty = 1.0 - (bearing_diff - 45) / 180.0
                            prob *= max(penalty, 0.1)
            except:
                # If bearing calculation fails, just use distance-based probability
                pass

        # Ensure minimum probability to avoid numerical issues
        return max(prob, 1e-10)

    def _calculate_bearing(self, point1, point2):
        """
        Calculate bearing (direction) from point1 to point2 in degrees (0-360).
        point1, point2 are (lat, lon) tuples.
        """
        import math

        lat1, lon1 = math.radians(point1[0]), math.radians(point1[1])
        lat2, lon2 = math.radians(point2[0]), math.radians(point2[1])

        d_lon = lon2 - lon1

        x = math.sin(d_lon) * math.cos(lat2)
        y = math.cos(lat1) * math.sin(lat2) - math.sin(lat1) * math.cos(lat2) * math.cos(d_lon)

        bearing = math.atan2(x, y)
        bearing = math.degrees(bearing)
        bearing = (bearing + 360) % 360

        return bearing

    """

    def _transition_probability(self, prev_edge, curr_edge, prev_gps, curr_gps, prev_time, curr_time):
        

        from geopy.distance import geodesic
        
        # Calculate GPS distance
        gps_dist = geodesic(prev_gps, curr_gps).meters
        
        u1, v1, k1 = prev_edge
        u2, v2, k2 = curr_edge
        
        # Check topological connectivity
        is_connected = (v1 == u2) or (v1 == v2) or (u1 == u2) or (u1 == v2)
        
        # Get edge metadata
        prev_data = self.road_graph.edges[prev_edge]
        curr_data = self.road_graph.edges[curr_edge]
        
        prev_osmid = prev_data.get('osmid')
        curr_osmid = curr_data.get('osmid')
        prev_name = prev_data.get('name')
        curr_name = curr_data.get('name')
        
        same_osmid = prev_osmid and curr_osmid and prev_osmid == curr_osmid
        same_name = prev_name and curr_name and prev_name == curr_name
        
        # Base probability by GPS distance
        if gps_dist < 50:
            if is_connected:
                base_prob = 0.9
            else:
                # FIXED: Heavy penalty for disconnected unless same road
                if same_osmid or same_name:
                    # Check if path exists and is reasonable
                    try:
                        path_length = nx.shortest_path_length(
                            self.road_graph, v1, u2, weight='length'
                        )
                        if path_length < gps_dist * 1.5:
                            base_prob = 0.7  # Reduced but acceptable
                        else:
                            base_prob = 1e-12  # Path too long
                    except (nx.NetworkXNoPath, nx.NodeNotFound):
                        base_prob = 1e-15  # No path
                else:
                    base_prob = 1e-15
        elif gps_dist < 150:
            if is_connected:
                base_prob = 0.8
            else:
                base_prob = 1e-12 if same_name else 1e-15
        elif gps_dist < 300:
            if is_connected:
                base_prob = 0.7
            else:
                # Check actual path for long distances
                try:
                    path_length = nx.shortest_path_length(
                        self.road_graph, v1, u2, weight='length'
                    )
                    if path_length < gps_dist * 2.0:
                        base_prob = 0.5
                    else:
                        base_prob = 1e-10
                except (nx.NetworkXNoPath, nx.NodeNotFound):
                    base_prob = 1e-12
        else:
            base_prob = 0.6 if is_connected else 0.01
        
        # Apply same-road bonus ONLY if connected or path exists
        if is_connected and (same_osmid or same_name):
            base_prob *= self.connectivity_bonus  # 3.0x
        
        return base_prob
    
    
    """
    
    def _transition_probability(self, prev_edge, curr_edge, prev_gps, curr_gps,
                                prev_time, curr_time, beta=0.05,
                                connectivity_bonus=3.0, disconnected_penalty=0.05):
        """
        FIXED: Check path connectivity for same-road edges, not just direct node sharing
        """
        try:
            import networkx as nx
            
            gps_distance = geodesic(
                (prev_gps[1], prev_gps[0]),
                (curr_gps[1], curr_gps[0])
            ).meters
    
            if prev_edge == curr_edge:
                return 0.95
    
            u1, v1, k1 = prev_edge
            u2, v2, k2 = curr_edge
            
            # Direct node sharing - definitely connected
            directly_connected = (v1 == u2) or (v1 == v2) or (u1 == u2) or (u1 == v2)
            
            if directly_connected:
                return 0.95
            
            # ============ SAME ROAD PATH CHECK ============
            # If same osmid, check if there's a short path between edges
            prev_osmid = self._get_osmid_for_edge(prev_edge)
            curr_osmid = self._get_osmid_for_edge(curr_edge)
            
            if prev_osmid and curr_osmid and prev_osmid == curr_osmid:
                # Same road - check path length
                try:
                    path_length = nx.shortest_path_length(
                        self.road_graph, v1, u2, weight='length'
                    )
                    # If path is reasonable (less than 2x GPS distance), treat as connected
                    if path_length < gps_distance * 2.0:
                        return 0.92  # High probability - same road, reasonable path
                    elif path_length < gps_distance * 3.0:
                        return 0.8   # Still good - same road
                except nx.NetworkXNoPath:
                    pass  # No path, fall through to distance-based logic
            # ============ END SAME ROAD PATH CHECK ============
            
            # Distance-based fallback
            curr_dist = getattr(self, '_current_candidate_distance', None)
            
            if curr_dist is not None:
                if curr_dist < 3.0:
                    return 0.9
                elif curr_dist < 7.0:
                    return 0.85
                elif curr_dist < 15.0:
                    return 0.7
                elif curr_dist < 25.0:
                    return 0.4
                else:
                    return 0.1
            
            return 0.1
    
        except Exception as e:
            return 0.1
            
    def _build_edge_spatial_index(self):
        """Build spatial index with proper geometry-to-edge mapping"""
        from shapely.strtree import STRtree
        from shapely.geometry import LineString
    
        if self.road_graph is None:
            return
    
        try:
            print("  🔍 Building spatial index for fast edge queries...")
            
            self._edge_geometries = []
            self._edge_keys_list = []  # Parallel list: same index = same edge
    
            for u, v, k, data in self.road_graph.edges(keys=True, data=True):
                edge_key = (u, v, k)
    
                if 'geometry' in data and data['geometry']:
                    geom = data['geometry']
                else:
                    u_coord = (self.road_graph.nodes[u]['x'], self.road_graph.nodes[u]['y'])
                    v_coord = (self.road_graph.nodes[v]['x'], self.road_graph.nodes[v]['y'])
                    geom = LineString([u_coord, v_coord])
    
                self._edge_geometries.append(geom)
                self._edge_keys_list.append(edge_key)
    
            self._edge_spatial_index = STRtree(self._edge_geometries)
            
            print(f"    ✅ Spatial index built for {len(self._edge_keys_list)} edges")
    
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"    ⚠️  Spatial index build failed: {e}")
            self._edge_spatial_index = None
    
    
    def _find_candidates_with_spatial_index(self, lon, lat, max_candidates, search_radius):
        """Fast candidate finding - use query_indices if available"""
        from shapely.geometry import Point
        
        point = Point(lon, lat)
        candidates = []
        seen_edges = set()
        
        buffer_deg = search_radius / 111000.0
        search_geom = point.buffer(buffer_deg)
        
        # Try to get indices directly (Shapely 2.0+)
        try:
            indices = self._edge_spatial_index.query(search_geom, predicate='intersects')
            for idx in indices:
                edge_key = self._edge_keys_list[idx]
                if edge_key in seen_edges:
                    continue
                
                geom = self._edge_geometries[idx]
                dist_m = point.distance(geom) * 111000
                
                if dist_m <= search_radius:
                    candidates.append((edge_key, dist_m))
                    seen_edges.add(edge_key)
        except TypeError:
            # Older Shapely - query returns geometries, need to find indices
            nearby_geoms = self._edge_spatial_index.query(search_geom)
            for geom in nearby_geoms:
                try:
                    idx = self._edge_geometries.index(geom)
                    edge_key = self._edge_keys_list[idx]
                    if edge_key in seen_edges:
                        continue
                    dist_m = point.distance(geom) * 111000
                    if dist_m <= search_radius:
                        candidates.append((edge_key, dist_m))
                        seen_edges.add(edge_key)
                except ValueError:
                    continue
        
        candidates.sort(key=lambda x: x[1])
        
        if not candidates:
            import osmnx as ox
            nearest_edge, nearest_dist = ox.distance.nearest_edges(
                self.road_graph, lon, lat, return_dist=True
            )
            candidates = [(nearest_edge, nearest_dist)]
        
        return candidates[:max_candidates]            
    
    def _load_cached_graph(self, area_hash):
        """Load cached road graph if available"""
        cache_file = os.path.join(self.cache_dir, f"road_graph_{area_hash}.pkl")
        
        if False:
            print("  📋 Loading cached road graph...")
            try:
                with open(cache_file, 'rb') as f:
                    graph_data = pickle.load(f)
                self.road_graph = graph_data['graph']
                # Restore any cached indices
                if 'geometry_cache' in graph_data:
                    self.geometry_cache = graph_data['geometry_cache']
                return True
            except Exception as e:
                print(f"  ⚠️  Cache load failed: {e}")
                return False
        return False
        
    def _select_best_edge(self, nearest_edge, current_road_id, previous_edge,
                          gps_point=None, previous_gps=None, gps_bearing=None):
        """
        Enhanced edge selection with routing logic to avoid adjacent road errors
        Uses spatial proximity + routing feasibility + bearing alignment
        """
        # If no previous edge context, use nearest
        if not previous_edge:
            return nearest_edge

        try:
            # Get candidate edges within reasonable distance (30m)
            if gps_point:
                lon, lat = gps_point
                candidates = self._get_edges_near_point(lon, lat, max_dist=30)
            else:
                candidates = [(nearest_edge, 0)]

            # If only one candidate or no valid candidates, use nearest
            if len(candidates) <= 1:
                return nearest_edge

            best_edge = None
            best_score = -float('inf')

            # Score each candidate
            for candidate_edge, spatial_dist in candidates[:5]:  # Top 5 closest
                score = 0

                # === 1. SPATIAL PROXIMITY COMPONENT (weight: -1.0 per meter) ===
                spatial_score = -spatial_dist * 1.0

                # === 2. ROUTING FEASIBILITY COMPONENT (weight: 50.0) ===
                routing_score = self._calculate_routing_score(
                    previous_edge, candidate_edge, previous_gps, gps_point
                )

                # === 3. BEARING ALIGNMENT COMPONENT (weight: 30.0) ===
                bearing_score = 0
                if gps_bearing is not None:
                    edge_bearing = self._get_edge_bearing(candidate_edge)
                    if edge_bearing is not None:
                        bearing_similarity = self._bearing_similarity(gps_bearing, edge_bearing)
                        bearing_score = bearing_similarity * 30.0

                # === 4. ROAD CONTINUITY BONUS (weight: 20.0) ===
                continuity_score = 0
                candidate_road_id = self._get_road_id(candidate_edge)
                if candidate_road_id and candidate_road_id == current_road_id:
                    continuity_score = 20.0

                # === 5. TOPOLOGICAL CONNECTIVITY BONUS (weight: 25.0) ===
                connectivity_score = 0
                if self._are_edges_connected(previous_edge, candidate_edge):
                    connectivity_score = 25.0

                # Combine scores
                score = (spatial_score + routing_score + bearing_score +
                        continuity_score + connectivity_score)

                if score > best_score:
                    best_score = score
                    best_edge = candidate_edge

            return best_edge if best_edge else nearest_edge

        except Exception as e:
            # Fallback to nearest on any error
            return nearest_edge

    def _calculate_routing_score(self, prev_edge, next_edge, prev_gps, next_gps):
        """
        Calculate routing feasibility score based on path distance vs GPS distance
        Returns positive score for realistic routes, negative for unrealistic detours
        """
        try:
            from geopy.distance import geodesic
            import networkx as nx

            if not prev_gps or not next_gps:
                return 0

            # Calculate great circle distance between GPS points
            gps_dist = geodesic((prev_gps[1], prev_gps[0]), (next_gps[1], next_gps[0])).meters

            # If GPS points are very close, routing is less critical
            if gps_dist < 10:
                return 10  # Small bonus for nearby points

            # Calculate shortest path distance on road network
            try:
                prev_u, prev_v, prev_k = prev_edge
                next_u, next_v, next_k = next_edge

                # Try to find shortest path between edges
                path_length = nx.shortest_path_length(
                    self.road_graph, prev_v, next_u, weight='length'
                )

                # Add the edge lengths themselves
                prev_edge_length = self.road_graph.edges[prev_edge].get('length', 0)
                next_edge_length = self.road_graph.edges[next_edge].get('length', 0)

                total_route_dist = path_length + prev_edge_length/2 + next_edge_length/2

            except (nx.NetworkXNoPath, nx.NodeNotFound):
                # No path exists - heavily penalize disconnected roads
                return -100

            # Calculate route distance to GPS distance ratio
            if gps_dist > 0:
                ratio = total_route_dist / gps_dist

                # Ideal ratio is close to 1.0 (route matches GPS trajectory)
                # Penalize large detours that would require unrealistic routing
                if 0.8 <= ratio <= 1.3:
                    # Good match - route distance matches GPS distance
                    score = 50 - abs(1.0 - ratio) * 50
                elif ratio < 0.8:
                    # Route is shorter than GPS distance (possible due to GPS noise)
                    score = 30 - abs(0.8 - ratio) * 30
                else:
                    # Route is much longer than GPS distance (unlikely path)
                    score = -50 * (ratio - 1.3)

                return max(score, -100)  # Cap minimum at -100

            return 0

        except Exception:
            return 0

    def _are_edges_connected(self, edge1, edge2):
        """Check if two edges are topologically connected (share a node)"""
        try:
            if not edge1 or not edge2:
                return False

            u1, v1, k1 = edge1
            u2, v2, k2 = edge2

            # Edges are connected if they share a node
            return v1 == u2 or u1 == u2 or v1 == v2 or u1 == v2

        except Exception:
            return False

    def _trajectory_smoothing(self, matched_edges, X, Y, window_size=3):
        """
        Post-process matched edges to fix isolated erroneous edge switches
        This removes "jumps" to adjacent roads that don't make sense in trajectory context
        """
        try:
            if len(matched_edges) < 3:
                return matched_edges

            smoothed = matched_edges.copy()
            corrections_made = 0

            # Sliding window to detect and fix isolated edge switches
            for i in range(1, len(matched_edges) - 1):
                if not matched_edges[i]:
                    continue

                # Get road IDs for context window
                prev_road = self._get_road_id(matched_edges[i-1]) if matched_edges[i-1] else None
                curr_road = self._get_road_id(matched_edges[i])
                next_road = self._get_road_id(matched_edges[i+1]) if matched_edges[i+1] else None

                # Pattern 1: Isolated single-point jump (prev == next, but curr is different)
                if prev_road and next_road and prev_road == next_road and curr_road != prev_road:
                    # Check if staying on previous road is geometrically feasible
                    if self._is_point_feasible_on_edge(X[i], Y[i], matched_edges[i-1], max_dist=30):
                        smoothed[i] = matched_edges[i-1]
                        corrections_made += 1
                        continue

                # Pattern 2: Brief 2-point excursion to adjacent road
                if i < len(matched_edges) - 2:
                    next_next_road = self._get_road_id(matched_edges[i+2]) if i+2 < len(matched_edges) and matched_edges[i+2] else None

                    if (prev_road and next_next_road and
                        prev_road == next_next_road and
                        curr_road != prev_road and
                        next_road == curr_road):
                        # Two-point excursion - check if both points can stay on main road
                        if (self._is_point_feasible_on_edge(X[i], Y[i], matched_edges[i-1], max_dist=30) and
                            self._is_point_feasible_on_edge(X[i+1], Y[i+1], matched_edges[i-1], max_dist=30)):
                            smoothed[i] = matched_edges[i-1]
                            smoothed[i+1] = matched_edges[i-1]
                            corrections_made += 2

            if corrections_made > 0:
                print(f"  ✅ Trajectory smoothing corrected {corrections_made} erroneous edge assignments")

            return smoothed

        except Exception as e:
            print(f"  ⚠️  Trajectory smoothing failed: {e}")
            return matched_edges

    def _is_point_feasible_on_edge(self, lon, lat, edge_key, max_dist=30):
        """Check if a GPS point can feasibly belong to an edge"""
        try:
            if not edge_key:
                return False

            distance = self._get_distance_to_edge(lon, lat, edge_key)
            return distance <= max_dist

        except Exception:
            return False

    def _get_edge_geometry_fast(self, edge_key):
        """Fast geometry lookup with caching"""
        if not edge_key:
            return None
        
        try:
            edge_data = self.road_graph.edges[edge_key]
            if 'geometry' in edge_data and edge_data['geometry'] is not None:
                return edge_data['geometry']
            else:
                # Create simple line between nodes
                u, v, k = edge_key
                nodes = self.road_graph.nodes
                ux, uy = nodes[u]['x'], nodes[u]['y']
                vx, vy = nodes[v]['x'], nodes[v]['y']
                return LineString([(ux, uy), (vx, vy)])
        except Exception:
            return None
    
    def _is_likely_roundabout(self, edge_key, current_idx, X, Y, look_ahead=4):
        """
        Detect if we're likely in a roundabout based on curvature and road type
        """
        try:
            # Check edge attributes for roundabout indicators
            edge_data = self.road_graph.edges[edge_key]
            
            # Check for roundabout tags
            if 'junction' in edge_data and edge_data['junction'] == 'roundabout':
                return True
            
            # Check for circular geometry pattern
            if current_idx >= 2 and current_idx < len(X) - look_ahead:
                # Calculate curvature over several points
                curvature = self._calculate_curvature(current_idx, X, Y, window=look_ahead)
                # High curvature + short segments often indicate roundabouts
                return curvature > 0.3  # Adjust threshold as needed
                
        except Exception:
            pass
        
        return False
    
    def _select_roundabout_edge(self, nearest_edge, current_idx, X, Y, previous_edge):
        """
        Special logic for selecting edges in roundabouts
        """
        if not previous_edge:
            return nearest_edge
        
        try:
            # Get direction of travel
            if current_idx >= 2:
                travel_bearing = self._calculate_bearing(
                    (Y[current_idx-2], X[current_idx-2]),
                    (Y[current_idx], X[current_idx])
                )
                
                # Find edges that align with travel direction
                candidates = self._get_edges_near_point(X[current_idx], Y[current_idx], max_dist=30)
                
                best_edge = nearest_edge
                best_alignment = -1
                
                for candidate_edge, dist in candidates[:5]:  # Check top 5 closest
                    edge_bearing = self._get_edge_bearing(candidate_edge)
                    if edge_bearing is not None:
                        alignment = self._bearing_similarity(travel_bearing, edge_bearing)
                        if alignment > best_alignment:
                            best_alignment = alignment
                            best_edge = candidate_edge
                
                return best_edge
                
        except Exception:
            pass
        
        return nearest_edge
    
    def _calculate_curvature(self, current_idx, X, Y, window=4):
        """
        Calculate path curvature around current point
        """
        try:
            if current_idx < window or current_idx >= len(X) - window:
                return 0
            
            # Get points before and after
            x1, y1 = X[current_idx - window], Y[current_idx - window]
            x2, y2 = X[current_idx], Y[current_idx]
            x3, y3 = X[current_idx + window], Y[current_idx + window]
            
            # Calculate angle change
            bearing1 = self._calculate_bearing((y1, x1), (y2, x2))
            bearing2 = self._calculate_bearing((y2, x2), (y3, x3))
            
            angle_change = abs(bearing2 - bearing1)
            if angle_change > 180:
                angle_change = 360 - angle_change
                
            # Normalize by distance to get curvature
            dist = ((x3-x1)**2 + (y3-y1)**2)**0.5
            return angle_change / (dist * 111000) if dist > 0 else 0  # Convert to curvature per meter
            
        except Exception:
            return 0
    
    def _get_edge_bearing(self, edge_key):
        """Calculate bearing (direction) of an edge in degrees"""
        if not edge_key:
            return None

        try:
            import math
            u, v, k = edge_key
            edge_data = self.road_graph.edges[edge_key]
            nodes = self.road_graph.nodes

            # Get start and end coordinates
            if 'geometry' in edge_data and edge_data['geometry'] is not None:
                geom = edge_data['geometry']
                coords = list(geom.coords)
                x1, y1 = coords[0]
                x2, y2 = coords[-1]
            else:
                # Use node coordinates
                x1, y1 = nodes[u]['x'], nodes[u]['y']
                x2, y2 = nodes[v]['x'], nodes[v]['y']

            # Calculate bearing
            dx = x2 - x1
            dy = y2 - y1
            bearing = math.degrees(math.atan2(dy, dx))
            return (bearing + 360) % 360

        except Exception:
            return None
    
    def _bearing_similarity(self, bearing1, bearing2):
        """
        Calculate similarity between two bearings (0-1, where 1 is identical)
        """
        diff = abs(bearing1 - bearing2)
        if diff > 180:
            diff = 360 - diff
        return 1 - (diff / 180)
    
    def _get_distance_to_edge(self, lon, lat, edge_key):
        """
        Calculate distance from point to edge in meters
        """
        try:
            edge_data = self.road_graph.edges[edge_key]
            point = Point(lon, lat)
            
            if 'geometry' in edge_data and edge_data['geometry']:
                geom = edge_data['geometry']
            else:
                u, v, k = edge_key
                u_data = self.road_graph.nodes[u]
                v_data = self.road_graph.nodes[v]
                geom = LineString([(u_data['x'], u_data['y']), (v_data['x'], v_data['y'])])
            
            dist_deg = point.distance(geom)
            return dist_deg * 111000  # Convert to meters
            
        except Exception:
            return float('inf')
    
    # Keep your existing helper methods
    def _get_road_id(self, edge_key):
        """Extract road identifier (name or osmid) from edge"""
        if not edge_key:
            return None
        try:
            edge_data = self.road_graph.edges[edge_key]
            # Try road name first, then osmid
            if 'name' in edge_data and edge_data['name']:
                return edge_data['name']
            elif 'osmid' in edge_data and edge_data['osmid']:
                return edge_data['osmid']
            else:
                return f"edge_{edge_key[0]}_{edge_key[1]}"  # fallback ID
        except:
            return None
    
    def _find_same_road_nearby(self, lon, lat, target_road_id, max_dist=40):
        """Find edge with same road ID near the point"""
        try:
            candidates = self._get_edges_near_point(lon, lat, max_dist)
            
            for edge_key, dist in candidates:
                road_id = self._get_road_id(edge_key)
                if road_id == target_road_id:
                    return edge_key
        except Exception:
            pass
        return None
    
    def _get_edges_near_point(self, lon, lat, max_dist=50):
        """
        OPTIMIZED: Get all edges within distance of point using bounding box filter
        """
        candidates = []
        seen_edges = set()

        try:
            point = Point(lon, lat)

            # Convert max_dist to degrees (approximate)
            # At mid-latitudes, 1 degree ≈ 111 km
            buffer_deg = max_dist / 111000.0 * 2  # Double for safety

            # Define bounding box
            min_lon, max_lon = lon - buffer_deg, lon + buffer_deg
            min_lat, max_lat = lat - buffer_deg, lat + buffer_deg

            # Filter nodes within bounding box first (much faster)
            nearby_nodes = []
            for node_id, node_data in self.road_graph.nodes(data=True):
                node_x, node_y = node_data['x'], node_data['y']
                if min_lon <= node_x <= max_lon and min_lat <= node_y <= max_lat:
                    nearby_nodes.append(node_id)

            # Get edges from nearby nodes
            for node in nearby_nodes:
                for neighbor in self.road_graph.neighbors(node):
                    for key in self.road_graph[node][neighbor].keys():
                        edge_key = (node, neighbor, key)

                        # Skip if already checked
                        if edge_key in seen_edges:
                            continue
                        seen_edges.add(edge_key)

                        # Calculate distance to edge
                        try:
                            edge_data = self.road_graph.edges[edge_key]
                            if 'geometry' in edge_data and edge_data['geometry']:
                                geom = edge_data['geometry']
                            else:
                                u_coord = (self.road_graph.nodes[node]['x'], self.road_graph.nodes[node]['y'])
                                v_coord = (self.road_graph.nodes[neighbor]['x'], self.road_graph.nodes[neighbor]['y'])
                                geom = LineString([u_coord, v_coord])

                            dist_deg = point.distance(geom)
                            dist_m = dist_deg * 111000

                            if dist_m <= max_dist:
                                candidates.append((edge_key, dist_m))
                        except Exception:
                            continue

        except Exception:
            pass

        # Sort by distance
        candidates.sort(key=lambda x: x[1])
        return candidates
        


    # ========================================================================
    # COMPATIBILITY METHODS - Maintain interface compatibility
    # ========================================================================
    
    def get_data_for_analysis(self):
        """Get appropriate data based on mode (single vs multi-week)"""
        if self.is_multi_week_mode and self.combined_data is not None:
            return self.combined_data
        elif self.processed_df is not None:
            return self.processed_df
        else:
            return None
    
    def get_appropriate_bounding_box(self):
        """Get bounding box appropriate for current mode"""
        if self.is_multi_week_mode:
            return self.get_weekly_bounding_boxes()
        else:
            return self.get_bounding_box()
    
    def is_single_trip_mode(self):
        """Check if processor is in single trip mode"""
        return not self.is_multi_week_mode
    
    def is_multi_week_mode_active(self):
        """Check if processor is in multi-week mode"""
        return self.is_multi_week_mode

# ============================================================================
# UPDATED MAIN ANALYZER - Uses Unified Processor
# ============================================================================

class UnifiedRoadContextGPSAnalyzer:
    """Updated main analyzer using the unified processor"""
    
    def __init__(self):
        session_id = f"monthly_analysis_{int(time.time())}"
        self.map_manager = UnifiedOfflineMapDataManager(session_id=session_id)
        self.processor = UnifiedGPSDataProcessor(self.map_manager)  # Use unified processor
        self.analyzer = None
        self.reporter = None
       
    def analyze_with_road_context(self, csv_files_or_pattern, driver_name=None, output_prefix="analysis"):
        """
        Universal analysis method that handles both single and multi-week data
        """
        print(f"🛣️  Universal GPS Analysis Starting...")
        print("="*60)
        
        # Step 1: Determine if single or multi-week analysis
        is_multi_week = self._is_multi_week_input(csv_files_or_pattern)
        
        if is_multi_week:
            return self._analyze_multi_week(csv_files_or_pattern, driver_name, output_prefix)
        else:
            return self._analyze_single_trip(csv_files_or_pattern, driver_name, output_prefix)
    
    def _is_multi_week_input(self, input_data):
        """Determine if input suggests multi-week analysis"""
        if isinstance(input_data, list) and len(input_data) > 1:
            return True
        if isinstance(input_data, str) and '*' in input_data:
            return True
        return False
    
    # ********************* updated claude single_trip

# ============================================================================
# MODIFIED ANALYSIS METHODS - Clean Weekly Results Structure
# ============================================================================

    def _analyze_single_trip(self, csv_file, driver_name, output_prefix):
        """Analyze single trip with road context - Returns weekly_results structure"""
        print("📊 Single-trip analysis mode")

         #Handle single file from list

            
        # Load single CSV
        if not self.processor.load_and_process_csv(csv_file):
            print("❌ Failed to load GPS data")
            return False
        
        # Get bounding box and download map data
        bbox = self.processor.get_bounding_box()
        if bbox:
            area_name = f"road_context_{driver_name or 'driver'}"
            self.map_manager.download_area_data(bbox, area_name)

        self.processor._perform_map_matching(self.processor.processed_df)
        # Set road context in map manager
        driven_road_ids = self.processor.get_driven_road_ids()
        route_geometry = self.processor.route_geometry
        
        # FIX #3: Capture driven_edges ✅
        driven_edges = self.processor.driven_edges.copy() if hasattr(self.processor, 'driven_edges') else set()
        
        self.map_manager.set_road_context(driven_road_ids, route_geometry)
        
        # Run analysis
        self.analyzer = UnifiedBehaviorAnalyzer(
            self.processor.processed_df, 
            self.map_manager
        )
        results = self.analyzer.analyze_all_behaviors()
        
        # CAPTURE geometry data (like multi-week does!)
        driven_edges = self.processor.driven_edges.copy() if self.processor.driven_edges else set()
        
        # Create copy of route_geometry to prevent overwriting
        if self.processor.route_geometry is not None:
            from shapely.geometry import LineString
            route_geometry_copy = LineString(self.processor.route_geometry.coords)
        else:
            route_geometry_copy = None
        
        # CREATE WEEKLY STRUCTURE for single trip WITH GEOMETRY DATA
        weekly_results = {
            "Week 1": {
                **results,
                '_geometry_data': {
                    'route_geometry': route_geometry_copy,
                    'driven_edges': driven_edges
                }
            }
        }

        # Generate reports using unified generator
        driver_info = {
            'name': driver_name or 'Driver',
            'report_type': 'Single Trip Analysis',
            'analysis_date': datetime.now().strftime('%Y-%m-%d')
        }
        
        # Create unified reporter
        self.reporter = UnifiedReportGenerator(
            weekly_results=weekly_results,
            processor=self.processor,
            driver_info=driver_info,
            map_manager=self.map_manager
        )
        
        self.reporter.print_summary()
        
        return {
            'type': 'single_trip',
            'files_processed': 1,
            'weekly_results': weekly_results,
            'report_file': "generated",
            'weeks_analyzed': 1
        }
    
    def _analyze_multi_week(self, csv_files_or_pattern, driver_name, output_prefix):
        """Analyze multiple weeks - Returns same weekly_results structure"""
        print("📊 Multi-week analysis mode")
        
        # Load multiple CSVs
        if not self.processor.load_multiple_csvs(csv_files_or_pattern, driver_name):
            print("❌ Failed to load multi-week data")
            return False
        
        # ALWAYS use combined download to ensure all GPS points have road coverage
        print("📍 Downloading map data for ALL weeks combined...")
        bbox = self.processor.get_weekly_bounding_boxes()  # Gets combined bbox
        print(f"   Combined bbox: lat [{bbox[0]:.4f}, {bbox[2]:.4f}], lon [{bbox[1]:.4f}, {bbox[3]:.4f}]")
        success = self.map_manager.download_area_data(bbox, f"combined_{driver_name}")
                    
        # Run multi-week analysis
        multi_analyzer = MultiWeekBehaviorAnalyzer(self.processor, self.map_manager)
        multi_results = multi_analyzer.analyze_all_weeks()
        
        # EXTRACT weekly_results (already in correct format from multi_analyzer)
        weekly_results = multi_results['weekly_results']  # This has actual week labels
        
        # Generate reports using same unified generator
        driver_info = {
            'name': driver_name or 'Driver',
            'report_type': 'Multi-Week Analysis',
            'analysis_date': datetime.now().strftime('%Y-%m-%d')
        }
        
        # Create unified reporter with clean weekly structure
        self.reporter = UnifiedReportGenerator(
            weekly_results=weekly_results,
            processor=self.processor,
            driver_info=driver_info,
            map_manager=self.map_manager
        )
        
        # Generate reports
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"{output_prefix}_MultiWeek_{driver_name or 'Driver'}_{timestamp}.xlsx"
        
        # Print summary  
        self.reporter.print_summary()
        
        print(f"\n🎉 Multi-Week Analysis Complete!")

        
        return {
            'type': 'multi_week',
            'files_processed': len(self.processor.weekly_data),
            'weekly_results': weekly_results,  # UNIFIED STRUCTURE
            'report_file': "no report file found",
            'weeks_analyzed': len(self.processor.weekly_data)
        }

# ============================================================================
# NEW: MULTI-WEEK ANALYZER
# ============================================================================

class MultiWeekBehaviorAnalyzer:
    """Analyze driving behavior across multiple weeks"""
      
    def __init__(self, multi_week_processor, map_manager):
        self.processor = multi_week_processor
        self.map_manager = map_manager
        self.weekly_results = {}
        self.overall_results = {}
        self.trends = {}
        
    def analyze_all_weeks(self):
        """Analyze each week individually and calculate trends - WITH COMBINED GRAPH"""
        print("🎯 Running multi-week behavior analysis...")
        
        # Check if we already have combined map data loaded
        existing_roads = len(self.map_manager.map_data.get('roads', []))
        use_combined_data = existing_roads > 5000
        
        if use_combined_data:
            print(f"   ✅ Using pre-loaded combined map data ({existing_roads} roads)")
        
        # ========== BUILD COMBINED ROAD GRAPH ONCE ==========
        print(f"\n🛣️ Building combined road graph for ALL weeks...")
        
        # Create combined dataframe with all GPS points from all weeks
        all_dfs = []
        for week_label, week_data in self.processor.weekly_data.items():
            all_dfs.append(week_data.copy())
        
        combined_df = pd.concat(all_dfs, ignore_index=True)
        print(f"   📊 Combined GPS data: {len(combined_df)} points from {len(self.processor.weekly_data)} weeks")
        
        # Build the road graph using combined data (covers ALL weeks)
        self.processor._perform_map_matching(combined_df, build_graph_only=True, rebuild_graph=True)
        
        # Verify coverage
        if self.processor.road_graph is not None:
            graph_nodes = self.processor.road_graph.number_of_nodes()
            graph_edges = self.processor.road_graph.number_of_edges()
            graph_lats = [data['y'] for _, data in self.processor.road_graph.nodes(data=True) if 'y' in data]
            graph_lons = [data['x'] for _, data in self.processor.road_graph.nodes(data=True) if 'x' in data]
            
            print(f"   ✅ Combined road graph built:")
            print(f"      Nodes: {graph_nodes}, Edges: {graph_edges}")
            print(f"      Lat range: [{min(graph_lats):.4f}, {max(graph_lats):.4f}]")
            print(f"      Lon range: [{min(graph_lons):.4f}, {max(graph_lons):.4f}]")
        
        # Store combined graph reference
        combined_road_graph = self.processor.road_graph
        # ========== END COMBINED GRAPH BUILDING ==========
        
        # Analyze each week using the SAME combined graph
        for week_label, week_data in self.processor.weekly_data.items():
            print(f"\n  📊 Analyzing {week_label}...")
            
            try:
                # Use combined map data (don't re-download)
                print(f"  📍 Using combined map data and road graph for {week_label}...")
                
                # Clear previous week's driven edges
                self.processor.driven_edges.clear()
                week_df = week_data.copy()
                
                # Ensure we're using the combined graph (don't rebuild!)
                self.processor.road_graph = combined_road_graph
        
                # Map matching using the combined graph (rebuild_graph=False)
                print(f"    🗺️  Map matching for {week_label}...")
                self.processor._perform_map_matching(week_df, build_graph_only=False, rebuild_graph=False)
                
                self.map_manager.road_graph = self.processor.road_graph
    
                #print("Testing road graph speed data:")
                if hasattr(self.processor, 'road_graph') and self.processor.road_graph:
                    edge_count = 0
                    for u, v, key, edge_data in self.processor.road_graph.edges(data=True, keys=True):
                        print(f"Edge {edge_count}: u={u}, v={v}, key={key}")
                        print(f"  Data: {dict(edge_data)}")
                        edge_count += 1
                        if edge_count >= 5:
                            break
                    #print(f"Total edges in road graph: {len(self.processor.road_graph.edges)}")
                else:
                    print("No road graph available")
                
                # Set road context BEFORE creating analyzer
                driven_edges = self.processor.driven_edges.copy() if self.processor.driven_edges else set()
                #route_geometry = self.processor.route_geometry
                if self.processor.route_geometry is not None:
                    # Create a copy of the geometry so it doesn't get overwritten
                    route_geometry = LineString(self.processor.route_geometry.coords)
                else:
                    route_geometry = None
                
                #print(f"  🔍 DEBUG: Captured route_geometry for {week_label}: {route_geometry is not None}")
                if route_geometry:
                    print(f"  🔍 DEBUG: Geometry type: {type(route_geometry)}, length: {route_geometry.length}")
            
                if route_geometry is not None:
                    self.map_manager.set_road_context(driven_edges, route_geometry)
                    print(f"    ✅ Road context set for {week_label}")
    
                # Coordinate comparison
                original_coords = list(zip(week_df['lat'], week_df['lon']))
                matched_coords = list(zip(week_df['road_matched_lat'], week_df['road_matched_lon']))
                
                # Create coordinate comparison DataFrame
                coord_comparison_data = []
                for i in range(len(original_coords)):
                    orig = original_coords[i]
                    matched = matched_coords[i]
                    
                    if pd.notna(matched[0]) and pd.notna(matched[1]):
                        distance = geodesic(orig, matched).meters
                        match_status = "Matched"
                    else:
                        distance = None
                        match_status = "No Match"
                    
                    coord_comparison_data.append({
                        'Index': i,
                        'Original_Lat': orig[0],
                        'Original_Lon': orig[1],
                        'Matched_Lat': matched[0] if pd.notna(matched[0]) else None,
                        'Matched_Lon': matched[1] if pd.notna(matched[1]) else None,
                        'Distance_Meters': distance,
                        'Match_Status': match_status,
                        'Timestamp': week_df.iloc[i]['timestamp'] if 'timestamp' in week_df.columns else None
                    })
                
                coord_df = pd.DataFrame(coord_comparison_data)
                
                # Export to Excel
                filename = f"coordinate_comparison_{week_label}.xlsx"
                coord_df.to_excel(filename, index=False)
                print(f"Coordinate comparison exported to {filename}")
                
                # Show summary statistics
                print(f"\nCoordinate Analysis Summary for {week_label}:")
                print(f"Total points: {len(coord_df)}")
                print(f"Successfully matched: {len(coord_df[coord_df['Match_Status'] == 'Matched'])}")
                print(f"No match: {len(coord_df[coord_df['Match_Status'] == 'No Match'])}")
                
                if len(coord_df[coord_df['Match_Status'] == 'Matched']) > 0:
                    matched_distances = coord_df[coord_df['Match_Status'] == 'Matched']['Distance_Meters']
                    print(f"Distance statistics (meters):")
                    print(f"  Mean: {matched_distances.mean():.1f}")
                    print(f"  Median: {matched_distances.median():.1f}")
                    print(f"  Max: {matched_distances.max():.1f}")
                    print(f"  Points >100m off: {len(matched_distances[matched_distances > 100])}")
                
                print("\n=== COORDINATE COMPARISON ===")
                #print("Index | Original Lat,Lon | Matched Lat,Lon | Distance (m)")
                #print("-" * 65)
                
                for i in range(min(20, len(original_coords))):
                    orig = original_coords[i]
                    matched = matched_coords[i]
                    
                    if pd.notna(matched[0]) and pd.notna(matched[1]):
                        distance = geodesic(orig, matched).meters
                        #print(f"{i:5d} | {orig[0]:.6f},{orig[1]:.6f} | {matched[0]:.6f},{matched[1]:.6f} | {distance:6.1f}")
                    else:
                        print(f"{i:5d} | {orig[0]:.6f},{orig[1]:.6f} | NO MATCH          | N/A")
                
                # Test speed lookups
                test_idx = None
                for i, coord in enumerate(original_coords):
                    if abs(coord[0] - 48.72127333) < 0.0001 and abs(coord[1] - 7.6942) < 0.0001:
                        test_idx = i
                        break
                
                if test_idx is not None:
                    print(f"\n=== SPEED LOOKUP COMPARISON FOR POINT {test_idx} ===")
                    matched_edge = week_df.iloc[test_idx]['matched_edge']
                    
                    if matched_edge is not None:
                        edge_speed = self.map_manager.get_speed_limit(matched_edge)
                        print(f"Point {test_idx}: {original_coords[test_idx]} → Edge speed: {edge_speed} km/h")
                        print(f"Matched edge: {matched_edge}")
                    else:
                        print(f"Point {test_idx}: {original_coords[test_idx]} → No matched edge")
    
                if route_geometry is not None:
                    # Create analyzer
                    analyzer = UnifiedBehaviorAnalyzer(week_df, self.map_manager, self.processor)
                    print("Testing road-linked filtering...")
                    stop_signs = self.map_manager.map_data['stop_signs'] 
                    result = self.map_manager.get_osmid_road_linked_features(stop_signs, "stop_signs")
                    print(f"Total stop signs: {len(stop_signs)}")
                    print(f"Road-linked stop signs: {result['statistics']['linked_count']}")
                    print(f"Filter percentage: {result['statistics']['filter_percentage']:.1f}%")
                                  
                    # Behavior analysis
                    analysis_results = analyzer.analyze_all_behaviors()
                
                    # Store results
                    self.weekly_results[week_label] = {
                        **analysis_results,
                        '_geometry_data': {
                            'route_geometry': route_geometry,
                            'driven_edges': driven_edges
                        }
                    }        
                else:
                    print(f"    ⚠️  No route geometry created for {week_label}")
                    self.weekly_results[week_label] = {
                        'speed_analysis': {},
                        'acceleration_analysis': {},
                        'cornering_analysis': {},
                        'braking_analysis': {},
                        'road_context_analysis': {},
                        'error': 'No route geometry available',
                        '_geometry_data': None
                    }
                    
            except Exception as e:
                print(f"    ❌ Failed to analyze {week_label}: {e}")
                import traceback
                traceback.print_exc()
                self.weekly_results[week_label] = {
                    'speed_analysis': {},
                    'acceleration_analysis': {},
                    'cornering_analysis': {},
                    'braking_analysis': {},
                    'road_context_analysis': {},
                    'error': str(e),
                    '_geometry_data': None
                }
    
        return {
            'weekly_results': self.weekly_results,
            'overall_results': self.overall_results,
            'trends': self.trends
        }
        
    
    def _calculate_overall_results(self):
        """Calculate overall results across all weeks"""
        if not self.processor.combined_data.empty:
            # Overall analysis using combined data
            analyzer = UnifiedBehaviorAnalyzer(self.processor.combined_data, self.map_manager)
            self.overall_results = analyzer.analyze_all_behaviors()
 
# ============================================================================
# UNIFIED BEHAVIOR ANALYZER - Combines all three analyzers
# ============================================================================

class UnifiedBehaviorAnalyzer:
    """
    Unified Behavior Analyzer that combines functionality of:
   
    Maintains ALL original functionality and method names.
    """
    def __init__(self, gps_data, map_manager, processor=None, max_workers=4, enable_parallel=True, enable_road_context=True):
        # Core data (from all analyzers)
        self.df = gps_data
        self.map_manager = map_manager
        self.results = {}
        
        # Enhanced functionality flags and settings
        self.max_workers = max_workers
        self.enable_parallel = enable_parallel
        self.enable_road_context = enable_road_context
        self.processor = processor  # ADD THIS LINE

        
        # Initialize speed calculator if parallel is enabled
        if self.enable_parallel:
                    gps_bounds = (
                        gps_data['lat'].min() - 0.001,
                        gps_data['lon'].min() - 0.001,
                        gps_data['lat'].max() + 0.001,
                        gps_data['lon'].max() + 0.001
                    )
                    self.speed_calculator = ParallelSpeedLimitCalculator(map_manager, max_workers)
                    self.speed_calculator.gps_bounds = gps_bounds
        else:
            self.speed_calculator = None
    
    # ========================================================================
    # MAIN ANALYSIS METHOD - Unified from all three classes
        # ========================================================================
    
    def analyze_all_behaviors(self):
        """
        Run road-context behavior analysis only - no fast/inaccurate methods
        Requires route geometry for accurate analysis
        """
        if hasattr(self, 'map_manager') and self.map_manager.map_data:
            print("🔍 Analyzing PBF ID relationships...")
            #id_relationships = self.analyze_pbf_id_relationships()
            #self.results['pbf_id_analysis'] = id_relationships


        
        # Validate route geometry is available
        if not hasattr(self.map_manager, 'route_geometry') or self.map_manager.route_geometry is None:
            raise ValueError("Route geometry is required for accurate behavior analysis. "
                            "Cannot perform analysis without route context.")
        
        # Debug info for road context
        print(f"🔧 DEBUG INFO:")
        print(f"  - GPS points: {len(self.df)}")
        print(f"  - Route geometry: {'Yes' if hasattr(self.map_manager, 'route_geometry') and self.map_manager.route_geometry else 'No'}")
        print(f"  - Driven roads: {len(getattr(self.map_manager, 'driven_road_ids', set()))}")
        print(f"  - Available features:")
        print(f"    * Roundabouts: {len(self.map_manager.map_data.get('roundabouts', []))}")
        print(f"    * Schools: {len(self.map_manager.map_data.get('schools', []))}")
        print(f"    * Stop signs: {len(self.map_manager.map_data.get('stop_signs', []))}")
        
        # Determine analysis type based on parallel processing capability
        analysis_type = "road-context enhanced" if self.enable_parallel else "road-context"
        print(f"🎯 Running {analysis_type} behavior analysis...")
        
        # Speed limit calculation timing
        #print(" [[Edge Based Method]] Pre-calculating speed limits...")
        speed_limits = []
        for _, row in self.df.iterrows():
            if pd.notna(row.get('matched_edge')):
                speed_limit = self.map_manager.get_speed_limit(row['matched_edge'])
            else:
                speed_limit = 50  # Default for unmatched points
            speed_limits.append(speed_limit)
        
        self.df['speed_limit'] = speed_limits
        #print(f"DataFrame columns after speed calculation: {list(self.df.columns)}")
        #print(f"Sample speed_limit values: {self.df['speed_limit'].head(10) if 'speed_limit' in self.df.columns else 'MISSING'}")
              

        #self.test_id_extraction()
            
        # Individual behavior analysis timing
        print("  • Analyzing speeding violations...")
        self._analyze_speeding_on_road()

        print("  • Analyzing roundabout approaches...")
        self._analyze_roundabouts_with_context()
        

        #self.build_road_feature_lookup_timed()
        print("  • Analyzing stop signs...")
        self._analyze_stop_signs_with_context()
       
       
        print("  • Analyzing school zones...")
        self._analyze_school_zones_with_context()

        print("\n  • Analyzing traffic lights...")
        self._analyze_traffic_lights_with_context()
        
        print("\n  • Analyzing harsh driving events...")
        self._analyze_harsh_events()

        
        
        return self.results
    
    # ========================================================================
    # SPEEDING ANALYSIS - Same for all analyzers



    
    def build_road_feature_lookup_timed(self):
        start = time.time()
        
        road_way_ids = set()
        road_osmids = set() 
        road_node_refs = set()
        
        print(f"Starting lookup build... {time.time() - start:.2f}s")
        
        for i, road in enumerate(self.map_manager.get('roads', [])):
            if i == 100:
                print(f"First 100 roads processed: {time.time() - start:.2f}s")
            
            road_way_ids.add(road.get('id'))
            road_osmids.add(road.get('osmid'))
            
            # TIME THIS SPECIFIC PART
            node_start = time.time()
            if 'node_refs' in road:
                road_node_refs.update(road['node_refs'])
            node_time = time.time() - node_start
            
            if node_time > 0.01:  # If any single road takes >10ms
                print(f"Road {road.get('id')} node extraction took {node_time:.3f}s")
                break
        
        print(f"Total time: {time.time() - start:.2f}s")

    

 # ========================================================================
     
    def _analyze_speeding_on_road(self):
                
        violations = []
        episodes_list = []  # ADD THIS LINE HERE

        MIN_EPISODE_GAP_SECONDS = 15
        
        # Ensure speed_limit column exists
        if 'speed_limit' not in self.df.columns:
            print("  ⚠️  Speed limits not pre-calculated, using fallback...")
            speed_limits = []
            for _, row in self.df.iterrows():
                if 'road_matched_lat' in row and pd.notna(row['road_matched_lat']):
                    limit = self.map_manager.get_speed_limit(row['matched_edge'])
                else:
                    limit = 50  # Default fallback
                speed_limits.append(limit)
            self.df['speed_limit'] = speed_limits
            # Add this right after speed limits calculation
                
        # Sort DataFrame by timestamp
        self.df = self.df.sort_values(by='timestamp').reset_index(drop=True)
        
        if not pd.api.types.is_datetime64_any_dtype(self.df['timestamp']):
            self.df['timestamp'] = pd.to_datetime(self.df['timestamp'])
        
        print(f"  📊 Analyzing {len(self.df)} GPS points across {len(self.df['speed_limit'].unique())} unique speed limits")
        
        # ========================================================================
        # STEP 1: CREATE GEOGRAPHIC SEGMENTS (Keep original logic)
        # ========================================================================
        geographic_segments = []
        if len(self.df) > 0:
            current_segment = {
                'segment_id': 1,
                'start_idx': 0,
                'speed_limit': self.df.iloc[0]['speed_limit'],
                'start_time': self.df.iloc[0]['timestamp'],
                'rows': []
            }
            
            for idx, row in self.df.iterrows():
                if row['speed_limit'] == current_segment['speed_limit']:
                    current_segment['rows'].append({
                        'idx': idx,
                        'timestamp': row['timestamp'],
                        'lat': row['lat'],
                        'lon': row['lon'],
                        'speed_kmh': row['speed_kmh'],
                        'speed_limit': row['speed_limit']
                    })
                else:
                    # Finalize current segment
                    current_segment['end_idx'] = idx - 1
                    current_segment['end_time'] = self.df.iloc[idx-1]['timestamp']
                    current_segment['total_points'] = len(current_segment['rows'])
                    current_segment['duration_seconds'] = (current_segment['end_time'] - current_segment['start_time']).total_seconds()
                    geographic_segments.append(current_segment)
                    
                    # Start new segment
                    current_segment = {
                        'segment_id': len(geographic_segments) + 1,
                        'start_idx': idx,
                        'speed_limit': row['speed_limit'],
                        'start_time': row['timestamp'],
                        'rows': [{
                            'idx': idx,
                            'timestamp': row['timestamp'],
                            'lat': row['lat'],
                            'lon': row['lon'],
                            'speed_kmh': row['speed_kmh'],
                            'speed_limit': row['speed_limit']
                        }]
                    }
            
            # Add final segment
            if current_segment['rows']:
                current_segment['end_idx'] = len(self.df) - 1
                current_segment['end_time'] = self.df.iloc[-1]['timestamp']
                current_segment['total_points'] = len(current_segment['rows'])
                current_segment['duration_seconds'] = (current_segment['end_time'] - current_segment['start_time']).total_seconds()
                geographic_segments.append(current_segment)
        
        print(f"  📍 Created {len(geographic_segments)} geographic road segments")

        # ========================================================================
        # STEP 1.5: FILTER OUT SINGLE-POINT SEGMENTS (NEW)
        # ========================================================================
        print(f"  🔍 Filtering single-point segments...")
        print(f"    Before filtering: {len(geographic_segments)} segments")
        
        # Filter out segments with only 1 GPS point
        meaningful_segments = []
        single_point_segments = []
        
        for segment in geographic_segments:
            if segment['total_points'] > 1:
                meaningful_segments.append(segment)
            else:
                single_point_segments.append(segment)
        
        print(f"    After filtering: {len(meaningful_segments)} segments")
        print(f"    Filtered out: {len(single_point_segments)} single-point segments")
        
        # Replace the original list with filtered segments
        geographic_segments = meaningful_segments
        
        # Filter out micro-segments (optional - keeps segments but notes them)
        micro_segments = [seg for seg in geographic_segments if seg['total_points'] <= 3 or seg['duration_seconds'] <= 10]
        substantial_segments = [seg for seg in geographic_segments if seg['total_points'] > 3 and seg['duration_seconds'] > 10]
        
        if micro_segments:
            print(f"  ⚠️  Found {len(micro_segments)} micro-segments (≤3 points or ≤10s) - keeping but will merge for analysis")
        
        # ========================================================================
        # STEP 2: DETECT VIOLATION EPISODES WITHIN SEGMENTS
        # ========================================================================
        violation_episodes = []
        episode_id = 0
        last_episode_by_speed_limit = {}
        
        print(f"  🔍 Detecting violation episodes across segments...")
        
        for segment in geographic_segments:
            current_episode = None
            segment_limit = segment['speed_limit']
            violations_in_segment = 0
            
            for point in segment['rows']:
                is_violating = point['speed_kmh'] > segment_limit
                
                if is_violating:
                    violations_in_segment += 1
                    
                    if current_episode is None:
                        # Check if we can merge with previous episode for same speed limit
                        can_merge = (segment_limit in last_episode_by_speed_limit and
                                    (point['timestamp'] - last_episode_by_speed_limit[segment_limit]['end_time']).total_seconds() <= MIN_EPISODE_GAP_SECONDS)
                        
                        if can_merge:
                            current_episode = last_episode_by_speed_limit[segment_limit]
                        else:
                            episode_id += 1
                            current_episode = {
                                'episode_id': episode_id,
                                'start_time': point['timestamp'],
                                'speed_limit': segment_limit,
                                'max_speed': point['speed_kmh'],
                                'max_excess': point['speed_kmh'] - segment_limit,
                                'max_excess_idx': point['idx'],
                                'points_data': []
                            }
                        
                        current_episode['points_data'].append({
                            'idx': point['idx'],
                            'timestamp': point['timestamp'],
                            'lat': point['lat'],
                            'lon': point['lon'],
                            'speed_kmh': point['speed_kmh'],
                            'excess_speed': point['speed_kmh'] - segment_limit
                        })
                    else:
                        # Continue current episode
                        current_episode['points_data'].append({
                            'idx': point['idx'],
                            'timestamp': point['timestamp'],
                            'lat': point['lat'],
                            'lon': point['lon'],
                            'speed_kmh': point['speed_kmh'],
                            'excess_speed': point['speed_kmh'] - segment_limit
                        })
                    
                    # Update max if needed
                    current_excess = point['speed_kmh'] - segment_limit
                    if current_excess > current_episode['max_excess']:
                        current_episode['max_speed'] = point['speed_kmh']
                        current_episode['max_excess'] = current_excess
                        current_episode['max_excess_idx'] = point['idx']
                else:
                    if current_episode is not None:
                        # End current episode
                        last_point = current_episode['points_data'][-1]
                        current_episode['end_time'] = last_point['timestamp']
                        current_episode['duration_seconds'] = (current_episode['end_time'] - current_episode['start_time']).total_seconds()
                        current_episode['total_gps_points'] = len(current_episode['points_data'])
                        
                        if current_episode not in violation_episodes:
                            violation_episodes.append(current_episode)
                        
                        last_episode_by_speed_limit[segment_limit] = current_episode
                        current_episode = None
            
            # Handle ongoing episode at segment end
            if current_episode is not None:
                last_point = current_episode['points_data'][-1]
                current_episode['end_time'] = last_point['timestamp']
                current_episode['duration_seconds'] = (current_episode['end_time'] - current_episode['start_time']).total_seconds()
                current_episode['total_gps_points'] = len(current_episode['points_data'])
                if current_episode not in violation_episodes: 
                    violation_episodes.append(current_episode)
                last_episode_by_speed_limit[segment_limit] = current_episode
        
        print(f"  🚨 Detected {len(violation_episodes)} violation episodes")
        
        # ========================================================================
        # STEP 3: DEFINE SPEED ZONE CATEGORIES (6 static categories)
        # ========================================================================
        def get_speed_zone(speed_limit):
            """Categorize speed limit into zone"""
            if pd.isna(speed_limit) or speed_limit is None:
                return "INVALID"
            if speed_limit <= 30:
                return "30"
            elif speed_limit <= 50:
                return "50"
            elif speed_limit <= 70:
                return "70"
            elif speed_limit <= 80:
                return "80"  # NEW ZONE
            elif speed_limit <= 90:
                return "90"
            elif speed_limit <= 110:
                return "110"
            elif speed_limit <= 130:
                return "130"
            else:
                return ">130"
        
        # ========================================================================
        # STEP 4: AGGREGATE SEGMENTS BY SPEED ZONE CATEGORIES
        # ========================================================================
        # Group segments by speed zone category
        segments_by_category = {}
        for segment in geographic_segments:
            category = get_speed_zone(segment['speed_limit'])
            if category not in segments_by_category:
                segments_by_category[category] = []
            segments_by_category[category].append(segment)
        
        print(f"  📋 Segments grouped into {len(segments_by_category)} speed zone categories:")
        for category, segs in segments_by_category.items():
            total_points = sum(seg['total_points'] for seg in segs)
            total_duration = sum(seg['duration_seconds'] for seg in segs)
            print(f"    {category}: {len(segs)} segments, {total_points} GPS points, {total_duration:.0f}s total")
        
        # ========================================================================
        # STEP 5: AGGREGATE VIOLATIONS BY SPEED ZONE CATEGORIES
        # ========================================================================
        episodes_by_category = {}
        for episode in violation_episodes:
            category = get_speed_zone(episode['speed_limit'])
            if category not in episodes_by_category:
                episodes_by_category[category] = []
            episodes_by_category[category].append(episode)
        
        print(f"  📊 Violation episodes by speed zone category:")
        for category, eps in episodes_by_category.items():
            print(f"    {category}: {len(eps)} episodes")
        
        # ========================================================================
        # STEP 6: CREATE VIOLATION RECORDS (Keep original format)
        # ========================================================================
        for episode in violation_episodes:
            worst_point_data = None
            max_excess_in_episode = 0
            
            for point in episode['points_data']:
                if point['excess_speed'] > max_excess_in_episode:
                    max_excess_in_episode = point['excess_speed']
                    worst_point_data = point
            
            if worst_point_data:
                violations.append({
                    'timestamp': worst_point_data['timestamp'],
                    'lat': worst_point_data['lat'],
                    'lon': worst_point_data['lon'],
                    'speed_kmh': worst_point_data['speed_kmh'],
                    'speed_limit': episode['speed_limit'],
                    'excess_speed': worst_point_data['excess_speed'],
                    'episode_id': episode['episode_id'],
                    'episode_duration_seconds': episode['duration_seconds'],
                    'episode_gps_points': episode['total_gps_points']
                })
        
        # Create violation DataFrame (original format)
        violation_data = pd.DataFrame(violations) if violations else pd.DataFrame(columns=[
            'timestamp', 'lat', 'lon', 'speed_kmh', 'speed_limit', 'excess_speed'
        ])
        
        if len(violation_data) > 0:
            violation_data['severity'] = pd.cut(
                violation_data['excess_speed'],
                bins=[0, 10, 20, 30, float('inf')],
                labels=['minor', 'moderate', 'major', 'severe']
            )
        else:
            violation_data['severity'] = pd.Series(dtype='category')
        
        worst_violations = violation_data.nlargest(3, 'excess_speed')[
            ['timestamp', 'lat', 'lon', 'speed_kmh', 'speed_limit', 'excess_speed', 'severity']
        ].to_dict('records')
        
        severity_counts = violation_data['severity'].value_counts().to_dict()
        
        # ========================================================================
        # STEP 7: CREATE SPEED ZONE ANALYSIS (CORRECTED - 6 categories)
        # ========================================================================
        total_driving_time = (self.df['timestamp'].max() - self.df['timestamp'].min()).total_seconds()
        
        # Calculate zone analysis for each of the 6 speed zone categories
        zone_analysis = {}
        
        # Ensure all 6 categories exist in results even if no data
        all_categories = ["30", "50", "70", "80", "90", "110", "130"]

        # REPLACE THE COMPLIANCE CALCULATION IN YOUR SPEED ZONE ANALYSIS
# Find this section in _analyze_speeding_on_road where zone_analysis is calculated
        
        for category in all_categories:
            # Get segments for this category
            category_segments = segments_by_category.get(category, [])
            
            # CORRECTED: Count segments and violating segments (like school zone logic)
            total_segments_in_category = len(category_segments)
            
            # Count segments that actually have violations (using the corrected temporal overlap logic)
            violating_segments_count = 0
            total_violation_episodes_in_category = 0
            
            for segment in category_segments:
                # Check if this segment has violation episodes (using corrected temporal overlap)
                segment_violations = []
                for episode in violation_episodes:
                    episode_overlaps = (
                        episode['start_time'] >= segment['start_time'] and episode['start_time'] <= segment['end_time']
                    ) or (
                        episode['end_time'] >= segment['start_time'] and episode['end_time'] <= segment['end_time']
                    ) or (
                        episode['start_time'] <= segment['start_time'] and episode['end_time'] >= segment['end_time']
                    )
                    
                    if episode_overlaps:
                        segment_violations.append(episode)
                
                if len(segment_violations) > 0:
                    violating_segments_count += 1
                
                total_violation_episodes_in_category += len(segment_violations)
            
            # Calculate compliance based on segments (like school zones)
            compliant_segments = total_segments_in_category - violating_segments_count
            compliance_percentage = (compliant_segments / total_segments_in_category * 100) if total_segments_in_category > 0 else 100.0
            
            # Calculate total time in this category (keep for other metrics)
            total_time_in_category = sum(seg['duration_seconds'] for seg in category_segments)
            total_points_in_category = sum(seg['total_points'] for seg in category_segments)
            
            # Calculate violation time (sum of all episodes - keep for reporting)
            violation_time_seconds = sum(ep['duration_seconds'] for ep in violation_episodes 
                                        if get_speed_zone(ep['speed_limit']) == category)
            violation_time_minutes = violation_time_seconds / 60.0
            
            # Find worst speeds in this category
            worst_speed_kmh = 0.0
            excess_speed_kmh = 0.0
            speed_limit_kmh = 0
            
            category_episodes = [ep for ep in violation_episodes if get_speed_zone(ep['speed_limit']) == category]
            if category_episodes:
                worst_episode = max(category_episodes, key=lambda ep: ep['max_excess'])
                worst_speed_kmh = worst_episode['max_speed']
                excess_speed_kmh = worst_episode['max_excess'] 
                speed_limit_kmh = worst_episode['speed_limit']
            elif category_segments:
                speed_limit_kmh = category_segments[0]['speed_limit']
            
            zone_analysis[category] = {
                'violation_episodes': total_violation_episodes_in_category,  # Total violation episodes in this zone
                'safe_driving_time': total_time_in_category - violation_time_seconds,  # Keep for reference
                'violation_time_minutes': round(violation_time_minutes, 1),  # Keep for reference
                'compliance_percentage': round(compliance_percentage, 1),  # NOW SEGMENT-BASED like school zones
                'worst_speed_kmh': round(worst_speed_kmh, 1),
                'speed_limit_kmh': speed_limit_kmh,
                'excess_speed_kmh': round(excess_speed_kmh, 1),
                'total_time_seconds': total_time_in_category,  # Keep for reference
                'total_gps_points': total_points_in_category,
                'segments_count': total_segments_in_category,  # Total segments in this zone
                'top_speed': round(worst_speed_kmh, 1),
                'violating_segments_count': violating_segments_count  # NEW: count of segments with violations
            }
        
        # Calculate total unique speed zones encountered (should be <= 6)
        total_zones_encountered = len([cat for cat in all_categories if zone_analysis[cat]['total_gps_points'] > 0])
        
        speed_zones_data = {
            'total_zones_encountered': total_zones_encountered,
            'zone_breakdown': zone_analysis
        }
        
        # Calculate overall compliance
        total_violation_time = sum(zone_analysis[cat]['violation_time_minutes'] * 60 for cat in all_categories)
        safe_driving_time = total_driving_time - total_violation_time
        compliance_percentage = (safe_driving_time / total_driving_time * 100) if total_driving_time > 0 else 100

        all_episodes_list = []

        
        
        # REPLACE THIS SECTION IN _analyze_speeding_on_road METHOD
        
        for segment in geographic_segments:
            # CORRECTED: Check if violation episodes actually overlap with this segment's time period
            segment_violations = []
            for episode in violation_episodes:
                # Check if episode overlaps with this segment's timeframe
                episode_overlaps = (
                    episode['start_time'] >= segment['start_time'] and episode['start_time'] <= segment['end_time']
                ) or (
                    episode['end_time'] >= segment['start_time'] and episode['end_time'] <= segment['end_time']
                ) or (
                    episode['start_time'] <= segment['start_time'] and episode['end_time'] >= segment['end_time']
                )
                
                if episode_overlaps:
                    segment_violations.append(episode)
            
            has_violations = len(segment_violations) > 0
            
            # Find max speed in this segment
            max_speed_in_segment = max(point['speed_kmh'] for point in segment['rows']) if segment['rows'] else 0
        
            
            segment_episode = {
                'episode_id': segment['segment_id'],
                'start_time': segment['start_time'],
                'end_time': segment['end_time'],
                'duration_seconds': segment['duration_seconds'],
                'speed_limit': segment['speed_limit'],
                'max_speed': max_speed_in_segment,
                'max_excess': max(0, max_speed_in_segment - segment['speed_limit']),
                'total_gps_points': segment['total_points'],
                'is_violation': has_violations,
                'violation_episodes_count': len(segment_violations),
                'segment_data': segment['rows']
            }
            all_episodes_list.append(segment_episode)
        
        for episode in violation_episodes:
            episode_dict = {
                'episode_id': episode['episode_id'],
                'start_time': episode['start_time'],
                'end_time': episode['end_time'],
                'duration_seconds': episode['duration_seconds'],
                'speed_limit': episode['speed_limit'],
                'max_speed': episode['max_speed'],
                'max_excess': episode['max_excess'],
                'total_gps_points': episode['total_gps_points'],
                'points_data': episode['points_data']
            }
            episodes_list.append(episode_dict)
        # ========================================================================
        # STEP 8: STORE RESULTS (Keep original structure)
        # ========================================================================
    
        self.results['speeding'] = {
            'total_violations': len(violation_data),
            'violations_by_severity': {
                'minor': severity_counts.get('minor', 0),
                'moderate': severity_counts.get('moderate', 0),
                'major': severity_counts.get('major', 0),
                'severe': severity_counts.get('severe', 0)
            },
            'worst_violations': worst_violations,
            'compliance_percentage': compliance_percentage,
            'total_violation_time': total_violation_time,
            'safe_driving_time': safe_driving_time,
            'speed_zones': speed_zones_data,
            'all_episodes': episodes_list,
            'all_segments': all_episodes_list
        }

       

        
        # Add episodes to results
        
        # Debug summary
        print(f"  ✅ Speed zone analysis complete:")
        print(f"    Geographic segments: {len(geographic_segments)}")
        print(f"    Speed zone categories: {total_zones_encountered}")
        print(f"    Total violation episodes: {len(violation_episodes)}")
        print(f"    Overall compliance: {compliance_percentage:.1f}%")
        
        print(f"  📊 Final speed zones breakdown:")
        for category, data in zone_analysis.items():
            if data['total_gps_points'] > 0:
                print(f"    {category}: {data['segments_count']} segments, {data['violation_episodes']} violations, {data['compliance_percentage']:.1f}% compliance")

    # Added new Roundabouts Logic Phase 2
    # ===========================================================
    def _analyze_roundabouts_with_context(self):
        """
        ENHANCED IMPLEMENTATION: Robust roundabout analysis for sparse GPS data
        
        Features:
        - 3-layer detection (direct evidence, indicators, segment analysis)
        - Adaptive thresholds for sparse GPS (6-second intervals)
        - Uses IO parameters (io_movement, io_engine_load) when available
        - Trend-based anticipation detection (not strict per-zone)
        - Proper approach vs exit direction detection
        - Confidence scoring for all results
        - Entry/Exit speed tracking per client requirements
        
        Zones analyzed:
        - Zone 1: 150-100m before roundabout (early anticipation)
        - Zone 2: 100-65m before roundabout (mid anticipation)  
        - Zone 3: 65-0m before roundabout (late anticipation)
        - Entry: 0-15m (entry compliance ≤34 km/h)
        - Inside: Within roundabout geometry
        - Exit: First points after roundabout
        
        Anticipation Categories:
        - Good: Deceleration starts 150-100m before
        - Moderate: Deceleration starts 100-65m before
        - Late: Deceleration starts <65m before
        - No Deceleration: No significant speed reduction detected
        """
        
        print(f"    🔍 Analyzing roundabout approaches (ENHANCED METHOD)...")
        
        # ================================================================
        # STEP 0: Initialize results structure
        # ================================================================
        results = {
            'total_roundabouts': 0,
            'total_approaches': 0,
            'anticipation_stats': {
                'good': 0,
                'moderate': 0,
                'late': 0,
                'no_deceleration': 0
            },
            'entry_compliance': {
                'compliant': 0,
                'non_compliant': 0,
                'unknown': 0
            },
            'approaches_detailed': [],
            'detection_mode': 'basic',
            'avg_confidence': 0,
            'format': 'multi_zone'
        }
        
        # ================================================================
        # STEP 1: Validate prerequisites
        # ================================================================
        roundabouts = self.map_manager.map_data.get('roundabouts', [])
        print(f"    📊 Total roundabouts in map: {len(roundabouts)}")
        
        if not roundabouts:
            print(f"    ❌ No roundabouts in map data")
            self.results['roundabouts'] = results
            return
        
        # Check route geometry
        if not hasattr(self.map_manager, 'route_geometry') or self.map_manager.route_geometry is None:
            print(f"    ❌ No route geometry available")
            self.results['roundabouts'] = results
            return
        
        route_geom = self.map_manager.route_geometry
        
        # Debug: Show route geometry range
        route_coords = list(route_geom.coords)
        route_lats = [c[1] for c in route_coords]
        print(f"    📍 Route geometry: {len(route_coords)} points, lat range: {min(route_lats):.4f} to {max(route_lats):.4f}")
        
        # ================================================================
        # STEP 2: Detect GPS quality and available features
        # ================================================================
        has_io_params = 'has_io_params' in self.df.columns and self.df['has_io_params'].any()
        has_movement_status = 'io_movement' in self.df.columns and self.df['io_movement'].notna().any()
        has_engine_data = 'io_engine_load' in self.df.columns and self.df['io_engine_load'].notna().any()
        
        # Calculate GPS interval
        avg_interval = self.df['time_diff_s'].mean() if 'time_diff_s' in self.df.columns else 6.0
        is_sparse_gps = avg_interval > 4
        
        # Set detection mode
        if has_movement_status and has_engine_data:
            results['detection_mode'] = 'enhanced'
        elif has_io_params:
            results['detection_mode'] = 'standard'
        else:
            results['detection_mode'] = 'basic'
        
        print(f"    🎯 Detection mode: {results['detection_mode']}")
        print(f"    📡 GPS interval: {avg_interval:.1f}s {'(SPARSE)' if is_sparse_gps else '(OK)'}")
        
        # Adaptive thresholds based on GPS quality
        if is_sparse_gps:
            MIN_APPROACH_POINTS = 2  # Relaxed for sparse GPS
            MIN_ZONE_POINTS = 1
            APPROACH_ZONE_RADIUS = 200  # meters - larger zone to capture more points
        else:
            MIN_APPROACH_POINTS = 3
            MIN_ZONE_POINTS = 1
            APPROACH_ZONE_RADIUS = 150
        
        # ================================================================
        # STEP 3: Setup coordinate transformation
        # ================================================================
        from pyproj import Transformer
        from shapely.ops import transform
        from shapely.geometry import Point, LineString
        
        transformer = Transformer.from_crs("EPSG:4326", "EPSG:2154", always_xy=True)
        route_proj = transform(lambda x, y: transformer.transform(x, y), route_geom)
        
        # ================================================================
        # STEP 4: Filter roundabouts on driven route (DISTANCE-BASED)
        # ================================================================
        
        # Setup coordinate transformation
        transformer = Transformer.from_crs("EPSG:4326", "EPSG:2154", always_xy=True)
        route_proj = transform(lambda x, y: transformer.transform(x, y), route_geom)
        
        # Get GPS bounding box for quick filtering
        lat_min, lat_max = self.df['lat'].min(), self.df['lat'].max()
        lon_min, lon_max = self.df['lon'].min(), self.df['lon'].max()
        bbox_buffer = 0.005  # ~500m buffer
        
        # Distance threshold - stricter to avoid false positives
        MAX_DISTANCE_TO_ROUTE = 30  # meters (was 50m, now 30m)
        
        relevant_roundabouts = []
        
        for roundabout in roundabouts:
            rb_lat = roundabout.get('lat')
            rb_lon = roundabout.get('lon')
            
            if rb_lat is None or rb_lon is None:
                continue
            
            # Quick bounding box filter first
            if not (lat_min - bbox_buffer <= rb_lat <= lat_max + bbox_buffer):
                continue
            if not (lon_min - bbox_buffer <= rb_lon <= lon_max + bbox_buffer):
                continue
            
            # Precise distance check
            try:
                rb_point = Point(rb_lon, rb_lat)
                rb_point_proj = transform(lambda x, y: transformer.transform(x, y), rb_point)
                distance_to_route = route_proj.distance(rb_point_proj)
                
                if distance_to_route <= MAX_DISTANCE_TO_ROUTE:
                    route_position = route_proj.project(rb_point_proj)
                    relevant_roundabouts.append({
                        'roundabout': roundabout,
                        'lat': rb_lat,
                        'lon': rb_lon,
                        'id': roundabout.get('id', roundabout.get('osmid', 'unknown')),
                        'osmid': roundabout.get('osmid'),
                        'geometry': roundabout.get('geometry'),
                        'distance_to_route': distance_to_route,
                        'route_position': route_position
                    })
                    print(f"    ✅ Roundabout {roundabout.get('id', 'unknown')}: {distance_to_route:.1f}m from route")
            except Exception as e:
                continue
        
        print(f"    ✅ Found {len(relevant_roundabouts)} roundabouts within {MAX_DISTANCE_TO_ROUTE}m of route")
        # ================================================================
        # STEP 5: Group roundabouts by proximity (same physical location)
        # ================================================================
        grouped_roundabouts = self._group_roundabouts_by_proximity(
            [rb['roundabout'] for rb in relevant_roundabouts],
            max_distance=100
        )
        
        results['total_roundabouts'] = len(grouped_roundabouts)
        print(f"    🔄 Grouped into {len(grouped_roundabouts)} physical roundabout locations")
        
        # ================================================================
        # STEP 6: Analyze each roundabout group
        # ================================================================
        all_confidences = []
        
        for group_idx, roundabout_group in enumerate(grouped_roundabouts):
            approach_result = self._analyze_single_roundabout_approach(
                roundabout_group=roundabout_group,
                route_proj=route_proj,
                transformer=transformer,
                is_sparse_gps=is_sparse_gps,
                has_io_params=has_io_params,
                has_movement_status=has_movement_status,
                has_engine_data=has_engine_data,
                min_approach_points=MIN_APPROACH_POINTS,
                approach_zone_radius=APPROACH_ZONE_RADIUS,
                group_idx=group_idx
            )
            
            if approach_result is None:
                continue
            
            # Count results
            results['total_approaches'] += 1
            results['approaches_detailed'].append(approach_result)
            
            # Track anticipation stats
            anticipation = approach_result.get('anticipation_category', 'no_deceleration')
            if anticipation in results['anticipation_stats']:
                results['anticipation_stats'][anticipation] += 1
            else:
                results['anticipation_stats']['no_deceleration'] += 1
            
            # Track entry compliance
            entry_compliant = approach_result.get('entry_compliant')
            if entry_compliant is True:
                results['entry_compliance']['compliant'] += 1
            elif entry_compliant is False:
                results['entry_compliance']['non_compliant'] += 1
            else:
                results['entry_compliance']['unknown'] += 1
            
            # Track confidence
            if approach_result.get('confidence'):
                all_confidences.append(approach_result['confidence'])
        
        # Calculate average confidence
        if all_confidences:
            results['avg_confidence'] = sum(all_confidences) / len(all_confidences)
        
        # Calculate compliance percentage
        total_evaluated = results['entry_compliance']['compliant'] + results['entry_compliance']['non_compliant']
        if total_evaluated > 0:
            results['entry_compliance_percentage'] = (results['entry_compliance']['compliant'] / total_evaluated) * 100
        else:
            results['entry_compliance_percentage'] = None
        
        # Print summary
        print(f"\n    📊 Roundabout Analysis Complete:")
        print(f"       Total roundabouts: {results['total_roundabouts']}")
        print(f"       Valid approaches: {results['total_approaches']}")
        print(f"       Anticipation - Good: {results['anticipation_stats']['good']}, "
              f"Moderate: {results['anticipation_stats']['moderate']}, "
              f"Late: {results['anticipation_stats']['late']}, "
              f"No decel: {results['anticipation_stats']['no_deceleration']}")
        print(f"       Entry compliance: {results['entry_compliance']['compliant']}/{total_evaluated} "
              f"({results['entry_compliance_percentage']:.0f}%)" if total_evaluated > 0 else "       Entry compliance: N/A")
        
        self.results['roundabouts'] = results
    
    
    def _analyze_single_roundabout_approach(self, roundabout_group, route_proj, transformer,
                                             is_sparse_gps, has_io_params, has_movement_status,
                                             has_engine_data, min_approach_points, 
                                             approach_zone_radius, group_idx):
        """
        Analyze a single roundabout approach with 3-layer detection.
        
        Returns detailed approach result or None if insufficient data.
        """
        from shapely.geometry import Point
        from shapely.ops import transform
        from geopy.distance import geodesic
        
        # Calculate group centroid
        group_lat = sum(rb['lat'] for rb in roundabout_group) / len(roundabout_group)
        group_lon = sum(rb['lon'] for rb in roundabout_group) / len(roundabout_group)
        
        print(f"\n    📍 Roundabout {group_idx + 1}: ({group_lat:.6f}, {group_lon:.6f})")
        
        # Project roundabout point
        rb_point = Point(group_lon, group_lat)
        try:
            rb_point_proj = transform(lambda x, y: transformer.transform(x, y), rb_point)
            roundabout_route_position = route_proj.project(rb_point_proj)
            distance_to_route = route_proj.distance(rb_point_proj)
        except Exception as e:
            print(f"       ❌ Projection failed: {e}")
            return None
        
        # ================================================================
        # Find GPS points near this roundabout
        # ================================================================
        nearby_points = []
        
        for idx, row in self.df.iterrows():
            # Use road-matched coordinates if available, else raw GPS
            if 'road_matched_lat' in self.df.columns and pd.notna(row.get('road_matched_lat')):
                point_lat = row['road_matched_lat']
                point_lon = row['road_matched_lon']
            else:
                point_lat = row['lat']
                point_lon = row['lon']
            
            try:
                gps_point = Point(point_lon, point_lat)
                gps_point_proj = transform(lambda x, y: transformer.transform(x, y), gps_point)
                
                # Distance along route from roundabout
                gps_route_position = route_proj.project(gps_point_proj)
                signed_distance = gps_route_position - roundabout_route_position
                
                # Perpendicular distance from route
                perp_distance = route_proj.distance(gps_point_proj)
                
                # Include points within approach zone
                if abs(signed_distance) <= approach_zone_radius and perp_distance <= 100:
                    nearby_points.append({
                        'idx': idx,
                        'lat': point_lat,
                        'lon': point_lon,
                        'signed_distance': signed_distance,  # Positive = before roundabout
                        'abs_distance': abs(signed_distance),
                        'perp_distance': perp_distance,
                        'speed_kmh': row['speed_kmh'],
                        'timestamp': row['timestamp'],
                        'io_movement': row.get('io_movement'),
                        'io_engine_load': row.get('io_engine_load'),
                        'io_green_driving_type': row.get('io_green_driving_type')
                    })
            except:
                continue
        
        print(f"       Found {len(nearby_points)} GPS points within {approach_zone_radius}m")
        
        if len(nearby_points) < min_approach_points:
            print(f"       ❌ Insufficient data ({len(nearby_points)} < {min_approach_points} points)")
            return None
        
        # Sort by timestamp to determine direction
        nearby_points.sort(key=lambda x: x['timestamp'])
        
        # ================================================================
        # Determine approach direction (approaching vs leaving)
        # ================================================================
        # If signed_distance decreases over time → approaching
        # If signed_distance increases over time → leaving
        
        if len(nearby_points) >= 2:
            first_dist = nearby_points[0]['signed_distance']
            last_dist = nearby_points[-1]['signed_distance']
            
            is_approaching = first_dist > last_dist  # Distance to roundabout decreasing
            direction = "approaching" if is_approaching else "leaving"
        else:
            is_approaching = True
            direction = "unknown"
        
        print(f"       Direction: {direction}")
        
        # ================================================================
        # Separate approach and exit points
        # ================================================================
        # Approach points: signed_distance > 0 (before roundabout)
        # Exit points: signed_distance < 0 (after roundabout)
        # Entry zone: -15m to +15m
                
        # ================================================================
        # ENTRY/EXIT DETECTION USING ROUNDABOUT GEOMETRY
        # ================================================================
        geometry_result = self._detect_entry_exit_using_geometry(
            roundabout_group=roundabout_group,
            nearby_points=nearby_points
        )
        
        # Use geometry-based entry/exit if available
        if geometry_result['entry_point'] or geometry_result['inside_points']:
            print(f"       ✅ Geometry-based detection: Entry={geometry_result['entry_speed']}, "
                  f"Inside={len(geometry_result['inside_points'])} points, Exit={geometry_result['exit_speed']}")
            
            # Entry speed from geometry detection
            entry_speed_geo = geometry_result['entry_speed']
            exit_speed_geo = geometry_result['exit_speed']
            inside_points = geometry_result['inside_points']
            time_inside = geometry_result['time_inside_seconds']
            stopped_inside = geometry_result['stopped_inside']
        else:
            print(f"       ⚠️ Geometry detection failed, using distance-based fallback")
            entry_speed_geo = None
            exit_speed_geo = None
            inside_points = []
            time_inside = None
            stopped_inside = False
        
        # ================================================================
        # FALLBACK: Distance-based separation (if geometry detection fails)
        # ================================================================
        approach_points = [p for p in nearby_points if p['signed_distance'] > 15]
        entry_points = [p for p in nearby_points if -15 <= p['signed_distance'] <= 15]
        exit_points = [p for p in nearby_points if p['signed_distance'] < -15]
        
        print(f"       Distance-based: Approach={len(approach_points)}, Entry zone={len(entry_points)}, Exit={len(exit_points)}")  
        # If we're leaving the roundabout, we can still analyze approach if we have approach points
        if not approach_points and not is_approaching:
            print(f"       ❌ No approach points (driver leaving roundabout)")

        
        # ================================================================
        # ZONE ANALYSIS: Calculate speeds in each zone
        # ================================================================
        # Zone 1: 150-100m (Good anticipation)
        # Zone 2: 100-65m (Moderate anticipation)
        # Zone 3: 65-0m (Late anticipation)
        
        zone_150_100 = [p for p in approach_points if 100 < p['signed_distance'] <= 150]
        zone_100_65 = [p for p in approach_points if 65 < p['signed_distance'] <= 100]
        zone_65_0 = [p for p in approach_points if 0 < p['signed_distance'] <= 65]
        
        def calc_zone_speed(points):
            if not points:
                return None
            speeds = [p['speed_kmh'] for p in points if p['speed_kmh'] is not None]
            return sum(speeds) / len(speeds) if speeds else None
        
        zone_150_speed = calc_zone_speed(zone_150_100)
        zone_100_speed = calc_zone_speed(zone_100_65)
        zone_65_speed = calc_zone_speed(zone_65_0)
                
        # ================================================================
        # ENTRY/EXIT SPEED: Prefer geometry-based, fallback to distance-based
        # ================================================================
        
        # Entry speed: Geometry > Distance-based > Zone-based
        if entry_speed_geo is not None:
            entry_speed = entry_speed_geo
            entry_method = 'geometry'
        elif entry_points:
            closest_entry = min(entry_points, key=lambda x: x['abs_distance'])
            entry_speed = closest_entry['speed_kmh']
            entry_method = 'distance'
        elif zone_65_0:
            closest_approach = min(zone_65_0, key=lambda x: x['signed_distance'])
            entry_speed = closest_approach['speed_kmh']
            entry_method = 'zone'
        else:
            entry_speed = None
            entry_method = 'none'
        
        # Exit speed: Geometry > Distance-based
        if exit_speed_geo is not None:
            exit_speed = exit_speed_geo
            exit_method = 'geometry'
        elif exit_points:
            first_exit = min(exit_points, key=lambda x: abs(x['signed_distance']))
            exit_speed = first_exit['speed_kmh']
            exit_method = 'distance'
        else:
            exit_speed = None
            exit_method = 'none'
        
        print(f"       Entry: {entry_speed:.0f} km/h ({entry_method})" if entry_speed else "       Entry: Unknown")
        print(f"       Exit: {exit_speed:.0f} km/h ({exit_method})" if exit_speed else "       Exit: Unknown")
        # ================================================================
        # ANTICIPATION DETECTION (3-Layer Approach)
        # ================================================================
        anticipation_result = self._detect_roundabout_anticipation(
            approach_points=approach_points,
            zone_150_speed=zone_150_speed,
            zone_100_speed=zone_100_speed,
            zone_65_speed=zone_65_speed,
            entry_speed=entry_speed,
            has_engine_data=has_engine_data,
            is_sparse_gps=is_sparse_gps
        )
        
        # ================================================================
        # ENTRY COMPLIANCE CHECK
        # ================================================================
        ENTRY_SPEED_LIMIT = 34  # km/h
        
        entry_compliant = None
        if entry_speed is not None:
            entry_compliant = entry_speed <= ENTRY_SPEED_LIMIT
        
        # ================================================================
        # CONFIDENCE SCORING
        # ================================================================
        confidence = self._calculate_roundabout_confidence(
            approach_points=approach_points,
            entry_points=entry_points,
            has_io_params=has_io_params,
            is_sparse_gps=is_sparse_gps,
            zone_150_speed=zone_150_speed,
            zone_100_speed=zone_100_speed,
            zone_65_speed=zone_65_speed
        )
        

        # ================================================================
        # BUILD RESULT
        # ================================================================
        result = {
            # Location
            'roundabout_lat': group_lat,
            'roundabout_lon': group_lon,
            'roundabout_id': roundabout_group[0].get('id', 'unknown'),
            'osm_entries': len(roundabout_group),
            
            # Zone speeds
            'zone_150m_avg_speed': zone_150_speed,
            'zone_100m_avg_speed': zone_100_speed,
            'zone_65m_avg_speed': zone_65_speed,
            
            # Entry/Exit (ENHANCED with geometry detection)
            'entry_speed': entry_speed,
            'entry_method': entry_method,  # NEW: 'geometry', 'distance', or 'zone'
            'exit_speed': exit_speed,
            'exit_method': exit_method,    # NEW: 'geometry' or 'distance'
            'entry_compliant': entry_speed <= ENTRY_SPEED_LIMIT if entry_speed else None,
            
            # Inside roundabout data (NEW)
            'inside_points_count': len(inside_points),
            'time_inside_seconds': time_inside,
            'stopped_inside': stopped_inside,
            
            # Anticipation
            'anticipation_category': anticipation_result['category'],
            'anticipation_distance': anticipation_result['distance'],
            'deceleration_detected': anticipation_result['deceleration_detected'],
            'total_speed_drop': anticipation_result['total_speed_drop'],
            
            # Data quality
            'approach_points': len(approach_points),
            'entry_points': len(entry_points),
            'exit_points': len(exit_points),
            'direction': direction,
            'confidence': confidence,
            
            # Backward compatibility
            'roundabout_avg_speed': entry_speed,
            'roundabout_compliant': entry_speed <= ENTRY_SPEED_LIMIT if entry_speed else None
        }

        
        print(f"       ✅ Anticipation: {anticipation_result['category']} "
              f"(drop: {anticipation_result['total_speed_drop']:.1f} km/h)" 
              if anticipation_result['total_speed_drop'] else 
              f"       ✅ Anticipation: {anticipation_result['category']}")
        print(f"       ✅ Entry: {entry_speed:.0f} km/h → {'Compliant' if entry_compliant else 'Non-compliant'}" 
              if entry_speed else "       ⚠️ Entry speed: Unknown")
        
        return result

        
    
    def _detect_roundabout_anticipation(self, approach_points, zone_150_speed, zone_100_speed,
                                         zone_65_speed, entry_speed, has_engine_data, is_sparse_gps):
        """
        Detect anticipation using trend-based analysis.
        
        Returns dict with:
        - category: 'good', 'moderate', 'late', 'no_deceleration'
        - distance: Distance where deceleration started (meters)
        - deceleration_detected: Boolean
        - total_speed_drop: Total speed reduction in approach
        """
        
        result = {
            'category': 'no_deceleration',
            'distance': None,
            'deceleration_detected': False,
            'total_speed_drop': 0
        }
        
        # Collect available speeds in order (far to near)
        speeds_by_distance = []
        
        if zone_150_speed is not None:
            speeds_by_distance.append(('150m', 125, zone_150_speed))  # Midpoint of zone
        if zone_100_speed is not None:
            speeds_by_distance.append(('100m', 82, zone_100_speed))
        if zone_65_speed is not None:
            speeds_by_distance.append(('65m', 32, zone_65_speed))
        if entry_speed is not None:
            speeds_by_distance.append(('entry', 0, entry_speed))
        
        if len(speeds_by_distance) < 2:
            # Not enough data points
            return result
        
        # Calculate total speed drop (first zone to entry)
        first_speed = speeds_by_distance[0][2]
        last_speed = speeds_by_distance[-1][2]
        total_drop = first_speed - last_speed
        
        result['total_speed_drop'] = total_drop
        
        # ================================================================
        # DECELERATION DETECTION THRESHOLDS
        # ================================================================
        # Client requirement: MAX(12 km/h, 10% of approach speed)
        min_drop_absolute = 12  # km/h
        min_drop_percentage = 0.10  # 10%
        min_required_drop = max(min_drop_absolute, first_speed * min_drop_percentage)
        
        # For sparse GPS, be more lenient
        if is_sparse_gps:
            min_required_drop = max(8, first_speed * 0.08)  # 8 km/h or 8%
        
        # Check if significant deceleration occurred
        if total_drop >= min_required_drop:
            result['deceleration_detected'] = True
            
            # Find WHERE deceleration started
            # Walk through zones from far to near, find first significant drop
            
            for i in range(len(speeds_by_distance) - 1):
                current_zone, current_dist, current_speed = speeds_by_distance[i]
                next_zone, next_dist, next_speed = speeds_by_distance[i + 1]
                
                zone_drop = current_speed - next_speed
                
                # Check if this zone has significant deceleration
                # Use lower threshold for detecting START of deceleration
                zone_threshold = max(5, current_speed * 0.05)  # 5 km/h or 5%
                
                if zone_drop >= zone_threshold:
                    # Deceleration started in this zone
                    result['distance'] = current_dist
                    
                    # Categorize based on where deceleration started
                    if current_dist >= 100:
                        result['category'] = 'good'
                    elif current_dist >= 65:
                        result['category'] = 'moderate'
                    else:
                        result['category'] = 'late'
                    
                    break
            
            # If we detected deceleration but couldn't find start point
            if result['category'] == 'no_deceleration' and result['deceleration_detected']:
                # Use overall trend to categorize
                if zone_150_speed and zone_150_speed > last_speed + min_required_drop:
                    result['category'] = 'good'
                    result['distance'] = 125
                elif zone_100_speed and zone_100_speed > last_speed + min_required_drop * 0.7:
                    result['category'] = 'moderate'
                    result['distance'] = 82
                else:
                    result['category'] = 'late'
                    result['distance'] = 32
        
        # ================================================================
        # SPECIAL CASE: Already slow (no deceleration needed)
        # ================================================================
        if first_speed < 40 and last_speed < 35:
            # Driver was already going slow, minimal deceleration needed
            # This is actually GOOD behavior
            if total_drop >= 3:  # Even small drop counts
                result['deceleration_detected'] = True
                result['category'] = 'good'
                result['distance'] = 125
        
        return result
    
    
    def _calculate_roundabout_confidence(self, approach_points, entry_points, has_io_params,
                                          is_sparse_gps, zone_150_speed, zone_100_speed, zone_65_speed):
        """
        Calculate confidence score (0-100) for roundabout analysis.
        """
        
        confidence = 50  # Base confidence
        
        # Data quantity bonuses
        if len(approach_points) >= 5:
            confidence += 15
        elif len(approach_points) >= 3:
            confidence += 10
        elif len(approach_points) >= 2:
            confidence += 5
        
        if len(entry_points) >= 1:
            confidence += 10
        
        # Zone coverage bonuses
        zones_with_data = sum([
            zone_150_speed is not None,
            zone_100_speed is not None,
            zone_65_speed is not None
        ])
        confidence += zones_with_data * 5
        
        # IO params bonus
        if has_io_params:
            confidence += 10
        
        # Sparse GPS penalty
        if is_sparse_gps:
            confidence -= 10
        
        # Cap at 0-100
        confidence = max(0, min(100, confidence))
        
        return confidence

    def _detect_entry_exit_using_geometry(self, roundabout_group, nearby_points):
        """
        Detect precise entry/exit points using roundabout polygon geometry.
        
        Returns:
            dict with entry_point, exit_point, inside_points, entry_speed, exit_speed
        """
        from shapely.geometry import Point, Polygon, LineString
        
        result = {
            'entry_point': None,
            'exit_point': None,
            'inside_points': [],
            'entry_speed': None,
            'exit_speed': None,
            'time_inside_seconds': None,
            'stopped_inside': False
        }
        
        # Get roundabout geometry
        # Try to build polygon from geometry data
        roundabout_polygon = None
        
        for rb in roundabout_group:
            geom = rb.get('geometry')
            
            if geom is not None:
                try:
                    if isinstance(geom, Polygon):
                        roundabout_polygon = geom
                        break
                    elif isinstance(geom, list) and len(geom) >= 3:
                        # List of coordinates - create polygon
                        # Check format: [(lon, lat), ...] or [(lat, lon), ...]
                        if all(isinstance(p, (list, tuple)) and len(p) >= 2 for p in geom):
                            # Assume (lon, lat) format
                            roundabout_polygon = Polygon(geom)
                            break
                except Exception as e:
                    continue
        
        # Fallback: Create circular buffer around centroid
        if roundabout_polygon is None:
            centroid_lat = sum(rb['lat'] for rb in roundabout_group) / len(roundabout_group)
            centroid_lon = sum(rb['lon'] for rb in roundabout_group) / len(roundabout_group)
            
            # Create approximate circular polygon (15m radius ≈ 0.00015 degrees)
            center = Point(centroid_lon, centroid_lat)
            roundabout_polygon = center.buffer(0.00015)  # ~15m radius
            print(f"       ⚠️ Using circular approximation (no geometry data)")
        
        if roundabout_polygon is None:
            return result
        
        # Sort points by timestamp
        sorted_points = sorted(nearby_points, key=lambda x: x['timestamp'])
        
        # Classify each point: inside or outside roundabout
        point_classifications = []
        
        for point in sorted_points:
            gps_point = Point(point['lon'], point['lat'])
            is_inside = roundabout_polygon.contains(gps_point)
            
            point_classifications.append({
                **point,
                'inside_roundabout': is_inside
            })
            
            if is_inside:
                result['inside_points'].append(point)
        
        # Find entry point: last OUTSIDE point before first INSIDE point
        # Find exit point: first OUTSIDE point after last INSIDE point
        
        prev_point = None
        entry_found = False
        last_inside_idx = None
        
        for i, point in enumerate(point_classifications):
            if point['inside_roundabout']:
                if not entry_found and prev_point is not None:
                    # This is entry: prev_point is entry point
                    result['entry_point'] = prev_point
                    result['entry_speed'] = prev_point['speed_kmh']
                    entry_found = True
                last_inside_idx = i
            
            prev_point = point
        
        # Find exit point
        if last_inside_idx is not None and last_inside_idx < len(point_classifications) - 1:
            exit_point = point_classifications[last_inside_idx + 1]
            if not exit_point['inside_roundabout']:
                result['exit_point'] = exit_point
                result['exit_speed'] = exit_point['speed_kmh']
        
        # Calculate time inside
        if result['inside_points']:
            timestamps = [p['timestamp'] for p in result['inside_points']]
            if len(timestamps) >= 2:
                time_diff = (max(timestamps) - min(timestamps)).total_seconds()
                result['time_inside_seconds'] = time_diff
            
            # Check if stopped inside (using io_movement or speed)
            for p in result['inside_points']:
                if p.get('io_movement') == 0:
                    result['stopped_inside'] = True
                    break
                if p.get('speed_kmh', 100) < 3:
                    result['stopped_inside'] = True
                    break
        
        return result
    
    def _group_roundabouts_by_proximity(self, roundabouts, max_distance=100):
        """
        Group roundabouts that are close together (same physical roundabout).
        
        Args:
            roundabouts: List of roundabout dictionaries
            max_distance: Maximum distance in meters to consider same group
        
        Returns:
            List of roundabout groups
        """
        from geopy.distance import geodesic
        
        if not roundabouts:
            return []
        
        groups = []
        ungrouped = roundabouts.copy()
        
        while ungrouped:
            current_group = [ungrouped.pop(0)]
            
            added = True
            while added:
                added = False
                
                for i in range(len(ungrouped) - 1, -1, -1):
                    candidate = ungrouped[i]
                    
                    for member in current_group:
                        try:
                            distance = geodesic(
                                (candidate['lat'], candidate['lon']),
                                (member['lat'], member['lon'])
                            ).meters
                            
                            if distance <= max_distance:
                                current_group.append(ungrouped.pop(i))
                                added = True
                                break
                        except:
                            continue
            
            groups.append(current_group)
        
        return groups

    # ========================================================================
    # STOP SIGN ANALYSIS - Two versions based on road context
    # ========================================================================    
    def _analyze_stop_signs_with_context(self, df=None, stop_signs=None, route_geometry=None):
        """
        ENHANCED IMPLEMENTATION: Stop sign compliance analysis using road-based filtering.
        
        Features:
        - Filters stop signs by OSM ID matching (stop sign node in driven edges)
        - 3-layer detection (direct evidence, indicators, segment analysis)
        - Uses IO parameters (io_movement, io_engine_load) when available
        - Adaptive thresholds for sparse GPS
        - Confidence scoring
        
        Args:
            df: GPS DataFrame (optional, defaults to self.df)
            stop_signs: List of stop signs (optional, defaults to map_manager data)
            route_geometry: Route geometry (optional, defaults to map_manager)
        """
        
        # Use defaults if not provided
        if df is None:
            df = self.df
        if stop_signs is None:
            stop_signs = self.map_manager.map_data.get('stop_signs', [])
        if route_geometry is None:
            route_geometry = getattr(self.map_manager, 'route_geometry', None)
        
        print(f"    🛑 Analyzing stop sign compliance (ENHANCED METHOD)...")
        print(f"    📊 Total stop signs in map: {len(stop_signs)}")
        
        # ================================================================
        # STEP 1: Initialize results structure
        # ================================================================
        results = {
            'total_stop_signs': 0,
            'total_approaches': 0,
            'compliance': {
                'stop_ok': 0,
                'stop_ko': 0,
                'uncertain': 0
            },
            'compliance_percentage': None,
            'approaches_detailed': [],
            'detection_mode': 'basic',
            'avg_confidence': 0
        }
        
        if not stop_signs:
            print(f"    ❌ No stop signs in map data")
            self.results['stop_signs'] = results
            return
        
        # ================================================================
        # STEP 2: Detect GPS quality and available features
        # ================================================================
        has_io_params = 'has_io_params' in df.columns and df['has_io_params'].any()
        has_movement_status = 'io_movement' in df.columns and df['io_movement'].notna().any()
        has_engine_data = 'io_engine_load' in df.columns and df['io_engine_load'].notna().any()
        
        # Calculate GPS interval
        avg_interval = df['time_diff_s'].mean() if 'time_diff_s' in df.columns else 6.0
        is_sparse_gps = avg_interval > 4
        
        # Set detection mode
        if has_movement_status and has_engine_data:
            results['detection_mode'] = 'enhanced'
        elif has_io_params:
            results['detection_mode'] = 'standard'
        else:
            results['detection_mode'] = 'basic'
        
        print(f"    🎯 Detection mode: {results['detection_mode']}")
        print(f"    📡 GPS interval: {avg_interval:.1f}s {'(SPARSE)' if is_sparse_gps else '(OK)'}")
        
        # ================================================================
        # STEP 3: Extract driven node IDs from edges (for OSM ID filtering)
        # ================================================================
        driven_node_ids = set()
        driven_edges = getattr(self.map_manager, 'driven_road_ids', set())
        
        for edge in driven_edges:
            if isinstance(edge, tuple) and len(edge) >= 2:
                driven_node_ids.add(edge[0])  # Start node
                driven_node_ids.add(edge[1])  # End node
        
        print(f"    📊 Driven edges: {len(driven_edges)}, Driven nodes: {len(driven_node_ids)}")
        
        # ================================================================
        # STEP 4: Filter stop signs on driven route (OSM ID MATCHING)
        # ================================================================
        
        # PRIMARY METHOD: OSM ID matching
        relevant_stop_signs_osmid = []
        
        for stop_sign in stop_signs:
            stop_osmid = stop_sign.get('osmid', stop_sign.get('osm_id'))
            
            if stop_osmid and stop_osmid in driven_node_ids:
                stop_lat = stop_sign.get('lat')
                stop_lon = stop_sign.get('lon')
                
                if stop_lat and stop_lon:
                    relevant_stop_signs_osmid.append({
                        'stop_sign': stop_sign,
                        'lat': stop_lat,
                        'lon': stop_lon,
                        'id': stop_sign.get('id', stop_osmid),
                        'osmid': stop_osmid,
                        'match_method': 'osmid'
                    })
        
        print(f"    ✅ Found {len(relevant_stop_signs_osmid)} stop signs by OSM ID match")
        
        # FALLBACK: If OSM ID matching finds very few, add distance-based
        # (Only if we have route geometry and found less than expected)
        relevant_stop_signs = relevant_stop_signs_osmid.copy()
        
        if len(relevant_stop_signs) < 3 and route_geometry is not None:
            print(f"    ⚠️ Few OSM matches, checking distance-based fallback...")
            
            from pyproj import Transformer
            from shapely.ops import transform
            from shapely.geometry import Point
            
            transformer = Transformer.from_crs("EPSG:4326", "EPSG:2154", always_xy=True)
            route_proj = transform(lambda x, y: transformer.transform(x, y), route_geometry)
            
            # Get GPS bounding box
            lat_min, lat_max = df['lat'].min(), df['lat'].max()
            lon_min, lon_max = df['lon'].min(), df['lon'].max()
            bbox_buffer = 0.003  # ~300m
            
            # Already matched OSM IDs (to avoid duplicates)
            matched_osmids = {s['osmid'] for s in relevant_stop_signs if s.get('osmid')}
            
            for stop_sign in stop_signs:
                stop_osmid = stop_sign.get('osmid', stop_sign.get('osm_id'))
                
                # Skip if already matched
                if stop_osmid in matched_osmids:
                    continue
                
                stop_lat = stop_sign.get('lat')
                stop_lon = stop_sign.get('lon')
                
                if stop_lat is None or stop_lon is None:
                    continue
                
                # Bounding box filter
                if not (lat_min - bbox_buffer <= stop_lat <= lat_max + bbox_buffer):
                    continue
                if not (lon_min - bbox_buffer <= stop_lon <= lon_max + bbox_buffer):
                    continue
                
                # Distance check (strict: 1m)
                try:
                    stop_point = Point(stop_lon, stop_lat)
                    stop_point_proj = transform(lambda x, y: transformer.transform(x, y), stop_point)
                    distance_to_route = route_proj.distance(stop_point_proj)
                    
                    if distance_to_route <= 1:  # Strict 5m threshold
                        relevant_stop_signs.append({
                            'stop_sign': stop_sign,
                            'lat': stop_lat,
                            'lon': stop_lon,
                            'id': stop_sign.get('id', stop_osmid),
                            'osmid': stop_osmid,
                            'distance_to_route': distance_to_route,
                            'match_method': 'distance'
                        })
                except:
                    continue
            
            distance_matches = len(relevant_stop_signs) - len(relevant_stop_signs_osmid)
            print(f"    ✅ Added {distance_matches} stop signs by distance fallback (≤15m)")
        
        print(f"    ✅ Total stop signs on route: {len(relevant_stop_signs)}")
        
        if not relevant_stop_signs:
            print(f"    ❌ No stop signs found on driven route")
            self.results['stop_signs'] = results
            return
        
        results['total_stop_signs'] = len(relevant_stop_signs)
        
        # ================================================================
        # STEP 5: Setup coordinate transformation for analysis
        # ================================================================
        from pyproj import Transformer
        from shapely.ops import transform
        from shapely.geometry import Point
        from geopy.distance import geodesic
        
        transformer = Transformer.from_crs("EPSG:4326", "EPSG:2154", always_xy=True)
        
        if route_geometry:
            route_proj = transform(lambda x, y: transformer.transform(x, y), route_geometry)
        else:
            route_proj = None
        
        # ================================================================
        # STEP 6: Analyze each stop sign
        # ================================================================
        all_confidences = []
        
        for stop_info in relevant_stop_signs:
            stop_lat = stop_info['lat']
            stop_lon = stop_info['lon']
            stop_id = stop_info['id']
            match_method = stop_info.get('match_method', 'unknown')
            
            print(f"\n    🛑 Stop sign {stop_id} ({match_method}):")
            
            # Find GPS points near this stop sign
            approach_result = self._analyze_single_stop_approach(
                df=df,
                stop_lat=stop_lat,
                stop_lon=stop_lon,
                stop_id=stop_id,
                route_proj=route_proj,
                transformer=transformer,
                is_sparse_gps=is_sparse_gps,
                has_movement_status=has_movement_status,
                has_engine_data=has_engine_data
            )
            
            if approach_result is None:
                continue
            
            # Add match method to result
            approach_result['match_method'] = match_method
            
            # Count results
            results['total_approaches'] += 1
            results['approaches_detailed'].append(approach_result)
            
            # Track compliance
            compliance = approach_result.get('compliance', 'uncertain')
            if compliance == 'STOP_OK':
                results['compliance']['stop_ok'] += 1
            elif compliance == 'STOP_KO':
                results['compliance']['stop_ko'] += 1
            else:
                results['compliance']['uncertain'] += 1
            
            # Track confidence
            if approach_result.get('confidence'):
                all_confidences.append(approach_result['confidence'])
        
        # Calculate averages
        if all_confidences:
            results['avg_confidence'] = sum(all_confidences) / len(all_confidences)
        
        # Calculate compliance percentage
        total_evaluated = results['compliance']['stop_ok'] + results['compliance']['stop_ko']
        if total_evaluated > 0:
            results['compliance_percentage'] = (results['compliance']['stop_ok'] / total_evaluated) * 100
        
        # Print summary
        print(f"\n    📊 Stop Sign Analysis Complete:")
        print(f"       Total stop signs: {results['total_stop_signs']}")
        print(f"       Valid approaches: {results['total_approaches']}")
        print(f"       STOP_OK: {results['compliance']['stop_ok']}, "
              f"STOP_KO: {results['compliance']['stop_ko']}, "
              f"Uncertain: {results['compliance']['uncertain']}")
        if results['compliance_percentage'] is not None:
            print(f"       Compliance: {results['compliance_percentage']:.0f}%")
        
        self.results['stop_signs'] = results
    
    def _analyze_single_stop_approach(self, df, stop_lat, stop_lon, stop_id, 
                                       route_proj, transformer, is_sparse_gps,
                                       has_movement_status, has_engine_data):
        """
        Analyze a single stop sign approach using 3-layer detection.
        
        Returns detailed approach result or None if insufficient data.
        """
        from shapely.geometry import Point
        from shapely.ops import transform
        from geopy.distance import geodesic
        
        # ================================================================
        # Find GPS points near this stop sign
        # ================================================================
        APPROACH_RADIUS = 50  # meters
        nearby_points = []
        
        for idx, row in df.iterrows():
            gps_lat = row['lat']
            gps_lon = row['lon']
            
            # Calculate distance to stop sign
            distance = geodesic((stop_lat, stop_lon), (gps_lat, gps_lon)).meters
            
            if distance <= APPROACH_RADIUS:
                nearby_points.append({
                    'idx': idx,
                    'lat': gps_lat,
                    'lon': gps_lon,
                    'distance': distance,
                    'speed_kmh': row['speed_kmh'],
                    'timestamp': row['timestamp'],
                    'io_movement': row.get('io_movement'),
                    'io_engine_load': row.get('io_engine_load'),
                    'io_green_driving_type': row.get('io_green_driving_type')
                })
        
        print(f"       Found {len(nearby_points)} GPS points within {APPROACH_RADIUS}m")
        
        if len(nearby_points) < 1:
            print(f"       ❌ No GPS points near stop sign")
            return None
        
        # Sort by timestamp
        nearby_points.sort(key=lambda x: x['timestamp'])
        
        # ================================================================
        # 3-LAYER DETECTION
        # ================================================================
        
        compliance = None
        confidence = 50
        detection_layer = None
        evidence = []
        
        # Get speeds and minimum speed
        speeds = [p['speed_kmh'] for p in nearby_points if p['speed_kmh'] is not None]
        min_speed = min(speeds) if speeds else None
        closest_point = min(nearby_points, key=lambda x: x['distance'])
        closest_speed = closest_point['speed_kmh']
        
        # ================================================================
        # LAYER 1: Direct Evidence (90-100% confidence)
        # ================================================================
        
        # Check io_movement (vehicle stationary)
        movement_values = [p['io_movement'] for p in nearby_points if p['io_movement'] is not None]
        has_stationary = 0 in movement_values
        
        if has_stationary:
            compliance = 'STOP_OK'
            confidence = 95
            detection_layer = 'layer1_io_movement'
            evidence.append('io_movement=0 (stationary)')
            print(f"       ✅ Layer 1: io_movement=0 detected → STOP_OK (95%)")
        
        # Check speed = 0
        elif min_speed is not None and min_speed < 2:
            compliance = 'STOP_OK'
            confidence = 92
            detection_layer = 'layer1_speed_zero'
            evidence.append(f'speed={min_speed:.0f} km/h (near zero)')
            print(f"       ✅ Layer 1: Speed near zero ({min_speed:.0f} km/h) → STOP_OK (92%)")
        
        # Check harsh brake event
        harsh_brake = any(p.get('io_green_driving_type') == 2 for p in nearby_points)
        if harsh_brake and min_speed is not None and min_speed < 10:
            if compliance is None:  # Don't override if already set
                compliance = 'STOP_OK'
                confidence = 88
                detection_layer = 'layer1_harsh_brake'
                evidence.append('harsh brake + low speed')
                print(f"       ✅ Layer 1: Harsh brake + low speed → STOP_OK (88%)")
        
        # Check maintained high speed (definite violation)
        if compliance is None and min_speed is not None and min_speed > 25:
            compliance = 'STOP_KO'
            confidence = 90
            detection_layer = 'layer1_high_speed'
            evidence.append(f'min_speed={min_speed:.0f} km/h (too high)')
            print(f"       ❌ Layer 1: High speed maintained ({min_speed:.0f} km/h) → STOP_KO (90%)")
        
        # ================================================================
        # LAYER 2: Strong Indicators (70-85% confidence)
        # ================================================================
        
        if compliance is None:
            # Low speed near stop sign
            if min_speed is not None and min_speed < 8:
                compliance = 'STOP_OK'
                confidence = 80
                detection_layer = 'layer2_low_speed'
                evidence.append(f'min_speed={min_speed:.0f} km/h (very low)')
                print(f"       ✅ Layer 2: Very low speed ({min_speed:.0f} km/h) → STOP_OK (80%)")
            
            # Multiple points clustered (driver slowed significantly)
            elif len(nearby_points) >= 3 and min_speed is not None and min_speed < 15:
                # Check clustering - multiple points in small area suggests slowing
                distances = [p['distance'] for p in nearby_points]
                distance_spread = max(distances) - min(distances)
                
                if distance_spread < 20:  # Points clustered within 20m
                    compliance = 'STOP_OK'
                    confidence = 75
                    detection_layer = 'layer2_clustering'
                    evidence.append(f'clustered points + low speed')
                    print(f"       ✅ Layer 2: Clustered points + low speed → STOP_OK (75%)")
            
            # Engine load drop (if available)
            if compliance is None and has_engine_data:
                engine_loads = [p['io_engine_load'] for p in nearby_points if p.get('io_engine_load') is not None]
                if len(engine_loads) >= 2:
                    load_drop = max(engine_loads) - min(engine_loads)
                    if load_drop > 40:  # Significant engine load drop
                        compliance = 'STOP_OK'
                        confidence = 72
                        detection_layer = 'layer2_engine_load'
                        evidence.append(f'engine load drop: {load_drop:.0f}%')
                        print(f"       ✅ Layer 2: Engine load drop ({load_drop:.0f}%) → STOP_OK (72%)")
            
            # Moderate speed maintained (likely violation)
            if compliance is None and min_speed is not None and min_speed > 15:
                compliance = 'STOP_KO'
                confidence = 75
                detection_layer = 'layer2_moderate_speed'
                evidence.append(f'min_speed={min_speed:.0f} km/h (too fast)')
                print(f"       ❌ Layer 2: Moderate speed ({min_speed:.0f} km/h) → STOP_KO (75%)")
        
        # ================================================================
        # LAYER 3: Segment Analysis (50-70% confidence)
        # ================================================================
        
        if compliance is None and len(nearby_points) >= 2:
            # Check for V-shaped speed pattern (approach → slow → exit)
            if len(speeds) >= 3:
                first_speed = speeds[0]
                last_speed = speeds[-1]
                
                if first_speed > min_speed and last_speed > min_speed:
                    speed_drop = first_speed - min_speed
                    if speed_drop > 10:
                        compliance = 'STOP_OK'
                        confidence = 65
                        detection_layer = 'layer3_v_pattern'
                        evidence.append(f'V-pattern: {first_speed:.0f}→{min_speed:.0f}→{last_speed:.0f}')
                        print(f"       ✅ Layer 3: V-shaped pattern → STOP_OK (65%)")
            
            # Sparse GPS with low-ish speed - give benefit of doubt
            if compliance is None and is_sparse_gps and min_speed is not None and min_speed < 20:
                compliance = 'STOP_OK'
                confidence = 60
                detection_layer = 'layer3_sparse_benefit'
                evidence.append(f'sparse GPS + moderate speed ({min_speed:.0f} km/h)')
                print(f"       ⚠️ Layer 3: Sparse GPS, giving benefit of doubt → STOP_OK (60%)")
        
        # ================================================================
        # FALLBACK: Uncertain
        # ================================================================
        
        if compliance is None:
            compliance = 'UNCERTAIN'
            confidence = 50
            detection_layer = 'fallback'
            evidence.append('insufficient data for determination')
            print(f"       ❓ Fallback: Insufficient evidence → UNCERTAIN (50%)")
        
        # ================================================================
        # BUILD RESULT
        # ================================================================
        
        result = {
            # Location
            'stop_sign_lat': stop_lat,
            'stop_sign_lon': stop_lon,
            'stop_sign_id': stop_id,
            
            # Compliance
            'compliance': compliance,
            'confidence': confidence,
            'detection_layer': detection_layer,
            'evidence': evidence,
            
            # Speed data
            'min_speed': min_speed,
            'closest_speed': closest_speed,
            'speeds': speeds,
            
            # Data quality
            'gps_points': len(nearby_points),
            'closest_distance': closest_point['distance'],
            
            # IO data used
            'had_io_movement': has_movement_status,
            'had_engine_data': has_engine_data
        }
        
        return result    
                
    
    
    def _find_stop_signs_on_route(self, df, stop_signs, route_geometry=None, max_distance=30):
        """
        Find stop signs that are on or near the driven route.
        
        Args:
            df: DataFrame with GPS points
            stop_signs: List of all stop signs in area
            route_geometry: Optional route geometry (LineString)
            max_distance: Maximum distance from route to consider (meters)
            
        Returns:
            List of stop signs on route with additional metadata
        """
        
        stop_signs_on_route = []
        
        # Build route line from GPS points if no geometry provided
        if route_geometry is None:
            route_coords = list(zip(df['lon'], df['lat']))
            if len(route_coords) >= 2:
                route_geometry = LineString(route_coords)
            else:
                return []
        
        for stop_sign in stop_signs:
            # Get stop sign location
            if isinstance(stop_sign, dict):
                stop_lat = stop_sign.get('lat', stop_sign.get('latitude'))
                stop_lon = stop_sign.get('lon', stop_sign.get('longitude'))
            else:
                # Assume tuple (lat, lon)
                stop_lat, stop_lon = stop_sign[0], stop_sign[1]
            
            if stop_lat is None or stop_lon is None:
                continue
            
            # Check distance to route
            stop_point = Point(stop_lon, stop_lat)
            distance_to_route = self._point_to_line_distance_meters(stop_point, route_geometry)
            
            if distance_to_route <= max_distance:
                stop_signs_on_route.append({
                    'lat': stop_lat,
                    'lon': stop_lon,
                    'distance_to_route': distance_to_route,
                    'original_data': stop_sign
                })
        
        return stop_signs_on_route
    
    
    def _point_to_line_distance_meters(self, point, line):
        """Calculate distance from point to line in meters (approximate)."""
        try:
            # Get nearest point on line
            nearest = line.interpolate(line.project(point))
            
            # Calculate geodesic distance
            distance = geodesic(
                (point.y, point.x),  # lat, lon
                (nearest.y, nearest.x)
            ).meters
            
            return distance
        except:
            return float('inf')
        
    def _check_layer1_direct_evidence(self, approach_points, zone_points, closest_point,
                                       has_movement_status, has_harsh_events, evidence):
        """
        Layer 1: Check for direct evidence of stop.
        - io_movement == 0 (vehicle stationary)
        - speed_kmh == 0
        - io_green_driving_type == 2 (harsh brake)
        """
        
        result = {'conclusive': False, 'result': None, 'confidence': None, 'reason': ''}
        
        # Check 1: Movement status (io240 == 0)
        if has_movement_status and 'io_movement' in zone_points.columns:
            stationary_points = zone_points[zone_points['io_movement'] == 0]
            if len(stationary_points) > 0:
                evidence['stationary_points'] = len(stationary_points)
                evidence['detection_method'] = 'io_movement'
                result['conclusive'] = True
                result['result'] = 'STOP_OK'
                result['confidence'] = 95
                result['reason'] = f'Vehicle stationary detected (io240=0) at {len(stationary_points)} point(s)'
                return result
        
        # Check 2: Speed == 0 in zone
        if 'speed_kmh' in zone_points.columns:
            zero_speed_points = zone_points[zone_points['speed_kmh'] == 0]
            if len(zero_speed_points) > 0:
                evidence['zero_speed_points'] = len(zero_speed_points)
                evidence['detection_method'] = 'zero_speed'
                result['conclusive'] = True
                result['result'] = 'STOP_OK'
                result['confidence'] = 95
                result['reason'] = f'Zero speed recorded at {len(zero_speed_points)} point(s) in stop zone'
                return result
        
        # Check 3: Speed == 0 at closest point
        if closest_point['speed_kmh'] == 0:
            evidence['closest_speed'] = 0
            evidence['detection_method'] = 'zero_speed_closest'
            result['conclusive'] = True
            result['result'] = 'STOP_OK'
            result['confidence'] = 90
            result['reason'] = f'Zero speed at closest point ({evidence["closest_distance"]:.1f}m from stop sign)'
            return result
        
        # Check 4: Harsh brake event in approach zone
        if has_harsh_events and 'io_green_driving_type' in approach_points.columns:
            harsh_brake_points = approach_points[approach_points['io_green_driving_type'] == 2]
            if len(harsh_brake_points) > 0:
                evidence['harsh_brake_detected'] = True
                evidence['detection_method'] = 'harsh_brake'
                # Harsh brake suggests stop, but check if followed by low speed
                min_speed_after_brake = approach_points['speed_kmh'].min()
                if min_speed_after_brake < 10:
                    result['conclusive'] = True
                    result['result'] = 'STOP_OK'
                    result['confidence'] = 85
                    result['reason'] = f'Harsh brake detected with speed drop to {min_speed_after_brake:.0f} km/h'
                    return result
        
        # Check 5: High speed maintained (clear violation)
        min_speed_in_zone = zone_points['speed_kmh'].min() if len(zone_points) > 0 else closest_point['speed_kmh']
        if min_speed_in_zone > 25:
            evidence['min_speed_in_zone'] = min_speed_in_zone
            evidence['detection_method'] = 'high_speed_violation'
            result['conclusive'] = True
            result['result'] = 'STOP_KO'
            result['confidence'] = 90
            result['reason'] = f'High speed maintained through stop zone (min {min_speed_in_zone:.0f} km/h)'
            return result
        
        # Not conclusive at Layer 1
        evidence['layer1_min_speed'] = min_speed_in_zone
        return result
    
    
    def _check_layer2_strong_indicators(self, approach_points, zone_points, closest_point,
                                         has_engine_data, is_sparse_gps, evidence):
        """
        Layer 2: Check strong indicators of stop.
        - Low speed (< 5 km/h) near stop sign
        - Multiple points clustered in zone
        - Engine load pattern (if available)
        """
        
        result = {'conclusive': False, 'result': None, 'confidence': None, 'reason': ''}
        
        # Get minimum speed in approach
        min_speed_approach = approach_points['speed_kmh'].min()
        min_speed_zone = zone_points['speed_kmh'].min() if len(zone_points) > 0 else min_speed_approach
        closest_speed = closest_point['speed_kmh']
        
        evidence['min_speed_approach'] = min_speed_approach
        evidence['min_speed_zone'] = min_speed_zone
        evidence['closest_speed'] = closest_speed
        
        # Check 1: Very low speed (< 5 km/h) near stop sign
        if min_speed_zone < 5 or closest_speed < 5:
            evidence['detection_method'] = 'low_speed'
            result['conclusive'] = True
            result['result'] = 'STOP_OK'
            result['confidence'] = 85 if min_speed_zone < 3 else 75
            result['reason'] = f'Very low speed detected ({min(min_speed_zone, closest_speed):.0f} km/h)'
            return result
        
        # Check 2: Multiple points clustered in zone (suggests slowing/stopping)
        if len(zone_points) >= 3:
            # Calculate cluster spread
            if len(zone_points) >= 2:
                lats = zone_points['lat'].values
                lons = zone_points['lon'].values
                cluster_spread = max(
                    geodesic((lats.min(), lons.min()), (lats.max(), lons.max())).meters,
                    0.1
                )
                evidence['cluster_spread'] = cluster_spread
                evidence['cluster_points'] = len(zone_points)
                
                # Tight cluster suggests stop
                if cluster_spread < 10:
                    evidence['detection_method'] = 'point_cluster'
                    result['conclusive'] = True
                    result['result'] = 'STOP_OK'
                    result['confidence'] = 75
                    result['reason'] = f'{len(zone_points)} points clustered within {cluster_spread:.1f}m'
                    return result
        
        # Check 3: Engine load drop (if available)
        if has_engine_data and 'io_engine_load' in approach_points.columns:
            engine_loads = approach_points['io_engine_load'].dropna()
            if len(engine_loads) >= 2:
                max_load = engine_loads.max()
                min_load = engine_loads.min()
                load_drop = max_load - min_load
                evidence['engine_load_drop'] = load_drop
                
                # Significant load drop suggests deceleration
                if load_drop > 50 and min_load < 30:
                    evidence['detection_method'] = 'engine_load'
                    result['conclusive'] = True
                    result['result'] = 'STOP_OK'
                    result['confidence'] = 70
                    result['reason'] = f'Engine load dropped {load_drop:.0f}% (from {max_load:.0f}% to {min_load:.0f}%)'
                    return result
        
        # Check 4: Low speed (5-15 km/h) - probable stop for sparse GPS
        if is_sparse_gps and min_speed_zone < 15:
            evidence['detection_method'] = 'low_speed_sparse'
            result['conclusive'] = True
            result['result'] = 'STOP_OK'
            result['confidence'] = 65
            result['reason'] = f'Low speed ({min_speed_zone:.0f} km/h) with sparse GPS - probable stop'
            return result
        
        # Check 5: Moderate-high speed maintained (probable violation)
        if min_speed_zone > 15 and closest_speed > 15:
            evidence['detection_method'] = 'moderate_speed_violation'
            
            # Higher confidence if we have good data
            if not is_sparse_gps and len(zone_points) >= 2:
                result['conclusive'] = True
                result['result'] = 'STOP_KO'
                result['confidence'] = 80
                result['reason'] = f'Speed maintained above 15 km/h (min {min_speed_zone:.0f} km/h)'
                return result
            elif is_sparse_gps:
                # Less confident with sparse GPS
                result['conclusive'] = True
                result['result'] = 'STOP_KO'
                result['confidence'] = 60
                result['reason'] = f'Speed likely maintained (min {min_speed_zone:.0f} km/h) - sparse GPS'
                return result
        
        # Not conclusive at Layer 2
        return result
    
    
    def _check_layer3_segment_analysis(self, df, approach_points, closest_point, closest_idx,
                                        is_sparse_gps, avg_interval, evidence):
        """
        Layer 3: Segment-based analysis.
        - Entry speed vs Exit speed pattern
        - Time spent in approach zone
        - Physics-based inference
        """
        
        result = {'conclusive': False, 'result': None, 'confidence': None, 'reason': ''}
        
        # Get points before and after closest point
        before_points = df[df.index < closest_idx].tail(5)
        after_points = df[df.index > closest_idx].head(5)
        
        if len(before_points) == 0 or len(after_points) == 0:
            return result
        
        # Calculate entry and exit speeds
        entry_speed = before_points['speed_kmh'].iloc[-1] if len(before_points) > 0 else None
        exit_speed = after_points['speed_kmh'].iloc[0] if len(after_points) > 0 else None
        
        if entry_speed is None or exit_speed is None:
            return result
        
        evidence['entry_speed'] = entry_speed
        evidence['exit_speed'] = exit_speed
        evidence['speed_at_closest'] = closest_point['speed_kmh']
        
        # Calculate speed differential
        speed_drop = entry_speed - closest_point['speed_kmh']
        evidence['speed_drop'] = speed_drop
        
        # Check 1: Significant speed drop pattern (entry -> low -> exit)
        # Pattern: High -> Low -> Rising suggests stop
        if entry_speed > 20 and closest_point['speed_kmh'] < 15 and exit_speed > 10:
            # V-shaped pattern suggests stop
            if speed_drop > 15:
                evidence['detection_method'] = 'speed_pattern'
                result['conclusive'] = True
                result['result'] = 'STOP_OK'
                result['confidence'] = 65
                result['reason'] = f'Speed pattern suggests stop (entry {entry_speed:.0f} → {closest_point["speed_kmh"]:.0f} → exit {exit_speed:.0f} km/h)'
                return result
        
        # Check 2: Calculate expected distance vs actual distance
        # If driver traveled less distance than expected at constant speed, they slowed/stopped
        if len(before_points) >= 1 and len(after_points) >= 1:
            # Time between entry and exit points
            time_before = before_points['timestamp'].iloc[-1]
            time_after = after_points['timestamp'].iloc[0]
            time_elapsed = (time_after - time_before).total_seconds()
            
            if time_elapsed > 0:
                # Expected distance at entry speed
                expected_distance = (entry_speed / 3.6) * time_elapsed  # km/h to m/s
                
                # Actual distance traveled
                actual_distance = 0
                segment_df = df[(df.index >= before_points.index[-1]) & (df.index <= after_points.index[0])]
                if 'distance_m' in segment_df.columns:
                    actual_distance = segment_df['distance_m'].sum()
                
                evidence['time_elapsed'] = time_elapsed
                evidence['expected_distance'] = expected_distance
                evidence['actual_distance'] = actual_distance
                
                # If actual << expected, driver slowed significantly
                if expected_distance > 0:
                    distance_ratio = actual_distance / expected_distance
                    evidence['distance_ratio'] = distance_ratio
                    
                    if distance_ratio < 0.5:
                        evidence['detection_method'] = 'physics_inference'
                        result['conclusive'] = True
                        result['result'] = 'STOP_OK'
                        result['confidence'] = 60
                        result['reason'] = f'Physics suggests stop (traveled {actual_distance:.0f}m vs expected {expected_distance:.0f}m)'
                        return result
                    elif distance_ratio > 0.85 and entry_speed > 20:
                        evidence['detection_method'] = 'physics_violation'
                        result['conclusive'] = True
                        result['result'] = 'STOP_KO'
                        result['confidence'] = 55
                        result['reason'] = f'Physics suggests no stop (traveled {actual_distance:.0f}m at ~{entry_speed:.0f} km/h)'
                        return result
        
        # Check 3: Time in zone (for non-sparse GPS)
        if not is_sparse_gps and len(approach_points) >= 2:
            time_in_approach = (approach_points['timestamp'].max() - approach_points['timestamp'].min()).total_seconds()
            evidence['time_in_approach'] = time_in_approach
            
            # Long time in approach suggests stopping
            if time_in_approach > 10:
                evidence['detection_method'] = 'time_in_zone'
                result['conclusive'] = True
                result['result'] = 'STOP_OK'
                result['confidence'] = 60
                result['reason'] = f'Extended time in approach zone ({time_in_approach:.0f}s)'
                return result
        
        # Not conclusive at Layer 3
        return result
    
    
    def _create_stop_result(self, result, confidence, reason, stop_sign, evidence, layer='unknown'):
        """
        Create standardized stop analysis result.
        """
        
        return {
            'result': result,
            'confidence': confidence,
            'reason': reason,
            'layer': layer,
            'stop_sign_lat': stop_sign['lat'],
            'stop_sign_lon': stop_sign['lon'],
            'closest_distance': evidence.get('closest_distance'),
            'points_in_zone': evidence.get('points_in_zone', 0),
            'points_in_approach': evidence.get('points_in_approach', 0),
            'min_speed_in_zone': evidence.get('min_speed_zone', evidence.get('layer1_min_speed')),
            'entry_speed': evidence.get('entry_speed'),
            'exit_speed': evidence.get('exit_speed'),
            'detection_method': evidence.get('detection_method', 'unknown'),
            'is_sparse_gps': evidence.get('is_sparse_gps', False),
            # Backward compatibility
            'stopped_properly': result == 'STOP_OK',
            'min_speed_at_stop': evidence.get('min_speed_zone', evidence.get('closest_speed'))
        }
        
        
    # ========================================================================
    # SCHOOL ZONE ANALYSIS - Two versions based on road context
    # ========================================================================
      
        
    def _group_schools_by_proximity(self, schools, max_distance=100):
        """
        Group schools that are within max_distance meters of each other
        This handles cases where one school campus has multiple OSM entries
        
        Args:
            schools: List of school dictionaries
            max_distance: Maximum distance in meters to group schools together
        
        Returns:
            List of school groups, each group represents one logical school zone
        """
        from geopy.distance import geodesic
        
        if not schools:
            return []
        
        groups = []
        ungrouped = schools.copy()
        
        while ungrouped:
            # Start new group with first ungrouped school
            current_group = [ungrouped.pop(0)]
            
            # Find all schools close to any school in current group
            added_to_group = True
            while added_to_group:
                added_to_group = False
                
                for i in range(len(ungrouped) - 1, -1, -1):  # Iterate backwards
                    candidate = ungrouped[i]
                    
                    # Check if candidate is close to any school in current group
                    for group_member in current_group:
                        distance = geodesic(
                            (candidate['lat'], candidate['lon']),
                            (group_member['lat'], group_member['lon'])
                        ).meters
                        
                        if distance <= max_distance:
                            current_group.append(ungrouped.pop(i))
                            added_to_group = True
                            break
            
            groups.append(current_group)
        
        return groups
    
    def _analyze_school_zones_with_context(self):
        """
        Fixed school zone analysis with school grouping to eliminate duplicates
        """
        schools = self.map_manager.map_data['schools']
        
        print(f"    🔍 Analyzing school zone passages...")
        print(f"    🔍 Total schools in area: {len(schools)}")
        
        if not hasattr(self.map_manager, 'route_geometry') or self.map_manager.route_geometry is None:
            raise ValueError("Route geometry is required for accurate school zone analysis.")
        
        route_geom = self.map_manager.route_geometry
        
        # STEP 1: Filter schools near route (same as before)
        relevant_schools = []
        transformer = Transformer.from_crs("EPSG:4326", "EPSG:2154", always_xy=True)
        route_proj = transform(lambda x, y: transformer.transform(x, y), route_geom)
        
        for school in schools:
            school_point = Point(school['lon'], school['lat'])
            
            try:
                point_proj = transform(lambda x, y: transformer.transform(x, y), school_point)
                distance_to_route = route_proj.distance(point_proj)
                
                if distance_to_route <= 150:  # 150m detection radius
                    relevant_schools.append({
                        'school': school,
                        'distance_to_route': distance_to_route
                    })
                    
            except Exception as e:
                print(f"    ⚠️ Projection failed for school: {e}")
                continue
        
        print(f"    ✅ Found {len(relevant_schools)} schools near route")
        
        # STEP 2: Group schools by proximity (NEW - eliminates duplicates)
        school_objects = [rs['school'] for rs in relevant_schools]
        school_groups = self._group_schools_by_proximity(school_objects, max_distance=100)
        
        print(f"    🔄 Grouped {len(relevant_schools)} schools into {len(school_groups)} school zones")
        
        # STEP 3: Analyze passages for each school group (not individual schools)
        all_school_passages = []
        
        for group_idx, school_group in enumerate(school_groups):
            # Calculate group centroid and representative name
            group_lat = sum(school['lat'] for school in school_group) / len(school_group)
            group_lon = sum(school['lon'] for school in school_group) / len(school_group)
            
            # Use most detailed name from group
            group_names = [school.get('name', 'Unknown') for school in school_group if school.get('name')]
            group_name = max(group_names, key=len) if group_names else f"School Zone {group_idx + 1}"
            
            print(f"    📍 Analyzing school zone '{group_name}' ({len(school_group)} buildings)")
            
            # Find GPS points within 200m of group centroid (French standard)
            zone_points = []
            SCHOOL_ZONE_RADIUS = 200  # French standard
            
            for i, row in self.df.iterrows():
                distance_to_zone = geodesic(
                    (row['lat'], row['lon']), 
                    (group_lat, group_lon)
                ).meters
                
                if distance_to_zone <= SCHOOL_ZONE_RADIUS:
                    zone_points.append({
                        'gps_index': i,
                        'timestamp': row['timestamp'],
                        'distance_to_zone': distance_to_zone,
                        'speed_kmh': row['speed_kmh'],
                        'row_data': row
                    })
            
            if not zone_points:
                continue
                
            # Group zone points into episodes (separate passages)
            zone_points.sort(key=lambda x: x['timestamp'])
            episodes = []
            current_episode = [zone_points[0]]
            
            for i in range(1, len(zone_points)):
                current_point = zone_points[i]
                previous_point = zone_points[i-1]
                
                time_gap = (current_point['timestamp'] - previous_point['timestamp']).total_seconds()
                
                if time_gap > 600:  # 10-minute gap = new passage
                    episodes.append(current_episode)
                    current_episode = [current_point]
                else:
                    current_episode.append(current_point)
            
            episodes.append(current_episode)
            
            # Create passage records
            for episode_idx, episode_points in enumerate(episodes):
                first_point = episode_points[0]
                row_data = first_point['row_data']
                
                # French school zone speed logic
                regular_speed_limit = row_data.get('speed_limit', 50)
                school_zone_limit = max(30, regular_speed_limit - 20)
                
                passage_record = {
                    'school_zone_id': f"zone_{group_idx}",
                    'school_zone_name': group_name,
                    'school_zone_lat': group_lat,
                    'school_zone_lon': group_lon,
                    'schools_in_zone': len(school_group),
                    'episode_number': episode_idx + 1,
                    'timestamp': first_point['timestamp'],
                    'gps_index': first_point['gps_index'],
                    'distance_to_zone_center': first_point['distance_to_zone'],
                    'zone_radius': SCHOOL_ZONE_RADIUS,
                    'speed_kmh': first_point['speed_kmh'],
                    'regular_speed_limit': regular_speed_limit,
                    'school_zone_limit': school_zone_limit,
                    'exceeded_school_limit': first_point['speed_kmh'] > school_zone_limit,
                    'exceeded_regular_limit': first_point['speed_kmh'] > regular_speed_limit,
                    'episode_duration_points': len(episode_points)
                }
                
                all_school_passages.append(passage_record)
            
            if episodes:
                print(f"      📊 {len(episodes)} passage episodes detected")
        
        # Build results
        total_passages = len(all_school_passages)
        school_violations = len([p for p in all_school_passages if p['exceeded_school_limit']])
        regular_violations = len([p for p in all_school_passages if p['exceeded_regular_limit']])
        unique_zones = len(set(p['school_zone_id'] for p in all_school_passages))
        
        self.results['school_zones'] = {
            'total_passages': total_passages,
            'school_zone_violations': school_violations,
            'regular_speed_violations': regular_violations,
            'school_zone_violation_percentage': (school_violations / total_passages * 100) if total_passages > 0 else 0,
            'school_zones_detected': len(school_groups),
            'school_zones_with_passages': unique_zones,
            'zone_radius_used': 200,  # French standard
            'passages': all_school_passages
        }
        
        print(f"    ✅ School zone analysis complete:")
        print(f"        📊 {total_passages} passages through {unique_zones} school zones")
        print(f"        🚨 {school_violations} school zone violations")
        print(f"        📏 Using 200m zone radius, 100m grouping distance")
    
        # ========================================================================
        # UTILITY METHODS - Same for all analyzers
        # ========================================================================
        
    def _get_approach_speeds(self, current_idx, lookback=5):
        start_idx = max(0, current_idx - lookback)
        return self.df.iloc[start_idx:current_idx + 1]['speed_kmh'].tolist()

    def _get_speed_limit_at_location(self, row_data):
        """
        CORRECTED: Get the actual speed limit at a specific GPS location
        """
        # Method 1: Use speed_limit column if available and valid
        if 'speed_limit' in row_data and pd.notna(row_data['speed_limit']) and row_data['speed_limit'] > 0:
            return row_data['speed_limit']
        
        # Method 2: Use matched edge if available
        if 'matched_edge' in row_data and pd.notna(row_data['matched_edge']):
            speed_limit = self.map_manager.get_speed_limit(row_data['matched_edge'])
            if speed_limit and speed_limit > 0:
                return speed_limit
        
        # Method 3: Spatial lookup at coordinates
        if hasattr(self.map_manager, 'get_speed_limit') and callable(self.map_manager.get_speed_limit):
            try:
                # If get_speed_limit accepts coordinates
                speed_limit = self.map_manager.get_speed_limit(row_data['lat'], row_data['lon'])
                if speed_limit and speed_limit > 0:
                    return speed_limit
            except:
                pass
        
        # Method 4: Default based on area type (conservative)
        # Most stop signs are in urban/residential areas
        return 30  # Default 30 km/h for stop sign areas



    def _analyze_traffic_lights_with_context(self, df=None, traffic_signals=None, route_geometry=None):
        """
        Analyze traffic light compliance using OSM ID matching.
        Uses existing map_manager filtering method.
        """
        from geopy.distance import geodesic
        
        if df is None:
            df = self.df
        if traffic_signals is None:
            traffic_signals = self.map_manager.map_data.get('traffic_lights', [])
        
        print(f"    🚦 Analyzing traffic light compliance...")
        print(f"    📊 Total traffic signals in map: {len(traffic_signals)}")
        
        # Initialize results
        results = {
            'total_traffic_signals': 0,
            'total_approaches': 0,
            'compliance': {
                'stopped': 0,
                'slowed': 0,
                'passed_through': 0,
                'uncertain': 0
            },
            'stop_percentage': None,
            'approaches_detailed': [],
            'detection_mode': 'basic',
            'avg_confidence': 0
        }
        
        if not traffic_signals:
            print(f"    ❌ No traffic signals in map data")
            self.results['traffic_lights'] = results
            return
        
        # Detect GPS quality
        has_movement_status = 'io_movement' in df.columns and df['io_movement'].notna().any()
        avg_interval = df['time_diff_s'].mean() if 'time_diff_s' in df.columns else 6.0
        is_sparse_gps = avg_interval > 4
        
        if has_movement_status:
            results['detection_mode'] = 'enhanced'
        
        print(f"    🎯 Detection mode: {results['detection_mode']}")
        print(f"    📡 GPS interval: {avg_interval:.1f}s {'(SPARSE)' if is_sparse_gps else '(OK)'}")
        
        # ================================================================
        # USE EXISTING FILTERING METHOD
        # ================================================================
        relevant_signals_raw = self.map_manager.get_driven_road_linked_features(
            traffic_signals, 
            feature_type="traffic_lights"
        )
        
        relevant_signals = []
        for signal in relevant_signals_raw:
            relevant_signals.append({
                'signal': signal,
                'lat': signal.get('lat'),
                'lon': signal.get('lon'),
                'id': signal.get('id'),
                'osmid': signal.get('id'),
                'signal_type': signal.get('tags', {}).get('traffic_signals', 'standard') if signal.get('tags') else 'standard',
                'match_method': 'osmid'
            })
        
        print(f"    ✅ Found {len(relevant_signals)} traffic signals on driven route")
        
        if not relevant_signals:
            print(f"    ❌ No traffic signals found on driven route")
            self.results['traffic_lights'] = results
            return
        
        results['total_traffic_signals'] = len(relevant_signals)
        
        # ================================================================
        # ANALYZE EACH TRAFFIC SIGNAL
        # ================================================================
        all_confidences = []
        APPROACH_RADIUS = 50
        
        for signal_info in relevant_signals:
            signal_lat = signal_info['lat']
            signal_lon = signal_info['lon']
            signal_id = signal_info['id']
            
            # Find GPS points near this traffic signal
            nearby_points = []
            
            for idx, row in df.iterrows():
                distance = geodesic((signal_lat, signal_lon), (row['lat'], row['lon'])).meters
                
                if distance <= APPROACH_RADIUS:
                    nearby_points.append({
                        'idx': idx,
                        'lat': row['lat'],
                        'lon': row['lon'],
                        'distance': distance,
                        'speed_kmh': row['speed_kmh'],
                        'timestamp': row['timestamp'],
                        'io_movement': row.get('io_movement')
                    })
            
            if len(nearby_points) < 1:
                continue
            
            nearby_points.sort(key=lambda x: x['timestamp'])
            
            # Analyze behavior
            approach_result = self._analyze_traffic_signal_behavior(
                nearby_points=nearby_points,
                signal_lat=signal_lat,
                signal_lon=signal_lon,
                signal_id=signal_id,
                is_sparse_gps=is_sparse_gps,
                has_movement_status=has_movement_status
            )
            
            if approach_result is None:
                continue
            
            results['total_approaches'] += 1
            results['approaches_detailed'].append(approach_result)
            
            behavior = approach_result.get('behavior', 'uncertain')
            if behavior in results['compliance']:
                results['compliance'][behavior] += 1
            
            if approach_result.get('confidence'):
                all_confidences.append(approach_result['confidence'])
        
        # Calculate averages
        if all_confidences:
            results['avg_confidence'] = sum(all_confidences) / len(all_confidences)
        
        total_evaluated = results['compliance']['stopped'] + results['compliance']['slowed'] + results['compliance']['passed_through']
        if total_evaluated > 0:
            results['stop_percentage'] = ((results['compliance']['stopped'] + results['compliance']['slowed']) / total_evaluated) * 100
        
        print(f"\n    📊 Traffic Light Analysis Complete:")
        print(f"       Total signals on route: {results['total_traffic_signals']}")
        print(f"       Approaches analyzed: {results['total_approaches']}")
        print(f"       Stopped: {results['compliance']['stopped']}, "
              f"Slowed: {results['compliance']['slowed']}, "
              f"Passed through: {results['compliance']['passed_through']}")
        
        self.results['traffic_lights'] = results    
    
    def _analyze_traffic_signal_behavior(self, nearby_points, signal_lat, signal_lon, 
                                          signal_id, is_sparse_gps, has_movement_status):
        """
        Analyze driver behavior at a traffic signal.
        
        Categories:
        - STOPPED: Vehicle came to complete stop (likely red light)
        - SLOWED: Vehicle slowed significantly (likely yellow or cautious)
        - PASSED_THROUGH: Vehicle maintained speed (likely green light)
        """
        
        behavior = None
        confidence = 50
        evidence = []
        
        speeds = [p['speed_kmh'] for p in nearby_points if p['speed_kmh'] is not None]
        min_speed = min(speeds) if speeds else None
        closest_point = min(nearby_points, key=lambda x: x['distance'])
        closest_speed = closest_point['speed_kmh']
        
        # Check io_movement (vehicle stationary)
        movement_values = [p['io_movement'] for p in nearby_points if p['io_movement'] is not None]
        has_stationary = 0 in movement_values
        
        # STOPPED: Clear stop detected
        if has_stationary:
            behavior = 'stopped'
            confidence = 95
            evidence.append('io_movement=0 (stationary)')
            print(f"       ✅ STOPPED: io_movement=0 detected (95%)")
        
        elif min_speed is not None and min_speed < 3:
            behavior = 'stopped'
            confidence = 90
            evidence.append(f'speed={min_speed:.0f} km/h (near zero)')
            print(f"       ✅ STOPPED: Speed near zero ({min_speed:.0f} km/h) (90%)")
        
        # SLOWED: Significant slowdown
        elif min_speed is not None and min_speed < 15:
            behavior = 'slowed'
            confidence = 75
            evidence.append(f'min_speed={min_speed:.0f} km/h (slowed significantly)')
            print(f"       🟡 SLOWED: Low speed ({min_speed:.0f} km/h) (75%)")
        
        # PASSED_THROUGH: Maintained speed (likely green light)
        elif min_speed is not None and min_speed >= 15:
            behavior = 'passed_through'
            confidence = 70
            evidence.append(f'min_speed={min_speed:.0f} km/h (maintained speed - likely green)')
            print(f"       🟢 PASSED THROUGH: Maintained speed ({min_speed:.0f} km/h) - likely green light (70%)")
        
        else:
            behavior = 'uncertain'
            confidence = 50
            evidence.append('insufficient data')
            print(f"       ❓ UNCERTAIN: Insufficient data (50%)")
        
        # Calculate dwell time if stopped
        dwell_time = None
        if behavior == 'stopped' and len(nearby_points) >= 2:
            low_speed_points = [p for p in nearby_points if p['speed_kmh'] < 5]
            if len(low_speed_points) >= 2:
                timestamps = [p['timestamp'] for p in low_speed_points]
                dwell_time = (max(timestamps) - min(timestamps)).total_seconds()
        
        result = {
            'signal_lat': signal_lat,
            'signal_lon': signal_lon,
            'signal_id': signal_id,
            'behavior': behavior,
            'confidence': confidence,
            'evidence': evidence,
            'min_speed': min_speed,
            'closest_speed': closest_speed,
            'speeds': speeds,
            'dwell_time_seconds': dwell_time,
            'gps_points': len(nearby_points),
            'closest_distance': closest_point['distance']
        }
        
        return result

    def _analyze_harsh_events(self, df=None):
        """
        Detect harsh driving events:
        - Harsh braking (sudden deceleration)
        - Harsh acceleration (sudden speed increase)
        - Sharp turns (lateral acceleration via heading change)
        
        Uses:
        1. io_green_driving_type (if available) - direct from device
        2. Speed/heading calculations (fallback)
        """
        import pandas as pd
        
        if df is None:
            df = self.df
        
        print(f"    ⚠️ Analyzing harsh driving events...")
        
        # ================================================================
        # STEP 1: Initialize results
        # ================================================================
        results = {
            'total_harsh_events': 0,
            'harsh_braking': {
                'count': 0,
                'events': []
            },
            'harsh_acceleration': {
                'count': 0,
                'events': []
            },
            'sharp_turns': {
                'count': 0,
                'events': []
            },
            'detection_method': 'unknown',
            'events_per_100km': None,
            'severity_summary': {}
        }
        
        # ================================================================
        # STEP 2: Check available detection methods
        # ================================================================
        has_io_green_driving = 'io_green_driving_type' in df.columns and df['io_green_driving_type'].notna().any()
        
        # Check for heading column (might be named differently)
        heading_col = None
        for col_name in ['heading', 'course', 'bearing', 'direction']:
            if col_name in df.columns and df[col_name].notna().any():
                heading_col = col_name
                break
        has_heading = heading_col is not None
        
        has_speed = 'speed_kmh' in df.columns
        has_time_diff = 'time_diff_s' in df.columns
        
        # Calculate GPS interval for threshold adjustment
        avg_interval = df['time_diff_s'].mean() if has_time_diff else 6.0
        is_sparse_gps = avg_interval > 4
        
        if has_io_green_driving:
            results['detection_method'] = 'device_io'
            print(f"    🎯 Using io_green_driving_type (device-reported events)")
        elif has_speed and has_time_diff:
            results['detection_method'] = 'calculated'
            print(f"    🎯 Using speed-based calculations")
        else:
            print(f"    ❌ Insufficient data for harsh event detection")
            self.results['harsh_events'] = results
            return
        
        # ================================================================
        # STEP 3: Set thresholds (adjust for sparse GPS)
        # ================================================================
        if is_sparse_gps:
            HARSH_BRAKE_THRESHOLD = 2.5    # More lenient for sparse GPS
            HARSH_ACCEL_THRESHOLD = 2.0
            SHARP_TURN_THRESHOLD = 2.0
            print(f"    ⚠️ Sparse GPS ({avg_interval:.1f}s) - using adjusted thresholds")
        else:
            HARSH_BRAKE_THRESHOLD = 3.9    # 0.4g standard
            HARSH_ACCEL_THRESHOLD = 3.4    # 0.35g standard
            SHARP_TURN_THRESHOLD = 2.9     # 0.3g standard
        
        # ================================================================
        # STEP 4: METHOD 1 - Device-reported events (io_green_driving_type)
        # ================================================================
        if has_io_green_driving:
            # io_green_driving_type values (verify with your device):
            # 1 = Harsh acceleration
            # 2 = Harsh braking
            # 3 = Harsh cornering
            
            for idx, row in df.iterrows():
                event_type = row.get('io_green_driving_type')
                
                if pd.isna(event_type) or event_type == 0:
                    continue
                
                event_data = {
                    'timestamp': row['timestamp'],
                    'lat': row['lat'],
                    'lon': row['lon'],
                    'speed_kmh': row['speed_kmh'],
                    'source': 'device_io',
                    'event_type_raw': event_type
                }
                
                if event_type == 1:  # Harsh acceleration
                    results['harsh_acceleration']['count'] += 1
                    results['harsh_acceleration']['events'].append(event_data)
                    
                elif event_type == 2:  # Harsh braking
                    results['harsh_braking']['count'] += 1
                    results['harsh_braking']['events'].append(event_data)
                    
                elif event_type == 3:  # Harsh cornering
                    results['sharp_turns']['count'] += 1
                    results['sharp_turns']['events'].append(event_data)
        
        # ================================================================
        # STEP 5: METHOD 2 - Calculated from speed/heading (fallback/supplement)
        # ================================================================
        if has_speed and has_time_diff:
            df_calc = df.copy()
            df_calc['speed_ms'] = df_calc['speed_kmh'] / 3.6  # Convert to m/s
            
            # Avoid division by zero
            df_calc['time_diff_safe'] = df_calc['time_diff_s'].replace(0, float('nan'))
            df_calc['acceleration_ms2'] = df_calc['speed_ms'].diff() / df_calc['time_diff_safe']
            
            # Detect harsh braking (negative acceleration exceeding threshold)
            harsh_brake_mask = df_calc['acceleration_ms2'] < -HARSH_BRAKE_THRESHOLD
            harsh_brake_mask = harsh_brake_mask & df_calc['acceleration_ms2'].notna()
            
            for idx in df_calc[harsh_brake_mask].index:
                row = df_calc.loc[idx]
                
                # Skip if already detected by device
                if has_io_green_driving:
                    device_event = row.get('io_green_driving_type')
                    if pd.notna(device_event) and device_event == 2:
                        continue
                
                event_data = {
                    'timestamp': row['timestamp'],
                    'lat': row['lat'],
                    'lon': row['lon'],
                    'speed_kmh': row['speed_kmh'],
                    'deceleration_ms2': abs(row['acceleration_ms2']),
                    'deceleration_g': abs(row['acceleration_ms2']) / 9.81,
                    'source': 'calculated'
                }
                
                results['harsh_braking']['count'] += 1
                results['harsh_braking']['events'].append(event_data)
            
            # Detect harsh acceleration (positive acceleration exceeding threshold)
            harsh_accel_mask = df_calc['acceleration_ms2'] > HARSH_ACCEL_THRESHOLD
            harsh_accel_mask = harsh_accel_mask & df_calc['acceleration_ms2'].notna()
            
            for idx in df_calc[harsh_accel_mask].index:
                row = df_calc.loc[idx]
                
                # Skip if already detected by device
                if has_io_green_driving:
                    device_event = row.get('io_green_driving_type')
                    if pd.notna(device_event) and device_event == 1:
                        continue
                
                event_data = {
                    'timestamp': row['timestamp'],
                    'lat': row['lat'],
                    'lon': row['lon'],
                    'speed_kmh': row['speed_kmh'],
                    'acceleration_ms2': row['acceleration_ms2'],
                    'acceleration_g': row['acceleration_ms2'] / 9.81,
                    'source': 'calculated'
                }
                
                results['harsh_acceleration']['count'] += 1
                results['harsh_acceleration']['events'].append(event_data)
            
            # ================================================================
            # SHARP TURN DETECTION (using heading change)
            # ================================================================
            if has_heading:
                df_calc['heading_change'] = df_calc[heading_col].diff().abs()
                
                # Normalize heading change (handle 359 -> 1 wraparound)
                df_calc['heading_change'] = df_calc['heading_change'].apply(
                    lambda x: min(x, 360 - x) if pd.notna(x) and x <= 360 else (0 if pd.isna(x) else x)
                )
                
                # Calculate angular velocity (degrees per second)
                df_calc['angular_velocity'] = df_calc['heading_change'] / df_calc['time_diff_safe']
                
                # Calculate lateral acceleration: a = v * ω (in radians)
                df_calc['lateral_accel_ms2'] = (
                    df_calc['speed_ms'] * 
                    (df_calc['angular_velocity'] * 3.14159 / 180)
                )
                
                # Detect sharp turns
                sharp_turn_mask = df_calc['lateral_accel_ms2'].abs() > SHARP_TURN_THRESHOLD
                sharp_turn_mask = sharp_turn_mask & df_calc['lateral_accel_ms2'].notna()
                
                for idx in df_calc[sharp_turn_mask].index:
                    row = df_calc.loc[idx]
                    
                    # Skip if already detected by device
                    if has_io_green_driving:
                        device_event = row.get('io_green_driving_type')
                        if pd.notna(device_event) and device_event == 3:
                            continue
                    
                    event_data = {
                        'timestamp': row['timestamp'],
                        'lat': row['lat'],
                        'lon': row['lon'],
                        'speed_kmh': row['speed_kmh'],
                        'heading_change_deg': row['heading_change'],
                        'lateral_accel_ms2': abs(row['lateral_accel_ms2']),
                        'lateral_accel_g': abs(row['lateral_accel_ms2']) / 9.81,
                        'source': 'calculated'
                    }
                    
                    results['sharp_turns']['count'] += 1
                    results['sharp_turns']['events'].append(event_data)
        
        # ================================================================
        # STEP 6: Calculate totals and rates
        # ================================================================
        results['total_harsh_events'] = (
            results['harsh_braking']['count'] +
            results['harsh_acceleration']['count'] +
            results['sharp_turns']['count']
        )
        
        # Calculate events per 100km
        total_distance_km = df['distance_m'].sum() / 1000 if 'distance_m' in df.columns else 0
        if total_distance_km > 0:
            results['events_per_100km'] = round(
                (results['total_harsh_events'] / total_distance_km) * 100, 2
            )
        
        # Categorize severity
        results['severity_summary'] = self._categorize_harsh_event_severity(results)
        
        # Print summary
        print(f"\n    📊 Harsh Events Analysis Complete:")
        print(f"       Detection method: {results['detection_method']}")
        print(f"       Total harsh events: {results['total_harsh_events']}")
        print(f"       - Harsh braking: {results['harsh_braking']['count']}")
        print(f"       - Harsh acceleration: {results['harsh_acceleration']['count']}")
        print(f"       - Sharp turns: {results['sharp_turns']['count']}")
        if results['events_per_100km'] is not None:
            print(f"       Events per 100km: {results['events_per_100km']}")
        print(f"       Severity: {results['severity_summary'].get('level', 'unknown')}")
        
        self.results['harsh_events'] = results
    
    
    def _categorize_harsh_event_severity(self, results):
        """Categorize overall harsh driving severity"""
        
        events_per_100km = results.get('events_per_100km')
        
        if events_per_100km is None or events_per_100km == 0:
            return {
                'level': 'excellent',
                'description': 'No harsh events detected',
                'score': 100
            }
        elif events_per_100km <= 2:
            return {
                'level': 'good',
                'description': 'Minimal harsh events',
                'score': 85
            }
        elif events_per_100km <= 5:
            return {
                'level': 'moderate',
                'description': 'Some harsh events detected',
                'score': 70
            }
        elif events_per_100km <= 10:
            return {
                'level': 'concerning',
                'description': 'Frequent harsh events',
                'score': 50
            }
        else:
            return {
                'level': 'critical',
                'description': 'Very frequent harsh events - coaching needed',
                'score': 25
            }



    # ===================================Debug Traffic Lights


    # ============================================================================
    # DEBUG: Traffic Lights and Harsh Events Detection
    # ============================================================================
    
    def debug_traffic_lights_and_harsh_events(analyzer, map_manager, df=None):
        """
        Debug function to verify traffic light and harsh events detection.
        Run this AFTER your main pipeline has completed.
        
        Usage:
            debug_traffic_lights_and_harsh_events(analyzer, map_manager)
        """
        
        print("=" * 70)
        print("🔍 DEBUG: Traffic Lights and Harsh Events Detection")
        print("=" * 70)
        
        # Use analyzer's dataframe if not provided
        if df is None:
            df = analyzer.df if hasattr(analyzer, 'df') else None
        
        if df is None:
            print("❌ ERROR: No DataFrame available")
            return
        
        print(f"\n📊 DataFrame Info:")
        print(f"   Total GPS points: {len(df)}")
        print(f"   Columns: {list(df.columns)}")
        
        # ========================================================================
        # PART 1: DEBUG TRAFFIC LIGHTS
        # ========================================================================
        print("\n" + "=" * 70)
        print("🚦 PART 1: TRAFFIC LIGHTS DEBUG")
        print("=" * 70)
        
        # Check 1: Traffic lights in map_data
        traffic_lights = map_manager.map_data.get('traffic_lights', [])
        print(f"\n1️⃣ Traffic lights in map_data: {len(traffic_lights)}")
        
        if traffic_lights:
            print(f"   Sample traffic light structure:")
            sample = traffic_lights[0]
            for key, value in sample.items():
                print(f"      {key}: {value}")
            
            # Check if osmid exists
            has_osmid = 'osmid' in sample
            print(f"\n   ✅ Has 'osmid' field: {has_osmid}")
            if not has_osmid:
                print("   ⚠️  WARNING: Traffic lights missing 'osmid' field!")
                print("   ⚠️  Add 'osmid': n.id to traffic light creation in _MasterCacheHandler")
        else:
            print("   ❌ No traffic lights found in map_data!")
            print("   Check: map_manager.map_data['traffic_lights']")
        
        # Check 2: Driven road IDs
        driven_road_ids = getattr(map_manager, 'driven_road_ids', set())
        print(f"\n2️⃣ Driven road IDs: {len(driven_road_ids)}")
        
        if driven_road_ids:
            sample_edges = list(driven_road_ids)[:3]
            print(f"   Sample driven edges: {sample_edges}")
            
            # Extract node IDs
            driven_node_ids = set()
            for edge in driven_road_ids:
                if isinstance(edge, tuple) and len(edge) >= 2:
                    driven_node_ids.add(edge[0])
                    driven_node_ids.add(edge[1])
            print(f"   Extracted node IDs: {len(driven_node_ids)}")
            print(f"   Sample node IDs: {list(driven_node_ids)[:5]}")
        else:
            print("   ❌ No driven road IDs found!")
        
        # Check 3: Filter traffic lights using existing method
        print(f"\n3️⃣ Filtering traffic lights on driven route...")
        
        if hasattr(map_manager, 'get_driven_road_linked_features'):
            try:
                filtered_lights = map_manager.get_driven_road_linked_features(
                    traffic_lights, 
                    feature_type="traffic_lights"
                )
                print(f"   ✅ Filtered traffic lights: {len(filtered_lights)}")
                
                if filtered_lights:
                    print(f"   Sample filtered light:")
                    for key, value in filtered_lights[0].items():
                        print(f"      {key}: {value}")
            except Exception as e:
                print(f"   ❌ Error filtering: {e}")
        else:
            print("   ❌ get_driven_road_linked_features() method not found!")
        
        # Check 4: GPS points near traffic lights
        print(f"\n4️⃣ Checking GPS coverage near traffic lights...")
        
        from geopy.distance import geodesic
        
        lights_with_gps = 0
        lights_without_gps = 0
        APPROACH_RADIUS = 50
        
        for light in traffic_lights[:10]:  # Check first 10
            light_lat = light.get('lat')
            light_lon = light.get('lon')
            
            if light_lat is None or light_lon is None:
                continue
            
            nearby_count = 0
            for idx, row in df.iterrows():
                distance = geodesic((light_lat, light_lon), (row['lat'], row['lon'])).meters
                if distance <= APPROACH_RADIUS:
                    nearby_count += 1
            
            if nearby_count > 0:
                lights_with_gps += 1
            else:
                lights_without_gps += 1
        
        print(f"   Traffic lights with nearby GPS points: {lights_with_gps}/10")
        print(f"   Traffic lights without nearby GPS points: {lights_without_gps}/10")
        
        # Check 5: Test traffic light analysis
        print(f"\n5️⃣ Testing traffic light analysis...")
        
        if hasattr(analyzer, '_analyze_traffic_lights_with_context'):
            try:
                analyzer._analyze_traffic_lights_with_context()
                
                results = analyzer.results.get('traffic_lights', {})
                print(f"   ✅ Analysis completed!")
                print(f"   Results:")
                print(f"      Total signals: {results.get('total_traffic_signals', 0)}")
                print(f"      Approaches analyzed: {results.get('total_approaches', 0)}")
                print(f"      Compliance: {results.get('compliance', {})}")
                print(f"      Stop percentage: {results.get('stop_percentage', 'N/A')}")
            except Exception as e:
                print(f"   ❌ Analysis failed: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("   ❌ _analyze_traffic_lights_with_context() method not found!")
        
        # ========================================================================
        # PART 2: DEBUG HARSH EVENTS
        # ========================================================================
        print("\n" + "=" * 70)
        print("⚠️  PART 2: HARSH EVENTS DEBUG")
        print("=" * 70)
        
        # Check 1: Required columns
        print(f"\n1️⃣ Checking required columns...")
        
        required_cols = {
            'io_green_driving_type': 'Device harsh events (primary)',
            'speed_kmh': 'Speed (required for calculation)',
            'time_diff_s': 'Time difference (required for calculation)',
            'heading': 'Heading (for sharp turn detection)',
            'course': 'Course (alternative to heading)',
            'bearing': 'Bearing (alternative to heading)',
            'distance_m': 'Distance (for events per 100km)'
        }
        
        for col, description in required_cols.items():
            exists = col in df.columns
            has_data = df[col].notna().any() if exists else False
            status = "✅" if exists and has_data else ("⚠️ exists but empty" if exists else "❌")
            print(f"   {status} {col}: {description}")
        
        # Check 2: io_green_driving_type values
        print(f"\n2️⃣ Checking io_green_driving_type values...")
        
        if 'io_green_driving_type' in df.columns:
            value_counts = df['io_green_driving_type'].value_counts(dropna=False)
            print(f"   Value distribution:")
            for value, count in value_counts.items():
                event_type = {
                    0: 'No event',
                    1: 'Harsh acceleration',
                    2: 'Harsh braking',
                    3: 'Harsh cornering'
                }.get(value, f'Unknown ({value})')
                print(f"      {value}: {count} ({event_type})")
            
            # Count actual events
            events_mask = df['io_green_driving_type'].notna() & (df['io_green_driving_type'] != 0)
            total_device_events = events_mask.sum()
            print(f"\n   Total device-reported events: {total_device_events}")
        else:
            print("   ❌ io_green_driving_type column not found!")
        
        # Check 3: Calculate acceleration manually
        print(f"\n3️⃣ Testing acceleration calculation...")
        
        if 'speed_kmh' in df.columns and 'time_diff_s' in df.columns:
            df_test = df.copy()
            df_test['speed_ms'] = df_test['speed_kmh'] / 3.6
            df_test['time_diff_safe'] = df_test['time_diff_s'].replace(0, float('nan'))
            df_test['acceleration_ms2'] = df_test['speed_ms'].diff() / df_test['time_diff_safe']
            
            print(f"   Acceleration stats:")
            print(f"      Min: {df_test['acceleration_ms2'].min():.2f} m/s²")
            print(f"      Max: {df_test['acceleration_ms2'].max():.2f} m/s²")
            print(f"      Mean: {df_test['acceleration_ms2'].mean():.2f} m/s²")
            
            # Check against thresholds
            HARSH_BRAKE_THRESHOLD = 2.5  # Adjusted for sparse GPS
            HARSH_ACCEL_THRESHOLD = 2.0
            
            harsh_braking_calc = (df_test['acceleration_ms2'] < -HARSH_BRAKE_THRESHOLD).sum()
            harsh_accel_calc = (df_test['acceleration_ms2'] > HARSH_ACCEL_THRESHOLD).sum()
            
            print(f"\n   Calculated harsh events (threshold adjusted for sparse GPS):")
            print(f"      Harsh braking (< -{HARSH_BRAKE_THRESHOLD} m/s²): {harsh_braking_calc}")
            print(f"      Harsh acceleration (> {HARSH_ACCEL_THRESHOLD} m/s²): {harsh_accel_calc}")
        else:
            print("   ❌ Missing speed_kmh or time_diff_s columns!")
        
        # Check 4: Heading/turn detection
        print(f"\n4️⃣ Checking heading for turn detection...")
        
        heading_col = None
        for col_name in ['heading', 'course', 'bearing', 'direction']:
            if col_name in df.columns and df[col_name].notna().any():
                heading_col = col_name
                break
        
        if heading_col:
            print(f"   ✅ Found heading column: '{heading_col}'")
            print(f"   Heading stats:")
            print(f"      Min: {df[heading_col].min():.1f}°")
            print(f"      Max: {df[heading_col].max():.1f}°")
            print(f"      Non-null values: {df[heading_col].notna().sum()}")
            
            # Calculate heading changes
            df_test = df.copy()
            df_test['heading_change'] = df_test[heading_col].diff().abs()
            df_test['heading_change'] = df_test['heading_change'].apply(
                lambda x: min(x, 360 - x) if pd.notna(x) and x <= 360 else 0
            )
            
            large_turns = (df_test['heading_change'] > 30).sum()
            print(f"   Large heading changes (>30°): {large_turns}")
        else:
            print("   ⚠️ No heading column found - turn detection will be limited")
        
        # Check 5: Test harsh events analysis
        print(f"\n5️⃣ Testing harsh events analysis...")
        
        if hasattr(analyzer, '_analyze_harsh_events'):
            try:
                analyzer._analyze_harsh_events()
                
                results = analyzer.results.get('harsh_events', {})
                print(f"   ✅ Analysis completed!")
                print(f"   Results:")
                print(f"      Detection method: {results.get('detection_method', 'unknown')}")
                print(f"      Total events: {results.get('total_harsh_events', 0)}")
                print(f"      Harsh braking: {results.get('harsh_braking', {}).get('count', 0)}")
                print(f"      Harsh acceleration: {results.get('harsh_acceleration', {}).get('count', 0)}")
                print(f"      Sharp turns: {results.get('sharp_turns', {}).get('count', 0)}")
                print(f"      Events per 100km: {results.get('events_per_100km', 'N/A')}")
                print(f"      Severity: {results.get('severity_summary', {}).get('level', 'unknown')}")
                
                # Show sample events
                if results.get('harsh_braking', {}).get('events'):
                    print(f"\n   Sample harsh braking event:")
                    sample = results['harsh_braking']['events'][0]
                    for key, value in sample.items():
                        print(f"      {key}: {value}")
            except Exception as e:
                print(f"   ❌ Analysis failed: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("   ❌ _analyze_harsh_events() method not found!")
        
        # ========================================================================
        # SUMMARY
        # ========================================================================
        print("\n" + "=" * 70)
        print("📋 DEBUG SUMMARY")
        print("=" * 70)
        
        issues = []
        
        # Traffic lights issues
        if len(traffic_lights) == 0:
            issues.append("❌ No traffic lights in map_data")
        if traffic_lights and 'osmid' not in traffic_lights[0]:
            issues.append("⚠️ Traffic lights missing 'osmid' field")
        if len(driven_road_ids) == 0:
            issues.append("❌ No driven road IDs set")
        
        # Harsh events issues
        if 'io_green_driving_type' not in df.columns:
            issues.append("⚠️ No device harsh events (io_green_driving_type missing)")
        if 'speed_kmh' not in df.columns or 'time_diff_s' not in df.columns:
            issues.append("❌ Missing columns for acceleration calculation")
        if not heading_col:
            issues.append("⚠️ No heading column for turn detection")
        
        if issues:
            print("\n🔴 Issues found:")
            for issue in issues:
                print(f"   {issue}")
        else:
            print("\n🟢 All checks passed!")
        
        print("\n" + "=" * 70)
        print("DEBUG COMPLETE")
        print("=" * 70)
        
        return {
            'traffic_lights_count': len(traffic_lights),
            'driven_road_ids_count': len(driven_road_ids),
            'has_osmid': traffic_lights and 'osmid' in traffic_lights[0] if traffic_lights else False,
            'has_io_green_driving': 'io_green_driving_type' in df.columns,
            'has_heading': heading_col is not None,
            'issues': issues
        }
    
    
    # ============================================================================
    # QUICK TEST FUNCTIONS
    # ============================================================================
    
    def quick_test_traffic_lights(map_manager):
        """Quick test for traffic lights data"""
        print("\n🚦 Quick Traffic Lights Test")
        print("-" * 40)
        
        traffic_lights = map_manager.map_data.get('traffic_lights', [])
        print(f"Total traffic lights: {len(traffic_lights)}")
        
        if traffic_lights:
            sample = traffic_lights[0]
            print(f"Sample: {sample}")
            print(f"Has osmid: {'osmid' in sample}")
        
        return len(traffic_lights)
    
    
    def quick_test_harsh_events(df):
        """Quick test for harsh events data availability"""
        print("\n⚠️ Quick Harsh Events Test")
        print("-" * 40)
        
        # Check io_green_driving_type
        if 'io_green_driving_type' in df.columns:
            events = df[df['io_green_driving_type'].notna() & (df['io_green_driving_type'] != 0)]
            print(f"Device-reported harsh events: {len(events)}")
            print(f"Value counts:\n{df['io_green_driving_type'].value_counts()}")
        else:
            print("io_green_driving_type: NOT FOUND")
        
        # Check speed data
        if 'speed_kmh' in df.columns:
            print(f"Speed range: {df['speed_kmh'].min():.1f} - {df['speed_kmh'].max():.1f} km/h")
        
        return 'io_green_driving_type' in df.columns
    
    
    # ============================================================================
    # USAGE
    # ============================================================================
    # 
    # After running your main pipeline:
    #
    # Option 1: Full debug
    # debug_traffic_lights_and_harsh_events(analyzer, map_manager)
    #
    # Option 2: Quick tests
    # quick_test_traffic_lights(map_manager)
    # quick_test_harsh_events(df)
    #
# ============================================================================
# UNIFIED REPORT GENERATOR - Clean & Simple Data Access
# ============================================================================

class UnifiedReportGenerator:
    """
    Unified report generator - SINGLE SOURCE OF TRUTH for all calculations.
    
    This class:
    1. Extracts ALL data from analysis results
    2. Performs ALL calculations (scores, aggregations, trends)
    3. Provides clean data structure for FeelGoodReportGenerator (display only)
    
    FeelGoodReportGenerator should ONLY display data from this class - NO calculations.
    """
    
    def __init__(self, weekly_results, processor, driver_info=None, map_manager=None):
        self.weekly_results = weekly_results  # {'Week 1': {...}, 'Week 2': {...}}
        self.processor = processor
        self.driver_info = driver_info or {}
        self.map_manager = map_manager
        
        # Determine analysis type
        self.is_single_week = len(weekly_results) == 1
        self.week_labels = list(weekly_results.keys())

    # ========================================================================
    # MAIN DATA EXTRACTION METHOD
    # ========================================================================
    
    def extract_all_data(self):
        """
        COMPREHENSIVE DATA EXTRACTION - SINGLE SOURCE OF TRUTH
        ========================================================
        
        Returns complete, pre-calculated data structure for reporting.
        FeelGoodReportGenerator should use this directly - NO recalculations.
        """
        
        extracted_data = {
            'metadata': self._extract_metadata(),
            'trip_summary': self._extract_trip_summary(),  # NEW: Trip summary
            'weekly_data': {},
            'aggregated_metrics': {},
            'trends': {},  # NEW: Week-over-week trends
            'map_data': self._extract_map_data(),
            'gps_stats': self._extract_gps_stats(),
            'weekly_geometries': {}
        }
        
        # Extract data for each week
        for week_label in self.week_labels:
            week_data = self.weekly_results[week_label]
            extracted_data['weekly_data'][week_label] = self._extract_week_data(week_data, week_label)
            
            # Extract geometry for plotting
            geometry_data = week_data.get('_geometry_data')
            if geometry_data is not None:
                extracted_data['weekly_geometries'][week_label] = {
                    'route_geometry': geometry_data.get('route_geometry'),
                    'driven_edges': geometry_data.get('driven_edges')
                }
            else:
                extracted_data['weekly_geometries'][week_label] = {
                    'route_geometry': None,
                    'driven_edges': None
                }
        
        # Add GPS DataFrame references
        extracted_data['gps_data'] = {}
        if hasattr(self.processor, 'weekly_data') and self.processor.weekly_data:
            extracted_data['gps_data'] = {
                week: df.copy() for week, df in self.processor.weekly_data.items()
            }
        elif hasattr(self.processor, 'processed_df') and self.processor.processed_df is not None:
            # Single-trip case - GPS data is in processed_df, not weekly_data
            extracted_data['gps_data'] = {
                "Week 1": self.processor.processed_df.copy()
            }
        
        # Calculate aggregated metrics (for multi-week)
        if not self.is_single_week:
            extracted_data['aggregated_metrics'] = self._calculate_aggregated_metrics(extracted_data['weekly_data'])
            extracted_data['trends'] = self._calculate_weekly_trends(extracted_data['weekly_data'])
        else:
            # Single week - use week data as aggregated
            week_data = list(extracted_data['weekly_data'].values())[0]
            extracted_data['aggregated_metrics'] = {
                'total_violations': week_data['calculated_metrics']['total_violations'],
                'combined_metrics': {
                    'avg_overall_score': week_data['calculated_metrics']['overall_score'],
                    'avg_speed_compliance': week_data['speeding']['compliance_percentage'] or 0,
                    'avg_roundabout_performance': week_data['roundabouts'].get('entry_compliance_percentage') or 0,
                    'avg_stop_sign_compliance': week_data['stop_signs']['compliance_percentage'] or 0,
                    'avg_school_zone_compliance': week_data['school_zones']['compliance_percentage'] or 0,
                    'avg_harsh_events_per_100km': week_data['harsh_events'].get('events_per_100km') or 0,
                    'avg_traffic_light_compliance': week_data['traffic_lights'].get('stop_percentage') or 0
                }
            }
        
        return extracted_data

    # ========================================================================
    # METADATA EXTRACTION
    # ========================================================================
    
    def _extract_metadata(self):
        """Extract metadata and analysis information"""
        return {
            'driver_name': self.driver_info.get('name', 'Not Specified'),
            'analysis_date': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'analysis_type': 'Single Week Analysis' if self.is_single_week else 'Multi-Week Analysis',
            'weeks_analyzed': len(self.weekly_results),
            'week_labels': self.week_labels,
            'is_single_week': self.is_single_week,
            'report_type': self.driver_info.get('report_type', 'Road-Context Analysis')
        }

    # ========================================================================
    # NEW: TRIP SUMMARY EXTRACTION
    # ========================================================================
    
    def _extract_trip_summary(self):
        """
        Extract trip summary for all weeks and totals.
        Provides distance, duration, trip count per week.
        """
        
        trip_summary = {
            'weekly': {},
            'totals': {
                'total_distance_km': 0,
                'total_duration_hours': 0,
                'total_duration_minutes': 0,
                'total_trips': 0,
                'total_gps_points': 0,
                'avg_speed_kmh': 0,
                'max_speed_kmh': 0
            }
        }
        
        all_speeds = []
        
        for week_label in self.week_labels:
            week_trip_data = self._extract_trip_summary_for_week(week_label)
            trip_summary['weekly'][week_label] = week_trip_data
            
            # Accumulate totals
            trip_summary['totals']['total_distance_km'] += week_trip_data['distance_km']
            trip_summary['totals']['total_duration_hours'] += week_trip_data['duration_hours']
            trip_summary['totals']['total_duration_minutes'] += week_trip_data['duration_minutes']
            trip_summary['totals']['total_trips'] += week_trip_data['trip_count']
            trip_summary['totals']['total_gps_points'] += week_trip_data['total_points']
            
            if week_trip_data['max_speed_kmh'] > trip_summary['totals']['max_speed_kmh']:
                trip_summary['totals']['max_speed_kmh'] = week_trip_data['max_speed_kmh']
            
            if week_trip_data['avg_speed_kmh'] > 0:
                all_speeds.append(week_trip_data['avg_speed_kmh'])
        
        # Calculate overall average speed
        if all_speeds:
            trip_summary['totals']['avg_speed_kmh'] = round(sum(all_speeds) / len(all_speeds), 1)
        
        # Round totals
        trip_summary['totals']['total_distance_km'] = round(trip_summary['totals']['total_distance_km'], 2)
        trip_summary['totals']['total_duration_hours'] = round(trip_summary['totals']['total_duration_hours'], 2)
        trip_summary['totals']['total_duration_minutes'] = round(trip_summary['totals']['total_duration_minutes'], 1)
        
        return trip_summary
    
    def _extract_trip_summary_for_week(self, week_label):
        """Extract trip summary for a specific week"""
        
        default_summary = {
            'week_label': week_label,
            'distance_km': 0,
            'duration_hours': 0,
            'duration_minutes': 0,
            'trip_count': 0,
            'total_points': 0,
            'avg_speed_kmh': 0,
            'max_speed_kmh': 0,
            'start_time': None,
            'end_time': None
        }
        
        if not hasattr(self.processor, 'weekly_data'):
            return default_summary
        
        week_df = self.processor.weekly_data.get(week_label)
        if week_df is None or week_df.empty:
            return default_summary
        
        # Calculate metrics
        distance_km = week_df['distance_m'].sum() / 1000 if 'distance_m' in week_df.columns else 0
        duration_seconds = week_df['time_diff_s'].sum() if 'time_diff_s' in week_df.columns else 0
        duration_hours = duration_seconds / 3600
        duration_minutes = duration_seconds / 60
        
        avg_speed = week_df['speed_kmh'].mean() if 'speed_kmh' in week_df.columns else 0
        max_speed = week_df['speed_kmh'].max() if 'speed_kmh' in week_df.columns else 0
        
        # Count trips (gaps > 30 minutes indicate new trip)
        trip_count = self._count_trips(week_df)
        
        # Get time range
        start_time = week_df['timestamp'].min() if 'timestamp' in week_df.columns else None
        end_time = week_df['timestamp'].max() if 'timestamp' in week_df.columns else None
        
        return {
            'week_label': week_label,
            'distance_km': round(distance_km, 2),
            'duration_hours': round(duration_hours, 2),
            'duration_minutes': round(duration_minutes, 1),
            'trip_count': trip_count,
            'total_points': len(week_df),
            'avg_speed_kmh': round(avg_speed, 1) if avg_speed else 0,
            'max_speed_kmh': round(max_speed, 1) if max_speed else 0,
            'start_time': start_time.strftime('%Y-%m-%d %H:%M') if start_time else None,
            'end_time': end_time.strftime('%Y-%m-%d %H:%M') if end_time else None
        }
    
    def _count_trips(self, df, gap_threshold_minutes=30):
        """Count number of trips based on time gaps"""
        
        if 'timestamp' not in df.columns or len(df) < 2:
            return 1
        
        df_sorted = df.sort_values('timestamp')
        time_diffs = df_sorted['timestamp'].diff()
        
        # Count gaps > threshold as new trips
        gap_threshold = pd.Timedelta(minutes=gap_threshold_minutes)
        trip_starts = (time_diffs > gap_threshold).sum()
        
        return trip_starts + 1  # +1 for first trip

    # ========================================================================
    # WEEKLY DATA EXTRACTION
    # ========================================================================
    
    def _extract_week_data(self, week_data, week_label):
        """Extract comprehensive data for a single week"""
        
        # Get trip summary for this week
        trip_summary = self._extract_trip_summary_for_week(week_label)
        
        return {
            # === TRIP SUMMARY === (NEW)
            'trip_summary': trip_summary,
            
            # === SPEEDING DATA ===
            'speeding': {
                'total_violations': self._safe_get_count(week_data, 'speeding', 'total_violations'),
                'violations_by_severity': week_data.get('speeding', {}).get('violations_by_severity', {
                    'minor': 0, 'moderate': 0, 'major': 0, 'severe': 0
                }),
                'worst_violations': week_data.get('speeding', {}).get('worst_violations', []),
                'compliance_percentage': self._safe_get_value(week_data, 'speeding', 'compliance_percentage', 0),
                'total_violation_time': week_data.get('speeding', {}).get('total_violation_time', 0),
                'safe_driving_time': week_data.get('speeding', {}).get('safe_driving_time', 0),
                'speed_zones': week_data.get('speeding', {}).get('speed_zones', {}),
                'all_episodes': week_data.get('speeding', {}).get('all_episodes', []),
                'all_segments': week_data.get('speeding', {}).get('all_segments', [])
            },
            
            # === ROUNDABOUT DATA ===
            'roundabouts': self._extract_roundabout_data(week_data),
            
            # === STOP SIGNS DATA ===
            'stop_signs': {
                'total_stop_signs': self._safe_get_count(week_data, 'stop_signs', 'total_stop_signs'),
                'total_approaches': self._safe_get_count(week_data, 'stop_signs', 'total_approaches'),
                'compliance': {
                    'stop_ok': week_data.get('stop_signs', {}).get('compliance', {}).get('stop_ok', 0) or 
                               self._safe_get_count(week_data, 'stop_signs', 'proper_stops'),
                    'stop_ko': week_data.get('stop_signs', {}).get('compliance', {}).get('stop_ko', 0) or
                               self._safe_get_count(week_data, 'stop_signs', 'violations'),
                    'uncertain': week_data.get('stop_signs', {}).get('compliance', {}).get('uncertain', 0)
                },
                'compliance_percentage': self._safe_get_value(week_data, 'stop_signs', 'compliance_percentage', 0),
                'detection_mode': week_data.get('stop_signs', {}).get('detection_mode', 'unknown'),
                'avg_confidence': week_data.get('stop_signs', {}).get('avg_confidence', 0),
                'approaches_detailed': week_data.get('stop_signs', {}).get('approaches_detailed', [])
            },
            
            # === SCHOOL ZONES DATA ===
            'school_zones': {
                'total_passages': self._safe_get_count(week_data, 'school_zones', 'total_passages'),
                'violations': self._safe_get_count(week_data, 'school_zones', 'violations'),
                'compliance_percentage': self._calculate_school_compliance(week_data),
                'violations_detail': week_data.get('school_zones', {}).get('violations_detail', []),
                'max_speed_recorded': week_data.get('school_zones', {}).get('max_speed_recorded', 0)
            },
            
            # === TRAFFIC LIGHTS DATA === (NEW)
            'traffic_lights': {
                'total_signals': self._safe_get_count(week_data, 'traffic_lights', 'total_traffic_signals'),
                'total_approaches': self._safe_get_count(week_data, 'traffic_lights', 'total_approaches'),
                'compliance': {
                    'stopped': week_data.get('traffic_lights', {}).get('compliance', {}).get('stopped', 0),
                    'slowed': week_data.get('traffic_lights', {}).get('compliance', {}).get('slowed', 0),
                    'passed_through': week_data.get('traffic_lights', {}).get('compliance', {}).get('passed_through', 0),
                    'uncertain': week_data.get('traffic_lights', {}).get('compliance', {}).get('uncertain', 0)
                },
                'stop_percentage': self._safe_get_value(week_data, 'traffic_lights', 'stop_percentage', 0),
                'detection_mode': week_data.get('traffic_lights', {}).get('detection_mode', 'unknown'),
                'avg_confidence': week_data.get('traffic_lights', {}).get('avg_confidence', 0),
                'approaches_detailed': week_data.get('traffic_lights', {}).get('approaches_detailed', [])
            },
            
            # === HARSH EVENTS DATA === (NEW)
            'harsh_events': {
                'total_events': self._safe_get_count(week_data, 'harsh_events', 'total_harsh_events'),
                'harsh_braking': {
                    'count': week_data.get('harsh_events', {}).get('harsh_braking', {}).get('count', 0),
                    'events': week_data.get('harsh_events', {}).get('harsh_braking', {}).get('events', [])
                },
                'harsh_acceleration': {
                    'count': week_data.get('harsh_events', {}).get('harsh_acceleration', {}).get('count', 0),
                    'events': week_data.get('harsh_events', {}).get('harsh_acceleration', {}).get('events', [])
                },
                'sharp_turns': {
                    'count': week_data.get('harsh_events', {}).get('sharp_turns', {}).get('count', 0),
                    'events': week_data.get('harsh_events', {}).get('sharp_turns', {}).get('events', [])
                },
                'events_per_100km': self._safe_get_value(week_data, 'harsh_events', 'events_per_100km', 0),
                'severity_summary': week_data.get('harsh_events', {}).get('severity_summary', {}),
                'detection_method': week_data.get('harsh_events', {}).get('detection_method', 'unknown')
            },
            
            # === CALCULATED METRICS ===
            'calculated_metrics': {
                'overall_score': self._calculate_overall_score(week_data),
                'safety_score': self._calculate_safety_score(week_data),
                'compliance_score': self._calculate_compliance_score(week_data),
                'total_violations': self._calculate_total_violations(week_data)
            }
        }

    # ========================================================================
    # ROUNDABOUT DATA EXTRACTION
    # ========================================================================
    
    def _extract_roundabout_data(self, week_data):
        """Extract roundabout data supporting all formats"""
        
        roundabout_data = week_data.get('roundabouts', {})
        
        if not isinstance(roundabout_data, dict):
            return self._empty_roundabout_result()
        
        # NEW ENHANCED FORMAT
        if 'total_roundabouts' in roundabout_data and 'anticipation_stats' in roundabout_data:
            
            anticipation_stats = roundabout_data.get('anticipation_stats', {})
            entry_compliance = roundabout_data.get('entry_compliance', {})
            approaches_detailed = roundabout_data.get('approaches_detailed', [])
            
            total_roundabouts = roundabout_data.get('total_roundabouts', 0)
            total_approaches = roundabout_data.get('total_approaches', 0)
            
            good_count = anticipation_stats.get('good', 0)
            moderate_count = anticipation_stats.get('moderate', 0)
            late_count = anticipation_stats.get('late', 0)
            no_decel_count = anticipation_stats.get('no_deceleration', 0)
            
            entry_compliant = entry_compliance.get('compliant', 0)
            entry_non_compliant = entry_compliance.get('non_compliant', 0)
            entry_unknown = entry_compliance.get('unknown', 0)
            
            # Calculate entry compliance percentage
            entry_total = entry_compliant + entry_non_compliant
            if entry_total > 0:
                entry_compliance_pct = (entry_compliant / entry_total) * 100
            else:
                entry_compliance_pct = 100
            
            # Calculate anticipation percentage
            total_with_anticipation = good_count + moderate_count + late_count + no_decel_count
            if total_with_anticipation > 0:
                good_anticipation_pct = ((good_count + moderate_count) / total_with_anticipation) * 100
            else:
                good_anticipation_pct = 100
            
            # Extract zone speeds
            zone_150_speeds = []
            zone_100_speeds = []
            zone_65_speeds = []
            entry_speeds = []
            exit_speeds = []
            
            for approach in approaches_detailed:
                if approach.get('zone_150m_avg_speed') is not None:
                    zone_150_speeds.append(approach['zone_150m_avg_speed'])
                if approach.get('zone_100m_avg_speed') is not None:
                    zone_100_speeds.append(approach['zone_100m_avg_speed'])
                if approach.get('zone_65m_avg_speed') is not None:
                    zone_65_speeds.append(approach['zone_65m_avg_speed'])
                if approach.get('entry_speed') is not None:
                    entry_speeds.append(approach['entry_speed'])
                if approach.get('exit_speed') is not None:
                    exit_speeds.append(approach['exit_speed'])
            
            return {
                'format': 'enhanced',
                'total_roundabouts': total_roundabouts,
                'total_approaches': total_approaches,
                'anticipation_stats': {
                    'good': good_count,
                    'moderate': moderate_count,
                    'late': late_count,
                    'no_deceleration': no_decel_count
                },
                'no_deceleration_count': no_decel_count,
                'good_anticipation_percentage': round(good_anticipation_pct, 1),
                'entry_compliance': {
                    'compliant': entry_compliant,
                    'non_compliant': entry_non_compliant,
                    'unknown': entry_unknown
                },
                'entry_compliance_percentage': round(entry_compliance_pct, 1),
                'approaches_detailed': approaches_detailed,
                'avg_zone_speeds': {
                    '150m': round(sum(zone_150_speeds) / len(zone_150_speeds), 1) if zone_150_speeds else None,
                    '100m': round(sum(zone_100_speeds) / len(zone_100_speeds), 1) if zone_100_speeds else None,
                    '65m': round(sum(zone_65_speeds) / len(zone_65_speeds), 1) if zone_65_speeds else None
                },
                'avg_entry_speed': round(sum(entry_speeds) / len(entry_speeds), 1) if entry_speeds else None,
                'avg_exit_speed': round(sum(exit_speeds) / len(exit_speeds), 1) if exit_speeds else None,
                'detection_mode': roundabout_data.get('detection_mode', 'unknown'),
                'avg_confidence': roundabout_data.get('avg_confidence', 0),
                'overall_percentage': round(good_anticipation_pct, 1),
                'roundabout_compliance_percentage': round(entry_compliance_pct, 1)
            }
        
        # OLD MULTI-ZONE FORMAT
        elif any(zone in roundabout_data for zone in ['150m', '100m', '50m']):
            detailed = roundabout_data.get('approaches_detailed', [])
            no_decel_count = len([a for a in detailed if a.get('anticipation_category') == 'no_deceleration'])
            
            return {
                'format': 'multi_zone',
                'data': {
                    '150m': {
                        'total_approaches': roundabout_data.get('150m', {}).get('total_approaches', 0),
                        'good_approaches': roundabout_data.get('150m', {}).get('good_approaches', 0),
                        'good_percentage': roundabout_data.get('150m', {}).get('good_percentage', 0)
                    },
                    '100m': {
                        'total_approaches': roundabout_data.get('100m', {}).get('total_approaches', 0),
                        'good_approaches': roundabout_data.get('100m', {}).get('good_approaches', 0),
                        'good_percentage': roundabout_data.get('100m', {}).get('good_percentage', 0)
                    },
                    '50m': {
                        'total_approaches': roundabout_data.get('50m', {}).get('total_approaches', 0),
                        'good_approaches': roundabout_data.get('50m', {}).get('good_approaches', 0),
                        'good_percentage': roundabout_data.get('50m', {}).get('good_percentage', 0)
                    }
                },
                'overall_percentage': roundabout_data.get('good_percentage', 0),
                'total_roundabouts': roundabout_data.get('total_roundabouts', 0),
                'roundabout_compliance_percentage': roundabout_data.get('roundabout_compliance_percentage', 100),
                'no_deceleration_count': no_decel_count,
                'approaches_detailed': detailed
            }
        
        # OLD SINGLE-ZONE FORMAT
        elif 'total_approaches' in roundabout_data:
            return {
                'format': 'single_zone',
                'data': {
                    'total_approaches': roundabout_data.get('total_approaches', 0),
                    'good_approaches': roundabout_data.get('good_approaches', 0),
                    'good_percentage': roundabout_data.get('good_percentage', 0)
                },
                'no_deceleration_count': 0
            }
        
        return self._empty_roundabout_result()
    
    def _empty_roundabout_result(self):
        """Return empty roundabout result structure"""
        return {
            'format': 'none',
            'total_roundabouts': 0,
            'total_approaches': 0,
            'anticipation_stats': {
                'good': 0, 'moderate': 0, 'late': 0, 'no_deceleration': 0
            },
            'no_deceleration_count': 0,
            'entry_compliance': {
                'compliant': 0, 'non_compliant': 0, 'unknown': 0
            },
            'entry_compliance_percentage': 100,
            'approaches_detailed': [],
            'avg_zone_speeds': {'150m': None, '100m': None, '65m': None},
            'avg_entry_speed': None,
            'avg_exit_speed': None,
            'good_anticipation_percentage': 100,
            'overall_percentage': 100,
            'roundabout_compliance_percentage': 100
        }

    # ========================================================================
    # AGGREGATED METRICS CALCULATION
    # ========================================================================
    
    def _calculate_aggregated_metrics(self, weekly_data):
        """Calculate aggregated metrics across all weeks"""
        
        aggregated = {
            'total_violations': 0,
            'average_violations_per_week': 0,
            'best_week': {'week': '', 'score': 0, 'violations': float('inf')},
            'worst_week': {'week': '', 'score': 100, 'violations': 0},
            'combined_metrics': {}
        }
        
        # Collect weekly metrics
        weekly_scores = []
        weekly_violations = []
        roundabout_scores = []
        stop_sign_scores = []
        school_zone_scores = []
        speed_scores = []
        traffic_light_scores = []
        harsh_event_rates = []
        
        for week_label in self.week_labels:
            week_extracted = weekly_data.get(week_label, {})
            
            overall_score = week_extracted.get('calculated_metrics', {}).get('overall_score', 0) or 0
            total_violations = week_extracted.get('speeding', {}).get('total_violations', 0) or 0
            
            weekly_scores.append(overall_score)
            weekly_violations.append(total_violations)
            
            # Individual metrics (with None protection)
            roundabout_scores.append(self._get_roundabout_score(week_extracted) or 0)
            stop_sign_scores.append(week_extracted.get('stop_signs', {}).get('compliance_percentage') or 0)
            school_zone_scores.append(week_extracted.get('school_zones', {}).get('compliance_percentage') or 0)
            speed_scores.append(week_extracted.get('speeding', {}).get('compliance_percentage') or 0)
            traffic_light_scores.append(week_extracted.get('traffic_lights', {}).get('stop_percentage') or 0)
            harsh_event_rates.append(week_extracted.get('harsh_events', {}).get('events_per_100km') or 0)
            
            # Track best/worst weeks
            if overall_score > aggregated['best_week']['score']:
                aggregated['best_week'] = {
                    'week': week_label,
                    'score': overall_score,
                    'violations': total_violations
                }
            
            if overall_score < aggregated['worst_week']['score']:
                aggregated['worst_week'] = {
                    'week': week_label,
                    'score': overall_score,
                    'violations': total_violations
                }
        
        # Calculate aggregated values
        aggregated['total_violations'] = sum(weekly_violations)
        aggregated['average_violations_per_week'] = round(sum(weekly_violations) / len(weekly_violations), 1) if weekly_violations else 0
        
        # Calculate averages (with None protection)
        aggregated['combined_metrics'] = {
            'avg_overall_score': round(sum(weekly_scores) / len(weekly_scores), 1) if weekly_scores else 0,
            'avg_roundabout_performance': round(sum(roundabout_scores) / len(roundabout_scores), 1) if roundabout_scores else 0,
            'avg_stop_sign_compliance': round(sum(stop_sign_scores) / len(stop_sign_scores), 1) if stop_sign_scores else 0,
            'avg_school_zone_compliance': round(sum(school_zone_scores) / len(school_zone_scores), 1) if school_zone_scores else 0,
            'avg_speed_compliance': round(sum(speed_scores) / len(speed_scores), 1) if speed_scores else 0,
            'avg_traffic_light_compliance': round(sum(traffic_light_scores) / len(traffic_light_scores), 1) if traffic_light_scores else 0,
            'avg_harsh_events_per_100km': round(sum(harsh_event_rates) / len(harsh_event_rates), 2) if harsh_event_rates else 0
        }
        
        return aggregated

    # ========================================================================
    # NEW: TREND CALCULATIONS
    # ========================================================================
    
    def _calculate_weekly_trends(self, weekly_data):
        """Calculate week-over-week trends with arrows"""
        
        trends = {
            'weekly_changes': {},
            'overall_trend': {}
        }
        
        week_labels = sorted(weekly_data.keys())
        
        if len(week_labels) < 2:
            return trends
        
        # Calculate changes between consecutive weeks
        for i in range(1, len(week_labels)):
            prev_week = week_labels[i - 1]
            curr_week = week_labels[i]
            
            prev_data = weekly_data[prev_week]
            curr_data = weekly_data[curr_week]
            
            trends['weekly_changes'][curr_week] = {
                'overall_score': self._get_trend_info(
                    prev_data.get('calculated_metrics', {}).get('overall_score', 0),
                    curr_data.get('calculated_metrics', {}).get('overall_score', 0)
                ),
                'speed_compliance': self._get_trend_info(
                    prev_data.get('speeding', {}).get('compliance_percentage', 0),
                    curr_data.get('speeding', {}).get('compliance_percentage', 0)
                ),
                'stop_sign_compliance': self._get_trend_info(
                    prev_data.get('stop_signs', {}).get('compliance_percentage', 0),
                    curr_data.get('stop_signs', {}).get('compliance_percentage', 0)
                ),
                'roundabout_performance': self._get_trend_info(
                    self._get_roundabout_score(prev_data),
                    self._get_roundabout_score(curr_data)
                ),
                'violations': self._get_trend_info(
                    prev_data.get('speeding', {}).get('total_violations', 0),
                    curr_data.get('speeding', {}).get('total_violations', 0),
                    invert=True  # Lower is better for violations
                ),
                'harsh_events': self._get_trend_info(
                    prev_data.get('harsh_events', {}).get('events_per_100km', 0),
                    curr_data.get('harsh_events', {}).get('events_per_100km', 0),
                    invert=True  # Lower is better
                )
            }
        
        # Calculate overall trend (first week to last week)
        first_week = week_labels[0]
        last_week = week_labels[-1]
        first_data = weekly_data[first_week]
        last_data = weekly_data[last_week]
        
        first_score = first_data.get('calculated_metrics', {}).get('overall_score', 0) or 0
        last_score = last_data.get('calculated_metrics', {}).get('overall_score', 0) or 0
        
        trends['overall_trend'] = {
            'direction': 'improving' if last_score > first_score else ('declining' if last_score < first_score else 'stable'),
            'change_percentage': round(((last_score - first_score) / first_score) * 100, 1) if first_score > 0 else 0,
            'first_score': first_score,
            'last_score': last_score
        }
        
        return trends
    
    def _get_trend_info(self, prev_value, curr_value, threshold=2, invert=False):
        """Get trend arrow and change info"""
        
        prev_value = prev_value or 0
        curr_value = curr_value or 0
        
        diff = curr_value - prev_value
        
        if invert:
            diff = -diff  # Invert for metrics where lower is better
        
        if abs(diff) < threshold:
            arrow = '→'
            direction = 'stable'
        elif diff > 0:
            arrow = '↑'
            direction = 'improving'
        else:
            arrow = '↓'
            direction = 'declining'
        
        return {
            'arrow': arrow,
            'direction': direction,
            'change': round(curr_value - prev_value, 1),
            'prev_value': round(prev_value, 1),
            'curr_value': round(curr_value, 1)
        }

    # ========================================================================
    # SCORE CALCULATIONS
    # ========================================================================
    
    def _calculate_overall_score(self, week_data):
        """
        Calculate overall driving score.
        
        Weights:
        - Speed compliance: 25%
        - Stop sign compliance: 20%
        - Roundabout performance: 20%
        - School zone compliance: 20%
        - Harsh events (inverted): 15%
        """
        
        speed_score = (self._safe_get_value(week_data, 'speeding', 'compliance_percentage', 0) or 0) * 0.25
        stop_score = (self._safe_get_value(week_data, 'stop_signs', 'compliance_percentage', 0) or 0) * 0.20
        roundabout_score = (self._get_roundabout_score_from_raw(week_data) or 0) * 0.20
        school_score = (self._calculate_school_compliance(week_data) or 0) * 0.20
        
        # Harsh events - convert to score (lower events = higher score)
        harsh_events_per_100km = self._safe_get_value(week_data, 'harsh_events', 'events_per_100km', 0) or 0
        harsh_score = max(0, 100 - (harsh_events_per_100km * 10)) * 0.15
        
        total = speed_score + stop_score + roundabout_score + school_score + harsh_score
        return round(total, 1)
    
    def _calculate_safety_score(self, week_data):
        """Calculate safety-focused score (school zones, stops, harsh events weighted higher)"""
        
        school_score = (self._calculate_school_compliance(week_data) or 0) * 0.35
        stop_score = (self._safe_get_value(week_data, 'stop_signs', 'compliance_percentage', 0) or 0) * 0.25
        speed_score = (self._safe_get_value(week_data, 'speeding', 'compliance_percentage', 0) or 0) * 0.20
        
        # Harsh events penalty
        harsh_events_per_100km = self._safe_get_value(week_data, 'harsh_events', 'events_per_100km', 0) or 0
        harsh_score = max(0, 100 - (harsh_events_per_100km * 10)) * 0.20
        
        return round(school_score + stop_score + speed_score + harsh_score, 1)
    
    def _calculate_compliance_score(self, week_data):
        """Calculate rule compliance score"""
        return self._calculate_overall_score(week_data)
    
    def _calculate_total_violations(self, week_data):
        """Calculate total violations across all categories"""
        
        speed_violations = self._safe_get_count(week_data, 'speeding', 'total_violations')
        stop_violations = week_data.get('stop_signs', {}).get('compliance', {}).get('stop_ko', 0) or \
                 self._safe_get_count(week_data, 'stop_signs', 'violations')
        school_violations = self._safe_get_count(week_data, 'school_zones', 'violations')
        
        return speed_violations + stop_violations + school_violations
    
    def _calculate_school_compliance(self, week_data):
        """Calculate school zone compliance percentage"""
        
        school_data = week_data.get('school_zones', {})
        
        if 'compliance_percentage' in school_data and school_data['compliance_percentage'] is not None:
            return school_data['compliance_percentage']
        
        if 'violation_percentage' in school_data:
            return 100.0 - (school_data['violation_percentage'] or 0)
        
        total_passages = school_data.get('total_passages', 0)
        violations = school_data.get('violations', 0)
        
        if total_passages > 0:
            return round(100 * (1 - violations / total_passages), 1)
        
        return 100.0
    
    def _get_roundabout_score(self, week_extracted):
        """Get roundabout score from extracted data"""
        
        roundabout_data = week_extracted.get('roundabouts', {})
        
        if roundabout_data.get('format') == 'enhanced':
            return roundabout_data.get('entry_compliance_percentage', 100)
        elif roundabout_data.get('format') == 'multi_zone':
            return roundabout_data.get('overall_percentage', 100)
        elif roundabout_data.get('format') == 'single_zone':
            return roundabout_data.get('data', {}).get('good_percentage', 100)
        
        return 100.0
    
    def _get_roundabout_score_from_raw(self, week_data):
        """Get roundabout score from raw week data"""
        
        roundabout_data = week_data.get('roundabouts', {})
        
        if 'entry_compliance_percentage' in roundabout_data:
            return roundabout_data['entry_compliance_percentage']
        elif 'good_percentage' in roundabout_data:
            return roundabout_data['good_percentage']
        elif '50m' in roundabout_data:
            return roundabout_data.get('50m', {}).get('good_percentage', 100)
        
        return 100.0

    # ========================================================================
    # MAP & GPS DATA EXTRACTION
    # ========================================================================
    
    def _extract_map_data(self):
        """Extract map verification data"""
        
        if not self.map_manager or not hasattr(self.map_manager, 'map_data'):
            return {'available': False, 'data': {}}
        
        map_data = self.map_manager.map_data
        
        return {
            'available': True,
            'data': {
                'schools': [
                    {
                        'index': i + 1,
                        'name': school.get('name', 'Unnamed'),
                        'school_type': school.get('school_type', 'school'),
                        'lat': round(school.get('lat', 0), 6),
                        'lon': round(school.get('lon', 0), 6),
                        'osm_id': school.get('id', 'N/A')
                    }
                    for i, school in enumerate(map_data.get('schools', []))
                ],
                'stop_signs': [
                    {
                        'index': i + 1,
                        'lat': round(stop.get('lat', 0), 6),
                        'lon': round(stop.get('lon', 0), 6),
                        'osm_id': stop.get('osmid', stop.get('id', 'N/A'))
                    }
                    for i, stop in enumerate(map_data.get('stop_signs', []))
                ],
                'roundabouts': [
                    {
                        'index': i + 1,
                        'type': rb.get('type', 'way'),
                        'lat': round(rb.get('lat', 0), 6),
                        'lon': round(rb.get('lon', 0), 6),
                        'osm_id': rb.get('osmid', 'N/A')
                    }
                    for i, rb in enumerate(map_data.get('roundabouts', []))
                ],
                'traffic_lights': [
                    {
                        'index': i + 1,
                        'lat': round(tl.get('lat', 0), 6),
                        'lon': round(tl.get('lon', 0), 6),
                        'osm_id': tl.get('osmid', tl.get('id', 'N/A'))
                    }
                    for i, tl in enumerate(map_data.get('traffic_lights', []))
                ]
            }
        }
    
    def _extract_gps_stats(self):
        """Extract GPS route statistics"""
        
        if not hasattr(self.processor, 'combined_data') or self.processor.combined_data is None:
            return {'available': False, 'stats': {}}
        
        df = self.processor.combined_data
        
        return {
            'available': True,
            'stats': {
                'total_points': len(df),
                'start_time': df['timestamp'].min().strftime('%Y-%m-%d %H:%M:%S') if 'timestamp' in df.columns else None,
                'end_time': df['timestamp'].max().strftime('%Y-%m-%d %H:%M:%S') if 'timestamp' in df.columns else None,
                'duration_minutes': round(df['time_diff_s'].sum() / 60, 1) if 'time_diff_s' in df.columns else 0,
                'total_distance_km': round(df['distance_m'].sum() / 1000, 2) if 'distance_m' in df.columns else 0,
                'average_speed_kmh': round(df['speed_kmh'].mean(), 1) if 'speed_kmh' in df.columns else 0,
                'max_speed_kmh': round(df['speed_kmh'].max(), 1) if 'speed_kmh' in df.columns else 0,
                'bounding_box': {
                    'min_lat': round(df['lat'].min(), 6),
                    'max_lat': round(df['lat'].max(), 6),
                    'min_lon': round(df['lon'].min(), 6),
                    'max_lon': round(df['lon'].max(), 6)
                }
            }
        }

    # ========================================================================
    # HELPER METHODS
    # ========================================================================
    
    def _safe_get_value(self, data, key1, key2=None, default=0):
        """Safely get nested value, never returns None"""
        
        if isinstance(data, dict):
            if key2 is not None:
                value = data.get(key1, {})
                if isinstance(value, dict):
                    result = value.get(key2, default)
                else:
                    result = default
            else:
                result = data.get(key1, default)
        else:
            result = default
        
        return result if result is not None else default
    
    def _safe_get_count(self, data, category, key, default=0):
        """Safely get count value"""
        value = data.get(category, {}).get(key, default)
        return value if value is not None else default
    
    def _safe_get_percentage(self, data, category, key, default=0.0):
        """Safely get percentage value"""
        value = data.get(category, {}).get(key, default)
        return value if value is not None else default

    # ========================================================================
    # SUMMARY OUTPUT
    # ========================================================================
    
    def print_summary(self):
        """Print unified summary"""
        
        print("\n" + "=" * 70)
        print("🛣️  UNIFIED ANALYSIS SUMMARY")
        print("=" * 70)
        
        all_data = self.extract_all_data()
        metadata = all_data['metadata']
        trip_summary = all_data['trip_summary']
        
        print(f"👤 Driver: {metadata['driver_name']}")
        print(f"📊 Analysis Type: {metadata['analysis_type']}")
        print(f"🗓️  Weeks Analyzed: {metadata['weeks_analyzed']}")
        
        # Trip summary
        print(f"\n📍 TRIP SUMMARY:")
        print(f"   Total Distance: {trip_summary['totals']['total_distance_km']:.1f} km")
        print(f"   Total Duration: {trip_summary['totals']['total_duration_hours']:.1f} hours")
        print(f"   Total Trips: {trip_summary['totals']['total_trips']}")
        print(f"   Avg Speed: {trip_summary['totals']['avg_speed_kmh']:.1f} km/h")
        
        # Performance
        aggregated = all_data['aggregated_metrics']
        print(f"\n🏆 PERFORMANCE:")
        print(f"   Overall Score: {aggregated['combined_metrics']['avg_overall_score']:.1f}%")
        print(f"   Speed Compliance: {aggregated['combined_metrics']['avg_speed_compliance']:.1f}%")
        print(f"   Stop Sign Compliance: {aggregated['combined_metrics']['avg_stop_sign_compliance']:.1f}%")
        print(f"   Roundabout Performance: {aggregated['combined_metrics']['avg_roundabout_performance']:.1f}%")
        print(f"   Harsh Events/100km: {aggregated['combined_metrics']['avg_harsh_events_per_100km']:.1f}")
        
        # Trends (if multi-week)
        if not self.is_single_week and all_data.get('trends', {}).get('overall_trend'):
            trend = all_data['trends']['overall_trend']
            print(f"\n📈 TREND: {trend['direction'].upper()}")
            print(f"   Change: {trend['change_percentage']:+.1f}%")
        
        print("=" * 70)
        
        return all_data
# =============================================================
# plot code start
#===========================================================
#"Plot version v2"
# =============================================================
# MODULAR PLOT CODE - v3
# =============================================================

import folium
from folium import plugins
import os


class DrivingBehaviorMapGenerator:
    """
    Modular map generator for driving behavior analysis validation.
    
    Features:
    - GPS points with speed color coding
    - Route geometry overlay
    - Speeding violations with detailed popups
    - Roundabout behavior markers
    - Stop sign compliance markers
    - Traffic light behavior markers (NEW)
    - Harsh events markers (NEW)
    - Trip summary info panel
    """
    
    # Color schemes
    COLORS = {
        # Speed colors
        'speed_high': 'darkred',      # > 80 km/h
        'speed_medium_high': 'red',   # 60-80 km/h
        'speed_medium': 'orange',     # 40-60 km/h
        'speed_low': 'green',         # < 40 km/h
        
        # Compliance colors
        'compliant': 'green',
        'partial': 'orange',
        'violation': 'red',
        'unknown': 'gray',
        
        # Severity colors
        'severe': 'darkred',
        'major': 'red',
        'moderate': 'orange',
        'minor': 'beige',
        
        # Feature colors
        'gps_point': 'red',
        'matched_point': 'blue',
        'gps_path': 'orange',
        'route_geometry': 'green',
        'school_zone': 'purple',
        'traffic_light': 'cadetblue',
        'harsh_event': 'darkred'
    }
    
    def __init__(self, extracted_data, week_label, week_gps_df):
        """
        Initialize map generator for a specific week.
        
        Args:
            extracted_data: Complete data from UnifiedReportGenerator
            week_label: Week identifier (e.g., "Week 1")
            week_gps_df: GPS DataFrame for this week
        """
        self.extracted_data = extracted_data
        self.week_label = week_label
        self.week_gps_df = week_gps_df
        self.week_analysis = extracted_data.get('weekly_data', {}).get(week_label, {})
        self.metadata = extracted_data.get('metadata', {})
        self.trip_summary = extracted_data.get('trip_summary', {}).get('weekly', {}).get(week_label, {})
        
        # Map object
        self.map = None
        self.feature_count = 0
    
    # ========================================================================
    # MAIN GENERATION METHOD
    # ========================================================================
    
    def generate_map(self, output_filename):
        """Generate complete map with all features"""
        
        print(f"   🗺️ Generating map for {self.week_label}...")
        
        # Create base map
        self._create_base_map()
        
        # Add layers
        self._add_info_panel()
        self._add_gps_points()
        self._add_gps_path()
        self._add_road_matched_points()
        self._add_route_geometry()
        
        # Add analysis features
        self._add_speeding_markers()
        self._add_roundabout_markers()
        self._add_stop_sign_markers()
        self._add_traffic_light_markers()
        self._add_harsh_event_markers()
        self._add_school_zone_markers()
        
        # Add extras
        self._add_speed_heatmap()
        self._add_legend()

        # Add layer control
        folium.LayerControl().add_to(self.map)
        
        # Save map
        self._save_map(output_filename)
        
        print(f"   ✅ Map generated with {self.feature_count} features")
        return output_filename
    
    # ========================================================================
    # BASE MAP CREATION
    # ========================================================================
    
    def _create_base_map(self):
        """Create base map centered on GPS data"""
        
        center_lat = self.week_gps_df['lat'].mean()
        center_lon = self.week_gps_df['lon'].mean()
        
        self.map = folium.Map(
            location=[center_lat, center_lon],
            zoom_start=15,
            tiles='https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png',
            attr='CartoDB Positron'
        )
        

    
    # ========================================================================
    # INFO PANEL
    # ========================================================================
    
    def _add_info_panel(self):
        """Add comprehensive info panel with trip summary and scores"""
        
        driver_name = self.metadata.get('driver_name', 'Driver')
        
        # Get metrics
        speeding = self.week_analysis.get('speeding', {})
        roundabouts = self.week_analysis.get('roundabouts', {})
        stop_signs = self.week_analysis.get('stop_signs', {})
        traffic_lights = self.week_analysis.get('traffic_lights', {})
        harsh_events = self.week_analysis.get('harsh_events', {})
        calculated = self.week_analysis.get('calculated_metrics', {})
        
        # Extract values
        overall_score = calculated.get('overall_score', 0)
        speed_violations = speeding.get('total_violations', 0)
        speed_compliance = speeding.get('compliance_percentage', 0)
        
        roundabout_compliance = roundabouts.get('entry_compliance_percentage', 0) or \
                               roundabouts.get('overall_percentage', 0)
        stop_compliance = stop_signs.get('compliance_percentage', 0)
        
        traffic_light_count = traffic_lights.get('total_signals', 0)
        traffic_light_stop_pct = traffic_lights.get('stop_percentage', 0)
        
        harsh_total = harsh_events.get('total_events', 0)
        harsh_per_100km = harsh_events.get('events_per_100km', 0)
        
        # Trip summary
        distance_km = self.trip_summary.get('distance_km', 0)
        duration_hours = self.trip_summary.get('duration_hours', 0)
        trip_count = self.trip_summary.get('trip_count', 0)
        avg_speed = self.trip_summary.get('avg_speed_kmh', 0)
        max_speed = self.trip_summary.get('max_speed_kmh', 0)
        
        info_html = f"""
        <div style="position: fixed; top: 10px; left: 50px; z-index: 1000; 
                    background: rgba(255,255,255,0.95); padding: 15px; border: 2px solid #1E3A8A;
                    border-radius: 8px; box-shadow: 0 4px 8px rgba(0,0,0,0.3); font-family: Arial;
                    max-width: 320px;">
            
            <h3 style="margin: 0 0 10px 0; color: #1E3A8A; border-bottom: 2px solid #1E3A8A; padding-bottom: 5px;">
                {self.week_label} - {driver_name}
            </h3>
            
            <!-- Trip Summary -->
            <div style="background: #f0f9ff; padding: 8px; border-radius: 4px; margin-bottom: 10px;">
                <h4 style="margin: 0 0 5px 0; color: #1E3A8A; font-size: 12px;">📍 Trip Summary</h4>
                <table style="width: 100%; font-size: 11px;">
                    <tr><td>Distance:</td><td style="text-align: right;"><b>{distance_km:.1f} km</b></td></tr>
                    <tr><td>Duration:</td><td style="text-align: right;"><b>{duration_hours:.1f} h</b></td></tr>
                    <tr><td>Trips:</td><td style="text-align: right;"><b>{trip_count}</b></td></tr>
                    <tr><td>Avg Speed:</td><td style="text-align: right;"><b>{avg_speed:.0f} km/h</b></td></tr>
                    <tr><td>Max Speed:</td><td style="text-align: right;"><b>{max_speed:.0f} km/h</b></td></tr>
                </table>
            </div>
            
            <!-- Scores -->
            <div style="background: #f0fdf4; padding: 8px; border-radius: 4px; margin-bottom: 10px;">
                <h4 style="margin: 0 0 5px 0; color: #166534; font-size: 12px;">🏆 Performance</h4>
                <table style="width: 100%; font-size: 11px;">
                    <tr>
                        <td>Overall Score:</td>
                        <td style="text-align: right;"><b style="color: {self._score_color(overall_score)};">{overall_score:.1f}%</b></td>
                    </tr>
                    <tr>
                        <td>Speed Compliance:</td>
                        <td style="text-align: right;"><b style="color: {self._score_color(speed_compliance)};">{speed_compliance:.1f}%</b></td>
                    </tr>
                    <tr>
                        <td>Roundabout:</td>
                        <td style="text-align: right;"><b style="color: {self._score_color(roundabout_compliance)};">{roundabout_compliance:.1f}%</b></td>
                    </tr>
                    <tr>
                        <td>Stop Signs:</td>
                        <td style="text-align: right;"><b style="color: {self._score_color(stop_compliance)};">{stop_compliance:.1f}%</b></td>
                    </tr>
                </table>
            </div>
            
            <!-- Violations -->
            <div style="background: #fef2f2; padding: 8px; border-radius: 4px; margin-bottom: 10px;">
                <h4 style="margin: 0 0 5px 0; color: #991b1b; font-size: 12px;">⚠️ Events</h4>
                <table style="width: 100%; font-size: 11px;">
                    <tr><td>Speed Violations:</td><td style="text-align: right;"><b>{speed_violations}</b></td></tr>
                    <tr><td>Traffic Lights:</td><td style="text-align: right;"><b>{traffic_light_count}</b> ({traffic_light_stop_pct:.0f}% stopped)</td></tr>
                    <tr><td>Harsh Events:</td><td style="text-align: right;"><b>{harsh_total}</b> ({harsh_per_100km:.1f}/100km)</td></tr>
                </table>
            </div>
            
            <!-- Legend -->
            <div style="font-size: 10px; color: #666; border-top: 1px solid #ddd; padding-top: 8px;">
                🔴 GPS points &nbsp; 🟠 GPS path<br>
                🔵 Road-matched &nbsp; 🟢 Route geometry
            </div>
        </div>
        """
        
        self.map.get_root().html.add_child(folium.Element(info_html))
    
    def _score_color(self, score):
        """Get color for score display"""
        if score >= 80:
            return '#16a34a'  # Green
        elif score >= 60:
            return '#ca8a04'  # Yellow
        elif score >= 40:
            return '#ea580c'  # Orange
        else:
            return '#dc2626'  # Red
    
    # ========================================================================
    # GPS DATA LAYERS
    # ========================================================================
    
    def _add_gps_points(self):
        """Add original GPS points with speed color coding"""
        
        gps_layer = folium.FeatureGroup(name='GPS Points', show=True)
        
        for idx, row in self.week_gps_df.iterrows():
            speed = row.get('speed_kmh', 0)
            
            # Color by speed
            if speed > 80:
                color = self.COLORS['speed_high']
            elif speed > 60:
                color = self.COLORS['speed_medium_high']
            elif speed > 40:
                color = self.COLORS['speed_medium']
            else:
                color = self.COLORS['speed_low']
            
            popup_html = f"""
            <div style="font-family: Arial; font-size: 11px; min-width: 150px;">
                <b>GPS Point #{idx}</b><br>
                <hr style="margin: 5px 0;">
                <b>Location:</b> {row['lat']:.6f}, {row['lon']:.6f}<br>
                <b>Speed:</b> {speed:.1f} km/h<br>
                <b>Time:</b> {row.get('timestamp', 'N/A')}<br>
                <b>Heading:</b> {row.get('heading', 'N/A')}°
            </div>
            """
            
            folium.CircleMarker(
                location=[row['lat'], row['lon']],
                radius=4,
                popup=folium.Popup(popup_html, max_width=200),
                color=self.COLORS['gps_point'],
                fillColor=color,
                fillOpacity=0.7,
                tooltip=f"Speed: {speed:.0f} km/h"
            ).add_to(gps_layer)
        
        gps_layer.add_to(self.map)
        print(f"      📍 Added {len(self.week_gps_df)} GPS points")
    
    def _add_gps_path(self):
        """Add GPS path line"""
        
        path_layer = folium.FeatureGroup(name='GPS Path', show=True)
        
        gps_coords = [[row['lat'], row['lon']] for _, row in self.week_gps_df.iterrows()]
        
        folium.PolyLine(
            locations=gps_coords,
            color=self.COLORS['gps_path'],
            weight=3,
            opacity=0.8,
            dash_array='5, 5',
            popup=f"GPS Path - {self.week_label}"
        ).add_to(path_layer)
        
        path_layer.add_to(self.map)
    
    def _add_road_matched_points(self):
        """Add road-matched points from route geometry"""
        
        matched_layer = folium.FeatureGroup(name='Road-Matched Points', show=False)
        
        route_geometry = self.extracted_data.get('weekly_geometries', {}).get(
            self.week_label, {}
        ).get('route_geometry')
        
        if route_geometry is not None:
            count = 0
            for i, (lon, lat) in enumerate(route_geometry.coords):
                folium.CircleMarker(
                    location=[lat, lon],
                    radius=3,
                    color=self.COLORS['matched_point'],
                    fillColor=self.COLORS['matched_point'],
                    fillOpacity=0.5,
                    popup=f"Matched point #{i+1}"
                ).add_to(matched_layer)
                count += 1
            
            print(f"      🛣️ Added {count} road-matched points")
        
        matched_layer.add_to(self.map)
    
    def _add_route_geometry(self):
        """Add route geometry line"""
        
        route_layer = folium.FeatureGroup(name='Route Geometry', show=True)
        
        route_geometry = self.extracted_data.get('weekly_geometries', {}).get(
            self.week_label, {}
        ).get('route_geometry')
        
        if route_geometry is not None:
            route_coords = [[y, x] for x, y in route_geometry.coords]
            
            folium.PolyLine(
                locations=route_coords,
                color=self.COLORS['route_geometry'],
                weight=4,
                opacity=1.0,
                popup=f"Map-Matched Route - {self.week_label}"
            ).add_to(route_layer)
            
            print(f"      🛤️ Added route geometry ({len(route_coords)} points)")
        
        route_layer.add_to(self.map)
    
    # ========================================================================
    # SPEEDING MARKERS
    # ========================================================================
    
    def _add_speeding_markers(self):
        """Add speeding violation markers with detailed popups"""
        
        speed_layer = folium.FeatureGroup(name='Speed Violations', show=True)
        
        speeding_data = self.week_analysis.get('speeding', {})
        worst_violations = speeding_data.get('worst_violations', [])
        all_episodes = speeding_data.get('all_episodes', [])
        
        # Use all_episodes if available, otherwise worst_violations
        violations_to_plot = all_episodes if all_episodes else worst_violations
        
        print(f"      🚨 Found {len(violations_to_plot)} speeding violations")
        
        for i, violation in enumerate(violations_to_plot[:20]):  # Limit to 20
            if 'lat' not in violation or 'lon' not in violation:
                continue
            
            severity = violation.get('severity', 'unknown')
            speed = violation.get('speed_kmh', 0)
            limit = violation.get('speed_limit', 0)
            excess = violation.get('excess_speed', 0)
            duration = violation.get('duration_seconds', 0)
            distance = violation.get('distance_m', 0)
            
            # Get severity color
            color = self.COLORS.get(severity, 'red')
            
            popup_html = f"""
            <div style="font-family: Arial; font-size: 11px; min-width: 220px;">
                <h4 style="margin: 0 0 8px 0; padding: 5px; background: {self._severity_bg_color(severity)}; 
                           color: white; border-radius: 4px;">
                    🚨 Speed Violation #{i+1}
                </h4>
                
                <table style="width: 100%; border-collapse: collapse;">
                    <tr style="background: #f5f5f5;">
                        <td colspan="2" style="padding: 4px; font-weight: bold;">📊 Speed Data</td>
                    </tr>
                    <tr>
                        <td style="padding: 2px 4px;">Speed:</td>
                        <td style="padding: 2px 4px;"><b>{speed:.1f} km/h</b></td>
                    </tr>
                    <tr>
                        <td style="padding: 2px 4px;">Limit:</td>
                        <td style="padding: 2px 4px;">{limit} km/h</td>
                    </tr>
                    <tr>
                        <td style="padding: 2px 4px;">Excess:</td>
                        <td style="padding: 2px 4px; color: red;"><b>+{excess:.1f} km/h</b></td>
                    </tr>
                    <tr>
                        <td style="padding: 2px 4px;">Severity:</td>
                        <td style="padding: 2px 4px;"><b>{severity.upper()}</b></td>
                    </tr>
                    
                    <tr style="background: #fff3e0;">
                        <td colspan="2" style="padding: 4px; font-weight: bold;">⏱️ Episode Details</td>
                    </tr>
                    <tr>
                        <td style="padding: 2px 4px;">Duration:</td>
                        <td style="padding: 2px 4px;">{duration:.0f} seconds</td>
                    </tr>
                    <tr>
                        <td style="padding: 2px 4px;">Distance:</td>
                        <td style="padding: 2px 4px;">{distance:.0f} m</td>
                    </tr>
                    <tr>
                        <td style="padding: 2px 4px;">Time:</td>
                        <td style="padding: 2px 4px;">{violation.get('timestamp', 'N/A')}</td>
                    </tr>
                    
                    <tr style="background: #e3f2fd;">
                        <td colspan="2" style="padding: 4px; font-weight: bold;">📍 Location</td>
                    </tr>
                    <tr>
                        <td style="padding: 2px 4px;">Coordinates:</td>
                        <td style="padding: 2px 4px;">{violation['lat']:.6f}, {violation['lon']:.6f}</td>
                    </tr>
                </table>
            </div>
            """
            
            folium.Marker(
                location=[violation['lat'], violation['lon']],
                popup=folium.Popup(popup_html, max_width=260),
                icon=folium.Icon(color=color, icon='flash', prefix='fa'),
                tooltip=f"Speed: {speed:.0f} km/h (+{excess:.0f})"
            ).add_to(speed_layer)
            
            self.feature_count += 1
        
        speed_layer.add_to(self.map)
        print(f"      ✅ Added {min(len(violations_to_plot), 20)} speeding markers")
    
    def _severity_bg_color(self, severity):
        """Get background color for severity"""
        colors = {
            'severe': '#B91C1C',
            'major': '#DC2626',
            'moderate': '#F59E0B',
            'minor': '#10B981'
        }
        return colors.get(severity, '#6B7280')
    
    # ========================================================================
    # ROUNDABOUT MARKERS
    # ========================================================================
    
    def _add_roundabout_markers(self):
        """Add roundabout behavior markers with comprehensive popups"""
        
        rb_layer = folium.FeatureGroup(name='Roundabouts', show=True)
        
        roundabout_data = self.week_analysis.get('roundabouts', {})
        approaches_detailed = roundabout_data.get('approaches_detailed', [])
        
        print(f"      🔄 Found {len(approaches_detailed)} roundabout approaches")
        
        # Track plotted roundabouts to avoid duplicates
        plotted_roundabouts = set()
        
        for approach in approaches_detailed:
            if not isinstance(approach, dict):
                continue
            
            rb_lat = approach.get('roundabout_lat')
            rb_lon = approach.get('roundabout_lon')
            rb_id = approach.get('roundabout_id', 'Unknown')
            
            if rb_lat is None or rb_lon is None:
                continue
            
            # Skip duplicates
            location_key = f"{rb_lat:.5f},{rb_lon:.5f}"
            if location_key in plotted_roundabouts:
                continue
            plotted_roundabouts.add(location_key)
            
            # Extract data
            anticipation_category = approach.get('anticipation_category', 'unknown')
            entry_compliant = approach.get('entry_compliant')
            entry_speed = approach.get('entry_speed')
            exit_speed = approach.get('exit_speed')
            confidence = approach.get('confidence', 0)
            
            zone_150_speed = approach.get('zone_150m_avg_speed')
            zone_100_speed = approach.get('zone_100m_avg_speed')
            zone_65_speed = approach.get('zone_65m_avg_speed')
            
            total_speed_drop = approach.get('total_speed_drop', 0)
            inside_points = approach.get('inside_points_count', 0)
            time_inside = approach.get('time_inside_seconds')
            
            # Determine marker color
            if entry_compliant is None:
                marker_color = 'gray'
                status = 'Unknown'
            elif entry_compliant and anticipation_category in ['good', 'moderate']:
                marker_color = 'green'
                status = 'Good'
            elif entry_compliant or anticipation_category in ['good', 'moderate']:
                marker_color = 'orange'
                status = 'Partial'
            else:
                marker_color = 'red'
                status = 'Poor'
            
            # Build popup
            popup_html = self._build_roundabout_popup(
                rb_id, rb_lat, rb_lon, status, marker_color,
                anticipation_category, entry_compliant, entry_speed, exit_speed,
                zone_150_speed, zone_100_speed, zone_65_speed,
                total_speed_drop, inside_points, time_inside, confidence
            )
            
            folium.Marker(
                location=[rb_lat, rb_lon],
                popup=folium.Popup(popup_html, max_width=320),
                icon=folium.Icon(color=marker_color, icon='refresh', prefix='fa'),
                tooltip=f"Roundabout: {status} | Entry: {entry_speed:.0f} km/h" if entry_speed else f"Roundabout: {status}"
            ).add_to(rb_layer)
            
            self.feature_count += 1
        
        rb_layer.add_to(self.map)
        print(f"      ✅ Added {len(plotted_roundabouts)} roundabout markers")
    
    def _build_roundabout_popup(self, rb_id, lat, lon, status, color,
                                 anticipation, entry_compliant, entry_speed, exit_speed,
                                 zone_150, zone_100, zone_65,
                                 speed_drop, inside_points, time_inside, confidence):
        """Build detailed roundabout popup HTML"""
        
        def format_speed(speed):
            return f"{speed:.0f} km/h" if speed is not None else "N/A"
        
        def format_compliant(value):
            if value is None:
                return "❓ Unknown"
            return "✅ Yes" if value else "❌ No"
        
        def format_anticipation(cat):
            icons = {
                'good': '🟢 Good (150-100m)',
                'moderate': '🟡 Moderate (100-65m)',
                'late': '🟠 Late (<65m)',
                'no_deceleration': '🔴 No Deceleration'
            }
            return icons.get(cat, f'⚪ {cat}')
        
        header_color = {
            'green': '#16a34a',
            'orange': '#ea580c',
            'red': '#dc2626',
            'gray': '#6b7280'
        }.get(color, '#6b7280')
        
        return f"""
        <div style="font-family: Arial; font-size: 11px; min-width: 280px;">
            <h4 style="margin: 0 0 8px 0; padding: 5px; background: {header_color}; 
                       color: white; border-radius: 4px;">
                🔄 Roundabout: {rb_id}
            </h4>
            
            <table style="width: 100%; border-collapse: collapse;">
                <tr style="background: #f5f5f5;">
                    <td colspan="2" style="padding: 4px; font-weight: bold;">📍 Location</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Coordinates:</td>
                    <td style="padding: 2px 4px;">{lat:.6f}, {lon:.6f}</td>
                </tr>
                
                <tr style="background: #e3f2fd;">
                    <td colspan="2" style="padding: 4px; font-weight: bold;">🎯 Anticipation</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Category:</td>
                    <td style="padding: 2px 4px;">{format_anticipation(anticipation)}</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Speed Drop:</td>
                    <td style="padding: 2px 4px;">{speed_drop:.0f} km/h</td>
                </tr>
                
                <tr style="background: #fff3e0;">
                    <td colspan="2" style="padding: 4px; font-weight: bold;">📊 Zone Speeds</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">150-100m:</td>
                    <td style="padding: 2px 4px;">{format_speed(zone_150)}</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">100-65m:</td>
                    <td style="padding: 2px 4px;">{format_speed(zone_100)}</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">65-0m:</td>
                    <td style="padding: 2px 4px;">{format_speed(zone_65)}</td>
                </tr>
                
                <tr style="background: #e8f5e9;">
                    <td colspan="2" style="padding: 4px; font-weight: bold;">🚗 Entry / Exit</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Entry Speed:</td>
                    <td style="padding: 2px 4px;">{format_speed(entry_speed)}</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Entry Compliant (≤34):</td>
                    <td style="padding: 2px 4px;">{format_compliant(entry_compliant)}</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Exit Speed:</td>
                    <td style="padding: 2px 4px;">{format_speed(exit_speed)}</td>
                </tr>
                
                <tr style="background: #fce4ec;">
                    <td colspan="2" style="padding: 4px; font-weight: bold;">⭕ Inside</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Points Inside:</td>
                    <td style="padding: 2px 4px;">{inside_points}</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Time Inside:</td>
                    <td style="padding: 2px 4px;">{f'{time_inside:.1f}s' if time_inside else 'N/A'}</td>
                </tr>
                
                <tr style="background: #ede7f6;">
                    <td colspan="2" style="padding: 4px; font-weight: bold;">📈 Confidence: {confidence:.0f}%</td>
                </tr>
            </table>
        </div>
        """
    
    # ========================================================================
    # STOP SIGN MARKERS
    # ========================================================================
    
    def _add_stop_sign_markers(self):
        """Add stop sign compliance markers"""
        
        stop_layer = folium.FeatureGroup(name='Stop Signs', show=True)
        
        stop_data = self.week_analysis.get('stop_signs', {})
        approaches = stop_data.get('approaches_detailed', [])
        
        print(f"      🛑 Found {len(approaches)} stop sign approaches")
        
        for i, stop in enumerate(approaches):
            if not isinstance(stop, dict):
                continue
            
            stop_lat = stop.get('stop_sign_lat')
            stop_lon = stop.get('stop_sign_lon')
            stop_id = stop.get('stop_sign_id', f'Stop_{i+1}')
            
            if stop_lat is None or stop_lon is None:
                continue
            
            compliance = stop.get('compliance', 'UNCERTAIN')
            confidence = stop.get('confidence', 0)
            min_speed = stop.get('min_speed')
            closest_speed = stop.get('closest_speed')
            evidence = stop.get('evidence', [])
            gps_points = stop.get('gps_points', 0)
            
            # Determine color
            if compliance == 'STOP_OK':
                marker_color = 'green'
                status_icon = '✅'
            elif compliance == 'STOP_KO':
                marker_color = 'red'
                status_icon = '❌'
            else:
                marker_color = 'gray'
                status_icon = '❓'
            
            popup_html = self._build_stop_sign_popup(
                stop_id, stop_lat, stop_lon, compliance, status_icon,
                confidence, min_speed, closest_speed, evidence, gps_points
            )
            
            folium.Marker(
                location=[stop_lat, stop_lon],
                popup=folium.Popup(popup_html, max_width=280),
                icon=folium.Icon(color=marker_color, icon='hand-stop-o', prefix='fa'),
                tooltip=f"Stop: {compliance} | Min: {min_speed:.0f} km/h" if min_speed else f"Stop: {compliance}"
            ).add_to(stop_layer)
            
            self.feature_count += 1
        
        stop_layer.add_to(self.map)
        print(f"      ✅ Added {len(approaches)} stop sign markers")
    
    def _build_stop_sign_popup(self, stop_id, lat, lon, compliance, status_icon,
                                confidence, min_speed, closest_speed, evidence, gps_points):
        """Build stop sign popup HTML"""
        
        header_color = '#16a34a' if compliance == 'STOP_OK' else '#dc2626' if compliance == 'STOP_KO' else '#6b7280'
        evidence_html = '<br>'.join([f"• {e}" for e in evidence]) if evidence else 'N/A'
        
        return f"""
        <div style="font-family: Arial; font-size: 11px; min-width: 240px;">
            <h4 style="margin: 0 0 8px 0; padding: 5px; background: {header_color}; 
                       color: white; border-radius: 4px;">
                🛑 Stop Sign: {stop_id}
            </h4>
            
            <table style="width: 100%; border-collapse: collapse;">
                <tr style="background: #e8f5e9;">
                    <td colspan="2" style="padding: 4px; font-weight: bold;">✅ Compliance</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Result:</td>
                    <td style="padding: 2px 4px;"><b>{status_icon} {compliance}</b></td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Confidence:</td>
                    <td style="padding: 2px 4px;">{confidence:.0f}%</td>
                </tr>
                
                <tr style="background: #fff3e0;">
                    <td colspan="2" style="padding: 4px; font-weight: bold;">🚗 Speed Data</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Min Speed:</td>
                    <td style="padding: 2px 4px;">{f'{min_speed:.0f} km/h' if min_speed is not None else 'N/A'}</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Closest Speed:</td>
                    <td style="padding: 2px 4px;">{f'{closest_speed:.0f} km/h' if closest_speed is not None else 'N/A'}</td>
                </tr>
                
                <tr style="background: #e3f2fd;">
                    <td colspan="2" style="padding: 4px; font-weight: bold;">📊 Evidence</td>
                </tr>
                <tr>
                    <td colspan="2" style="padding: 2px 4px; font-size: 10px;">{evidence_html}</td>
                </tr>
                
                <tr style="background: #f5f5f5;">
                    <td style="padding: 2px 4px;">GPS Points:</td>
                    <td style="padding: 2px 4px;">{gps_points}</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Location:</td>
                    <td style="padding: 2px 4px;">{lat:.6f}, {lon:.6f}</td>
                </tr>
            </table>
        </div>
        """
    
    # ========================================================================
    # TRAFFIC LIGHT MARKERS (NEW)
    # ========================================================================
    
    def _add_traffic_light_markers(self):
        """Add traffic light behavior markers"""
        
        tl_layer = folium.FeatureGroup(name='Traffic Lights', show=True)
        
        tl_data = self.week_analysis.get('traffic_lights', {})
        approaches = tl_data.get('approaches_detailed', [])
        
        print(f"      🚦 Found {len(approaches)} traffic light approaches")
        
        for i, tl in enumerate(approaches):
            if not isinstance(tl, dict):
                continue
            
            tl_lat = tl.get('signal_lat')
            tl_lon = tl.get('signal_lon')
            tl_id = tl.get('signal_id', f'TL_{i+1}')
            
            if tl_lat is None or tl_lon is None:
                continue
            
            behavior = tl.get('behavior', 'uncertain')
            confidence = tl.get('confidence', 0)
            min_speed = tl.get('min_speed')
            dwell_time = tl.get('dwell_time_seconds')
            evidence = tl.get('evidence', [])
            gps_points = tl.get('gps_points', 0)
            
            # Determine color
            if behavior == 'stopped':
                marker_color = 'green'
                status_icon = '✅'
                status_text = 'Stopped'
            elif behavior == 'slowed':
                marker_color = 'orange'
                status_icon = '🟡'
                status_text = 'Slowed'
            elif behavior == 'passed_through':
                marker_color = 'lightgreen'
                status_icon = '🟢'
                status_text = 'Passed (Green)'
            else:
                marker_color = 'gray'
                status_icon = '❓'
                status_text = 'Uncertain'
            
            popup_html = self._build_traffic_light_popup(
                tl_id, tl_lat, tl_lon, behavior, status_icon, status_text,
                confidence, min_speed, dwell_time, evidence, gps_points
            )
            
            folium.Marker(
                location=[tl_lat, tl_lon],
                popup=folium.Popup(popup_html, max_width=280),
                icon=folium.Icon(color=marker_color, icon='traffic-light', prefix='fa'),
                tooltip=f"Traffic Light: {status_text} | {confidence:.0f}% conf"
            ).add_to(tl_layer)
            
            self.feature_count += 1
        
        tl_layer.add_to(self.map)
        print(f"      ✅ Added {len(approaches)} traffic light markers")
    
    def _build_traffic_light_popup(self, tl_id, lat, lon, behavior, status_icon, status_text,
                                    confidence, min_speed, dwell_time, evidence, gps_points):
        """Build traffic light popup HTML"""
        
        header_colors = {
            'stopped': '#16a34a',
            'slowed': '#ea580c',
            'passed_through': '#22c55e',
            'uncertain': '#6b7280'
        }
        header_color = header_colors.get(behavior, '#6b7280')
        evidence_html = '<br>'.join([f"• {e}" for e in evidence]) if evidence else 'N/A'
        
        return f"""
        <div style="font-family: Arial; font-size: 11px; min-width: 240px;">
            <h4 style="margin: 0 0 8px 0; padding: 5px; background: {header_color}; 
                       color: white; border-radius: 4px;">
                🚦 Traffic Light: {tl_id}
            </h4>
            
            <table style="width: 100%; border-collapse: collapse;">
                <tr style="background: #e8f5e9;">
                    <td colspan="2" style="padding: 4px; font-weight: bold;">🚗 Behavior</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Result:</td>
                    <td style="padding: 2px 4px;"><b>{status_icon} {status_text}</b></td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Confidence:</td>
                    <td style="padding: 2px 4px;">{confidence:.0f}%</td>
                </tr>
                
                <tr style="background: #fff3e0;">
                    <td colspan="2" style="padding: 4px; font-weight: bold;">📊 Data</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Min Speed:</td>
                    <td style="padding: 2px 4px;">{f'{min_speed:.0f} km/h' if min_speed is not None else 'N/A'}</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Dwell Time:</td>
                    <td style="padding: 2px 4px;">{f'{dwell_time:.1f}s' if dwell_time is not None else 'N/A'}</td>
                </tr>
                
                <tr style="background: #e3f2fd;">
                    <td colspan="2" style="padding: 4px; font-weight: bold;">📋 Evidence</td>
                </tr>
                <tr>
                    <td colspan="2" style="padding: 2px 4px; font-size: 10px;">{evidence_html}</td>
                </tr>
                
                <tr style="background: #f5f5f5;">
                    <td style="padding: 2px 4px;">GPS Points:</td>
                    <td style="padding: 2px 4px;">{gps_points}</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Location:</td>
                    <td style="padding: 2px 4px;">{lat:.6f}, {lon:.6f}</td>
                </tr>
            </table>
        </div>
        """
    
    # ========================================================================
    # HARSH EVENT MARKERS (NEW)
    # ========================================================================
    
    def _add_harsh_event_markers(self):
        """Add harsh driving event markers"""
        
        harsh_layer = folium.FeatureGroup(name='Harsh Events', show=True)
        
        harsh_data = self.week_analysis.get('harsh_events', {})
        
        # Collect all events
        all_events = []
        
        braking_events = harsh_data.get('harsh_braking', {}).get('events', [])
        for e in braking_events:
            e['event_type'] = 'braking'
            all_events.append(e)
        
        accel_events = harsh_data.get('harsh_acceleration', {}).get('events', [])
        for e in accel_events:
            e['event_type'] = 'acceleration'
            all_events.append(e)
        
        turn_events = harsh_data.get('sharp_turns', {}).get('events', [])
        for e in turn_events:
            e['event_type'] = 'turn'
            all_events.append(e)
        
        print(f"      ⚠️ Found {len(all_events)} harsh events")
        
        for i, event in enumerate(all_events[:30]):  # Limit to 30
            if 'lat' not in event or 'lon' not in event:
                continue
            
            event_type = event.get('event_type', 'unknown')
            speed = event.get('speed_kmh', 0)
            source = event.get('source', 'unknown')
            
            # Get event-specific data
            decel_g = event.get('deceleration_g')
            accel_g = event.get('acceleration_g')
            lateral_g = event.get('lateral_accel_g')
            heading_change = event.get('heading_change_deg')
            
            # Determine icon and color
            if event_type == 'braking':
                icon = 'arrow-down'
                color = 'red'
                event_name = 'Harsh Braking'
                g_force = decel_g
            elif event_type == 'acceleration':
                icon = 'arrow-up'
                color = 'orange'
                event_name = 'Harsh Acceleration'
                g_force = accel_g
            else:
                icon = 'rotate-right'
                color = 'blue'
                event_name = 'Sharp Turn'
                g_force = lateral_g
            
            popup_html = self._build_harsh_event_popup(
                i + 1, event_name, event['lat'], event['lon'],
                speed, g_force, heading_change, source, event.get('timestamp')
            )
            
            folium.Marker(
                location=[event['lat'], event['lon']],
                popup=folium.Popup(popup_html, max_width=260),
                icon=folium.Icon(color=color, icon=icon, prefix='fa'),
                tooltip=f"{event_name}: {speed:.0f} km/h"
            ).add_to(harsh_layer)
            
            self.feature_count += 1
        
        harsh_layer.add_to(self.map)
        print(f"      ✅ Added {min(len(all_events), 30)} harsh event markers")
    
    def _build_harsh_event_popup(self, num, event_name, lat, lon, speed, g_force, heading_change, source, timestamp):
        """Build harsh event popup HTML"""
        
        header_colors = {
            'Harsh Braking': '#dc2626',
            'Harsh Acceleration': '#ea580c',
            'Sharp Turn': '#2563eb'
        }
        header_color = header_colors.get(event_name, '#6b7280')
        
        g_force_str = f"{g_force:.2f}g" if g_force is not None else "N/A"
        heading_str = f"{heading_change:.0f}°" if heading_change is not None else "N/A"
        
        return f"""
        <div style="font-family: Arial; font-size: 11px; min-width: 220px;">
            <h4 style="margin: 0 0 8px 0; padding: 5px; background: {header_color}; 
                       color: white; border-radius: 4px;">
                ⚠️ {event_name} #{num}
            </h4>
            
            <table style="width: 100%; border-collapse: collapse;">
                <tr style="background: #fef2f2;">
                    <td colspan="2" style="padding: 4px; font-weight: bold;">📊 Event Data</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Speed:</td>
                    <td style="padding: 2px 4px;"><b>{speed:.1f} km/h</b></td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">G-Force:</td>
                    <td style="padding: 2px 4px;"><b>{g_force_str}</b></td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Heading Change:</td>
                    <td style="padding: 2px 4px;">{heading_str}</td>
                </tr>
                
                <tr style="background: #f5f5f5;">
                    <td colspan="2" style="padding: 4px; font-weight: bold;">📍 Details</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Source:</td>
                    <td style="padding: 2px 4px;">{source}</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Time:</td>
                    <td style="padding: 2px 4px;">{timestamp if timestamp else 'N/A'}</td>
                </tr>
                <tr>
                    <td style="padding: 2px 4px;">Location:</td>
                    <td style="padding: 2px 4px;">{lat:.6f}, {lon:.6f}</td>
                </tr>
            </table>
        </div>
        """
    
    # ========================================================================
    # SCHOOL ZONE MARKERS
    # ========================================================================
    
    def _add_school_zone_markers(self):
        """Add school zone markers"""
        
        school_layer = folium.FeatureGroup(name='School Zones', show=True)
        
        school_data = self.week_analysis.get('school_zones', {})
        passages = school_data.get('violations_detail', [])
        
        print(f"      🏫 Found {len(passages)} school zone passages")
        
        for i, passage in enumerate(passages):
            if not isinstance(passage, dict):
                continue
            
            if 'school_lat' not in passage or 'school_lon' not in passage:
                continue
            
            exceeded = passage.get('exceeded_limit', False)
            color = 'red' if exceeded else 'green'
            
            # Add school zone circle
            folium.Circle(
                location=[passage['school_lat'], passage['school_lon']],
                radius=passage.get('zone_radius', 150),
                color='purple',
                fillColor='purple',
                fillOpacity=0.1
            ).add_to(school_layer)
            
            # Add marker
            popup_html = f"""
            <div style="font-family: Arial; font-size: 11px;">
                <b>🏫 {passage.get('school_name', 'School Zone')}</b><br>
                <hr>
                Speed: {passage.get('speed_kmh', 0):.1f} km/h<br>
                Limit: {passage.get('school_zone_limit', 30)} km/h<br>
                Exceeded: {'❌ Yes' if exceeded else '✅ No'}
            </div>
            """
            
            folium.Marker(
                location=[passage['school_lat'], passage['school_lon']],
                popup=folium.Popup(popup_html, max_width=200),
                icon=folium.Icon(color=color, icon='graduation-cap', prefix='fa'),
                tooltip=f"School: {passage.get('school_name', 'School')}"
            ).add_to(school_layer)
            
            self.feature_count += 1
        
        school_layer.add_to(self.map)
    
    # ========================================================================
    # SPEED HEATMAP
    # ========================================================================
    
    def _add_speed_heatmap(self):
        """Add speed violation heatmap"""
        
        speeding_data = self.week_analysis.get('speeding', {})
        violations = speeding_data.get('worst_violations', [])
        
        if not violations:
            return
        
        heat_coords = []
        for v in violations:
            if 'lat' in v and 'lon' in v:
                severity = v.get('severity', 'minor')
                weight = {'minor': 1, 'moderate': 2, 'major': 3, 'severe': 4}.get(severity, 1)
                for _ in range(weight):
                    heat_coords.append([v['lat'], v['lon']])
        
        if heat_coords:
            heat_map = plugins.HeatMap(
                heat_coords,
                name='Speed Violation Density',
                radius=15,
                blur=10,
                max_zoom=18,
                show=False
            )
            heat_map.add_to(self.map)
    
    # ========================================================================
    # LEGEND
    # ========================================================================
    
    def _add_legend(self):
        """Add comprehensive map legend"""
        
        legend_html = """
        <div style="position: fixed; bottom: 10px; right: 10px; z-index: 1000; 
                    background: rgba(255,255,255,0.95); padding: 12px; border: 2px solid #1E3A8A;
                    border-radius: 8px; box-shadow: 0 4px 8px rgba(0,0,0,0.3); font-family: Arial;
                    max-width: 220px; font-size: 11px;">
            
            <h4 style="margin: 0 0 8px 0; color: #1E3A8A; font-size: 12px;">Map Legend</h4>
            
            <div style="margin-bottom: 8px;">
                <b style="color: #555;">GPS Data</b>
                <div style="display: flex; align-items: center; margin: 2px 0;">
                    <div style="width: 10px; height: 10px; background: red; border-radius: 50%; margin-right: 6px;"></div>
                    <span>GPS points</span>
                </div>
                <div style="display: flex; align-items: center; margin: 2px 0;">
                    <div style="width: 10px; height: 10px; background: blue; border-radius: 50%; margin-right: 6px;"></div>
                    <span>Road-matched</span>
                </div>
                <div style="display: flex; align-items: center; margin: 2px 0;">
                    <div style="width: 16px; height: 2px; background: orange; margin-right: 6px; border-style: dashed;"></div>
                    <span>GPS path</span>
                </div>
                <div style="display: flex; align-items: center; margin: 2px 0;">
                    <div style="width: 16px; height: 2px; background: green; margin-right: 6px;"></div>
                    <span>Route geometry</span>
                </div>
            </div>
            
            <div style="margin-bottom: 8px;">
                <b style="color: #555;">Features</b>
                <div style="margin: 2px 0;">🚨 Speed violations</div>
                <div style="margin: 2px 0;">🔄 Roundabouts</div>
                <div style="margin: 2px 0;">🛑 Stop signs</div>
                <div style="margin: 2px 0;">🚦 Traffic lights</div>
                <div style="margin: 2px 0;">⚠️ Harsh events</div>
                <div style="margin: 2px 0;">🏫 School zones</div>
            </div>
            
            <div>
                <b style="color: #555;">Colors</b>
                <div style="display: flex; align-items: center; margin: 2px 0;">
                    <div style="width: 10px; height: 10px; background: #16a34a; margin-right: 6px;"></div>
                    <span>Compliant/Good</span>
                </div>
                <div style="display: flex; align-items: center; margin: 2px 0;">
                    <div style="width: 10px; height: 10px; background: #ea580c; margin-right: 6px;"></div>
                    <span>Partial/Warning</span>
                </div>
                <div style="display: flex; align-items: center; margin: 2px 0;">
                    <div style="width: 10px; height: 10px; background: #dc2626; margin-right: 6px;"></div>
                    <span>Violation/Poor</span>
                </div>
            </div>
        </div>
        """
        
        self.map.get_root().html.add_child(folium.Element(legend_html))
    
    # ========================================================================
    # SAVE MAP
    # ========================================================================
    
    def _save_map(self, output_filename):
        """Save map to file"""
        
        output_dir = os.path.join(os.getcwd(), "output")
        os.makedirs(output_dir, exist_ok=True)
        
        output_path = os.path.join(output_dir, output_filename)
        self.map.save(output_path)
        
        print(f"      💾 Saved: {output_path}")


# =============================================================
# MAIN GENERATION FUNCTION
# =============================================================

def generate_plots_main(extracted_data, output_prefix="analysis"):
    """
    Generate HTML maps for all weeks using extracted data.
    
    Args:
        extracted_data: Complete data from UnifiedReportGenerator.extract_all_data()
        output_prefix: Prefix for output filenames
    
    Returns:
        List of generated filenames
    """
    
    print("🗺️ Generating weekly maps...")
    
    weekly_results = extracted_data.get('weekly_data', {})
    weekly_gps_data = extracted_data.get('gps_data', {})
    
    if not weekly_results:
        print("❌ No weekly results in extracted data")
        return []
    
    if not weekly_gps_data:
        print("❌ No GPS data in extracted data")
        return []
    
    print(f"📊 Found data for {len(weekly_results)} weeks")
    
    generated_files = []
    
    for week_label, week_gps_df in weekly_gps_data.items():
        try:
            if week_gps_df is None or len(week_gps_df) == 0:
                print(f"⚠️ No GPS data for {week_label}")
                continue
            
            safe_week = week_label.replace(' ', '_')
            filename = f"{output_prefix}_{safe_week}_map.html"
            
            # Create map generator
            map_gen = DrivingBehaviorMapGenerator(
                extracted_data=extracted_data,
                week_label=week_label,
                week_gps_df=week_gps_df
            )
            
            # Generate map
            map_gen.generate_map(filename)
            generated_files.append(filename)
            
        except Exception as e:
            print(f"❌ Failed to generate map for {week_label}: {e}")
            import traceback
            traceback.print_exc()
    
    print(f"🎉 Generated {len(generated_files)} maps")
    return generated_files


# =============================================================
# USAGE
# =============================================================
#
# # After running main pipeline
# extracted_data = analyzer.reporter.extract_all_data()
# generated_files = generate_plots_main(extracted_data, output_prefix="driver_analysis")
#

#==================================================================
class MapDataVerificationExcelGenerator:
    def __init__(self):
        pass
    
    def generate_verification_excel(self, all_data, output_file="map_verification.xlsx"):
        """
        Generate verification Excel workbook from all_data dictionary
        Creates one sheet per week with infrastructure detection data
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        weekly_data = all_data.get('weekly_data', {})
        
        # Create a sheet for each week
        for week_label, week_data in weekly_data.items():
            ws = wb.create_sheet(title=week_label)
            self._populate_week_sheet(ws, week_data, all_data)
        
        # If no weekly data, create at least one sheet
        if not weekly_data:
            ws = wb.create_sheet(title="Week1")
            self._populate_empty_sheet(ws)
        
        wb.save(output_file)
        print(f"Verification Excel generated: {output_file}")
        return output_file

    def _populate_week_sheet(self, ws, week_data, all_data):
        """Populate a single week sheet with all infrastructure data"""
        current_row = 1
        
        # Title
        ws['A1'] = "Extract Relevant Meta Data"
        current_row = 3

        # Speed Violations section
        current_row = self._add_speed_violations_section(ws, week_data, current_row)
        current_row += 2
        

        
        # Roundabouts section  
        current_row = self._add_roundabouts_section(ws, week_data, all_data, current_row)
        current_row += 2

        # Schools section
        current_row = self._add_schools_section(ws, week_data, all_data, current_row)
        current_row += 2

        
        # Stop Signs section
        current_row = self._add_stop_signs_section(ws, week_data, all_data, current_row)
    
    def _safe_extract_field(self, data_dict, primary_key, fallback_key=None, default=''):
        """Safely extract field with fallback options"""
        if primary_key in data_dict:
            return data_dict[primary_key]
        elif fallback_key and fallback_key in data_dict:
            return data_dict[fallback_key]
        return default
    
    def _add_schools_section(self, ws, week_data, all_data, start_row):
        """Add schools data section - CORRECTED"""
        # Headers
        ws[f'A{start_row}'] = "Index"
        ws[f'B{start_row}'] = "School Name"
        ws[f'C{start_row}'] = "Latitude"
        ws[f'D{start_row}'] = "Longitude"
        ws[f'E{start_row}'] = "Driver Speed"
        ws[f'F{start_row}'] = "Speed Limit"
        
        current_row = start_row + 1
        
        # Get school zone passages from week data
        school_data = week_data.get('school_zones', {})
        passages_detail = school_data.get('passages_detail', [])
        
        # Also get school infrastructure from map data
        map_schools = []
        if 'map_data' in all_data and all_data['map_data'].get('available', False):
            map_schools = all_data['map_data'].get('data', {}).get('schools', [])
        
        # Combine data from passages and map schools
        school_entries = []
        
        # Add entries from actual passages (these have speed data)
        for i, passage in enumerate(passages_detail):
            school_entries.append({
                'index': i + 1,
                'name': passage.get('school_name', 'Unknown School'),  # CORRECTED: was using _safe_extract_field
                'lat': passage.get('school_lat', ''),  # CORRECTED: was using _safe_extract_field
                'lon': passage.get('school_lon', ''),  # CORRECTED: was using _safe_extract_field
                'driver_speed': passage.get('speed_kmh', ''),  # CORRECTED: this one was already correct
                'speed_limit': passage.get('school_zone_limit', passage.get('speed_limit', ''))  # CORRECTED: prefer school_zone_limit
            })
            
        # Add map schools that weren't in passages (no speed data)
        existing_coords = {(entry['lat'], entry['lon']) for entry in school_entries}
        for school in map_schools:
            school_coord = (school.get('lat', ''), school.get('lon', ''))
            if school_coord not in existing_coords:
                school_entries.append({
                    'index': len(school_entries) + 1,
                    'name': school.get('name', 'Unnamed School'),
                    'lat': school.get('lat', ''),
                    'lon': school.get('lon', ''),
                    'driver_speed': '',
                    'speed_limit': ''
                })
        
        # Populate data
        for entry in school_entries:
            ws[f'A{current_row}'] = entry['index']
            ws[f'B{current_row}'] = entry['name']
            ws[f'C{current_row}'] = entry['lat']
            ws[f'D{current_row}'] = entry['lon']
            ws[f'E{current_row}'] = entry['driver_speed']
            ws[f'F{current_row}'] = entry['speed_limit']
            current_row += 1
        
        return current_row
    
    def _add_roundabouts_section(self, ws, week_data, all_data, start_row):
        """Add roundabouts data section - CORRECTED"""
        # Headers
        ws[f'A{start_row}'] = "Index"
        ws[f'B{start_row}'] = "Roundabouts"
        ws[f'C{start_row}'] = "Latitude"
        ws[f'D{start_row}'] = "Longitude"
        ws[f'E{start_row}'] = "Speed 150"
        ws[f'F{start_row}'] = "Speed 100"
        ws[f'G{start_row}'] = "Speed 50"
        
        current_row = start_row + 1
        
        roundabout_data = week_data.get('roundabouts', {})
        
        if roundabout_data.get('format') == 'multi_zone':
            # Multi-zone format - combine approaches from all zones
            zones_data = roundabout_data.get('data', {})
            all_approaches = {}
            
            # Collect all unique roundabouts
            for zone in ['150m', '100m', '50m']:
                zone_info = zones_data.get(zone, {})
                approaches = zone_info.get('approaches_detail', [])
                
                for approach in approaches:
                    # CORRECTED: Use roundabout_id as primary key
                    roundabout_key = approach.get('roundabout_id', 
                        f"{approach.get('roundabout_lat', '')}_{approach.get('roundabout_lon', '')}")
                    
                    if roundabout_key not in all_approaches:
                        all_approaches[roundabout_key] = {
                            'id': approach.get('roundabout_id', ''),  # CORRECTED: store ID instead of name
                            'lat': approach.get('roundabout_lat', ''),  # CORRECTED: use roundabout_lat
                            'lon': approach.get('roundabout_lon', ''),  # CORRECTED: use roundabout_lon
                            'speed_150': '',
                            'speed_100': '',
                            'speed_50': ''
                        }
                    
                    # CORRECTED: Add speed for this zone using speed_at_distance
                    speed_key = f'speed_{zone.replace("m", "")}'
                    all_approaches[roundabout_key][speed_key] = approach.get('speed_at_distance', '')
            
            # Populate rows
            for i, (_, data) in enumerate(all_approaches.items(), 1):
                ws[f'A{current_row}'] = i
                ws[f'B{current_row}'] = data['id']
                ws[f'C{current_row}'] = data['lat']
                ws[f'D{current_row}'] = data['lon']
                ws[f'E{current_row}'] = data['speed_150']
                ws[f'F{current_row}'] = data['speed_100']
                ws[f'G{current_row}'] = data['speed_50']
                current_row += 1
        
        elif roundabout_data.get('format') == 'single_zone':
            # Single zone format
            data = roundabout_data.get('data', {})
            approaches = data.get('approaches_detail', [])
            
            for i, approach in enumerate(approaches, 1):
                ws[f'A{current_row}'] = i
                ws[f'B{current_row}'] = approach.get('roundabout_id', '')  # CORRECTED: use roundabout_id directly
                ws[f'C{current_row}'] = approach.get('roundabout_lat', '')  # CORRECTED: use roundabout_lat
                ws[f'D{current_row}'] = approach.get('roundabout_lon', '')  # CORRECTED: use roundabout_lon
                ws[f'E{current_row}'] = ''  # No 150m data in single zone
                ws[f'F{current_row}'] = ''  # No 100m data in single zone
                ws[f'G{current_row}'] = approach.get('speed_at_distance', '')  # CORRECTED: use speed_at_distance
                current_row += 1
        
        return current_row
    
    def _add_speed_violations_section(self, ws, week_data, start_row):
        """Add ALL speed episodes data section - includes both violation and compliant segments"""
        
        # SECTION 1: TOP 3 WORST VIOLATIONS
        ws[f'A{start_row}'] = "TOP 3 WORST VIOLATIONS"
        ws[f'A{start_row}'].font = Font(bold=True, size=14)
        start_row += 2
        
        # Top violations headers
        top_violation_headers = [
            "Rank", "Timestamp", "Speed (km/h)", "Speed Limit", "Excess Speed", 
            "Severity", "Latitude", "Longitude", "Episode ID", "Episode Duration (sec)", 
            "Episode GPS Points"
        ]
        
        for i, header in enumerate(top_violation_headers):
            ws[f'{chr(65+i)}{start_row}'] = header
            ws[f'{chr(65+i)}{start_row}'].font = Font(bold=True)
        
        start_row += 1
        
        # Extract top violations data
        speeding_data = week_data.get('speeding', {})
        worst_violations = speeding_data.get('worst_violations', [])
        all_episodes = speeding_data.get('all_episodes', [])
        
        # Create a lookup for episode details
        episode_lookup = {ep.get('episode_id'): ep for ep in all_episodes}
        
        for rank, violation in enumerate(worst_violations[:3], 1):
            ws[f'A{start_row}'] = rank
            ws[f'B{start_row}'] = violation.get('timestamp', '')
            ws[f'C{start_row}'] = violation.get('speed_kmh', '')
            ws[f'D{start_row}'] = violation.get('speed_limit', '')
            ws[f'E{start_row}'] = violation.get('excess_speed', '')
            ws[f'F{start_row}'] = violation.get('severity', '')
            ws[f'G{start_row}'] = violation.get('lat', '')
            ws[f'H{start_row}'] = violation.get('lon', '')
            
            # Try to find matching episode details
            episode_id = violation.get('episode_id', '')
            episode_details = episode_lookup.get(episode_id, {})
            
            ws[f'I{start_row}'] = episode_id
            ws[f'J{start_row}'] = episode_details.get('duration_seconds', violation.get('episode_duration_seconds', ''))
            ws[f'K{start_row}'] = episode_details.get('total_gps_points', violation.get('episode_gps_points', ''))
            
            start_row += 1
        
        # Add spacing
        start_row += 2
        
        # SECTION 2: ALL SEGMENTS DATA
        ws[f'A{start_row}'] = "ALL SPEED EPISODES DATA"
        ws[f'A{start_row}'].font = Font(bold=True, size=14)
        start_row += 2
        
        # Updated headers to include more available info
        ws[f'A{start_row}'] = "Episode ID"
        ws[f'B{start_row}'] = "Start Time"
        ws[f'C{start_row}'] = "End Time"
        ws[f'D{start_row}'] = "Duration (sec)"
        ws[f'E{start_row}'] = "Speed Limit"
        ws[f'F{start_row}'] = "Max Speed"
        ws[f'G{start_row}'] = "Max Excess"
        ws[f'H{start_row}'] = "Total Points"
        ws[f'I{start_row}'] = "Is Violation"
        ws[f'J{start_row}'] = "Violation Episodes Count"
        ws[f'K{start_row}'] = "Start Lat"
        ws[f'L{start_row}'] = "Start Lon"
        ws[f'M{start_row}'] = "End Lat"
        ws[f'N{start_row}'] = "End Lon"
        ws[f'O{start_row}'] = "Speed Zone Category"
        
        current_row = start_row + 1
        
        # Extract ALL segments data (both violation and compliant)
        speeding_data = week_data.get('speeding', {})
        all_segments = speeding_data.get('all_segments', [])
        
        # Function to categorize speed limit into zones
        def get_speed_zone(speed_limit):
            """Map speed limit to fixed road speed limit categories"""
            if pd.isna(speed_limit) or speed_limit is None:
                return "INVALID"
            if speed_limit <= 30:
                return "30"
            elif speed_limit <= 50:
                return "50" 
            elif speed_limit <= 70:
                return "70"
            elif speed_limit <= 80:
                return "80"  # NEW ZONE
            elif speed_limit <= 90:
                return "90"
            elif speed_limit <= 110:
                return "110"
            elif speed_limit <= 130:
                return "130"
            else:
                return ">130"
        
        for segment in all_segments:
            ws[f'A{current_row}'] = segment.get('episode_id', '')
            ws[f'B{current_row}'] = segment.get('start_time', '')
            ws[f'C{current_row}'] = segment.get('end_time', '')
            ws[f'D{current_row}'] = segment.get('duration_seconds', '')
            ws[f'E{current_row}'] = segment.get('speed_limit', '')
            ws[f'F{current_row}'] = segment.get('max_speed', '')
            ws[f'G{current_row}'] = segment.get('max_excess', '')
            ws[f'H{current_row}'] = segment.get('total_gps_points', '')
            ws[f'I{current_row}'] = segment.get('is_violation', False)
            ws[f'J{current_row}'] = segment.get('violation_episodes_count', 0)
            
            # Extract start and end coordinates from segment_data
            segment_data = segment.get('segment_data', [])
            if segment_data:
                start_point = segment_data[0]
                end_point = segment_data[-1]
                ws[f'K{current_row}'] = start_point.get('lat', '')
                ws[f'L{current_row}'] = start_point.get('lon', '')
                ws[f'M{current_row}'] = end_point.get('lat', '')
                ws[f'N{current_row}'] = end_point.get('lon', '')
            else:
                ws[f'K{current_row}'] = ''
                ws[f'L{current_row}'] = ''
                ws[f'M{current_row}'] = ''
                ws[f'N{current_row}'] = ''
            
            # Add speed zone category
            speed_limit = segment.get('speed_limit', 0)
            ws[f'O{current_row}'] = get_speed_zone(speed_limit)
                
            current_row += 1
        
        return current_row

    
    
    def _add_stop_signs_section(self, ws, week_data, all_data, start_row):
        """Add stop signs data section - CORRECTED"""
        # Headers
        ws[f'A{start_row}'] = "Index"
        ws[f'B{start_row}'] = "Stop Signs"
        ws[f'C{start_row}'] = "Latitude" 
        ws[f'D{start_row}'] = "Longitude"
        ws[f'E{start_row}'] = "Driver Speed"
        ws[f'F{start_row}'] = "Allowed Speed"
        
        current_row = start_row + 1
        
        # Extract stop sign data directly
        stop_data = week_data.get('stop_signs', {})
        approaches_detail = stop_data.get('approaches_detail', [])
        
        for i, approach in enumerate(approaches_detail, 1):
            ws[f'A{current_row}'] = i
            ws[f'B{current_row}'] = approach.get('stop_sign_id', '')  # CORRECTED: use stop_sign_id directly
            ws[f'C{current_row}'] = approach.get('stop_sign_lat', '')  # CORRECTED: use stop_sign_lat
            ws[f'D{current_row}'] = approach.get('stop_sign_lon', '')  # CORRECTED: use stop_sign_lon
            ws[f'E{current_row}'] = approach.get('min_speed_at_stop', '')  # CORRECTED: use min_speed_at_stop instead of speed_kmh
            ws[f'F{current_row}'] = '0'  # CORRECTED: Stop signs should have 0 allowed speed
            current_row += 1
        
        # If no approaches, check if there are detected stop signs from map data
        if not approaches_detail:
            # This would require accessing the global map data through all_data parameter
            # Since this method doesn't have all_data, we'll leave empty for now
            pass
        
        return current_row
        # If no approaches, check if there are detected stop signs from map data
        if not approaches_detail:
            # This would require accessing the global map data through all_data parameter
            # Since this method doesn't have all_data, we'll leave empty for now
            pass
        
        return current_row  
        
    def _populate_empty_sheet(self, ws):
        """Populate empty sheet with headers only"""
        ws['A1'] = "Extract Relevant Meta Data"
        
        sections = [
            ("Schools", ["Index", "School Name", "Latitude", "Longitude", "Driver Speed", "Speed Limit"]),
            ("Roundabouts", ["Index", "Roundabouts", "Latitude", "Longitude", "Speed 150", "Speed 100", "Speed 50"]),
            ("Traffic Lights", ["Index", "Traffic Lights", "Latitude", "Longitude", "Driver Speed", "Allowed Speed"]),
            ("Road ID", ["Index", "Road ID", "Latitude", "Longitude", "Driver Speed", "Allowed Speed"]),
            ("Stop Signs", ["Index", "Stop Signs", "Latitude", "Longitude", "Driver Speed", "Allowed Speed"])
        ]
        
        current_row = 3
        for section_name, headers in sections:
            # Add section headers
            for i, header in enumerate(headers):
                ws[f'{chr(65+i)}{current_row}'] = header
            
            current_row += 4  # Space between sections


 # Integration function
    def generate_verification_excel_from_analyzer(self,analyzer, output_prefix="verification"):
        """
        Generate verification Excel from analyzer's data
        """
        if not hasattr(analyzer, 'reporter') or analyzer.reporter is None:
            print("Error: No reporter found in analyzer")
            return None
        
        # Extract all data
        all_data = analyzer.reporter.extract_all_data()
        
        # Create verification Excel generator
        verification_generator = MapDataVerificationExcelGenerator()
    
        base_path = f"{output_prefix}_verification_workbook"
        extension = ".xlsx"
        counter = 0
        excel_file = None
    
        while True:
            if counter == 0:
                output_file = base_path + extension
            else:
                output_file = f"{base_path} ({counter}){extension}"
            try:
                # Attempt to generate and save the Excel file
                excel_file = verification_generator.generate_verification_excel(all_data, output_file)
                print(f"✅ Successfully saved verification Excel to: {output_file}")
                break
            except PermissionError:
                print(f"⚠️ Permission denied for '{output_file}' (likely open in another program). Trying next suffix...")
                counter += 1
            except Exception as e:
                print(f"❌ Unexpected error while generating Excel: {str(e)}")
                return None
        
        return {
            'verification_file': excel_file,
            'sheets_created': len(all_data.get('weekly_data', {}))
        }


# ==========================================================# 
# ==========================================================
#   EXCEL FEEL GOOD Report Generator
# ==========================================================
""" 
French Translation

"""
FRENCH_STRINGS = {
    # Company and Headers
    'company_name': 'FeelGood Conduite',
    'report_title': 'Analyse du Comportement Routier',
    'driver_profile': 'Profil du Conducteur',

    # Driver Info Labels
    'driver_label': 'Conducteur :',
    'analysis_type_label': 'Type d’Analyse :',
    'report_date_label': 'Date du Rapport :',
    'report_type_label': 'Type de Rapport :',
    'report_type_value': 'Analyse Contextuelle du Comportement',

    # KPI Section
    'kpi_title': 'Indicateurs Clés',
    'overall_score': 'Score Global',
    'total_violations': 'Infractions Totales',
    'speed_compliance': 'Respect des Vitesses',
    'roundabout_performance': 'Comportement aux Rond-Points',
    'school_zone_safety': 'Sécurité en Zone Scolaire',
    'stop_signs_compliance': 'Respect des Stops',

    # RAG Status
    'excellent': 'Excellent',
    'good': 'Bon',
    'needs_work': 'À Améliorer',
    'critical': 'Critique',
    'no_data': 'Aucune Donnée',

    # Speed Analysis
    'top_violations_title': 'Top 3 Infractions de Vitesse (Toutes Semaines)',
    'speed_analysis_title': 'Analyse par Zone de Vitesse',
    'speed_violations_headers': ['Semaine', 'Vitesse (km/h)', 'Limite (km/h)', 'Excès (km/h)', 'Gravité'],
    'speed_zone_headers': ['Semaine', 'Zone', 'Nb. Infractions', 'Respect (%)', 'Limite (km/h)', 'Vitesse Max (km/h)'],
    'speed_chart_title': 'Respect Hebdomadaire par Zone de Vitesse (%)',

    # Severity levels
    'severe': 'Grave',
    'major': 'Majeur',
    'moderate': 'Modéré',
    'minor': 'Mineur',

    # Roundabout Section
    'roundabout_title': 'Approche des Rond-Points',
    'roundabout_headers': ['Semaine', 'Rond-Points', 'Approches Correctes (150m)', 
                           'Approches Correctes (100m)', 'Approches Correctes (50m)', 
                           'Respect du Bon Comportement (%)'],
    'roundabout_chart_title': 'Approche Hebdomadaire des Rond-Points',

    # School Zones
    'school_zone_title': 'Zones scolaires et zones sensibles (< 30 km/h)',
    'school_zone_headers': ['Semaine', 'Zones Sensibles', 'Respect < 30 km/h', 
                            'Bon Comportement (%)', 'Vitesse Max (km/h)', 'Limite (km/h)'],
    'school_chart_title': 'Comportement en Zone Scolaire - Tendance Hebdomadaire (%)',

    # Stop Signs
    'stop_sign_title': 'Respect des Panneaux Stop',
    'stop_sign_headers': ['Semaine', 'Stops', 'Vitesse à 20m', 
                          'Limite à 20m', 'Respect', 'Respect (%)'],
    'stop_chart_title': 'Respect des Stops par Semaine',

    # Improvement Section
    'improvement_title': 'Axes d’Amélioration',

    # Improvement Categories
    'speed_management': 'Gestion de la Vitesse',
    'roundabout_technique': 'Technique aux Rond-Points',
    'stop_sign_compliance': 'Respect des Stops',
    'school_zone_safety': 'Sécurité en Zone Scolaire',
    'overall_awareness': 'Sensibilisation Générale',
    'excellent_performance': 'Performance Exemplaire',

    # Recommendations
    'speed_recommendation': 'Maintenez une vitesse adaptée dans les zones de transition et à limitation variable.',
    'roundabout_recommendation': 'Réduisez la vitesse dès 150m avant un rond-point pour une approche fluide.',
    'stop_recommendation': 'Arrêtez-vous complètement aux stops. Comptez “mille-un” avant de repartir.',
    'school_recommendation': 'Ralentissez proactivement à l’approche des écoles, surtout aux heures d’entrée/sortie.',
    'awareness_recommendation': 'Renforcez votre attention aux limitations et panneaux. Envisagez une formation préventive.',
    'excellent_recommendation': 'Continuez à maintenir vos standards élevés dans toutes les catégories évaluées.',

    # Priority levels
    'high_priority': 'Priorité Élevée',
    'medium_priority': 'Priorité Moyenne',
    'low_priority': 'Priorité Faible',

    # Footer
    'footer_text': 'Rapport généré par la plateforme FeelGood Conduite',
    'generated_text': 'Rapport Généré :',

    # Chart axis labels
    'week_label': 'Semaine',
    'compliance_label': 'Respect (%)',
    'good_behavior_label': 'Bon Comportement (%)',
    'distance_label': 'Distance',

    # Common terms
    'week': 'Semaine',
    'na': 'N/D',
    'allowed_speed': '10 km/h',
    'stop_chart_y_axis': 'Conformité %',
    'good_behavior_chart_label': 'Bon Comportement %',
    'compliance_chart_label': 'Conformité %',
}




class FeelGoodDrivingReportGenerator:
    """
    Enterprise-grade driving behavior report generator.
    
    DESIGN PRINCIPLES:
    1. DISPLAY ONLY - All calculations come from UnifiedReportGenerator
    2. RAG-based color coding (Red/Amber/Green)
    3. French localization
    4. Professional Excel formatting
    5. Modular table/chart creation
    
    TABLES INCLUDED:
    - Trip Summary (NEW)
    - KPI Dashboard with Trends
    - Top Speed Violations
    - Speed Analysis by Zone
    - Roundabout Behavior
    - Stop Sign Compliance
    - Traffic Light Compliance (NEW)
    - Harsh Events Analysis (NEW)
    - Improvement Recommendations
    """
    
    # ========================================================================
    # FRENCH LOCALIZATION
    # ========================================================================
    
    FRENCH_STRINGS = {
        # Header
        'company_name': 'FeelGood Conduite',
        'report_title': 'Analyse du Comportement Routier',
        'driver_profile': 'Profil du Conducteur',
        'driver_label': 'Conducteur :',
        'analysis_type_label': "Type d'Analyse :",
        'report_date_label': 'Date du Rapport :',
        'report_type_label': 'Type de Rapport :',
        'report_type_value': 'Analyse Contextuelle du Comportement',
        
        # Trip Summary
        'trip_summary_title': 'Résumé des Trajets',
        'trip_headers': ['Semaine', 'Distance (km)', 'Durée (h:mm)', 'Trajets', 'Vitesse Moy.', 'Vitesse Max'],
        
        # KPI
        'kpi_title': 'Indicateurs Clés',
        'overall_score': 'Score Global',
        'total_violations': 'Infractions Totales',
        'speed_compliance': 'Respect des Vitesses',
        'roundabout_performance': 'Comportement Rond-Points',
        'stop_signs_compliance': 'Respect des Stops',
        'traffic_light_compliance': 'Respect des Feux',
        'harsh_events_rate': 'Événements Brusques',
        
        # Speed
        'top_violations_title': 'Top 3 Infractions de Vitesse (Toutes Semaines)',
        'speed_violations_headers': ['Semaine', 'Vitesse (km/h)', 'Limite (km/h)', 'Excès (km/h)', 'Gravité'],
        'speed_analysis_title': 'Analyse par Zone de Vitesse',
        'speed_zone_headers': ['Semaine', 'Zone', 'Nb. Infractions', 'Respect (%)', 'Limite (km/h)', 'Vitesse Max'],
        'speed_chart_title': 'Respect Hebdomadaire par Zone de Vitesse (%)',
        
        # Roundabouts
        'roundabout_title': 'Approche des Rond-Points',
        'roundabout_headers': ['Semaine', 'Rond-Points', 'Bonne\n(150-100m)', 'Modérée\n(100-65m)', 
                              'Tardive\n(<65m)', 'Pas de\nDécél.', 'Entrée\nConforme', 'Conformité\n(%)'],
        'roundabout_chart_title': 'Anticipation aux Rond-Points',
        
        # Stop Signs
        'stop_sign_title': 'Respect des Panneaux Stop',
        'stop_sign_headers': ['Semaine', 'Panneaux\nStop', 'Approches', 'Arrêt\nCorrect', 
                             'Arrêt Non-\nConforme', 'Incertain', 'Conformité\n(%)'],
        'stop_chart_title': 'Respect des Stops par Semaine',
        
        # Traffic Lights
        'traffic_light_title': 'Comportement aux Feux de Signalisation',
        'traffic_light_headers': ['Semaine', 'Feux\nDétectés', 'Approches', 'Arrêt\nComplet', 
                                  'Ralenti', 'Passage\n(Vert)', 'Taux Arrêt\n(%)'],
        'traffic_light_chart_title': 'Comportement aux Feux par Semaine',
        
        # Harsh Events
        'harsh_events_title': 'Événements de Conduite Brusque',
        'harsh_events_headers': ['Semaine', 'Total\nÉvénements', 'Freinages\nBrusques', 
                                 'Accélérations\nBrusques', 'Virages\nSerrés', 'Événements\n/100km', 'Sévérité'],
        'harsh_events_chart_title': 'Événements Brusques par Semaine',
        
        # Severity
        'severe': 'Sévère',
        'major': 'Majeur',
        'moderate': 'Modéré',
        'minor': 'Mineur',
        
        # Harsh Events Severity
        'excellent': 'Excellent',
        'good': 'Bon',
        'needs_work': 'À Améliorer',
        'critical': 'Critique',
        'no_data': 'N/A',
        
        # Improvements
        'improvement_title': "Axes d'Amélioration",
        'speed_management': 'Gestion de la Vitesse',
        'speed_recommendation': 'Respectez les limites de vitesse, particulièrement dans les zones à risque.',
        'roundabout_technique': 'Technique aux Rond-Points',
        'roundabout_recommendation': 'Réduisez la vitesse dès 150m avant un rond-point pour une approche fluide.',
        'stop_sign_compliance': 'Respect des Stops',
        'stop_recommendation': 'Arrêtez-vous complètement aux stops. Comptez "mille-un" avant de repartir.',
        'traffic_light_compliance': 'Respect des Feux',
        'traffic_light_recommendation': 'Anticipez les feux et préparez-vous à vous arrêter.',
        'harsh_driving': 'Conduite Souple',
        'harsh_recommendation': 'Adoptez une conduite plus progressive pour réduire les freinages et accélérations brusques.',
        'overall_awareness': 'Sensibilisation Générale',
        'awareness_recommendation': 'Renforcez votre attention aux limitations et panneaux. Envisagez une formation préventive.',
        'excellent_performance': 'Performance Excellente',
        'excellent_recommendation': 'Continuez ainsi ! Votre conduite est exemplaire.',
        
        # Priority
        'high_priority': 'Priorité Élevée',
        'medium_priority': 'Priorité Moyenne',
        'low_priority': 'Priorité Basse',
        
        # Footer
        'footer_text': 'Rapport généré par la plateforme FeelGood Conduite.',
        'generated_text': 'Rapport Généré: '
    }

    # ========================================================================
    # INITIALIZATION
    # ========================================================================
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Rapport FeelGood"
        
        # Remove extra sheets
        for sheet in self.wb.worksheets[1:]:
            self.wb.remove(sheet)
        
        # Enterprise color scheme
        self.colors = {
            # Brand colors
            'brand_primary': '1E3A8A',      # Deep blue
            'brand_secondary': '3B82F6',    # Medium blue
            'brand_light': 'DBEAFE',        # Light blue
            
            # RAG colors (Red/Amber/Green)
            'rag_green': '10B981',          # Excellent (≥80%)
            'rag_light_green': '34D399',    # Good (60-79%)
            'rag_amber': 'F59E0B',          # Needs work (40-59%)
            'rag_red': 'EF4444',            # Critical (<40%)
            'rag_dark_red': 'B91C1C',       # Severe
            
            # Neutral
            'gray_light': 'F3F4F6',
            'gray_medium': '9CA3AF',
            'gray_dark': '4B5563',
            'white': 'FFFFFF',
            'black': '000000'
        }
        
        # Professional fonts
        self.fonts = {
            'title_large': Font(name='Segoe UI', size=24, bold=True, color=self.colors['white']),
            'title_medium': Font(name='Segoe UI', size=16, bold=True, color=self.colors['brand_primary']),
            'section_header': Font(name='Segoe UI', size=14, bold=True, color=self.colors['brand_primary']),
            'table_header': Font(name='Segoe UI', size=10, bold=True, color=self.colors['white']),
            'kpi_value': Font(name='Segoe UI', size=24, bold=True),
            'kpi_label': Font(name='Segoe UI', size=10, bold=True),
            'normal': Font(name='Segoe UI', size=10),
            'bold': Font(name='Segoe UI', size=10, bold=True),
            'small': Font(name='Segoe UI', size=9),
            'trend_arrow': Font(name='Segoe UI', size=12, bold=True)
        }
        
        # Standard border
        self.thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        self._setup_page_layout()

    def _setup_page_layout(self):
        """Setup professional page layout"""
        
        # Column widths
        column_widths = {
            'A': 12, 'B': 12, 'C': 12, 'D': 12, 'E': 12,
            'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 12,
            'K': 12, 'L': 12
        }
        for col, width in column_widths.items():
            self.ws.column_dimensions[col].width = width
        
        # Default row height
        for row in range(1, 150):
            self.ws.row_dimensions[row].height = 18
        
        # Print settings
        self.ws.page_setup.orientation = 'landscape'
        self.ws.page_setup.fitToWidth = 1
        self.ws.page_setup.fitToHeight = False

    # ========================================================================
    # RAG COLOR HELPER
    # ========================================================================
    
    def _get_rag_color(self, score, invert=False):
        """
        Get RAG color based on score.
        
        Args:
            score: Value 0-100
            invert: If True, lower is better (for violations, harsh events)
        
        Returns:
            tuple: (background_color, text_color, status_text)
        """
        if score is None:
            return self.colors['gray_medium'], self.colors['white'], self.FRENCH_STRINGS['no_data']
        
        if invert:
            score = 100 - min(score, 100)
        
        if score >= 80:
            return self.colors['rag_green'], self.colors['white'], self.FRENCH_STRINGS['excellent']
        elif score >= 60:
            return self.colors['rag_light_green'], self.colors['white'], self.FRENCH_STRINGS['good']
        elif score >= 40:
            return self.colors['rag_amber'], self.colors['white'], self.FRENCH_STRINGS['needs_work']
        else:
            return self.colors['rag_red'], self.colors['white'], self.FRENCH_STRINGS['critical']

    def _apply_rag_fill(self, cell, score, invert=False):
        """Apply RAG coloring to a cell"""
        bg_color, text_color, _ = self._get_rag_color(score, invert)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.font = Font(name='Segoe UI', size=11, bold=True, color=text_color)

    # ========================================================================
    # REUSABLE TABLE HELPER
    # ========================================================================
    
    def _create_table(self, start_row, title, headers, data_rows, column_widths=None):
        """
        Create a standardized table with headers and data.
        
        Args:
            start_row: Starting row number
            title: Section title
            headers: List of header strings
            data_rows: List of dicts with 'values' and optional 'formats'
            column_widths: Optional list of column widths
        
        Returns:
            tuple: (next_row, header_row) for chart positioning
        """
        
        # Title
        self.ws[f'A{start_row}'] = title
        self.ws[f'A{start_row}'].font = self.fonts['section_header']
        start_row += 2
        
        header_row = start_row
        
        # Headers
        for i, header in enumerate(headers):
            cell = self.ws.cell(row=start_row, column=i+1, value=header)
            cell.font = self.fonts['table_header']
            cell.fill = PatternFill(start_color=self.colors['brand_secondary'], 
                                   end_color=self.colors['brand_secondary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.thin_border
        
        self.ws.row_dimensions[start_row].height = 40
        
        # Column widths
        if column_widths:
            for i, width in enumerate(column_widths):
                self.ws.column_dimensions[get_column_letter(i+1)].width = width
        
        start_row += 1
        
        # Data rows
        for row_data in data_rows:
            values = row_data.get('values', [])
            formats = row_data.get('formats', {})
            
            for j, value in enumerate(values):
                cell = self.ws.cell(row=start_row, column=j+1, value=value)
                cell.font = self.fonts['normal']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
                
                # Apply custom formatting if specified
                if j in formats:
                    fmt = formats[j]
                    if 'rag_score' in fmt:
                        self._apply_rag_fill(cell, fmt['rag_score'], fmt.get('invert', False))
                    if 'fill' in fmt:
                        cell.fill = PatternFill(start_color=fmt['fill'], end_color=fmt['fill'], fill_type='solid')
                    if 'font_color' in fmt:
                        cell.font = Font(name='Segoe UI', size=11, bold=fmt.get('bold', False), color=fmt['font_color'])
            
            start_row += 1
        
        return start_row + 1, header_row

    # ========================================================================
    # REUSABLE CHART HELPER
    # ========================================================================
    
    def _create_stacked_bar_chart(self, title, data_dict, categories, series_names, colors_list, 
                                   position, hidden_col_start=27):
        """
        Create a standardized stacked bar chart.
        
        Args:
            title: Chart title
            data_dict: Dict with category keys and series values
            categories: List of category labels (e.g., week names)
            series_names: List of series names
            colors_list: List of colors for each series
            position: Cell position for chart (e.g., 'H10')
            hidden_col_start: Column to store chart data (hidden area)
        """
        
        if not categories or not data_dict:
            return
        
        # Find next available row in hidden area
        chart_data_row = 1
        while self.ws.cell(row=chart_data_row, column=hidden_col_start).value is not None:
            chart_data_row += 1
        
        # Write chart data headers
        self.ws.cell(row=chart_data_row, column=hidden_col_start, value="Category")
        for i, series_name in enumerate(series_names):
            self.ws.cell(row=chart_data_row, column=hidden_col_start + i + 1, value=series_name)
        
        # Write chart data
        data_start_row = chart_data_row
        chart_data_row += 1
        
        for category in categories:
            self.ws.cell(row=chart_data_row, column=hidden_col_start, value=category)
            category_data = data_dict.get(category, {})
            
            for i, series_name in enumerate(series_names):
                value = category_data.get(series_name, 0) or 0
                self.ws.cell(row=chart_data_row, column=hidden_col_start + i + 1, value=value)
            
            chart_data_row += 1
        
        # Create chart
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.style = 10
        chart.title = title
        chart.y_axis.title = ""
        chart.x_axis.title = ""
        
        chart.width = 10
        chart.height = 6
        
        chart.layout = Layout(manualLayout=ManualLayout(x=0.02, y=0.05, w=0.75, h=0.85))
        chart.legend.position = "r"
        chart.legend.layout = Layout(manualLayout=ManualLayout(x=0.78, y=0.3, w=0.2, h=0.4))
        chart.y_axis.majorGridlines = None
        chart.gapWidth = 150
        
        # Data references
        num_categories = len(categories)
        num_series = len(series_names)
        
        data_ref = Reference(self.ws,
                            min_col=hidden_col_start + 1,
                            max_col=hidden_col_start + num_series,
                            min_row=data_start_row,
                            max_row=data_start_row + num_categories)
        
        cats_ref = Reference(self.ws,
                            min_col=hidden_col_start,
                            min_row=data_start_row + 1,
                            max_row=data_start_row + num_categories)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        # Apply colors
        for i, color in enumerate(colors_list):
            if i < len(chart.series):
                chart.series[i].graphicalProperties.solidFill = color
        
        self.ws.add_chart(chart, position)

    # ========================================================================
    # HEADER SECTION
    # ========================================================================
    
    def _create_pie_chart(self, title, data_dict, colors_dict, position, 
                          chart_style='doughnut', hidden_col_start=27):
        """
        Create a pie/doughnut chart for behavior distribution.
        
        Args:
            title: Chart title
            data_dict: Dict with category names as keys and counts as values
                      Example: {'Good': 15, 'Moderate': 8, 'Late': 3, 'No Decel': 2}
            colors_dict: Dict mapping category names to hex colors
            position: Cell position for chart (e.g., 'H10')
            chart_style: 'pie' for standard pie, 'doughnut' for donut chart
            hidden_col_start: Column to store chart data (hidden area)
        """
        from openpyxl.chart import PieChart, DoughnutChart
        from openpyxl.chart.series import DataPoint
        
        if not data_dict:
            return
        
        # Find next available row in hidden area
        chart_data_row = 1
        while self.ws.cell(row=chart_data_row, column=hidden_col_start).value is not None:
            chart_data_row += 1
        
        # Write chart data headers
        self.ws.cell(row=chart_data_row, column=hidden_col_start, value="Category")
        self.ws.cell(row=chart_data_row, column=hidden_col_start + 1, value="Count")
        
        data_start_row = chart_data_row
        chart_data_row += 1
        
        # Write chart data and prepare color list
        categories = []
        for category, value in data_dict.items():
            if value and value > 0:  # Only include non-zero values
                self.ws.cell(row=chart_data_row, column=hidden_col_start, value=category)
                self.ws.cell(row=chart_data_row, column=hidden_col_start + 1, value=value)
                categories.append(category)
                chart_data_row += 1
        
        if not categories:
            return  # No data to chart
        
        # Create appropriate chart type
        if chart_style == 'doughnut':
            chart = DoughnutChart()
            chart.holeSize = 50  # Size of center hole (0-90)
        else:
            chart = PieChart()
        
        chart.title = title
        chart.style = 10
        chart.width = 10
        chart.height = 7
        
        # Data references
        num_categories = len(categories)
        
        data_ref = Reference(self.ws,
                            min_col=hidden_col_start + 1,
                            max_col=hidden_col_start + 1,
                            min_row=data_start_row + 1,
                            max_row=data_start_row + num_categories)
        
        cats_ref = Reference(self.ws,
                            min_col=hidden_col_start,
                            min_row=data_start_row + 1,
                            max_row=data_start_row + num_categories)
        
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        
        # Apply colors to each slice
        if len(chart.series) > 0:
            series = chart.series[0]
            for i, category in enumerate(categories):
                if category in colors_dict:
                    pt = DataPoint(idx=i)
                    pt.graphicalProperties.solidFill = colors_dict[category]
                    series.dPt.append(pt)
        
        # Legend configuration
        chart.legend.position = "r"
        chart.legend.layout = Layout(manualLayout=ManualLayout(x=0.75, y=0.25, w=0.2, h=0.5))
        
        # Data labels showing percentages
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showCatName = False
        
        self.ws.add_chart(chart, position)
    
    
    def _create_multi_week_pie_charts(self, title, data_dict, colors_dict, 
                                       position_start, chart_style='doughnut'):
        """
        Create multiple pie charts, one for each week, arranged horizontally.
        
        Args:
            title: Base chart title
            data_dict: Dict with week labels as keys, each containing category counts
                      Example: {'Week1': {'Good': 15, 'Late': 3}, 'Week2': {...}}
            colors_dict: Dict mapping category names to hex colors
            position_start: Starting cell position (e.g., 'H10')
            chart_style: 'pie' or 'doughnut'
        """
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        
        if not data_dict:
            return
        
        # Parse starting position
        col_letter, row = coordinate_from_string(position_start)
        col_idx = column_index_from_string(col_letter)
        
        # Create one pie chart per week
        weeks = sorted(data_dict.keys())
        
        for i, week in enumerate(weeks):
            week_data = data_dict[week]
            
            # Calculate position for this chart (offset by 6 columns per chart)
            chart_col = col_idx + (i * 6)
            chart_position = f"{get_column_letter(chart_col)}{row}"
            
            # Create individual pie chart
            week_title = f"{title} - {week}"
            self._create_pie_chart(
                title=week_title,
                data_dict=week_data,
                colors_dict=colors_dict,
                position=chart_position,
                chart_style=chart_style
            )
    
    def _aggregate_chart_data(self, data_dict):
        """
        Aggregate data across multiple weeks.
        
        Args:
            data_dict: Dict with week labels as keys, each containing category counts
                      Example: {'Week1': {'Good': 15, 'Late': 3}, 'Week2': {...}}
        
        Returns:
            Dict with aggregated totals per category
            Example: {'Good': 30, 'Moderate': 15, 'Late': 8}
        """
        aggregated = {}
        
        for week_label, week_data in data_dict.items():
            for category, value in week_data.items():
                if category not in aggregated:
                    aggregated[category] = 0
                aggregated[category] += value or 0
        
        return aggregated
    
    def _create_header(self, all_data):
        """Create professional header with driver info"""
        
        metadata = all_data['metadata']
        
        # Company banner
        self.ws.merge_cells('A1:L3')
        self.ws['A1'] = self.FRENCH_STRINGS['company_name']
        self.ws['A1'].font = self.fonts['title_large']
        self.ws['A1'].fill = PatternFill(start_color=self.colors['brand_primary'], 
                                        end_color=self.colors['brand_primary'], fill_type='solid')
        self.ws['A1'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 20
        self.ws.row_dimensions[3].height = 20
        
        # Report title
        self.ws['A5'] = self.FRENCH_STRINGS['report_title']
        self.ws['A5'].font = self.fonts['title_medium']
        
        # Driver profile box
        self.ws.merge_cells('H5:K5')
        self.ws['H5'] = self.FRENCH_STRINGS['driver_profile']
        self.ws['H5'].font = Font(name='Segoe UI', size=11, bold=True, color=self.colors['white'])
        self.ws['H5'].fill = PatternFill(start_color=self.colors['brand_primary'],
                                        end_color=self.colors['brand_primary'], fill_type='solid')
        self.ws['H5'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Driver info rows
        driver_info = [
            (self.FRENCH_STRINGS['driver_label'], metadata['driver_name']),
            (self.FRENCH_STRINGS['analysis_type_label'], metadata['analysis_type']),
            (self.FRENCH_STRINGS['report_date_label'], metadata['analysis_date']),
            (self.FRENCH_STRINGS['report_type_label'], self.FRENCH_STRINGS['report_type_value'])
        ]
        
        for i, (label, value) in enumerate(driver_info):
            row = 6 + i
            self.ws[f'H{row}'] = label
            self.ws[f'H{row}'].font = self.fonts['bold']
            self.ws[f'I{row}'] = value
            self.ws[f'I{row}'].font = self.fonts['normal']
            self.ws.merge_cells(f'I{row}:K{row}')
            
            for col in ['H', 'I', 'J', 'K']:
                self.ws[f'{col}{row}'].fill = PatternFill(start_color=self.colors['gray_light'],
                                                         end_color=self.colors['gray_light'], fill_type='solid')
                self.ws[f'{col}{row}'].border = self.thin_border
        
        return 10

    # ========================================================================
    # TRIP SUMMARY TABLE
    # ========================================================================
    
    def _create_trip_summary_table(self, all_data, start_row):
        """Create trip summary table with distance, duration, trips per week"""
        
        trip_summary = all_data.get('trip_summary', {})
        weekly_trips = trip_summary.get('weekly', {})
        totals = trip_summary.get('totals', {})
        
        # Prepare data rows
        data_rows = []
        
        for week_label in sorted(weekly_trips.keys()):
            week_data = weekly_trips[week_label]
            
            # Format duration as h:mm
            duration_hours = week_data.get('duration_hours', 0)
            hours = int(duration_hours)
            minutes = int((duration_hours - hours) * 60)
            duration_str = f"{hours}:{minutes:02d}"
            
            data_rows.append({
                'values': [
                    week_label,
                    f"{week_data.get('distance_km', 0):.1f}",
                    duration_str,
                    week_data.get('trip_count', 0),
                    f"{week_data.get('avg_speed_kmh', 0):.0f} km/h",
                    f"{week_data.get('max_speed_kmh', 0):.0f} km/h"
                ]
            })
        
        # Add totals row
        total_hours = int(totals.get('total_duration_hours', 0))
        total_minutes = int((totals.get('total_duration_hours', 0) - total_hours) * 60)
        
        data_rows.append({
            'values': [
                'TOTAL',
                f"{totals.get('total_distance_km', 0):.1f}",
                f"{total_hours}:{total_minutes:02d}",
                totals.get('total_trips', 0),
                f"{totals.get('avg_speed_kmh', 0):.0f} km/h",
                f"{totals.get('max_speed_kmh', 0):.0f} km/h"
            ],
            'formats': {
                0: {'bold': True, 'fill': self.colors['gray_light']},
                1: {'bold': True, 'fill': self.colors['gray_light']},
                2: {'bold': True, 'fill': self.colors['gray_light']},
                3: {'bold': True, 'fill': self.colors['gray_light']},
                4: {'bold': True, 'fill': self.colors['gray_light']},
                5: {'bold': True, 'fill': self.colors['gray_light']}
            }
        })
        
        next_row, header_row = self._create_table(
            start_row=start_row,
            title=self.FRENCH_STRINGS['trip_summary_title'],
            headers=self.FRENCH_STRINGS['trip_headers'],
            data_rows=data_rows,
            column_widths=[10, 12, 12, 10, 12, 12]
        )
        
        return next_row

    # ========================================================================
    # KPI DASHBOARD WITH TRENDS
    # ========================================================================
    
    def _create_kpi_dashboard(self, all_data, start_row):
        """Create KPI dashboard with trend arrows"""
        
        aggregated = all_data.get('aggregated_metrics', {})
        combined = aggregated.get('combined_metrics', {})
        trends = all_data.get('trends', {})
        weekly_changes = trends.get('weekly_changes', {})
        
        # Get latest trend arrows
        latest_week = sorted(weekly_changes.keys())[-1] if weekly_changes else None
        latest_trends = weekly_changes.get(latest_week, {}) if latest_week else {}
        
        # Title
        self.ws[f'A{start_row}'] = self.FRENCH_STRINGS['kpi_title']
        self.ws[f'A{start_row}'].font = self.fonts['section_header']
        start_row += 2
        
        # KPI definitions
        kpis = [
            {
                'label': self.FRENCH_STRINGS['overall_score'],
                'value': combined.get('avg_overall_score', 0),
                'format': '{:.1f}%',
                'trend_key': 'overall_score'
            },
            {
                'label': self.FRENCH_STRINGS['total_violations'],
                'value': aggregated.get('total_violations', 0),
                'format': '{}',
                'invert': True
            },
            {
                'label': self.FRENCH_STRINGS['speed_compliance'],
                'value': combined.get('avg_speed_compliance', 0),
                'format': '{:.1f}%',
                'trend_key': 'speed_compliance'
            },
            {
                'label': self.FRENCH_STRINGS['roundabout_performance'],
                'value': combined.get('avg_roundabout_performance', 0),
                'format': '{:.1f}%',
                'trend_key': 'roundabout_performance'
            },
            {
                'label': self.FRENCH_STRINGS['stop_signs_compliance'],
                'value': combined.get('avg_stop_sign_compliance', 0),
                'format': '{:.1f}%',
                'trend_key': 'stop_sign_compliance'
            },
            {
                'label': self.FRENCH_STRINGS['harsh_events_rate'],
                'value': combined.get('avg_harsh_events_per_100km', 0),
                'format': '{:.1f}/100km',
                'invert': True
            }
        ]
        
        # Create 3x2 grid
        positions = [('A', 'B'), ('C', 'D'), ('E', 'F')]
        
        for i, kpi in enumerate(kpis):
            row = start_row if i < 3 else start_row + 3
            col_start, col_end = positions[i % 3]
            
            value = kpi['value'] or 0
            invert = kpi.get('invert', False)
            
            # Get RAG color
            if invert:
                # For violations/harsh events - convert to score-like value
                rag_score = max(0, 100 - (value * 5)) if 'violations' in kpi['label'].lower() else max(0, 100 - (value * 10))
            else:
                rag_score = value
            
            bg_color, text_color, _ = self._get_rag_color(rag_score)
            
            # Apply background to both rows
            for r in range(row, row + 2):
                for c in [col_start, col_end]:
                    cell = self.ws[f'{c}{r}']
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                    cell.border = self.thin_border
            
            # Label row
            self.ws.merge_cells(f'{col_start}{row}:{col_end}{row}')
            label_cell = self.ws[f'{col_start}{row}']
            label_cell.value = kpi['label']
            label_cell.font = Font(name='Segoe UI', size=10, bold=True, color=text_color)
            label_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Value row with trend arrow
            self.ws.merge_cells(f'{col_start}{row+1}:{col_end}{row+1}')
            value_cell = self.ws[f'{col_start}{row+1}']
            
            # Format value
            formatted_value = kpi['format'].format(value)
            
            # Add trend arrow if available
            trend_key = kpi.get('trend_key')
            if trend_key and trend_key in latest_trends:
                arrow = latest_trends[trend_key].get('arrow', '')
                formatted_value = f"{formatted_value} {arrow}"
            
            value_cell.value = formatted_value
            value_cell.font = Font(name='Segoe UI', size=20, bold=True, color=text_color)
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set row heights
        self.ws.row_dimensions[start_row].height = 25
        self.ws.row_dimensions[start_row + 1].height = 40
        self.ws.row_dimensions[start_row + 3].height = 25
        self.ws.row_dimensions[start_row + 4].height = 40
        
        return start_row + 6

    # ========================================================================
    # TOP SPEED VIOLATIONS TABLE
    # ========================================================================
    
    def _create_top_violations_table(self, all_data, start_row):
        """Create top 3 speed violations table"""
        
        # Collect all violations
        all_violations = []
        for week_label, week_data in all_data.get('weekly_data', {}).items():
            violations = week_data.get('speeding', {}).get('worst_violations', [])
            for v in violations:
                v_copy = v.copy()
                v_copy['week'] = week_label
                all_violations.append(v_copy)
        
        # Sort by excess speed and take top 3
        top_violations = sorted(all_violations, key=lambda x: x.get('excess_speed', 0), reverse=True)[:3]
        
        # Prepare data rows
        severity_map = {
            'severe': self.FRENCH_STRINGS['severe'],
            'major': self.FRENCH_STRINGS['major'],
            'moderate': self.FRENCH_STRINGS['moderate'],
            'minor': self.FRENCH_STRINGS['minor']
        }
        
        data_rows = []
        for v in top_violations:
            severity = v.get('severity', 'unknown')
            severity_french = severity_map.get(severity, severity.upper())
            excess = v.get('excess_speed', 0)
            
            # Determine severity color
            if severity == 'severe':
                sev_color = self.colors['rag_dark_red']
            elif severity == 'major':
                sev_color = self.colors['rag_red']
            elif severity == 'moderate':
                sev_color = self.colors['rag_amber']
            else:
                sev_color = self.colors['rag_green']
            
            data_rows.append({
                'values': [
                    v.get('week', 'N/A'),
                    f"{v.get('speed_kmh', 0):.0f}",
                    f"{v.get('speed_limit', 0)}",
                    f"{excess:.0f}",
                    severity_french
                ],
                'formats': {
                    3: {'rag_score': max(0, 100 - excess * 2)},
                    4: {'fill': sev_color, 'font_color': self.colors['white'], 'bold': True}
                }
            })
        
        next_row, _ = self._create_table(
            start_row=start_row,
            title=self.FRENCH_STRINGS['top_violations_title'],
            headers=self.FRENCH_STRINGS['speed_violations_headers'],
            data_rows=data_rows,
            column_widths=[10, 12, 12, 12, 12]
        )
        
        return next_row

    # ========================================================================
    # SPEED ANALYSIS TABLE
    # ========================================================================
    
    def _create_speed_analysis_table(self, all_data, start_row):
        """Create speed analysis by zone table with chart"""
        
        # Title
        self.ws[f'A{start_row}'] = self.FRENCH_STRINGS['speed_analysis_title']
        self.ws[f'A{start_row}'].font = self.fonts['section_header']
        start_row += 2
        header_row = start_row
        
        # Headers
        headers = self.FRENCH_STRINGS['speed_zone_headers']
        for i, header in enumerate(headers):
            cell = self.ws.cell(row=start_row, column=i+1, value=header)
            cell.font = self.fonts['table_header']
            cell.fill = PatternFill(start_color=self.colors['brand_secondary'],
                                   end_color=self.colors['brand_secondary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.thin_border
        
        self.ws.row_dimensions[start_row].height = 40
        start_row += 1
        
        # Zone name mapping
        # Zone name mapping - Fixed speed limit categories
        zone_name_map = {
                '30': 'Zone 30',
                '50': 'Zone 50',
                '70': 'Zone 70',
                '80': 'Zone 80',
                '90': 'Zone 90',
                '110': 'Zone 110',
                '130': 'Zone 130'
        }
        
        chart_data = {}
        
        for week_label in sorted(all_data.get('weekly_data', {}).keys()):
            week_data = all_data['weekly_data'][week_label]
            speed_zones = week_data.get('speeding', {}).get('speed_zones', {})
            zone_breakdown = speed_zones.get('zone_breakdown', {})
            
            chart_data[week_label] = {}
            
            for zone_key, zone_data in zone_breakdown.items():
                
                zone_name = zone_name_map.get(zone_key, zone_key)
                violations = zone_data.get('violation_episodes', 0)
                compliance = zone_data.get('compliance_percentage', 0)
                limit = zone_data.get('speed_limit_kmh', 0)
                max_speed = zone_data.get('top_speed', 0)
                
                chart_data[week_label][zone_name] = compliance
                
                # Write row
                row_values = [
                    week_label,
                    zone_name,
                    violations if violations > 0 else '-',
                    f"{compliance:.1f}%" if compliance else '-',
                    limit if limit > 0 else '-',
                    f"{max_speed:.0f}" if max_speed > 0 else '-'
                ]
                
                for j, value in enumerate(row_values):
                    cell = self.ws.cell(row=start_row, column=j+1, value=value)
                    cell.font = self.fonts['normal']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.thin_border
                    
                    # Color code violations
                    if j == 2 and isinstance(value, int) and value > 0:
                        if value >= 10:
                            self._apply_rag_fill(cell, 20)
                        elif value >= 5:
                            self._apply_rag_fill(cell, 50)
                        else:
                            self._apply_rag_fill(cell, 75)
                
                start_row += 1
        
        # Create chart
        if chart_data:
            categories = sorted(chart_data.keys())
            all_zones = set()
            for week_zones in chart_data.values():
                all_zones.update(week_zones.keys())
            series_names = sorted(all_zones)
            
            colors_list = [
                self.colors['brand_primary'],
                self.colors['rag_amber'],
                self.colors['rag_green'],
                self.colors['rag_red'],
                self.colors['brand_secondary']
            ]
            
            self._create_stacked_bar_chart(
                title=self.FRENCH_STRINGS['speed_chart_title'],
                data_dict=chart_data,
                categories=categories,
                series_names=series_names,
                colors_list=colors_list[:len(series_names)],
                position=f'H{header_row}'
            )
        
        return start_row + 2

    # ========================================================================
    # ROUNDABOUT BEHAVIOR TABLE
    # ========================================================================
    
    def _create_roundabout_table(self, all_data, start_row):
        """Create roundabout behavior table with chart"""
        
        # Title
        self.ws[f'A{start_row}'] = self.FRENCH_STRINGS['roundabout_title']
        self.ws[f'A{start_row}'].font = self.fonts['section_header']
        start_row += 2
        header_row = start_row
        
        # Headers
        headers = self.FRENCH_STRINGS['roundabout_headers']
        for i, header in enumerate(headers):
            cell = self.ws.cell(row=start_row, column=i+1, value=header)
            cell.font = self.fonts['table_header']
            cell.fill = PatternFill(start_color=self.colors['brand_secondary'],
                                   end_color=self.colors['brand_secondary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.thin_border
        
        self.ws.row_dimensions[start_row].height = 50
        start_row += 1
        
        chart_data = {}
        
        for week_label in sorted(all_data.get('weekly_data', {}).keys()):
            week_data = all_data['weekly_data'][week_label]
            rb_data = week_data.get('roundabouts', {})
            
            total_rb = rb_data.get('total_roundabouts', 0)
            anticipation = rb_data.get('anticipation_stats', {})
            entry_compliance = rb_data.get('entry_compliance', {})
            
            good = anticipation.get('good', 0)
            moderate = anticipation.get('moderate', 0)
            late = anticipation.get('late', 0)
            no_decel = anticipation.get('no_deceleration', 0)
            
            entry_compliant = entry_compliance.get('compliant', 0)
            entry_total = entry_compliant + entry_compliance.get('non_compliant', 0)
            
            compliance_pct = rb_data.get('entry_compliance_percentage', 0)
            
            # Store for chart
            chart_data[week_label] = {
                'Bonne': good,
                'Modérée': moderate,
                'Tardive': late,
                'Pas de Décél.': no_decel
            }
            
            # Write row
            row_values = [
                week_label,
                total_rb,
                good,
                moderate,
                late,
                no_decel,
                f"{entry_compliant}/{entry_total}" if entry_total > 0 else "N/A",
                f"{compliance_pct:.0f}%" if compliance_pct else "N/A"
            ]
            
            for j, value in enumerate(row_values):
                cell = self.ws.cell(row=start_row, column=j+1, value=value)
                cell.font = self.fonts['normal']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
                
                # Color coding
                if j == 2 and isinstance(value, int) and value > 0:  # Good
                    cell.fill = PatternFill(start_color=self.colors['rag_green'],
                                           end_color=self.colors['rag_green'], fill_type='solid')
                elif j == 3 and isinstance(value, int) and value > 0:  # Moderate
                    cell.fill = PatternFill(start_color=self.colors['rag_amber'],
                                           end_color=self.colors['rag_amber'], fill_type='solid')
                elif j == 4 and isinstance(value, int) and value > 0:  # Late
                    cell.fill = PatternFill(start_color=self.colors['rag_red'],
                                           end_color=self.colors['rag_red'], fill_type='solid')
                    cell.font = Font(name='Segoe UI', size=10, color=self.colors['white'])
                elif j == 5 and isinstance(value, int) and value > 0:  # No decel
                    cell.fill = PatternFill(start_color=self.colors['rag_dark_red'],
                                           end_color=self.colors['rag_dark_red'], fill_type='solid')
                    cell.font = Font(name='Segoe UI', size=10, color=self.colors['white'])
                elif j == 7:  # Compliance %
                    self._apply_rag_fill(cell, compliance_pct)
            
            start_row += 1
        
        # Create chart
        if chart_data:
            # Aggregate data across all weeks
            aggregated_data = self._aggregate_chart_data(chart_data)
            
            # Define color mapping
            colors_dict = {
                'Bonne': self.colors['rag_green'],
                'Modérée': self.colors['rag_amber'],
                'Tardive': self.colors['rag_red'],
                'Pas de Décél.': self.colors['rag_dark_red']
            }
            
            # Create single aggregated pie chart
            self._create_pie_chart(
                title=self.FRENCH_STRINGS['roundabout_chart_title'],
                data_dict=aggregated_data,
                colors_dict=colors_dict,
                position=f'H{header_row}',
                chart_style='doughnut'
            )
        
        return start_row + 2

    # ========================================================================
    # STOP SIGN COMPLIANCE TABLE
    # ========================================================================
    
    def _create_stop_sign_table(self, all_data, start_row):
        """Create stop sign compliance table with chart"""
        
        # Title
        self.ws[f'A{start_row}'] = self.FRENCH_STRINGS['stop_sign_title']
        self.ws[f'A{start_row}'].font = self.fonts['section_header']
        start_row += 2
        header_row = start_row
        
        # Headers
        headers = self.FRENCH_STRINGS['stop_sign_headers']
        for i, header in enumerate(headers):
            cell = self.ws.cell(row=start_row, column=i+1, value=header)
            cell.font = self.fonts['table_header']
            cell.fill = PatternFill(start_color=self.colors['brand_secondary'],
                                   end_color=self.colors['brand_secondary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.thin_border
        
        self.ws.row_dimensions[start_row].height = 50
        start_row += 1
        
        chart_data = {}
        
        for week_label in sorted(all_data.get('weekly_data', {}).keys()):
            week_data = all_data['weekly_data'][week_label]
            stop_data = week_data.get('stop_signs', {})
            
            total_stops = stop_data.get('total_stop_signs', 0)
            total_approaches = stop_data.get('total_approaches', 0)
            compliance = stop_data.get('compliance', {})
            
            stop_ok = compliance.get('stop_ok', 0)
            stop_ko = compliance.get('stop_ko', 0)
            uncertain = compliance.get('uncertain', 0)
            
            compliance_pct = stop_data.get('compliance_percentage', 0)
            
            # Store for chart
            chart_data[week_label] = {
                'STOP_OK': stop_ok,
                'STOP_KO': stop_ko,
                'Incertain': uncertain
            }
            
            # Write row
            row_values = [
                week_label,
                total_stops,
                total_approaches,
                stop_ok,
                stop_ko,
                uncertain,
                f"{compliance_pct:.0f}%" if compliance_pct else "N/A"
            ]
            
            for j, value in enumerate(row_values):
                cell = self.ws.cell(row=start_row, column=j+1, value=value)
                cell.font = self.fonts['normal']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
                
                # Color coding
                if j == 3 and isinstance(value, int) and value > 0:  # STOP_OK
                    cell.fill = PatternFill(start_color=self.colors['rag_green'],
                                           end_color=self.colors['rag_green'], fill_type='solid')
                elif j == 4 and isinstance(value, int) and value > 0:  # STOP_KO
                    cell.fill = PatternFill(start_color=self.colors['rag_red'],
                                           end_color=self.colors['rag_red'], fill_type='solid')
                    cell.font = Font(name='Segoe UI', size=10, color=self.colors['white'])
                elif j == 5 and isinstance(value, int) and value > 0:  # Uncertain
                    cell.fill = PatternFill(start_color=self.colors['gray_medium'],
                                           end_color=self.colors['gray_medium'], fill_type='solid')
                    cell.font = Font(name='Segoe UI', size=10, color=self.colors['white'])
                elif j == 6:  # Compliance %
                    self._apply_rag_fill(cell, compliance_pct)
            
            start_row += 1
        
        # Create chart
        if chart_data:
            # Aggregate data across all weeks
            aggregated_data = self._aggregate_chart_data(chart_data)
            
            # Define color mapping
            colors_dict = {
                'STOP_OK': self.colors['rag_green'],
                'STOP_KO': self.colors['rag_red'],
                'Incertain': self.colors['gray_medium']
            }
            
            # Create single aggregated pie chart
            self._create_pie_chart(
                title=self.FRENCH_STRINGS['stop_chart_title'],
                data_dict=aggregated_data,
                colors_dict=colors_dict,
                position=f'H{header_row}',
                chart_style='doughnut'
            )
        
        return start_row + 2

    # ========================================================================
    # TRAFFIC LIGHT TABLE
    # ========================================================================
    
    def _create_traffic_light_table(self, all_data, start_row):
        """Create traffic light compliance table with chart"""
        
        # Title
        self.ws[f'A{start_row}'] = self.FRENCH_STRINGS['traffic_light_title']
        self.ws[f'A{start_row}'].font = self.fonts['section_header']
        start_row += 2
        header_row = start_row
        
        # Headers
        headers = self.FRENCH_STRINGS['traffic_light_headers']
        for i, header in enumerate(headers):
            cell = self.ws.cell(row=start_row, column=i+1, value=header)
            cell.font = self.fonts['table_header']
            cell.fill = PatternFill(start_color=self.colors['brand_secondary'],
                                   end_color=self.colors['brand_secondary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.thin_border
        
        self.ws.row_dimensions[start_row].height = 50
        start_row += 1
        
        chart_data = {}
        has_data = False
        
        for week_label in sorted(all_data.get('weekly_data', {}).keys()):
            week_data = all_data['weekly_data'][week_label]
            tl_data = week_data.get('traffic_lights', {})
            
            total_signals = tl_data.get('total_signals', 0)
            total_approaches = tl_data.get('total_approaches', 0)
            compliance = tl_data.get('compliance', {})
            
            stopped = compliance.get('stopped', 0)
            slowed = compliance.get('slowed', 0)
            passed = compliance.get('passed_through', 0)
            
            stop_pct = tl_data.get('stop_percentage', 0)
            
            if total_signals > 0:
                has_data = True
            
            # Store for chart
            chart_data[week_label] = {
                'Arrêt': stopped,
                'Ralenti': slowed,
                'Passage': passed
            }
            
            # Write row
            row_values = [
                week_label,
                total_signals,
                total_approaches,
                stopped,
                slowed,
                passed,
                f"{stop_pct:.0f}%" if stop_pct else "N/A"
            ]
            
            for j, value in enumerate(row_values):
                cell = self.ws.cell(row=start_row, column=j+1, value=value)
                cell.font = self.fonts['normal']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
                
                # Color coding
                if j == 3 and isinstance(value, int) and value > 0:  # Stopped
                    cell.fill = PatternFill(start_color=self.colors['rag_green'],
                                           end_color=self.colors['rag_green'], fill_type='solid')
                elif j == 4 and isinstance(value, int) and value > 0:  # Slowed
                    cell.fill = PatternFill(start_color=self.colors['rag_amber'],
                                           end_color=self.colors['rag_amber'], fill_type='solid')
                elif j == 5 and isinstance(value, int) and value > 0:  # Passed
                    cell.fill = PatternFill(start_color=self.colors['rag_light_green'],
                                           end_color=self.colors['rag_light_green'], fill_type='solid')
                elif j == 6:  # Stop %
                    self._apply_rag_fill(cell, stop_pct)
            
            start_row += 1
        
        # Create chart only if data exists
        if has_data and chart_data:
            # Aggregate data across all weeks
            aggregated_data = self._aggregate_chart_data(chart_data)
            
            # Define color mapping
            colors_dict = {
                'Arrêt': self.colors['rag_green'],
                'Ralenti': self.colors['rag_amber'],
                'Passage': self.colors['rag_light_green']
            }
            
            # Create single aggregated pie chart
            self._create_pie_chart(
                title=self.FRENCH_STRINGS['traffic_light_chart_title'],
                data_dict=aggregated_data,
                colors_dict=colors_dict,
                position=f'H{header_row}',
                chart_style='doughnut'
            )
        
        return start_row + 2

    # ========================================================================
    # HARSH EVENTS TABLE
    # ========================================================================
    
    def _create_harsh_events_table(self, all_data, start_row):
        """Create harsh events analysis table with chart"""
        
        # Title
        self.ws[f'A{start_row}'] = self.FRENCH_STRINGS['harsh_events_title']
        self.ws[f'A{start_row}'].font = self.fonts['section_header']
        start_row += 2
        header_row = start_row
        
        # Headers
        headers = self.FRENCH_STRINGS['harsh_events_headers']
        for i, header in enumerate(headers):
            cell = self.ws.cell(row=start_row, column=i+1, value=header)
            cell.font = self.fonts['table_header']
            cell.fill = PatternFill(start_color=self.colors['brand_secondary'],
                                   end_color=self.colors['brand_secondary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.thin_border
        
        self.ws.row_dimensions[start_row].height = 50
        start_row += 1
        
        chart_data = {}
        
        # Severity mapping
        severity_map = {
            'excellent': self.FRENCH_STRINGS['excellent'],
            'good': self.FRENCH_STRINGS['good'],
            'moderate': self.FRENCH_STRINGS['needs_work'],
            'concerning': self.FRENCH_STRINGS['needs_work'],
            'critical': self.FRENCH_STRINGS['critical']
        }
        
        for week_label in sorted(all_data.get('weekly_data', {}).keys()):
            week_data = all_data['weekly_data'][week_label]
            harsh_data = week_data.get('harsh_events', {})
            
            total_events = harsh_data.get('total_events', 0)
            braking = harsh_data.get('harsh_braking', {}).get('count', 0)
            accel = harsh_data.get('harsh_acceleration', {}).get('count', 0)
            turns = harsh_data.get('sharp_turns', {}).get('count', 0)
            events_per_100km = harsh_data.get('events_per_100km', 0)
            
            severity_info = harsh_data.get('severity_summary', {})
            severity_level = severity_info.get('level', 'unknown')
            severity_french = severity_map.get(severity_level, severity_level)
            
            # Store for chart
            chart_data[week_label] = {
                'Freinages': braking,
                'Accélérations': accel,
                'Virages': turns
            }
            
            # Determine severity color
            if severity_level == 'excellent':
                sev_color = self.colors['rag_green']
            elif severity_level == 'good':
                sev_color = self.colors['rag_light_green']
            elif severity_level in ['moderate', 'concerning']:
                sev_color = self.colors['rag_amber']
            else:
                sev_color = self.colors['rag_red']
            
            # Write row
            row_values = [
                week_label,
                total_events,
                braking,
                accel,
                turns,
                f"{events_per_100km:.1f}" if events_per_100km else "0.0",
                severity_french
            ]
            
            for j, value in enumerate(row_values):
                cell = self.ws.cell(row=start_row, column=j+1, value=value)
                cell.font = self.fonts['normal']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
                
                # Color coding
                if j == 2 and isinstance(value, int) and value > 0:  # Braking
                    cell.fill = PatternFill(start_color=self.colors['rag_red'],
                                           end_color=self.colors['rag_red'], fill_type='solid')
                    cell.font = Font(name='Segoe UI', size=10, color=self.colors['white'])
                elif j == 3 and isinstance(value, int) and value > 0:  # Acceleration
                    cell.fill = PatternFill(start_color=self.colors['rag_amber'],
                                           end_color=self.colors['rag_amber'], fill_type='solid')
                elif j == 4 and isinstance(value, int) and value > 0:  # Turns
                    cell.fill = PatternFill(start_color=self.colors['brand_secondary'],
                                           end_color=self.colors['brand_secondary'], fill_type='solid')
                    cell.font = Font(name='Segoe UI', size=10, color=self.colors['white'])
                elif j == 6:  # Severity
                    cell.fill = PatternFill(start_color=sev_color, end_color=sev_color, fill_type='solid')
                    cell.font = Font(name='Segoe UI', size=10, bold=True, color=self.colors['white'])
            
            start_row += 1
        
        # Create chart
        if chart_data:
            # Aggregate data across all weeks
            aggregated_data = self._aggregate_chart_data(chart_data)
            
            # Define color mapping
            colors_dict = {
                'Freinages': self.colors['rag_red'],
                'Accélérations': self.colors['rag_amber'],
                'Virages': self.colors['brand_secondary']
            }
            
            # Create single aggregated pie chart
            self._create_pie_chart(
                title=self.FRENCH_STRINGS['harsh_events_chart_title'],
                data_dict=aggregated_data,
                colors_dict=colors_dict,
                position=f'H{header_row}',
                chart_style='doughnut'
            )
        
        return start_row + 2

    # ========================================================================
    # IMPROVEMENT RECOMMENDATIONS
    # ========================================================================
    
    def _create_recommendations(self, all_data, start_row):
        """Generate improvement recommendations based on performance"""
        
        aggregated = all_data.get('aggregated_metrics', {})
        combined = aggregated.get('combined_metrics', {})
        
        # Title
        self.ws[f'A{start_row}'] = self.FRENCH_STRINGS['improvement_title']
        self.ws[f'A{start_row}'].font = self.fonts['section_header']
        start_row += 2
        
        recommendations = []
        
        # Speed compliance
        speed_compliance = combined.get('avg_speed_compliance', 100)
        if speed_compliance < 80:
            recommendations.append({
                'category': self.FRENCH_STRINGS['speed_management'],
                'text': self.FRENCH_STRINGS['speed_recommendation'],
                'priority': self.FRENCH_STRINGS['high_priority'] if speed_compliance < 60 else self.FRENCH_STRINGS['medium_priority']
            })
        
        # Roundabout performance
        roundabout_perf = combined.get('avg_roundabout_performance', 100)
        if roundabout_perf < 70:
            recommendations.append({
                'category': self.FRENCH_STRINGS['roundabout_technique'],
                'text': self.FRENCH_STRINGS['roundabout_recommendation'],
                'priority': self.FRENCH_STRINGS['high_priority'] if roundabout_perf < 50 else self.FRENCH_STRINGS['medium_priority']
            })
        
        # Stop sign compliance
        stop_compliance = combined.get('avg_stop_sign_compliance', 100)
        if stop_compliance < 60:
            recommendations.append({
                'category': self.FRENCH_STRINGS['stop_sign_compliance'],
                'text': self.FRENCH_STRINGS['stop_recommendation'],
                'priority': self.FRENCH_STRINGS['high_priority']
            })
        
        # Traffic lights
        traffic_light_compliance = combined.get('avg_traffic_light_compliance', 100)
        if traffic_light_compliance < 70:
            recommendations.append({
                'category': self.FRENCH_STRINGS['traffic_light_compliance'],
                'text': self.FRENCH_STRINGS['traffic_light_recommendation'],
                'priority': self.FRENCH_STRINGS['medium_priority']
            })
        
        # Harsh events
        harsh_events_rate = combined.get('avg_harsh_events_per_100km', 0)
        if harsh_events_rate > 5:
            recommendations.append({
                'category': self.FRENCH_STRINGS['harsh_driving'],
                'text': self.FRENCH_STRINGS['harsh_recommendation'],
                'priority': self.FRENCH_STRINGS['high_priority'] if harsh_events_rate > 10 else self.FRENCH_STRINGS['medium_priority']
            })
        
        # Total violations
        total_violations = aggregated.get('total_violations', 0)
        if total_violations > 10:
            recommendations.append({
                'category': self.FRENCH_STRINGS['overall_awareness'],
                'text': self.FRENCH_STRINGS['awareness_recommendation'],
                'priority': self.FRENCH_STRINGS['high_priority']
            })
        
        # Default if excellent
        if not recommendations:
            recommendations.append({
                'category': self.FRENCH_STRINGS['excellent_performance'],
                'text': self.FRENCH_STRINGS['excellent_recommendation'],
                'priority': self.FRENCH_STRINGS['low_priority']
            })
        
        # Write recommendations
        for i, rec in enumerate(recommendations[:5]):
            # Category
            self.ws[f'A{start_row}'] = f"{i+1}. {rec['category']}"
            self.ws[f'A{start_row}'].font = Font(name='Segoe UI', size=11, bold=True, color=self.colors['brand_primary'])
            
            # Text
            self.ws[f'B{start_row}'] = rec['text']
            self.ws[f'B{start_row}'].font = self.fonts['normal']
            self.ws.merge_cells(f'B{start_row}:G{start_row}')
            
            # Priority
            priority = rec['priority']
            if 'Élevée' in priority:
                priority_color = self.colors['rag_red']
            elif 'Moyenne' in priority:
                priority_color = self.colors['rag_amber']
            else:
                priority_color = self.colors['rag_green']
            
            self.ws[f'H{start_row}'] = priority
            self.ws[f'H{start_row}'].font = Font(name='Segoe UI', size=10, bold=True, color=priority_color)
            
            start_row += 1
        
        return start_row + 2

    # ========================================================================
    # FOOTER
    # ========================================================================
    
    def _create_footer(self, start_row):
        """Create professional footer"""
        
        self.ws.merge_cells(f'A{start_row}:L{start_row}')
        self.ws[f'A{start_row}'] = self.FRENCH_STRINGS['footer_text']
        self.ws[f'A{start_row}'].font = Font(name='Segoe UI', size=9, italic=True, color=self.colors['gray_medium'])
        self.ws[f'A{start_row}'].alignment = Alignment(horizontal='center')
        
        self.ws.merge_cells(f'A{start_row+1}:L{start_row+1}')
        self.ws[f'A{start_row+1}'] = f"{self.FRENCH_STRINGS['generated_text']}{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        self.ws[f'A{start_row+1}'].font = Font(name='Segoe UI', size=9, italic=True, color=self.colors['gray_medium'])
        self.ws[f'A{start_row+1}'].alignment = Alignment(horizontal='center')
        
        return start_row + 2

    # ========================================================================
    # MAIN REPORT GENERATION
    # ========================================================================
    
    def generate_report(self, all_data, driver_name=None, output_filename="FeelGood_Report.xlsx"):
        """
        Generate complete FeelGood driving report.
        
        Args:
            all_data: Complete data from UnifiedReportGenerator.extract_all_data()
            driver_name: Optional driver name override
            output_filename: Output file path
        
        Returns:
            str: Path to generated report file
        """
        
        # Override driver name if provided
        if driver_name:
            all_data['metadata']['driver_name'] = driver_name
        
        print(f"Generating FeelGood Report...")
        
        try:
            # Build report sections
            current_row = self._create_header(all_data)
            current_row += 1
            
            current_row = self._create_trip_summary_table(all_data, current_row)
            current_row += 1
            
            current_row = self._create_kpi_dashboard(all_data, current_row)
            current_row += 1
            
            current_row = self._create_top_violations_table(all_data, current_row)
            current_row += 1
            
            current_row = self._create_speed_analysis_table(all_data, current_row)
            current_row += 1
            
            current_row = self._create_roundabout_table(all_data, current_row)
            current_row += 1
            
            current_row = self._create_stop_sign_table(all_data, current_row)
            current_row += 1
            
            current_row = self._create_traffic_light_table(all_data, current_row)
            current_row += 1
            
            current_row = self._create_harsh_events_table(all_data, current_row)
            current_row += 1
            
            current_row = self._create_recommendations(all_data, current_row)
            current_row += 1
            
            self._create_footer(current_row)
            
            # Save workbook
            self.wb.save(output_filename)
            print(f"✅ Report saved: {output_filename}")
            return output_filename
            
        except PermissionError:
            # File locked - try alternative name
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            alt_filename = output_filename.replace('.xlsx', f'_{timestamp}.xlsx')
            self.wb.save(alt_filename)
            print(f"✅ Report saved (alternative): {alt_filename}")
            return alt_filename
            
        except Exception as e:
            print(f"❌ Error generating report: {e}")
            import traceback
            traceback.print_exc()
            return None
# ==========================================================
# ==========================================================
# ==========================================================

def analyze_gps_data(csv_files_or_pattern, driver_name=None, output_prefix="gps_analysis", generate_excel=True):
    """Fixed main analyzer function"""
    print("Starting analyze_gps_data()...")
    
    analyzer = UnifiedRoadContextGPSAnalyzer()
    globals()["analyzer"] = analyzer 

    result = analyzer.analyze_with_road_context(csv_files_or_pattern, driver_name, output_prefix)
    # Create and run your analyzer

    # After GPS processing is complete, run the test
    globals()["drivenIds"] = analyzer.processor.test_feature_filtering_export()

    # Optionally specify custom filename
    extracted_data = analyzer.reporter.extract_all_data()

    
    globals()["result_from_gps_analysis"] = result 
    #print("        --> result_from_gps_analysis variable in global env")
    
    globals()["extracted_data_new"] = extracted_data 

    # Generate maps using the analyzer's data
    plot_results = generate_plots_main(extracted_data=extracted_data, output_prefix= "Plot")
    
    # Generate verification Excel
    print("Generating verification Excel...")
    verification_result = None
    try:
        verification_generator = MapDataVerificationExcelGenerator()
        verification_result = verification_generator.generate_verification_excel_from_analyzer(
            analyzer=analyzer, 
            output_prefix="verification_sheet"
        )
        if verification_result:
            print(f"Verification Excel: {verification_result['verification_file']}")
    except Exception as e:
        print(f"Error generating verification Excel: {str(e)}")
        verification_result = None
    
    # Generate FeelGood report with proper data extraction
    print("Generating FeelGood report...")
    feelgood_report_file = None
    try:
        # Create report generator and generate report directly
        report_generator = FeelGoodDrivingReportGenerator()
        feelgood_report_file = report_generator.generate_report(
            all_data=extracted_data,
            driver_name=driver_name,
            output_filename=f"{output_prefix}_feelgood_report.xlsx"
        )
        
        if feelgood_report_file:
            print(f"FeelGood report generated: {feelgood_report_file}")
        else:
            print("Failed to generate FeelGood report")
            
    except Exception as e:
        print(f"Error generating report: {str(e)}")
        
        traceback.print_exc()
        feelgood_report_file = None
    
    return {
        "analysis_result": result,
        "plot_results": plot_results,
        "verification_result": verification_result,
        "feelgood_report": feelgood_report_file
    }


def main():
    print("Starting GPS Analysis Pipeline...")
    
    # Get the directory where the script is running
    try:
        if getattr(sys, 'frozen', False):  # Running as .exe
            exe_dir = os.path.dirname(sys.executable)
        else:  # Running as .py script
            exe_dir = os.path.dirname(os.path.realpath(__file__))
    except (NameError, AttributeError):
        exe_dir = os.getcwd()
    
    # Create input and output folders next to exe
    input_folder = os.path.join(exe_dir, "input")
    output_folder = os.path.join(exe_dir, "output")

    print(f"DEBUG: exe_dir = {exe_dir}")
    print(f"DEBUG: input_folder = {input_folder}")
    print(f"DEBUG: output_folder = {output_folder}")
    
    # Create folders if they don't exist
    if not os.path.exists(input_folder):
        print(f"Creating input folder: {input_folder}")
        os.makedirs(input_folder)
        print(f"Please place your CSV files in the 'input' folder and run again.")
        # input("Press Enter to exit...")  # Fixed commented input
        return
    
    if not os.path.exists(output_folder):
        print(f"Creating output folder: {output_folder}")
        os.makedirs(output_folder)
    
    # Find all CSV and Excel files in input folder
    csv_files = glob.glob(os.path.join(input_folder, "*.csv"))
    excel_files = glob.glob(os.path.join(input_folder, "*.xlsx")) + glob.glob(os.path.join(input_folder, "*.xls"))
    all_files = csv_files + excel_files
    
    # Set input data BEFORE the check
    files = all_files
    input_data = files[0] if len(files) == 1 else files
    
    if not all_files:
        print(f"No CSV or Excel files found in '{input_folder}' folder.")
        print(f"Please place your CSV or Excel files in the 'input' folder and run again.")
        return
        
        
        print(f"Found {len(all_files)} file(s):")
        print(f"  - CSV files: {len(csv_files)}")
        print(f"  - Excel files: {len(excel_files)}")
        for file in all_files:
            file_type = "Excel" if file.lower().endswith(('.xlsx', '.xls')) else "CSV"
            print(f"  - {os.path.basename(file)} ({file_type})")
        

    
    # Get driver name from user - FIXED SYNTAX
    driver_name = input('Enter driver name (or press Enter for default): ').strip()  # Fixed syntax
    if not driver_name:
        driver_name = "Driver"
    
    force_excel = True 
    
    # Generate output prefix with output folder path
    if isinstance(input_data, list):
        base_filename = os.path.splitext(os.path.basename(input_data[0]))[0]
    else:
        base_filename = os.path.splitext(os.path.basename(input_data.replace('*', 'multi')))[0]
    
    output_prefix = os.path.join(output_folder, base_filename + "_analysis")
    
    # Run complete analysis pipeline
    try:
        print("Starting GPS Analysis Pipeline...")
        result = analyze_gps_data(
            csv_files_or_pattern=input_data,
            driver_name=driver_name,
            output_prefix=output_prefix,
            generate_excel=force_excel
        )
        print("Analysis completed successfully!")
        globals()["last_result"] = result  # keep in memory

        
        # Print summary of generated files
        print("\nGenerated Files:")
        if result.get("verification_result"):
            verification_file = result['verification_result'].get('verification_file')
            if verification_file:
                print(f"- Verification Excel: {verification_file}")
        
        if result.get("feelgood_report"):
            print(f"- FeelGood Report: {result['feelgood_report']}")
        
        if result.get("plot_results"):
            print(f"- Maps generated: {len(result['plot_results'])} files")
            for map_file in result['plot_results']:
                print(f"  - {map_file}")
        
        print(f"\nAll results saved in: {output_folder}")
        print("Analysis complete!")
        
    except Exception as e:
        print(f"Error during analysis: {str(e)}")
        
        traceback.print_exc()
        print("\nPlease contact support with the above error message.")
        # input("Press Enter to exit...")  # Fixed commented input
        return 1
    
    # Keep window open for user to see results
    print("\n" + "="*50)
    print("ANALYSIS COMPLETE!")
    print(f"Results saved in: {output_folder}")
    print("="*50)
    # input("Press Enter to close this window...")  # Fixed commented input
    
    return 0


if __name__ == "__main__":
    main()

